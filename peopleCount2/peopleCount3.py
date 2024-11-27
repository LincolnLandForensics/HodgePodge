#!/usr/bin/env python3
# coding: utf-8
'''
This is a clone of
https://github.com/ndaidong/people-detecting.git
it counts the approximate number of people in a folder of videos
and outputs it to .xlsx
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import cv2
import sys
import shutil
import datetime
import openpyxl
import argparse  # for menu system
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from moviepy.editor import VideoFileClip    # pip install moviepy
from imutils.object_detection import non_max_suppression

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'ndaidong'
coauthor = 'LincolnLandForensics'
description = "counts the approximate number of people in a folder of videos"
version = '1.0.3'

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global row
    row = 0  # defines arguments
    global data
    data = []

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-c', '--count', help='count people', required=False, action='store_true')
    parser.add_argument('-H', '--Help', help='help', required=False, action='store_true')

    args = parser.parse_args()

    global input_folder
    global outuput_xlsx

    if args.Help:
        usage()
        exit()
        
    if not args.input: 
        input_folder = "videos"        
    else:
        input_folder = args.input

    if not os.path.exists(input_folder):
        create_folder(input_folder)
        print(f'fill {input_folder} with videos')
        exit()

    # Check if the folder is empty
    elif not os.listdir(input_folder):
        print(f"\n\n\n\tNo videos found in '{input_folder}'.")
        exit()
    else:
        file_count = ''
        # Count the number of files in the folder
        file_count = len([name for name in os.listdir(input_folder) if os.path.isfile(os.path.join(input_folder, name))]) #task
        print(f'\n\tThere are {file_count} files in the {input_folder} folder\n')
   
    if not args.output: 
        outuput_xlsx = "videos.xlsx"        
    else:
        outuput_xlsx = args.output

    if args.count:

        folder_exists = os.path.exists(input_folder)
        if folder_exists == True:
            # data = []
            count_people()
            write_xlsx(data)  # disable xlsxwriter

            print(f'Writing to {outuput_xlsx} ')
        else:
            print(f'{input_xlsx} does not exist')
            exit()
 
    else:
        usage()
    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def count_people():
    hog = cv2.HOGDescriptor()
    hog.setSVMDetector(cv2.HOGDescriptor_getDefaultPeopleDetector())

    font = cv2.FONT_HERSHEY_SIMPLEX

    dest_folder = "nobody"

    # Iterate through all vidoes in the "videos" folder

    for file_name in os.listdir(input_folder):
        row_data = {}
        (max_people_detected, file_size, creation_time, access_time, modified_time) = ('', '', '', '', '')
        (video) = ('')

        video = os.path.join(input_folder, file_name)

        file_size = os.path.getsize(video)

        # get creation, access, modified times of the .plist
        file_info = os.stat(video)
        
        (creation_time, access_time, modified_time) = convert_to_iso(file_info)
        # (codec, duration, metadata) = get_video_info(input_folder, file_name)   # test
        (codec, duration, metadata, audio_found, video_duration, video_fps, video_size) = get_video_info(input_folder, file_name)   # test



        cap = cv2.VideoCapture(video)

        w = cap.get(3)
        h = cap.get(4)
        mx = int(w - 400)
        my = int(h - 24)

        results = []

        people_detected = 0
        max_people_detected = 0
            
        while(cap.isOpened()):
            ret, frame = cap.read()

            if ret is False:
                break
            
            k = cv2.waitKey(30) 
            if k == 27:
                break

            rects, weights = hog.detectMultiScale(
                frame,
                winStride=(4, 4), 
                padding=(8, 8), 
                scale=1.05
            )
            for (x, y, w, h) in rects:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)

            rects = np.array([[x, y, x + w, y + h] for (x, y, w, h) in rects])
            pick = non_max_suppression(rects, probs=None, overlapThresh=0.65)

            people_detected = 0
            
            for (xA, yA, xB, yB) in pick:
                people_detected += 1
                cv2.rectangle(frame, (xA, yA), (xB, yB), (0, 255, 0), 2)

            if people_detected > max_people_detected:
                max_people_detected = people_detected

            text = 'People detected: ' + str(people_detected)
            cv2.putText(
                frame, 
                text, 
                (mx, my), 
                font, 
                1, 
                (255, 255, 255), 
                1, 
                cv2.LINE_AA
            )
            cv2.imshow('Frame', frame)

        print(f'    {file_name}  {max_people_detected}')

        row_data["file_name"] = file_name
        row_data["max_people_detected"] = max_people_detected
        row_data["file_size"] = file_size
        row_data["creation_time"] = creation_time
        row_data["access_time"] = access_time
        row_data["modified_time"] = modified_time
        row_data["duration"] = duration
        row_data["audio_found"] = audio_found
        row_data["video_fps"] = video_fps
        row_data["video_size"] = video_size
        row_data["metadata"] = metadata
        row_data["codec"] = codec

        data.append(row_data)

        cap.release()
        cv2.destroyAllWindows()

        if max_people_detected == 0:
            move_file(video, dest_folder)


def convert_to_iso(file_info):
    # Convert timestamps to datetime objects
    ctime = datetime.datetime.fromtimestamp(file_info.st_ctime)
    atime = datetime.datetime.fromtimestamp(file_info.st_atime)
    mtime = datetime.datetime.fromtimestamp(file_info.st_mtime)

    # Format datetime objects to ISO standard format without microseconds
    iso_atime = atime.strftime('%Y-%m-%d %H:%M:%S')
    iso_mtime = mtime.strftime('%Y-%m-%d %H:%M:%S')
    iso_ctime = ctime.strftime('%Y-%m-%d %H:%M:%S')

    return iso_ctime, iso_atime, iso_mtime

def create_folder(folder_name):
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"Folder '{folder_name}' created successfully.")
    # else:
        # print(f"Folder '{folder_name}' already exists.")


def get_video_info(input_folder, file_name):
    # print(f'input_folder = {input_folder}   file_name   = {file_name}') # temp
    (codec, duration, metadata) = ('', '', '')
    (audio_found, video_duration, video_fps, video_size, duration) = ('', '', '', '', '')

    try:
        # fullpath = (f'{input_folder}/{file_name}')
        fullpath = (os.path.join(input_folder, file_name))
        
        
        # Load the video file
        clip = VideoFileClip(fullpath)
        # print(f'clip = {clip}') # temp
        # Extract duration
        duration = clip.duration
        # print(f'duration = {duration}') # temp

        # Extract codec
        # codec = clip.reader.infos['codec_name']   #task

        # Extract duration
        duration = clip.duration
        
        # Extract metadata
        # metadata = clip.reader.infos    #task
        
        # Extract audio information
        audio_found = clip.audio is not None
        
        # Extract frame rate (FPS)
        video_fps = clip.fps
        
        # Extract video size (width x height)
        # video_size = clip.size  #task
        
        # Close the clip
        clip.close()

    except Exception as e:
        print("Error:", e)
        codec = ('Non-Video-File')  #task
    if audio_found == 0:
        audio_found = "No"
    elif audio_found == 1:
        audio_found = "Yes"        
        
    return codec, duration, metadata, audio_found, video_duration, video_fps, video_size
    
def move_file(file_path, dest_folder):
    
    create_folder(dest_folder)
    
    if os.path.isfile(file_path):
        # file_size = os.path.getsize(file_path)
        if 1==1:
        # if file_size < 20 * 1024: # 20 KB
            try:
                shutil.move(file_path, dest_folder)
            except Exception as e:
                print("Error:", e)            
            print(f"File {file_path} moved to {dest_folder} because nobody was in it.")
        else:
            print(f"File {file_path} is too big to move")
    else:
        print(f"{file_path} is not a valid file")




def write_xlsx(data):
    '''
    The write_xlsx() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''

    # global workbook
    workbook = Workbook()
    # global worksheet
    worksheet = workbook.active

    worksheet.title = 'Videos'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = [
        "file_name", "max_people_detected", "file_size", "creation_time", "access_time"
        , "modified_time", "duration", "audio_found", "video_fps", "video_size", "metadata", "codec"
    ]

    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        # cell.font = header_format  # Apply the header format
        if col_index in [0, 1, 2, 3, 4, 5, 6, 7, 8]:  # Indices of columns A, C, D, E, F, G, H
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange header
            cell.fill = fill

    # Excel column width
    worksheet.column_dimensions['A'].width = 25 # video
    worksheet.column_dimensions['B'].width = 18  # People (approximate)
    worksheet.column_dimensions['C'].width = 9 # FileSize (bytes)
    worksheet.column_dimensions['D'].width = 18 # creation_time
    worksheet.column_dimensions['E'].width = 18 # access_time
    worksheet.column_dimensions['F'].width = 18 # modified_time
    worksheet.column_dimensions['G'].width = 7 # duration
    worksheet.column_dimensions['H'].width = 11 # audio_found
    worksheet.column_dimensions['I'].width = 9 # video_fps
    worksheet.column_dimensions['J'].width = 11 # video_size
    worksheet.column_dimensions['K'].width = 9 # metadata
    worksheet.column_dimensions['L'].width = 6 # codec

    for row_index, row_data in enumerate(data):
        # print(f'Processing row: {row_data}')  # Debugging output
        # print("row_index = %s" %(row_index))
        # print(f'row = {row_index+1}')        

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"Error printing line: {str(e)}")
    
    workbook.save(outuput_xlsx)


def usage():
    '''
    working examples of syntax 
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {description}')
    print(f'{file} Version: {version} by {author} and modified by {coauthor}')
    print(f'\ninsert your videos into videos folder')
    print(f'\nExample:')
    print(f'    {file} -c')
    print(f'    {file} -c -I videos videos.xlsx') 
 
                
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

1.0.1 - if video folder doesn't exist, create one

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
if it's not a video file mark codec as non-video-file

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""



"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

