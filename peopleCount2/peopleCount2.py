#!/usr/bin/env python3
# coding: utf-8

"""
This script counts the approximate number of people in a folder of videos
and outputs the results to an Excel file. Can specify how many to do in parallel (default is 4).
Can show video (-v) or not.
"""

import os
import sys
import cv2  # pip install opencv-python
import shutil
import datetime
import argparse  # for menu system
import numpy as np
from functools import cache
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from imutils.object_detection import non_max_suppression  # pip install imutils

author = 'ndaidong'
coauthor = 'LincolnLandForensics'
description = "counts the approximate number of people in a folder of videos"
version = '1.0.7'

@cache
def main():
    global data
    data = []

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='Input folder with videos', required=False)
    parser.add_argument('-O', '--output', help='Output Excel file', required=False)
    parser.add_argument('-c', '--count', help='Count people', required=False, action='store_true')
    parser.add_argument('-H', '--Help', help='Display help', required=False, action='store_true')
    parser.add_argument('-v', '--video', help='Display video while processing', required=False, action='store_true')
    parser.add_argument('-p', '--processes', help='Number of parallel processes', type=int, default=4)

    args = parser.parse_args()

    global input_folder
    global output_xlsx
    global display_video
    global num_processes

    if args.Help:
        usage()
        sys.exit()

    input_folder = args.input if args.input else "videos"
    output_xlsx = args.output if args.output else "videos.xlsx"
    display_video = args.video
    num_processes = args.processes

    if not os.path.exists(input_folder):
        create_folder(input_folder)
        print(f'Fill {input_folder} with videos')
        sys.exit()

    if not os.listdir(input_folder):
        print(f"\n\n\n\tNo videos found in '{input_folder}'.")
        sys.exit()

    file_count = len([name for name in os.listdir(input_folder) if os.path.isfile(os.path.join(input_folder, name))])
    print(f'\n\tThere are {file_count} files in the {input_folder} folder\n')

    if args.count:
        count_people_parallel()
        write_xlsx(data)
        print(f'Writing to {output_xlsx}')
    else:
        usage()

    input("Press Enter to continue...")

def count_people_parallel():
    videos = [os.path.join(input_folder, file_name) for file_name in os.listdir(input_folder)]
    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        futures = {executor.submit(process_video, video, display_video): video for video in videos}
        for future in as_completed(futures):
            video = futures[future]
            try:
                result = future.result()
                if result:
                    data.append(result)
            except Exception as e:
                print(f'Error processing {video}: {e}')

def process_video(video, display_video):
    hog = cv2.HOGDescriptor()
    hog.setSVMDetector(cv2.HOGDescriptor_getDefaultPeopleDetector())

    row_data = {}

    file_info = os.stat(video)
    creation_time, access_time, modified_time = convert_to_iso(file_info)

    cap = cv2.VideoCapture(video)
    if not cap.isOpened():
        print(f"Failed to open video file {video}")
        return None

    ret, prev_frame = cap.read()
    if not ret:
        print(f"Failed to read first frame of {video}")
        return None

    w, h = cap.get(3), cap.get(4)
    mx, my = int(w - 400), int(h - 24)

    max_people_detected = 0
    frame_count = 0

    prev_gray = cv2.cvtColor(prev_frame, cv2.COLOR_BGR2GRAY)

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        frame_delta = cv2.absdiff(prev_gray, gray)
        thresh = cv2.threshold(frame_delta, 25, 255, cv2.THRESH_BINARY)[1]

        if cv2.countNonZero(thresh) > 500:  # Adjust this threshold based on your requirements
            rects, _ = hog.detectMultiScale(frame, winStride=(4, 4), padding=(8, 8), scale=1.05)
            rects = np.array([[x, y, x + w, y + h] for (x, y, w, h) in rects])
            pick = non_max_suppression(rects, probs=None, overlapThresh=0.65)

            people_detected = len(pick)
            max_people_detected = max(max_people_detected, people_detected)

            if display_video:
                for (xA, yA, xB, yB) in pick:
                    cv2.rectangle(frame, (xA, yA), (xB, yB), (0, 255, 0), 2)

                text = 'People detected: ' + str(people_detected)
                cv2.putText(frame, text, (mx, my), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 1, cv2.LINE_AA)
                cv2.imshow('Frame', frame)

                if cv2.waitKey(1) == 27:  # Press 'Esc' to exit the video display
                    break

        prev_gray = gray
        frame_count += 1

    print(f'    {video}  {max_people_detected}')

    row_data.update({
        "file_name": os.path.basename(video),
        "max_people_detected": max_people_detected,
        "file_size": os.path.getsize(video),
        "creation_time": creation_time,
        "access_time": access_time,
        "modified_time": modified_time
    })

    cap.release()
    if display_video:
        cv2.destroyAllWindows()

    if max_people_detected == 0:
        move_file(video, "nobody")

    return row_data

def convert_to_iso(file_info):
    iso_ctime = datetime.datetime.fromtimestamp(file_info.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
    iso_atime = datetime.datetime.fromtimestamp(file_info.st_atime).strftime('%Y-%m-%d %H:%M:%S')
    iso_mtime = datetime.datetime.fromtimestamp(file_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
    return iso_ctime, iso_atime, iso_mtime

def create_folder(folder_name):
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"Folder '{folder_name}' created successfully.")

def move_file(file_path, dest_folder):
    create_folder(dest_folder)
    if os.path.isfile(file_path):
        try:
            shutil.move(file_path, dest_folder)
            print(f"File {file_path} moved to {dest_folder} because nobody was in it.")
        except Exception as e:
            print("Error:", e)

def write_xlsx(data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Videos'

    headers = [
        "file_name", "max_people_detected", "file_size", "creation_time", "access_time", "modified_time"
    ]

    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    col_widths = [25, 18, 9, 18, 18, 18]
    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F'], col_widths):
        worksheet.column_dimensions[col].width = width

    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            worksheet.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name)

    workbook.save(output_xlsx)

def usage():
    file = sys.argv[0].split('/')[-1]
    print(f'\nDescription: {description}')
    print(f'{file} Version: {version} by {author} and modified by {coauthor}')
    print(f'\nInsert your videos into the "videos" folder')
    print(f'\nExample:')
    print(f'    {file} -c')
    print(f'    {file} -c -I videos -O videos.xlsx')

if __name__ == '__main__':
    main()
