#!/usr/bin/env python3
import sys
import openpyxl
from bs4 import BeautifulSoup
import requests # pip install requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

author = 'LincolnLandForensics'
description = "parse a sherlock OSINT output"
version = '1.1.4'


'''
https://sherlockproject.xyz/installation
sherlock kevinrose
'''


headers_intel = [
    "query", "ranking", "fullname", "url", "email", "user", "phone",
    "business", "fulladdress", "city", "state", "country", "note", "AKA",
    "DOB", "SEX", "info", "misc", "firstname", "middlename", "lastname",
    "associates", "case", "sosfilenumber", "owner", "president", "sosagent",
    "managers", "Time", "Latitude", "Longitude", "Coordinate",
    "original_file", "Source", "Source file information", "Plate", "VIS", "VIN",
    "VYR", "VMA", "LIC", "LIY", "DLN", "DLS", "content", "referer", "osurl",
    "titleurl", "pagestatus", "ip", "dnsdomain", "Tag", "Icon", "Type"
]

data = []
output_xlsx = "sherlock_output.xlsx"


def parse_sherlock_output(path):
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    user = ''

    for line in lines:
        
        (content, referer, osurl, titleurl, pagestatus) = ('', '', '', '', '')
        fullname, ranking, url, query = '', '', '', ''
        if "Checking username " in line:
            user = line.split("Checking username ")[1].replace(" on:", "").strip()

            print(f'user = {user}') #    test

        elif ": " in line:
            line = line.replace('[+] ', '7 - ')
            
            if ":" in line:
                parts = line.split(': ')

                try:
                    ranking = parts[0].strip()
                    url = parts[1].strip()
                except:
                    pass
                # Print to screen
                
                (content, referer, osurl, titleurl, pagestatus) = request_url(url)

                if titleurl != '':
                    ranking = ranking.replace('7 - ', '5 - ')
                    if " " in titleurl and user not in titleurl:
                        fullname = titleurl
                        ranking = ranking.replace('5 - ', '4 - ')
                elif 'Success' in pagestatus:
                    ranking = ranking.replace('7 - ', '6 - ')
                elif 'Fail' in pagestatus:
                    ranking = ranking.replace('7 - ', '9 - ')                
                print(f"{user}\t{ranking}\t{url}\t{user}")

                # Build DFIR row
                row_data = {h: "" for h in headers_intel}
                row_data["query"] = user
                row_data["ranking"] = ranking
                row_data["fullname"] = fullname
                row_data["url"] = url
                row_data["user"] = user

                # row_data["content"] = content
                # row_data["referer"] = referer
                # row_data["osurl"] = osurl
                row_data["titleurl"] = titleurl
                row_data["pagestatus"] = pagestatus                

                data.append(row_data)

    write_intel(data)



def request_url(url):
    
    fake_referer = 'https://www.google.com/'
    headers_url = {'Referer': fake_referer}

    
    (content, referer, osurl, titleurl, pagestatus)= ('blank', '', '', '', '')
    (response) = ('')
    
    if url.lower().startswith('http'):
        blah = ''
    else:
        url = ("https://" +url)

    try:
        response = requests.get(url, verify=False, headers=headers_url)        
        response.raise_for_status()
        pagestatus  = response.status_code
        content = response.content.decode()
        content = BeautifulSoup(content, 'html.parser')
        


    except requests.exceptions.RequestException as e:
        # print(f"Could not fetch URL {url}: {str(e)}")
        # raise requests.exceptions.RequestException(str(e))
        (pagestatus) = ('Fail')

# osurl
    try:
        osurl = response.headers['Server']
    except:
    # except KeyError:
        pass

# titleurl
    try:
        titleurl = titleurl_get(content)
        # titleurl = titleurl_og(content)
    except KeyError:
        titleurl = ''


#pagestatus
    
    if str(pagestatus).startswith('2') :    
        pagestatus = (f'Success')
    elif str(pagestatus).startswith('3') :    
        pagestatus = (f'Redirect - {pagestatus}')
    elif str(pagestatus).startswith('4') :    
        pagestatus = (f'Fail - {pagestatus}')
    elif str(pagestatus).startswith('5') :    
        pagestatus = (f'Fail - {pagestatus}')
    elif str(pagestatus).startswith('1') :    
        pagestatus = (f'Info - {pagestatus}')
    try:
        pagestatus = str(pagestatus).strip()    
    except:pass
    
    return (content, referer, osurl, titleurl, pagestatus)


def titleurl_get(content):

    titleurl = ''
    try:
        # soup = BeautifulSoup(html, 'html.parser')
        title_tag = content.find('title')
        if title_tag is not None:
            title = title_tag.text.strip()
            # The full name is often included in parentheses after the username
            # Example: "Tom Anderson (myspacetom)"
            # We'll split the title at the first occurrence of '(' to extract the full name
            parts = title.split(' (', 1)
            if len(parts) > 1:
                titleurl = parts[0]
    except Exception as e:
        # print(f'Error parsing title: {str(e)}')
        pass


    return titleurl

def write_intel(data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Intel'
    worksheet.freeze_panes = 'B2'

    # Write headers
    for col_index, header in enumerate(headers_intel):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header

        # Orange highlight
        if col_index in [3, 4, 5, 6, 49, 50]:
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Yellow highlight
        elif col_index in [
            7, 8, 13, 14, 15, 29, 30, 35, 36, 37, 38, 39, 40, 41, 42, 43
        ]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Column widths
    for i in range(1, len(headers_intel) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    # Write data rows
    for row_index, row_data in enumerate(data, start=2):
        for col_index, col_name in enumerate(headers_intel, start=1):
            worksheet.cell(row=row_index, column=col_index).value = row_data.get(col_name, "")

    # ColorCode sheet
    color_sheet = workbook.create_sheet("ColorCode")
    color_sheet.freeze_panes = "B2"
    color_sheet.append(["Color", "Description"])
    color_sheet.append(["Red", "Bad Intel or dead link"])
    color_sheet.append(["Orange", "Research"])
    color_sheet.append(["Green", "Good Intel"])
    color_sheet.append(["Yellow", "Highlighted"])

    # Log sheet
    log_sheet = workbook.create_sheet("Log")
    log_sheet.freeze_panes = "B2"
    log_sheet.append(["Date", "Subject", "Requesting Agency", "Requesting Agent",
                      "Case", "Summary of Findings", "Notes"])

    workbook.save(output_xlsx)
    print(f"\nExcel written: {output_xlsx}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python sherlock_parser.py <file>")
        sys.exit(1)

    parse_sherlock_output(sys.argv[1])
