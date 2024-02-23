 #!/usr/bin/env python3
# coding: utf-8
'''
read GPS coordinates and convert them to addresses
or
read addresses and convert them to coordinates
and
read GPS coordinates (xlsx) and convert them to KML
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import io
import requests
from openpyxl.drawing.image import Image

import os
import re
import sys
import time
import openpyxl
import simplekml    # pip install simplekml
import geohash2    # pip install geohash2
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import argparse  # for menu system
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from geopy.geocoders import Nominatim   # pip install geopy
geolocator = Nominatim(user_agent="GeoTraxer")

# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = "convert GPS coordinates to addresses or visa versa & create a KML file"
version = '1.2.0'

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-c', '--create', help='create blank input sheet', required=False, action='store_true')
    parser.add_argument('-i', '--intel', help='convert intel sheet to locations', required=False, action='store_true')
    parser.add_argument('-k', '--kml', help='xlsx to kml with nothing else', required=False, action='store_true')
    parser.add_argument('-r', '--read', help='read xlsx', required=False, action='store_true')
    args = parser.parse_args()

    # global input_xlsx
    global outuput_xlsx
    global output_kml
    output_kml = 'gps.kml'

    if not args.input: 
        input_xlsx = "locations.xlsx"        
    else:
        input_xlsx = args.input

        
    if not args.output: 
        outuput_xlsx = "locations2addresses_.xlsx"        
    else:
        outuput_xlsx = args.output

    if args.create:
        data = []
        print(f'{color_green}Writing to {outuput_xlsx} {color_reset}')
        write_locations(data)

    elif args.kml:
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            print(f'{color_green}Reading {input_xlsx} {color_reset}')
            
            # data = read_xlsx(input_xlsx)
            data = read_locations(input_xlsx)

            # create kml file
            write_kml(data)
            
            
            # workbook.close()
            # print(f'{color_green}Writing to {outuput_xlsx} {color_reset}')
            print(f'{color_blue}Writing to gps.kml {color_reset}')
            print(f'''\n\n{color_yellow}
            visit https://earth.google.com/
            <file><Import KML> select gps.kml <open>
            {color_reset}\n''')
        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()
    elif args.intel:
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            print(f'{color_green}Reading {input_xlsx} {color_reset}')

            data = read_intel(input_xlsx)
            # write_kml(data)
            write_locations(data)
            print(f'{color_green}Writing to {outuput_xlsx} {color_reset}')
        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()
            
    elif args.read:
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            print(f'{color_green}Reading {input_xlsx} {color_reset}')
            
            # data = read_xlsx(input_xlsx)
            data = read_locations(input_xlsx)
            data = read_gps(data)
            write_locations(data)
            write_kml(data)

            workbook.close()
            print(f'{color_green}Writing to {outuput_xlsx} {color_reset}')
            print(f'''\n\n{color_yellow}
            visit https://earth.google.com/
            <file><Import KML> select gps.kml <open>
            {color_reset}\n''')
        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()

    else:
        usage()
    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>


def convert_timestamp(timestamp, time_orig, timezone):
    if timezone is None:
        timezone = ''
    if time_orig is None:
        time_orig = ''

    timestamp = str(timestamp)

    if time_orig == "":
        time_orig = timestamp

    # timestamp = timestamp.replace(' at ', ' ')
    if "(" in timestamp:
        timestamp = timestamp.split('(')
        timezone = timestamp[1].replace(")", '')
        timestamp = timestamp[0]
    elif " CDT" in timestamp:
        timezone = "CDT"
        timestamp = timestamp.replace(" CDT", "")
    elif " CST" in timestamp:
        timezone = "CST"
        timestamp = timestamp.replace(" CST", "")

    formats = [
        "%B %d, %Y, %I:%M:%S %p %Z",    # June 13, 2022, 9:41:33 PM CDT (Flock)
        "%Y:%m:%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",  # timestamps without seconds
        "%m/%d/%Y %H:%M:%S",  # timestamps in military time without seconds
        "%B %d, %Y at %I:%M:%S %p %Z",
        "%B %d, %Y at %I:%M:%S %p",
        "%B %d, %Y %I:%M:%S %p %Z",
        "%B %d, %Y %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p %Z",
        "%m-%d-%y %I:%M:%S %p",
        "%Y-%m-%dT%H:%M:%SZ",  # ISO 8601 format with UTC timezone
        "%Y/%m/%d %H:%M:%S",  # 2022/06/13 21:41:33
        "%d-%m-%Y %I:%M:%S %p",  # 13-06-2022 9:41:33 PM
        "%d/%m/%Y %H:%M:%S",  # 13/06/2022 21:41:33
        "%Y-%m-%d %I:%M:%S %p",  # 2022-06-13 9:41:33 PM
        "%Y%m%d%H%M%S",  # 20220613214133
        "%Y%m%d %H%M%S",  # 20220613 214133
        "%m/%d/%y %H:%M:%S",  # 06/13/22 21:41:33
        "%d-%b-%Y %I:%M:%S %p",  # 13-Jun-2022 9:41:33 PM
        "%d/%b/%Y %H:%M:%S",  # 13/Jun/2022 21:41:33
        "%Y/%b/%d %I:%M:%S %p",  # 2022/Jun/13 9:41:33 PM
        "%d %b %Y %H:%M:%S",  # 13 Jun 2022 21:41:33
        "%A, %B %d, %Y %I:%M:%S %p %Z"  # Monday, June 13, 2022 9:41:33 PM CDT
    ]
 
    for fmt in formats:
        try:
            dt_obj = datetime.strptime(timestamp.strip(), fmt)
            timestamp = dt_obj
            return timestamp, time_orig, timezone
                        
        except ValueError:
            pass

    raise ValueError(f"{time_orig} Timestamp format not recognized")

    
def read_gps(data): 

    """Read data and return as a list of dictionaries.
    It extracts headers from the 
    first row and then iterates through the data rows, creating dictionaries 
    for each row with headers as keys and cell values as values.
    
    """

    for row_index, row_data in enumerate(data):
        print(f'\n{row_index + 2} ______________________\n')
        (street, fulladdress) = ('', '')
        
        (zipcode, business, number, street, city, county) = ('', '', '', '', '', '')
        (state, latitude, longitude, query, coordinate) = ('', '', '', '', '')
        (Index, country, latitude, longitude, PlusCode) = ('', '', '', '', '')
        (county, query, type_data) = ('', '', '')
        (location, skip, address) = ('', '', '')
        (type_data, plate, country_code, state_ftk, city_ftk, hwy) = ('', '', '', '', '', '')
        (direction, name_data, no, account, container, time_local) = ('', '', '', '', '', '') 
        (deleted, service_id, carved, sighting_state, sighting_location, manually_decoded) = ('', '', '', '', '', '')

        fulladdress = row_data.get("fulladdress")
        if fulladdress is None:
            fulladdress = ''  

        name_data = row_data.get("Name")
        latitude = row_data.get("Latitude") # works
        latitude = str(latitude)
        if latitude is None:
            latitude = ''        
        longitude = row_data.get("Longitude") # works
        longitude = str(longitude)
        if longitude is None:
            longitude = ''        
        address = row_data.get("Address") # works
        type_data = row_data.get("Type") # works  
        business = row_data.get("business")
        number = row_data.get("number")
        street = row_data.get("street")
        city = row_data.get("city")
        county = row_data.get("county")
        state = row_data.get("state")
        zipcode = row_data.get("zipcode")
        query = row_data.get("query")
        country = row_data.get("country")
        Subgroup = row_data.get("Subgroup")
        
        coordinate = row_data.get("Coordinate")
        row_data["Index"] = (row_index + 2)
        PlusCode = row_data.get("PlusCode")


# skip lines with fulladdress
        # if fulladdress != '' or fulladdress == 'Soul Buoy':  
        if "," in fulladdress:        
            print(f'skipping fulladdress')    # temp
        # if len(fulladdress) > 2 or fulladdress == 'Soul Buoy':        
            if fulladdress == 'Soul Buoy':
                fulladdress = ''
        # if len(fulladdress) < 2 or fulladdress == 'Soul Buoy':        
            # fulladdress = ''
            skip == 'skip'
# GPS to fulladdress
        elif latitude != '' and isinstance(latitude, str) and len(latitude) > 5:
            print(f'trying to get full address')    # temp
            if latitude != '' and isinstance(latitude, str) and len(latitude) > 3:

                ##  if no fulladdress
                if len(fulladdress) < 2:
                    query = (f'{latitude}, {longitude}') # backwards

                    try:
                        # location = geolocator.reverse((longitude, latitude), language='en')

                        location = geolocator.reverse((latitude, longitude), language='en') #lat/long
                    except Exception as e:
                        print(f"{color_red}Error : {str(e)}{color_reset}") 
                    try:
                        fulladdress = location.address

                    except Exception as e:
                        print(f"{color_red}Error : {str(e)}{color_reset}") 

                    if fulladdress == 'Soul Buoy':        
                        fulladdress = ''
  

                    time.sleep(8)   ## Sleep x seconds

# address to gps / Full address
        elif address != '':
            if len(address) > 3:

                try:
                    location = geolocator.geocode(address)

                    if location:
                        latitude = location.latitude
                        longitude = location.longitude
                        query = (f'{address}')
                    else:
                        print(f"Location not found for address: {address}")

                except Exception as e:
                    print(f"Error: {str(e)}")

            try:
                location = geolocator.reverse((latitude, longitude), language='en')
                fulladdress = location.address

            except Exception as e:
                print(f"{color_red}Error: {str(e)}{color_reset}")

            if fulladdress == 'Soul Buoy':        
                fulladdress = ''
                skip == 'skip'

    # coordinate        
            if latitude != '' and longitude != '':
                coordinate = (f'{latitude},{longitude}')

                time.sleep(8)   # Sleep for x seconds

        elif PlusCode != '':
            print(f'testing PlusCode {PlusCode}')
            
            try:
                decoded_location = geohash2.decode(PlusCode)
                latitude = decoded_location[0]
                longitude = decoded_location[1]

            
                print(f"Coordinates for Geohash '{PlusCode}': {decoded_location[0]}, {decoded_location[1]}")
                
                
                
                # location = geolocator.geocode(PlusCode)
                if decoded_location is not None:
                    # latitude, longitude = location.latitude, location.longitude
                    print(f'PlusCode {PlusCode} = {latitude}, {longitude}')
                    fulladdress = location.address
                else:
                    print(f"Unable to find coordinates for address: {PlusCode}")
            except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} PlusCode = <{PlusCode}>")  

        else:
            print(f'{color_red}none of the above{color_reset}')

# parse fulladdress
        if len(fulladdress) > 5 and skip != 'skip' and "," in fulladdress:
            try:
                address_parts = fulladdress.split(', ')
                if country != '' and len(address_parts) >= 1:
                    country = address_parts[-1]
                if zipcode != '' and len(address_parts) >= 2:
                    zipcode = address_parts[-2]
                if state == '' and len(address_parts) >= 3:
                    state = address_parts[-3]
                if country != '' and len(address_parts) >= 4:
                    county = address_parts[-4]
                if len(address_parts) >= 6:
                    if "Township" in address_parts[-5]:
                        city = address_parts[-6]
                    else:    
                        county = address_parts[-5]
                        city = address_parts[-6]
                
                # if len(address_parts) == 7:
                    # print(f'there are 7 parts') # temp
                if fulladdress.count(',') == 7:    # Check if there are exactly 7 commas
                    # print(f'there are also 7 parts') # temp
                    if 'Township' in address_parts[3]:    # task
                             
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[0]
                        street = address_parts[1]

                    else:
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[1]
                        street = address_parts[2]
    
                elif fulladdress.count(',') == 5:
                    if 'Township' in address_parts[1]:    # task
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[1]
                        street = address_parts[2]

                    else:
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[1]
                        street = address_parts[2]

                if "United States" in country:
                    country = "US" 

                if street.endswith(" Township"):
                    street == ''                

                try:
                    if business == '' and address_parts[1].isdigit():
                        business = address_parts[0]
                except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} Business = <{business}>")  


                try:
                    if business.isdigit():
                        business = ''
                    elif business.endswith(" Street") or business.endswith(" Road") or business.endswith(" Tollway") or business.endswith(" Avenue"): 
                        business = ''
                except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} Business2 = <{business}>")  


                if address_parts[0].isdigit():
                    number = address_parts[0]
                    if street == '':
                        street = address_parts[1]
                if number is None:
                    number = ''
                    
                try:        
                    number = number if number.isdigit() else ''
                except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} number = <{number}>")  


            except Exception as e:
                print(f"{color_red}Error : {str(e)}{color_reset} Business = <{business}> Full address =<{address_parts}>")  

# Icon    
        Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''

        if Icon != "":
            Icon = Icon        
        elif type_data == "Calendar":
            Icon = "Calendar"
        elif type_data == "LPR":
            Icon = "LPR"
        elif type_data == "Images":
            Icon = "Images"
        elif type_data == "Intel":
            Icon = "Intel"
        elif type_data == "Locations":
            Icon = "Locations"
            if Subgroup == "SearchedPlaces":
                Icon = "Searched"
            elif Subgroup == "Shared":
                Icon = "Shared"   
            elif Subgroup == "Mentioned":
                Icon = "Locations"  # task
        elif type_data == "Searched Items":
            Icon = "Searched"
        elif type_data == "Toll":
            Icon = "Toll"
        elif type_data == "Videos":
            Icon = "Videos"




# write rows to data
        row_data["Latitude"] = latitude
        row_data["Longitude"] = longitude 
        row_data["Address"] = address
        row_data["business"] = business 
        row_data["number"] = number 
        row_data["street"] = street
        row_data["city"] = city 
        row_data["county"] = county 
        row_data["state"] = state 
        row_data["zipcode"] = zipcode
        row_data["country"] = country 
        row_data["fulladdress"] = fulladdress
        row_data["query"] = query
        row_data["Coordinate"] = coordinate
        row_data["PlusCode"] = PlusCode
        row_data["Icon"] = Icon
        
        print(f'\nName: {name_data}\nCoordinate: {coordinate}\naddress = {address}\nbusiness = {business}\nfulladdress = {fulladdress}\n')

    return data

def read_intel(input_xlsx):
    """Read intel_.xlsx sheet and convert it to locations format.    
    """

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []

    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        print(f"{color_red}No data found in the Excel file.{color_reset}")
        return None

    for row_index, row_data in enumerate(data):

        (fullname, phone, business, fulladdress, city, state) = ('', '', '', '', '', '')
        (note, sosagent, Time, Latitude, Longitude, Coordinate) = ('', '', '', '', '', '')
        (Source, description, type_data, Name, tag, source) = ('', '', 'Intel', '', '', '')
        (group, subgroup, number, street, county, zipcode) = ('', '', '', '', '', '')
        (country, query, plate, capture_time, hwy, direction) = ('', '', '', '', '', '')
        (end_time, category, time_orig, timezone, PlusCode, Icon) = ('', '', '', '', '', '')

# Time
        # Time = ''
        Time = row_data.get("Time")
        if Time is None:
            Time = ''

# Latitude
        # latitude = ''
        latitude = row_data.get("Latitude")
        latitude = str(latitude)
        if latitude is None or latitude == 'None':
            latitude = ''        

# Latitude
        # longitude = ''
        longitude = row_data.get("Longitude")
        longitude = str(longitude)
        if longitude is None or longitude == 'None':
            longitude = ''        

# Coordinate    
        Coordinate = ''
        Coordinate = row_data.get("Coordinate")
        if Coordinate is None:
            Coordinate = ''
            
# address
        address = row_data.get("fulladdress")
        if address is None:
            address = ''    
            
# Name    
        Name = ''
        Name = row_data.get("fullname")
        if Name is None:
            Name = ''

# source
        source = row_data.get("Source")
        if source is None:
            source = ''        
            
# source file
        source_file = row_data.get("Source file information")
        if source_file is None or source_file == '':
            source_file = input_xlsx

# origin_file
        origin_file = row_data.get("origin_file")
        if origin_file is None or origin_file == "":
            origin_file = input_xlsx
            
# sosagent    
        sosagent = ''
        sosagent = row_data.get("sosagent")
        if sosagent is None:
            sosagent = ''
            
# phone    
        phone = ''
        phone = row_data.get("phone")
        if phone is None:
            phone = ''
        else:
            description = (f'{description}\nPhone:{phone}')

# business    
        business = ''
        business = row_data.get("business")
        if business is None:
            business = ''
        else:
            description = (f'{description}\nBusiness:{business}')

        description = description.strip()

# city    
        city = ''
        city = row_data.get("city")
        if city is None:
            city = ''

# state    
        state = ''
        state = row_data.get("state")
        if state is None:
            state = ''

# Icon    
        Icon = ''
        Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''

      
# write rows to data
        row_data["Time"] = Time
        row_data["Latitude"] = latitude
        row_data["Longitude"] = longitude 
        row_data["Address"] = address
        # row_data["Group"] = group
        row_data["Subgroup"] = subgroup
        row_data["Description"] = description
        row_data["Type"] = type_data
        row_data["Tag"] = tag        
        row_data["Source"] = source
        row_data["Source file information"] = source_file
        row_data["Name"] = Name
        row_data["business"] = business 
        # row_data["number"] = number 
        # row_data["street"] = street
        row_data["city"] = city 
        # row_data["county"] = county 
        row_data["state"] = state 
        # row_data["zipcode"] = zipcode
        # row_data["country"] = country 
        row_data["fulladdress"] = fulladdress
        # row_data["query"] = query
        # row_data["Plate"] = plate
        # row_data["Capture Time"] = capture_time
        # row_data["Highway Name"] = hwy
        row_data["Coordinate"] = Coordinate
        # row_data["Direction"] = direction
        # row_data["End time"] = end_time
        # row_data["Category"] = category
        # row_data["Time Original"] = time_orig
        # row_data["Timezone"] = timezone
        # row_data["PlusCode"] = PlusCode
        row_data["Icon"] = Icon
        row_data["origin_file"] = origin_file # test
        # index
     
     
    return data



def read_locations(input_xlsx):

    """Read data from an xlsx file and return as a list of dictionaries.
    Read XLSX Function: The read_xlsx() function reads data from the input 
    Excel file using the openpyxl library. It extracts headers from the 
    first row and then iterates through the data rows.    
    """

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []

    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        print(f"{color_red}No data found in the Excel file.{color_reset}")
        return None

    for row_index, row_data in enumerate(data):
        (zipcode, business, number, street, city, county) = ('', '', '', '', '', '')
        (state, fulladdress, Latitude, Longitude, query, Coordinate) = ('', '', '', '', '', '')
        (Index, country, capture_time, PlusCode, time_orig, Icon) = ('', '', '', '', '', '')
        (description, group, subgroup, source, source_file, tag) = ('', '', '', '', '', '')
        (Time, capture_date, timezone, origin_file) = ('', '', '', '')  
        (end_time, category, latitude, longitude, coordinate, address) = ('', '', '', '', '', '')

        (type_data, plate, country_code, state_ftk, city_ftk, hwy) = ('', '', '', '', '', '')
        (direction, name_data, no, account, container, time_local) = ('', '', '', '', '', '') 
        (deleted, service_id, carved, sighting_state, sighting_location, manually_decoded) = ('', '', '', '', '', '')

        name_data = row_data.get("Name")
        if name_data is None:
            name_data = ''

# no 
        no = row_data.get("#")
        if no is None:
            no = ''
            
# Description    
        description = row_data.get("Description")
        if description is None:
            description = ''




# Time
        Time = row_data.get("Time")
        if Time is None:
            Time = ''

# time_orig
        time_orig = row_data.get("Time Original")
        if time_orig is None:
            time_orig = ''

# Capture Time
        capture_time  = row_data.get("Capture Time") 
        if capture_time is None:
            capture_time = ''     

        if Time == '':
            if capture_time != '':
                Time = capture_time

# Capture Date
        capture_date  = row_data.get("Capture Date") 
        if capture_date is None:
            capture_date = ''     

        if Time == '':
            if capture_date != '':
                Time = capture_date

# timezone
        timezone  = row_data.get("Timezone")
        if timezone is None:
            timezone = ''  

# time Gmail warrant retuirn
        time2  = row_data.get("Timestamp (UTC)")
        if time2 is None:
            time2 = ''  
        if Time == "" and time2 != "":
            Time = time2
            timezone = "UTC"
            

# convert time
        # output_format = "%m/%d/%Y %H:%M:%S"  # Changed to military time
        output_format = "%Y/%m/%d %H:%M:%S"  # Changed to ISO military time
        # output_format = "%Y-%m-%dT%H:%M:%SZ"    # Google Earth format

        # pattern = r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$'
        pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'  # ISO military time

        if time_orig == '' and Time != '': # copy the original time
            time_orig = Time
        try:
            (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
            Time = Time.strftime(output_format)
            
            # print(f'            Time = {Time}   time_orig = {time_orig}')   # temp
            if Time is None:
                Time = ''              
            
        except ValueError as e:
            print(f"Error time2: {e} - {Time}")
            Time = ''    # temp rem of this
            pass
        
# End time
        end_time = row_data.get("End time")
        if end_time is None:
            end_time = ''        

# Category
        category = row_data.get("Category")
        if category is None:
            category = ''        

# gps

        latitude = row_data.get("Latitude")
        latitude = str(latitude)
        if latitude is None or latitude == 'None':
            latitude = ''        

        longitude = row_data.get("Longitude")
        longitude = str(longitude)
        if longitude is None or longitude == 'None':
            longitude = ''        

        if latitude == '': 
            latitude = row_data.get("Capture Location Latitude")
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = ''        

        if longitude == '':
            longitude = row_data.get("Capture Location Longitude")
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = ''        

# coordinate        


        if latitude != '' and longitude != '':
            coordinate = (f'{latitude},{longitude}')
        elif row_data.get("Coordinate") != None:
            coordinate = row_data.get("Coordinate")
            
        elif row_data.get("Capture Location") != None:
            coordinate = row_data.get("Capture Location")
        elif row_data.get("Capture Location (Latitude,Longitude)") != None:
            coordinate = row_data.get("Capture Location (Latitude,Longitude)")

        if len(coordinate) > 6:
           
            coordinate = coordinate.replace('(', '').replace(')', '')
            if ',' in coordinate:
                coordinate = coordinate.split(',')
                latitude = coordinate[0].strip()
                longitude = coordinate[1].strip()
                longitude = coordinate[1].strip()
                coordinate = (f'{latitude},{longitude}')

# address
        address = row_data.get("Address")
        if address is None:
            address = ''        

# group
        group = row_data.get("Group")
        if group is None:
            group = ''

# subgroup
        subgroup = row_data.get("Subgroup")
        if subgroup is None:
            subgroup = ''

# type
        type_data = row_data.get("Type")
        if type_data is None:
            type_data = ''  

# tag
        tag = row_data.get("Tag")
        if tag is None:
            tag = ''


# tag from label
        label  = row_data.get("Label") 
        if label is None:
            label = ''     

        if tag == '':
            if label != '':
                tag = label

# source
        source = row_data.get("Source")
        if source is None:
            source = ''

# carved
        carved = row_data.get("Source file information")
        if carved is None or 'xlsx' in carved:
            carved = ''

# source file
        source_file = row_data.get("Source file information")
        if source_file is None:
            source_file = ''

# origin_file
        origin_file = row_data.get("origin_file")
        if origin_file is None or origin_file == "":
            origin_file = input_xlsx
            
# business  
        business = row_data.get("business")
        if business is None:
            business = ''

# fulladdress
        fulladdress  = row_data.get("fulladdress")
        if fulladdress is None:
            fulladdress = ''       

# query
        query  = row_data.get("query")
        if query is None:
            query = ''     

# Plate
        plate  = row_data.get("Plate")         # red
        if plate is None:
            plate = ''     

        if plate != '' and type_data == '':
            type_data = 'LPR'
            


# country
        country = row_data.get("country")
        if country is None:
            country = ''     

# Country Code
        country_code  = row_data.get("Country Code") 
        if country_code is None:
            country_code = ''     

        if country == '':
            if country_code != '':
                country = country_code

# county  
        county = row_data.get("county")
        if county is None:
            county = ''

# state  
        state = row_data.get("state")
        if state is None:
            state = ''
        
# state_ftk
        state_ftk  = row_data.get("Region") 
        if state_ftk is None:
            state_ftk = ''     
        if state == '':
            if state_ftk != '':
                state = state_ftk

# zipcode  
        zipcode = row_data.get("zipcode")
        if zipcode is None:
            zipcode = ''

# number  
        number = row_data.get("number")
        if number is None:
            number = ''
            
# street  
        street = row_data.get("street")
        if street is None:
            street = ''


# city  
        city = row_data.get("city")
        if city is None:
            city = ''

# City
        city_ftk  = row_data.get("City") 
        if city_ftk is None:
            city_ftk = ''     

        if city == '':
            if city_ftk != '':
                city = city_ftk 
            
# hwy
        hwy  = row_data.get("Highway Name")
        if hwy is None:
            hwy = ''           

        if hwy == '':
            hwy  = row_data.get("Capture Camera")
            if hwy is None:
                hwy = ''           

# Direction
        direction  = row_data.get("Direction")
        if direction is None:
            direction = ''    
        if direction == '':
            match = re.search(r'\((\w+)\)',hwy)
            if ' NB' in hwy:
                direction = 'N'
            elif ' EB' in hwy:
                direction = 'E'
            elif ' SB' in hwy:
                direction = 'S'
            elif ' WB' in hwy:
                direction = 'W'
            elif match:
                direction = match.group(1).replace('B','')

            if 'Northbound' in direction:
                direction = 'N'
            elif 'Eastbound' in direction:
                direction = 'E'
            elif 'Southbound' in direction:
                direction = 'S'
            elif 'Westbound' in direction:
                direction = 'W'
                


# PlusCode
        PlusCode  = row_data.get("PlusCode")
        if PlusCode is None:
            PlusCode = ''  

# Icon    
        Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''
        if Icon != "":
            Icon = Icon[0].upper() + Icon[1:].lower()
        
# write rows to data
        row_data["#"] = no
        row_data["Time"] = Time
        row_data["Latitude"] = latitude
        row_data["Longitude"] = longitude 
        row_data["Address"] = address
        row_data["Group"] = group
        row_data["Subgroup"] = subgroup
        row_data["Description"] = description
        row_data["Type"] = type_data
        row_data["Source"] = source
        # row_data["Deleted"] = deleted
        row_data["Tag"] = tag        
        row_data["Source file information"] = source_file
        # row_data["Service Identifier"] = service_id   
        # row_data["Carved"] = carved        
        row_data["Name"] = name_data
        row_data["business"] = business 
        row_data["number"] = number 
        row_data["street"] = street
        row_data["city"] = city 
        row_data["county"] = county 
        row_data["state"] = state 
        row_data["zipcode"] = zipcode
        row_data["country"] = country 
        row_data["fulladdress"] = fulladdress
        row_data["query"] = query
        # row_data["Sighting State"] = sighting_state
        row_data["Plate"] = plate
        row_data["Capture Time"] = capture_time
        row_data["Highway Name"] = hwy
        row_data["Coordinate"] = coordinate
        row_data["Capture Location Latitude"] = latitude        
        row_data["Capture Location Longitude"] = longitude       
        # row_data["Container"] = container       
        # row_data["Sighting Location"] = sighting_location        
        row_data["Direction"] = direction
        # row_data["Time Local"] = time_local
        row_data["End time"] = end_time
        row_data["Category"] = category
        # row_data["Manually decoded"] = manually_decoded        
        # row_data["Account"] = account
        row_data["PlusCode"] = PlusCode        
        row_data["Time Original"] = time_orig
        row_data["Timezone"] = timezone
        row_data["Icon"] = Icon        
        row_data["origin_file"] = origin_file
     
     
    return data


def write_kml(data):
    '''
    The write_kml() function receives the processed data as a list of 
    dictionaries and writes it to a kml using simplekml. 
    '''

    # Create KML object
    kml = simplekml.Kml()

    # Define different default icons
    # square_icon = 'http://maps.google.com/mapfiles/kml/shapes/square.png'
    # triangle_icon = 'http://maps.google.com/mapfiles/kml/shapes/triangle.png'
    # star_icon = 'http://maps.google.com/mapfiles/kml/shapes/star.png'
    # polygon_icon = 'http://maps.google.com/mapfiles/kml/shapes/polygon.png'
    # circle_icon = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'
    # yellow_circle_icon = 'http://maps.google.com/mapfiles/kml/paddle/ylw-circle.png'
    # red_circle_icon = 'http://maps.google.com/mapfiles/kml/paddle/red-circle.png'
    # white_circle_icon = 'http://maps.google.com/mapfiles/kml/paddle/wht-circle.png'

    default_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon13.png'   # yellow flag

    car_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon15.png'   # red car
    car2_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon47.png'  # yellow car
    car3_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon54.png'  # green car with circle
    car4_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon7.png'  # red car with circle

    truck_icon = 'https://maps.google.com/mapfiles/kml/shapes/truck.png'    # blue truck
    calendar_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon23.png' # paper
    chat_icon = 'https://maps.google.com/mapfiles/kml/shapes/post_office.png' # email
    locations_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon28.png'    # yellow paddle
    home_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon56.png'
    images_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon46.png'
    intel_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon44.png'
    office_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon21.png'
    searched_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon9.png'
    shared_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon20.png'
    
    toll_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-none.png'
    videos_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon30.png'
    n_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-0.png'
    e_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-4.png'
    s_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-8.png'
    w_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-12.png'
      
    

    for row_index, row_data in enumerate(data):
        index_data = row_index + 2  # excel row starts at 2, not 0
        
        Time = row_data.get("Time") #
        latitude = row_data.get("Latitude") #
        longitude = row_data.get("Longitude") #
        address = row_data.get("Address") #
        group_data = row_data.get("Group") 
        subgroup = row_data.get("Subgroup")   
        description = row_data.get("Description")
        type_data = row_data.get("Type")#
        tag = row_data.get("Tag")
        source_file = row_data.get("Source file information") #
        name_data = row_data.get("Name")
        business = row_data.get("business")
        fulladdress  = row_data.get("fulladdress")
        plate  = row_data.get("Plate")   
        hwy  = row_data.get("Highway Name") #
        coordinate = row_data.get("Coordinate")
        direction  = row_data.get("Direction")
        end_time = row_data.get("End time")
        category = row_data.get("Category")
        Icon = row_data.get("Icon")
        origin_file = row_data.get("origin_file")

        if name_data != '':
            (description) = (f'{description}\nNAME: {name_data}')
        
        if Time != '':
            (description) = (f'{description}\nTIME: {Time}')

        if end_time != '':
            (description) = (f'{description}\nEnd Time: {end_time}')

        if address != '':
            (description) = (f'{description}\n{address}')

        elif fulladdress != '':
            (description) = (f'{description}\nADDRESS: {fulladdress}')

        if hwy != '':
            (description) = (f'{description}\nHWY NAME: {hwy}')
            
        if direction != '':
            (description) = (f'{description}\nDIRECTION: {direction}')

        # if source_file != '' and source_file != None:
            # (description) = (f'{description}\nSOURCE: {source_file}')

        if tag != '':
            (description) = (f'{description}\nTAG: {tag}')    # test

        if type_data != '':
            (description) = (f'{description}\nTYPE: {type_data}')

        if group_data != '':
            (description) = (f'{description} / {group_data}')

        if subgroup != '' and subgroup != 'Unknown':
            (description) = (f'{description} / {subgroup}')

        if business != '':
            (description) = (f'{description}\nBusiness: {business}')
            
        if plate != '':
            (description) = (f'{description}\nPLATE: {plate}')

        point = ''  # Initialize point variable outside the block
        
        if latitude == '' or longitude == '' or latitude == None or longitude == None:
            print(f'skipping row {index_data} - No GPS')

        elif Icon == "Lpr" or Icon == "Car":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = car_icon  # red car

        elif Icon == "Car2":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = car2_icon  # yellow car

        elif Icon == "Car3":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = car3_icon  # green car

        elif Icon == "Car4":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = car4_icon  # red car (with circle)

        elif Icon == "Truck":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = truck_icon

        elif Icon == "Calendar":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = calendar_icon

        elif Icon == "Chat":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = chat_icon
            
        elif Icon == "Home":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = home_icon

        elif Icon == "Images":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = images_icon

        elif Icon == "Intel":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = intel_icon

        elif Icon == "Office":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = office_icon

        elif Icon == "Searched":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = searched_icon

        elif Icon == "Shared":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = shared_icon
            
        elif Icon == "Videos":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = videos_icon

        elif Icon == "Locations":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = locations_icon    # yellow paddle

        elif Icon == "Toll":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = toll_icon

        elif Icon == "N":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = n_icon

        elif Icon == "E":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = e_icon

        elif Icon == "S":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = s_icon

        elif Icon == "W":
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = w_icon

        else:
            point = kml.newpoint(
                name=f"{index_data}",
                description=f"{description}",
                coords=[(longitude, latitude)]
            )
            point.style.iconstyle.icon.href = default_icon    # orange paddle


        
        if tag != '':   # mark label yellow if tag is not blank
            try:
                point.style.labelstyle.color = simplekml.Color.yellow  # Set label text color
                # point.style.labelstyle.scale = 1.2  # Adjust label scale if needed    # task
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    
    kml.save(output_kml)    # Save the KML document to the specified output file

    print(f"KML file '{output_kml}' created successfully!")


def write_locations(data):
    '''
    The write_locations() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''
    print(f'write_locations')   # temp
    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Locations'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = [
        "#", "Time", "Latitude", "Longitude", "Address", "Group", "Subgroup"
        , "Description", "Type", "Source", "Deleted", "Tag", "Source file information"
        , "Service Identifier", "Carved", "Name", "business", "number", "street"
        , "city", "county", "state", "zipcode", "country", "fulladdress", "query"
        , "Sighting State", "Plate", "Capture Time", "Capture Network", "Highway Name"
        , "Coordinate", "Capture Location Latitude", "Capture Location Longitude"
        , "Container", "Sighting Location", "Direction", "Time Local", "End time"
        , "Category", "Manually decoded", "Account", "PlusCode", "Time Original", "Timezone"
        , "Icon", "origin_file", "Index"

    ]

    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [2, 3, 4]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [1, 5, 6, 7, 8, 9, 15, 16, 24, 30, 31, 36, 38]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill
        elif col_index == 27:  # Red for column 27
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
            cell.fill = fill

    ## Excel column width
    worksheet.column_dimensions['A'].width = 8# #
    worksheet.column_dimensions['B'].width = 19# Time
    worksheet.column_dimensions['C'].width = 18# Latitude
    worksheet.column_dimensions['D'].width = 18# Longitude
    worksheet.column_dimensions['E'].width = 45# Address
    worksheet.column_dimensions['F'].width = 14# Group
    worksheet.column_dimensions['G'].width = 13# Subgroup
    worksheet.column_dimensions['H'].width = 17# Description
    worksheet.column_dimensions['I'].width = 9# Type
    worksheet.column_dimensions['J'].width = 10# Source
    worksheet.column_dimensions['K'].width = 10# Deleted
    worksheet.column_dimensions['L'].width = 11# Tag
    worksheet.column_dimensions['M'].width = 20# Source file information
    worksheet.column_dimensions['N'].width = 15# Service Identifier
    worksheet.column_dimensions['O'].width = 7# Carved
    worksheet.column_dimensions['P'].width = 15# Name
    
    ## bonus
    worksheet.column_dimensions['Q'].width = 20# business 
    worksheet.column_dimensions['R'].width = 10# number
    worksheet.column_dimensions['S'].width = 20# street 
    worksheet.column_dimensions['T'].width = 15# city   
    worksheet.column_dimensions['Y'].width = 25# county    
    worksheet.column_dimensions['V'].width = 12# state   
    worksheet.column_dimensions['W'].width = 8# zipcode     
    worksheet.column_dimensions['X'].width = 6# country    
    worksheet.column_dimensions['Y'].width = 26# FullAddress   
    worksheet.column_dimensions['Z'].width = 26# query

    ##  Flock
    worksheet.column_dimensions['AA'].width = 11# Sighting State
    worksheet.column_dimensions['AB'].width = 11# Plate
    worksheet.column_dimensions['AC'].width = 22# Capture Time
    worksheet.column_dimensions['AD'].width = 15# Capture Network
    worksheet.column_dimensions['AE'].width = 21# Highway Name
    worksheet.column_dimensions['AF'].width = 30# Coordinate
    worksheet.column_dimensions['AG'].width = 20# Capture Location Latitude
    worksheet.column_dimensions['AH'].width = 20# Capture Location Longitude

    ##
    worksheet.column_dimensions['AI'].width = 10# Container
    worksheet.column_dimensions['AJ'].width = 14# Sighting Location
    worksheet.column_dimensions['AK'].width = 10# Direction
    worksheet.column_dimensions['AL'].width = 11# Time Local
    worksheet.column_dimensions['AM'].width = 25# End time
    worksheet.column_dimensions['AN'].width = 10# Category
    worksheet.column_dimensions['AO'].width = 18# Manually decoded
    worksheet.column_dimensions['AP'].width = 10# Account
    worksheet.column_dimensions['AQ'].width = 25 # PlusCode
    worksheet.column_dimensions['AR'].width = 21 # Time Original
    worksheet.column_dimensions['AS'].width = 9 # Timezone
    worksheet.column_dimensions['AT'].width = 10 # Icon   
    worksheet.column_dimensions['AU'].width = 7 # origin_file
    worksheet.column_dimensions['AV'].width = 6 # Index

    
    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")


    # Create a new worksheet for color codes
    color_worksheet = workbook.create_sheet(title='Icons')
    color_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    color_worksheet.column_dimensions['A'].width = 8# Icon sample
    color_worksheet.column_dimensions['B'].width = 9# Name
    color_worksheet.column_dimensions['C'].width = 29# Description

    # Excel row height
    color_worksheet.row_dimensions[2].height = 22  # Adjust the height as needed
    color_worksheet.row_dimensions[3].height = 22
    color_worksheet.row_dimensions[4].height = 23
    color_worksheet.row_dimensions[5].height = 23
    color_worksheet.row_dimensions[6].height = 40   # truck
    color_worksheet.row_dimensions[7].height = 6
    color_worksheet.row_dimensions[8].height = 24
    color_worksheet.row_dimensions[9].height = 22
    color_worksheet.row_dimensions[10].height = 22
    color_worksheet.row_dimensions[11].height = 22
    color_worksheet.row_dimensions[12].height = 23
    color_worksheet.row_dimensions[13].height = 23
    color_worksheet.row_dimensions[14].height = 25
    color_worksheet.row_dimensions[15].height = 25
    color_worksheet.row_dimensions[16].height = 23
    color_worksheet.row_dimensions[17].height = 6
    color_worksheet.row_dimensions[18].height = 38
    color_worksheet.row_dimensions[19].height = 38
    color_worksheet.row_dimensions[20].height = 38
    color_worksheet.row_dimensions[21].height = 38
    color_worksheet.row_dimensions[22].height = 38
    color_worksheet.row_dimensions[23].height = 6
    color_worksheet.row_dimensions[24].height = 15
    color_worksheet.row_dimensions[25].height = 6
    color_worksheet.row_dimensions[26].height = 15


    
    # Define color codes
    color_worksheet['A1'] = ' '
    color_worksheet['B1'] = 'Icon'
    color_worksheet['C1'] = 'Icon Description'

    icon_data = [

        ('', 'Car', 'Lpr red car (License Plate Reader)'),
        ('', 'Car2', 'Lpr yellow car'),
        ('', 'Car3', 'Lpr greeen car with circle'),
        ('', 'Car4', 'Lpr red car with circle'),
        ('', 'Truck', 'Lpr truck'),         
        ('', '', ''),
        ('', 'Calendar', 'Calendar'), 
        ('', 'Home', 'Home'),                
        ('', 'Images', 'Photo'),
        ('', 'Intel', 'I'),  
        ('', 'Locations', 'Reticle'),  
        ('', 'default', 'Yellow flag'),  
        ('', 'Office', 'Office'),         
        ('', 'Searched', 'Searched Item'),          
        ('', 'Videos', 'Video clip'),        
        ('', '', ''),
        ('', 'Toll', 'Blue square'), 
        ('', 'N', 'Northbound blue arrow'),
        ('', 'E', 'Eastbound blue arrow'),
        ('', 'S', 'Southbound blue arrow'),
        ('', 'W', 'Westbound blue arrow'),
        ('', '', ''),
        ('', 'Yellow font', 'Tagged'),
        ('', 'Chats', 'Chats'),   # 


        ('', '', ''),
        ('', 'NOTE', 'visit https://earth.google.com/ <file><Import KML> select gps.kml <open>'),
    ]

    for row_index, (icon, tag, description) in enumerate(icon_data):
        color_worksheet.cell(row=row_index + 2, column=1).value = icon
        color_worksheet.cell(row=row_index + 2, column=2).value = tag
        color_worksheet.cell(row=row_index + 2, column=3).value = description

    car_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon15.png'   # red car
    car2_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon47.png'  # yellow car
    car3_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon54.png'  # green car with circle
    car4_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon7.png'  # red car with circle
    truck_icon = 'https://maps.google.com/mapfiles/kml/shapes/truck.png'    # blue truck
    default_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon13.png'   # yellow flag
    calendar_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon23.png' # paper
    chat_icon = 'https://maps.google.com/mapfiles/kml/shapes/post_office.png' # email
    locations_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon28.png'    # yellow paddle
    home_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon56.png'
    images_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon46.png'
    intel_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon44.png'
    office_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon21.png'
    searched_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon0.png'  #  
    toll_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-none.png'
    videos_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon30.png'
    n_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-0.png'
    e_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-4.png'
    s_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-8.png'
    w_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-12.png'

    try:
        # Insert graphic from URL into cell of color_worksheet

        response = requests.get(car_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A2')

        response = requests.get(car2_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A3')

        response = requests.get(car3_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A4')

        response = requests.get(car4_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A5')

        response = requests.get(truck_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A6')

        response = requests.get(calendar_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A8')
        
        response = requests.get(home_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A9')

        response = requests.get(images_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A10')

        response = requests.get(intel_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A11')

        response = requests.get(locations_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A12')

        response = requests.get(default_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A13')

        response = requests.get(office_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A14')        

        response = requests.get(searched_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A15')

        response = requests.get(videos_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A16')
        
        response = requests.get(toll_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A18')

        response = requests.get(n_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A19')

        response = requests.get(e_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A20')

        response = requests.get(s_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A21')

        response = requests.get(w_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A22')

        response = requests.get(chat_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A24')     
        
    except:
        pass

    
    workbook.save(outuput_xlsx)


def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description2}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}insert your input into locations.xlsx')
    print(f'\nExample:')
    print(f'    {file} -c -O input_blank.xlsx') 
    print(f'    {file} -k -I locations.xlsx  # xlsx 2 kml with no internet processing')     
    print(f'    {file} -r')
    print(f'    {file} -r -I locations.xlsx -O locations2addresses_.xlsx') 
    print(f'    {file} -r -I locations_FTK.xlsx -O locations2addresses_.xlsx') 
    print(f'    {file} -r -I Flock.xlsx -O locations_Flock.xlsx')    
    print(f'    {file} -r -I MediaLocations_.xlsx')  
    print(f'    {file} -r -I PointsOfInterest.xlsx -O locations_PointsOfInterest.xlsx') 
    print(f'    {file} -r -I Tolls.xlsx -O locations_Tolls.xlsx')     
    print(f'    {file} -i -I intel_.xlsx -O intel2locations_.xlsx')  
    print(f'    {file} -i -I intel_SearchedItems_.xlsx')  
    print(f'    {file} -i -I intel_Chats_.xlsx')  

    
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
1.1.0 - color code sheet
1.0.1 - Color coded pins for gps.kml
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
fix Time
Country and zipcode are blank
if address / fulldata only, get lat/long
export a temp copy to output.txt
if it's less than 3000 skip the sleep timer

Add Group and Subgroup, color

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
connection timeout after about 4000 attempts
with the sleep timer set to 10 (sec) it doesn't crap out.

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

