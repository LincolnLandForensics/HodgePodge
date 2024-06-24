 #!/usr/bin/env python3
# coding: utf-8
'''
read pdf and convert it to pdf
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import re
import os
import sys
import PyPDF2   # splitting

from datetime import datetime
import argparse
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import pdfplumber
import pandas as pd

# Colorize section
global color_red, color_yellow, color_green, color_blue, color_purple, color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000:
        import colorama # pip install colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}')
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL
        
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>
author = 'LincolnLandForensics'
description = "read pdfs and convert their tables to a single .xlsx"
version = '1.0.2'



def main():
    # global row
    # row = 0
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-b', '--blank', help='blank sheet', required=False, action='store_true')
    parser.add_argument('-p', '--pdf', help='parse pdf', required=False, action='store_true')
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-t', '--tables', help='output tables', required=False, action='store_true')
    parser.add_argument('-s', '--subfolders', help='parse subfolders', required=False, action='store_true')
    parser.add_argument('-S', '--split', help='split pdfs', required=False, action='store_true')
    parser.add_argument('-d', '--table_dump_1', help='parse pdf type 2', required=False, action='store_true')

    args = parser.parse_args()

    global tables_out
    # Assuming args.tables is a list or other iterable
    tables_out = False  # Default value

    if args.tables:  # If args.tables is not empty or None
        tables_out = True

    global output_xlsx
    if not args.output: 
        output_xlsx = f"invoices_.xlsx"        
    else:
        output_xlsx = args.output    
    
    global input_folder
    if not args.input: 
        input_folder = "pdfs"        
    else:
        input_folder = args.input

    global output_split_folder
    output_split_folder = 'pdfs_single_page'

    global sub_folders
    if args.subfolders:
        sub_folders = True
    else:
        sub_folders = False
    # print(f"PDF's in subfolders are going to be processed")

    if args.blank:
        data = []
        write_xlsx(data)

    elif args.split:
        split_pdfs(input_folder, output_split_folder)
        
    elif args.pdf:
        # Directory containing PDF files
        # input_folder = 'pdfs'
        # Path where the output Excel file will be saved
        output_excel_path = 'output_pdfs.xlsx'


        if not os.path.exists(input_folder):
            print(f"Error: The directory '{input_folder}' does not exist.")
            # return
        else:
            msg_blurb = (f"Reading pdf's in {input_folder}")
            msg_blurb_square(msg_blurb, color_green)             
        # process_pdfs_in_directory_old(input_folder, output_excel_path) 
        process_pdfs_in_directory(input_folder, output_excel_path)
        # process_pdfs_in_directory(input_folder, output_excel_path, tables_out=True, sub_folders=True)

        if tables_out:    
            msg_blurb = (f'Tables have been extracted and saved to {output_excel_path}')
            msg_blurb_square(msg_blurb, color_green) 


    elif args.table_dump_1:
        output_excel_path = 'output_pdfs.xlsx'

        if not os.path.exists(input_folder):
            print(f"Error: The directory '{input_folder}' does not exist.")
        else:
            msg_blurb = (f"Reading pdf's in {input_folder}")
            msg_blurb_square(msg_blurb, color_green)             
        # process_pdfs_in_directory_old(input_folder, output_excel_path) 
        table_dump1(input_folder, output_excel_path)
        # process_pdfs_in_directory(input_folder, output_excel_path, tables_out=True, sub_folders=True)

        if tables_out:    
            msg_blurb = (f'Tables have been extracted and saved to {output_excel_path}')
            msg_blurb_square(msg_blurb, color_green) 

    else:
        usage()

    # return 0

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def extract_information(tables):
    """
    Extract specific information from the tables.
    
    Parameters:
    tables (list of pd.DataFrame): List of DataFrames representing tables.
    
    Returns:
    dict: Dictionary containing the extracted values.
    """
    data_tables = {
        "Date": None,
        "Invoice #": None,
        "Bill To:": None,
        "Ship To:": None,
        "Quantity": None,
        "Qty": None,
        "Description": None,
        "Price Each": None,
        "Amount": None,
        "Unit Price": None,
        "Total": None,
        "Product": None,
        "Qty": None,
        "Column1": None,
        "Column2": None,
        "Column3": None,
        "Column4": None,
        "Column5": None,
        "Column6": None,
        "Column7": None,
        "Column8": None,
        "Column9": None,
        "Column10": None
    }

    for table in tables:
        for key in data_tables.keys():
            if key in table.columns:
                data_tables[key] = table.loc[0, key]
                
        if all(data_tables.values()):
            break

    return data_tables
  
def extract_tables_from_pdf(pdf_path):
    """
    Extract tables from a given PDF file.
    
    Parameters:
    pdf_path (str): Path to the PDF file.
    
    Returns:
    list of pd.DataFrame: A list of DataFrames, one for each table found in the PDF.
    """
    tables = []
    text, total, date = '', '', ''
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:

            # extract tables
            extracted_tables = page.extract_tables()
            for table in extracted_tables:
                if table:
                    df = pd.DataFrame(table[1:], columns=table[0])
                    tables.append(df)

            # extract text
            text = page.extract_text()
            # print(f"Text on page {i+1}:\n{text}")
            # print(f"Text on page :\n{text}")  # temp


            
            
    return tables, text

def extract_info_from_text(text, total, date, grand_total, shipping, invoice):
    disclaimer = ''
    # total, date = '', ''
    # Regular expression to find the pattern 'Total $' followed by a number with commas
    match_time = re.search(r'Total:\s*\$\s*([\d,]+\.\d{2})', text)
    match_date = re.search(r'Date:\s*(\d{1,2}/\d{1,2}/\d{4})', text)
    match_total = re.search(r'Total\s*\$([\d,]+\.\d{2})', text)
    # match_total2 = re.search(r'Total:\s*\$([0-9]+\.[0-9]{2})', text)
    match_total2 = re.search(r'Total:\s*\$\s*([0-9,]+\.[0-9]{2})', text)

    match_shipping = re.search(r'Shipping charge 1 \$([0-9]+\.[0-9]{2})', text)
    match_invoice = re.search(r'Willow st, suite #B (\d+)', text)

    # todo one of the grand_total has less spaces when the amount is over $999
    if match_total:
        # Remove commas from the number and return it as a float
        grand_total1 = float(match_total.group(1).replace(',', ''))
        # print(f'text2 = {text}OOOOOOOOOOOOOOOOOOOOOOOOOOOOOO')  # temp
        if grand_total1 is None:
            grand_total1 = ''
        if grand_total == '':
            grand_total = grand_total1
        else:
            grand_total = (f'{grand_total}\t{grand_total1}')
        # print(f'grand_total3 = {grand_total}____________________________') # temp   # todo
    if match_total2:
        grand_total2 = float(match_total2.group(1).replace(',', ''))
        # print(f'grand_total2-1 = {grand_total2}____________________________') # temp   # todo

        if grand_total2 is None:
            grand_total2 = ''
        if grand_total == '':
            grand_total = grand_total2
        else:
            grand_total2 = (f'{grand_total}\t{grand_total2}')    # temp        
        # print(f'grand_total2 = {grand_total2}____________________________') # temp   # todo

    if match_date:
        if date is None or date == '':
            date = match_date.group(1)
            # print(f'date5 = {date}_______________________') # temp    # works

    if 'Customer is responsible for local tobacco tax' in text:
        disclaimer = 'Customer is responsible for local tobacco tax.'
    if 'purchased for delivery outside of Illinois' in text:
        disclaimer = (f'{disclaimer} Tobacco products purchased for delivery outside of Illinois.')

    if match_shipping:
        if shipping is None or shipping == '':
            shipping = match_shipping.group(1)
            
    if match_invoice:
        if invoice is None or invoice == '':
            invoice = match_invoice.group(1)
            # print(f'invoice = {invoice}_____________________')  # temp
            
    return disclaimer, total, date, grand_total, shipping, invoice


def msg_blurb_square(msg_blurb, color):
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')

def process_pdfs_in_directory(input_folder, output_excel_path):
    """
    Loop through all PDFs in the specified directory (and optionally in subdirectories), extract tables, and export them to an Excel sheet.
    
    Parameters:
    input_folder (str): Path to the directory containing PDF files.
    output_excel_path (str): Path where the output Excel file will be saved.
    tables_out (bool): Whether to save the extracted tables into separate sheets in the Excel file.
    sub_folders (bool): Whether to include PDFs from subdirectories.
    """
    if tables_out:
        writer = pd.ExcelWriter(output_excel_path, engine='openpyxl')
        
    all_tables = []
    data = []
    seller = input("SELLER: ")
    # Walk through directories and subdirectories if sub_folders is True
    for root, dirs, files in os.walk(input_folder) if sub_folders else [(input_folder, [], os.listdir(input_folder))]:
        for filename in files:
            if filename.lower().endswith('.pdf'):
                date, invoice, billTo, shipTo, quantity, description = '', '', '', '', '', ''
                priceEach, subtotal, amount, total, billToState, shipToState = '', '', '', '', '', ''
                customer, state, year, month, day, folder = '', '', '', '', '', ''
                grand_total, shipping = '', ''
                row_data = {}
                pdf_path = os.path.join(root, filename)
                folder = pdf_path
                tables, text = extract_tables_from_pdf(pdf_path)
                # print(f'text = {text}') # temp
                disclaimer, total, date, grand_total, shipping, invoice = extract_info_from_text(text, total, date, grand_total, shipping, invoice)
                if total:
                    print(f"Found total value : ${total}")
                # else:
                    # print(f"total not found")                

                all_tables.extend(tables)

                data_tables = extract_information(tables)
                # print(f'date2 = {date}') # temp
                for key, value in data_tables.items():
                    if value:
                        if key == "Amount":
                            amount = value
                        elif key == "Date":
                            date = value
                        elif key == "Invoice #":
                            invoice = value
                        elif key == "Bill To:":
                            billTo = value
                        elif key == "Ship To:":
                            shipTo = value
                        elif key == "Quantity":
                            quantity = value
                            if quantity is None:
                                quantity == ''
                        elif key == "Qty" and quantity == '' :
                            quantity = value                            
                        elif key == "Qty":
                            quantity = value
                        elif key == "Description":
                            description = value
                            if description is None:
                                description == ''                            
                        elif key == "Product" and description == '':
                            description = value
                        elif key == "Price Each":
                            priceEach = value
                            if priceEach is None:
                                priceEach == ''
                        elif key == "Unit Price" and priceEach == '':
                            priceEach = value

                        elif key == "Total":
                            total = value


                pattern = r',\s([A-Z]{2})\s\d{5}'

                match1 = re.search(pattern, billTo)
                if match1:
                    billToState = match1.group(1)
                else:
                    billToState = ''

                match2 = re.search(pattern, shipTo)
                if match2:
                    shipToState = match2.group(1)
                else:
                    shipToState = ''

                match3 = re.search(r'^(.*)$', shipTo, re.MULTILINE)
                if match3:
                    customer = match3.group(1)
                else:
                    customer = ''

# total
                # calculate "total" by adding up all the amounts in amount
                pattern4 = r'\$([0-9,]+\.[0-9]{2})'
                match4 = re.findall(pattern4, amount)

                total = sum(float(value4.replace(',', '')) for value4 in match4)    
                total = round(float(total), 2)
                
                try:
                    if date.count('/') == 2:
                        # print(f'date3 = {date}') # temp
                        input_date = datetime.strptime(date, '%m/%d/%Y')
                        date = input_date.strftime('%Y-%m-%d')
                        year = input_date.strftime('%Y')
                        month = input_date.strftime('%m')
                        day = input_date.strftime('%d')
                except Exception as e:
                    print(f"Error converting date: {str(e)}")
                # print(f'date4 = {date}') # temp
                row_data["filename"] = filename
                row_data["date"] = date
                row_data["invoice"] = invoice
                row_data["billTo"] = billTo
                row_data["shipTo"] = shipTo
                row_data["quantity"] = quantity            
                row_data["description"] = description
                row_data["priceEach"] = priceEach
                row_data["subtotal"] = subtotal
                row_data["amount"] = amount
                row_data["total"] = total
                row_data["billToState"] = billToState
                row_data["shipToState"] = shipToState
                row_data["customer"] = customer
                row_data["state"] = state
                row_data["year"] = year
                row_data["month"] = month
                row_data["day"] = day
                row_data["folder"] = folder
                row_data["seller"] = seller
                row_data["disclaimer"] = disclaimer
                row_data["grand_total"] = grand_total
                row_data["shipping"] = shipping

                print(f'{filename}    {total}')   # test
                data.append(row_data)

    if tables_out:
        for idx, table in enumerate(all_tables):
            sheet_name = f'Table{idx+1}'
            table.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
        writer.close()

    write_xlsx(data)

def split_pdfs(input_folder, output_split_folder):
    '''
    Function to split PDFs into single-page PDFs
    '''
    # Create the output folder if it doesn't exist
    if not os.path.exists(output_split_folder):
        os.makedirs(output_split_folder)

        msg_blurb = (f'Creating split pdf folder: {output_split_folder}')
        msg_blurb_square(msg_blurb, color_green)

    msg_blurb = (f"Splitting PDFs from {input_folder}")
    msg_blurb_square(msg_blurb, color_green)
        
    for root, dirs, files in os.walk(input_folder):
        for filename in files:
            if filename.lower().endswith('.pdf'):  # Check for .pdf extension in a case-insensitive manner
                pdf_path = os.path.join(root, filename)
                with open(pdf_path, 'rb') as pdf_file:
                    reader = PyPDF2.PdfFileReader(pdf_file)
                    for page_num in range(reader.numPages):
                        writer = PyPDF2.PdfFileWriter()
                        writer.addPage(reader.getPage(page_num))
                        
                        # Construct output path maintaining subdirectory structure
                        relative_path = os.path.relpath(root, input_folder)
                        folder = relative_path
                        output_dir = os.path.join(output_split_folder, relative_path)
                        if not os.path.exists(output_dir):
                            os.makedirs(output_dir)

                        output_filename = f"{os.path.splitext(filename)[0]}_page_{page_num+1}.pdf"
                        output_path = os.path.join(output_dir, output_filename)
                        with open(output_path, 'wb') as output_pdf:
                            writer.write(output_pdf)
                        
                        print(f"Created: {output_path}")

    msg_blurb = (f"writing split pdfs to {output_split_folder}")
    msg_blurb_square(msg_blurb, color_green)

def table_dump1(input_folder, output_excel_path):
    """
    Loop through all PDFs in the specified directory (and optionally in subdirectories), extract tables, and export them to an Excel sheet.
    
    Parameters:
    input_folder (str): Path to the directory containing PDF files.
    output_excel_path (str): Path where the output Excel file will be saved.
    tables_out (bool): Whether to save the extracted tables into separate sheets in the Excel file.
    sub_folders (bool): Whether to include PDFs from subdirectories.
    """
    if tables_out:
        writer = pd.ExcelWriter(output_excel_path, engine='openpyxl')
        
    all_tables = []
    # data = []
    seller = input("SELLER: ")
    # Walk through directories and subdirectories if sub_folders is True
    for root, dirs, files in os.walk(input_folder) if sub_folders else [(input_folder, [], os.listdir(input_folder))]:
        for filename in files:
            if filename.lower().endswith('.pdf'):
                date, invoice, billTo, shipTo, quantity, description = '', '', '', '', '', ''
                priceEach, subtotal, amount, total, billToState, shipToState = '', '', '', '', '', ''
                customer, state, year, month, day, folder = '', '', '', '', '', ''
                
                row_data = {}
                pdf_path = os.path.join(root, filename)
                folder = pdf_path
                # tables = extract_tables_from_pdf(pdf_path)
                tables, text = extract_tables_from_pdf(pdf_path)
                all_tables.extend(tables)


    if tables_out:
        for idx, table in enumerate(all_tables):
            sheet_name = f'Table{idx+1}'
            table.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
        writer.close()

    # write_xlsx(data)    

def write_xlsx(data):

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Invoices'
    header_format = {'bold': True, 'border': True}
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = [
        "filename", "date", "invoice", "customer", "total", "billTo", "shipTo", "quantity", "description"
        , "priceEach", "amount", "billToState", "shipToState", "year", "month", "day", "folder", "seller"
        , "disclaimer", "grand_total", "shipping"
    ]
 
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in range(21):  # range(18) generates numbers from 0 to 25 inclusive orange
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange
            cell.fill = fill

    # Excel column width
    worksheet.column_dimensions['A'].width = 27 # 
    worksheet.column_dimensions['B'].width = 11 # 
    worksheet.column_dimensions['C'].width = 11 # 
    worksheet.column_dimensions['D'].width = 30 # 
    worksheet.column_dimensions['E'].width = 13 # 
    worksheet.column_dimensions['F'].width = 13 #   
    worksheet.column_dimensions['G'].width = 10 # 
    worksheet.column_dimensions['H'].width = 9 # 
    worksheet.column_dimensions['I'].width = 11 # 
    worksheet.column_dimensions['J'].width = 18 # 
    worksheet.column_dimensions['K'].width = 11  # 
    worksheet.column_dimensions['L'].width = 12  # 
    worksheet.column_dimensions['M'].width = 12  # 
    worksheet.column_dimensions['N'].width = 12  # 
    worksheet.column_dimensions['O'].width = 12  # 
    worksheet.column_dimensions['P'].width = 12  # 
    worksheet.column_dimensions['Q'].width = 30  # folder
    worksheet.column_dimensions['R'].width = 15  # seller
    worksheet.column_dimensions['S'].width = 11  # disclaimer
    worksheet.column_dimensions['T'].width = 11  # grand_total
    worksheet.column_dimensions['U'].width = 9  # shipping

    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                if isinstance(cell_data, list):
                    cell_data = str(cell_data)  # Convert lists to strings
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    msg_blurb = (f'Writing to {output_xlsx}')
    msg_blurb_square(msg_blurb, color_green)

    workbook.save(output_xlsx)

def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}insert your pdfs into the {input_folder} folder')
    print(f'\nExample:')
    print(f'    {file} -p') 
    print(f'    {file} -p -s')       
    # print(f'    {file} -p -t')     
    print(f'    {file} -p -I pdfs -O invoices_.xlsx ')     
    print(f'    {file} -S   # split pdfs into single page into the {output_split_folder} folder')     

if __name__ == "__main__":

    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
1.0.0 - added -S split and -s sub directories
0.1.5 - working prototype
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
-S split pdfs into seperate pages before you process them (-p)
can be adapated to different table column names
-t will optionaly print out the tables in different sheet
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
