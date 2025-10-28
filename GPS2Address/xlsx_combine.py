import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from copy import copy

def sanitize_sheet_name(filename):
    name = os.path.splitext(filename)[0]
    name = name.replace("FCM-Dump-com.", "")
    return name[:30]  # Truncate to 30 characters

def combine_xlsx_files(input_folder, output_filename, message_box):
    if not input_folder:
        input_folder = os.getcwd()
    if not output_filename:
        output_filename = "combined.xlsx"

    message_box.insert(tk.END, f"📁 Combining XLSX files in {input_folder}\n")

    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)  # Remove default sheet

    count = 0
    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx") and filename != output_filename:
            filepath = os.path.join(input_folder, filename)
            try:
                wb = load_workbook(filepath)
            except Exception as e:
                message_box.insert(tk.END, f"❌ Skipped (corrupted): {filename} — {str(e)}")
                continue

            for sheetname in wb.sheetnames:
                src_sheet = wb[sheetname]
                new_sheet_name = sanitize_sheet_name(filename)
                new_sheet = combined_wb.create_sheet(title=new_sheet_name)

                for row in src_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.fill = copy(cell.fill)
                            new_cell.border = copy(cell.border)
                            new_cell.number_format = cell.number_format
                        # Set alignment to top-left
                        new_cell.alignment = Alignment(vertical="top", horizontal="left")

                        # Format header row
                        if cell.row == 1:
                            new_cell.font = Font(bold=True)
                            new_cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

                for col in src_sheet.column_dimensions:
                    new_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width

                # Freeze cell B2
                new_sheet.freeze_panes = "B2"

            count += 1
            message_box.insert(tk.END, f"✅ Combined: {filename}")

    combined_wb.save(os.path.join(input_folder, output_filename))
    message_box.insert(tk.END, f"\n📊 Total files processed: {count}")

def start_combining():
    input_folder = input_entry.get()
    output_filename = output_entry.get()
    message_box.delete(0, tk.END)
    combine_xlsx_files(input_folder, output_filename, message_box)

# GUI setup
root = tk.Tk()
root.title("XLSX Combiner")

tk.Label(root, text="📎 Combine xlsx files into one sheet.", font=("Arial", 10, "bold")).pack(pady=(5, 0))
tk.Label(root, text="Input Folder defaults to current folder", font=("Arial", 9)).pack(pady=(0, 5))

frame = tk.Frame(root)
frame.pack(pady=5)

tk.Label(frame, text="Input Folder:").grid(row=0, column=0, sticky="e")
input_entry = tk.Entry(frame, width=50)
input_entry.grid(row=0, column=1)
tk.Button(frame, text="Browse", command=lambda: input_entry.insert(0, filedialog.askdirectory())).grid(row=0, column=2)

tk.Label(frame, text="Output File Name:").grid(row=1, column=0, sticky="e")
output_entry = tk.Entry(frame, width=50)
output_entry.insert(0, "combined.xlsx")
output_entry.grid(row=1, column=1)

tk.Button(root, text="Combine", command=start_combining, bg="lightblue").pack(pady=10)

message_box = tk.Listbox(root, width=80, height=15)
message_box.pack(pady=5)

root.mainloop()