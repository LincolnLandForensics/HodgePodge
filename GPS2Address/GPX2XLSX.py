#!/usr/bin/env python3
# coding: utf-8
"""
Parse (Garmin) GPX files and write results to Excel. 
Author: LincolnLandForensics
Version: 1.0.0
"""


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import re
import sys
import zipfile
import argparse
import gpxpy  # pip install gpxpy
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import pandas as pd
import xml.etree.ElementTree as ET

# Optional: Color support for Windows terminal
color_red = color_yellow = color_green = color_blue = color_purple = color_reset = ''
if sys.version_info > (3, 7, 9) and os.name == "nt":
    from colorama import Fore, Back, Style
    print(Back.BLACK)
    color_red, color_yellow, color_green = Fore.RED, Fore.YELLOW, Fore.GREEN
    color_blue, color_purple, color_reset = Fore.BLUE, Fore.MAGENTA, Style.RESET_ALL

# Global variables
gpx_folder = "GPX/"
outuput_xlsx = "gpx_output.xlsx"
data = []
input_details = 'no'
row = 0

def main():
    banner_print()

    parser = argparse.ArgumentParser(description="Parse GPX/KML and export to Excel")
    parser.add_argument('-g', '--gpx', help='Input .gpx file')
    parser.add_argument('-O', '--output', help='Output .xlsx file')
    parser.add_argument('-G', '--gpx_folder', help='Parse all .gpx files in Logs/ folder', action='store_true')
    parser.add_argument('-k', '--kml', help='Input .kml or .kmz file')

    args = parser.parse_args()

    global input_file, outuput_xlsx, log_type

    outuput_xlsx = args.output if args.output else "output.xlsx"

    if args.gpx_folder:
        log_type = 'folder'
        parse_gpx()

    elif args.gpx:
        input_file = args.gpx
        log_type = 'file'
        parse_gpx()

    elif args.kml:
        input_file = args.kml
        log_type = 'file'

        if input_file.lower().endswith('.kmz'):
            print("[*] Extracting KML from KMZ...")
            try:
                kml_file = extract_kml_from_kmz(input_file)
                data = parse_kml(kml_file)
            except Exception as e:
                print(f"{color_red}Error extracting KML: {e}{color_reset}")
                return
        else:
            data = parse_kml(input_file)

        write_xlsx(data, outuput_xlsx)
        print(f"{color_green}Exported {len(data)} KML records to {outuput_xlsx}{color_reset}")

    else:
        usage()

    
def banner_print():
    art = """  
  ______________________  ___ ________          .__                  
 /  _____/\______   \   \/  / \_____  \  ___  __|  |   _________  ___
/   \  ___ |     ___/\     /   /  ____/  \  \/  /  |  /  ___/\  \/  /
\    \_\  \|    |    /     \  /       \   >    <|  |__\___ \  >    < 
 \______  /|____|   /___/\  \ \_______ \ /__/\_ \____/____  >/__/\_ \

    """
    print(f"{color_blue}{art}{color_reset}")

def case_number_prompt():
    return input("Please enter the case: ").strip()

def extract_kml_from_kmz(kmz_path):
    with zipfile.ZipFile(kmz_path, 'r') as zf:
        for file in zf.namelist():
            if file.endswith('.kml'):
                zf.extract(file, path='.')
                return file
    raise Exception("No KML file found in KMZ.")

def gpx_extract(filename, caseNumber):
    """
    Extract data from a GPX file (1.0 or 1.1) including trkpt, wpt, rtept.
    Supports speed and heading if present.
    """
    print(f'{color_blue}{filename}{color_reset}')
    
    # Sanitize the GPX file by stripping invalid attributes
    with open(filename, 'r', encoding='utf-8') as f:
        raw_xml = f.read()

    # Optional: Remove problematic <script> blocks or malformed tags
    
    raw_xml = re.sub(r'<script.*?>.*?</script>', '', raw_xml, flags=re.DOTALL)
    raw_xml = re.sub(r'crossorigin\b(?!\=)', 'crossorigin="anonymous"', raw_xml)

    try:
        gpx = gpxpy.parse(raw_xml)
    except gpxpy.gpx.GPXXMLSyntaxException as e:
        print(f"{color_red}Invalid GPX format in {filename}: {e}{color_reset}")
        return []

    with open(filename, 'r', encoding='utf-8') as gpx_file:
        gpx = gpxpy.parse(gpx_file)

    data = []

    def process_point(pt, point_type):
        time_str = pt.time.isoformat() if pt.time else ''
        time_no_T = time_str.replace('T', ' ')
        time_part, Timezone = time_no_T.rsplit('-', 1) if '-' in time_no_T else (time_no_T, '')
        time_readable = pt.time.strftime('%Y-%m-%d %H:%M:%S') if pt.time else ''

        # Initialize
        street = Description = speed = heading = desc = name = address = city = state = zipcode = ''

        # These are standard GPX fields
        Description = pt.description or ''
        Name = pt.name or ''
        address = getattr(pt, 'address', '')  # GPX doesn't standardize 'address', but some exports include it



        if pt.extensions:
            for ext in pt.extensions:
                speed_elem = ext.find(".//{*}speed")
                heading_elem = ext.find(".//{*}course")
                # desc_elem = ext.find(".//{*}desc")
                # name_elem = ext.find(".//{*}name")
                address_elem = ext.find(".//{*}address")
                city_elem = ext.find(".//{*}City")
                state_elem = ext.find(".//{*}State")
                zip_elem = ext.find(".//{*}PostalCode")
                street_elem = ext.find(".//{*}StreetAddress")
                
                if speed_elem is not None:
                    speed = speed_elem.text
                if heading_elem is not None:
                    heading = heading_elem.text
                # if desc_elem is not None:
                    # desc = desc_elem.text
                # if name_elem is not None:
                    # name = name_elem.text
                if address_elem is not None:
                    address = address_elem.text
                if city_elem is not None:
                    city = city_elem.text
                if state_elem is not None:
                    state = state_elem.text
                if zip_elem is not None:
                    zipcode = zip_elem.text
                if street_elem is not None:
                    street = street_elem.text

                # if desc_elem is not None:
                    # Description = desc_elem.text
                    # print(f'Description = {Description}')   # temp

        row_data = {
            "Name": Name,
            "Time": time_readable,
            "Latitude": pt.latitude,
            "Longitude": pt.longitude,
            "Address": address,
            "street": street,            
            "city": city,
            "state": state,
            "zipcode": zipcode,
            "Description": Description,
            "Type": "GPX-" + point_type,
            "Direction": heading,
            "Timezone": Timezone,
            "speed": speed,
            "Altitude": pt.elevation,
            "Coordinate": f"{pt.latitude}, {pt.longitude}",
            # "Icon": "Yellow",
            "original_file": os.path.basename(filename),
            "case": caseNumber,
        }
        
        data.append(row_data)
        if address:
            print(f'address = {address}')   # temp



    # Trackpoints
    for track in gpx.tracks:
        for segment in track.segments:
            for pt in segment.points:
                process_point(pt, "trkpt")

    # Waypoints
    for wpt in gpx.waypoints:
        process_point(wpt, "wpt")

    # Routepoints
    for rte in gpx.routes:
        for rtept in rte.points:
            process_point(rtept, "rtept")

    return data


def msg_blurb_square(msg, color):
    border = f"+{'-' * (len(msg) + 2)}+"
    print(f"{color}{border}\n| {msg} |\n{border}{color_reset}")

def parse_gpx():
    '''
    Parse one or more GPX files and export to XLSX.
    '''
    caseNumber = case_number_prompt()
    logs_list = []

    if log_type == 'file':
        logs_list = [input_file]
        
    elif log_type == 'folder':
        if not os.path.exists(gpx_folder):
            print(f"{color_red}{gpx_folder} does not exist.{color_reset}")
            os.makedirs(gpx_folder)
            print(f"{color_yellow}Created {gpx_folder}. Add .gpx files and re-run.{color_reset}")
            return

        gpx_files = [f for f in os.listdir(gpx_folder) if f.lower().endswith('.gpx')]
        if not gpx_files:
            print(f"{color_red}Error: No .gpx files found in {gpx_folder}.{color_reset}")
            sys.exit(1)

        logs_list = [os.path.join(gpx_folder, f) for f in gpx_files]
       
    all_data = []

    for logFile in logs_list:
        # msg_blurb_square(f"Reading {logFile}", color_green)
        parsed = gpx_extract(logFile, caseNumber)
        if parsed:
            all_data.extend(parsed)
        else:
            print(f"{color_red}  No points found in {logFile}{color_reset}")

    if all_data:
        write_xlsx(all_data, outuput_xlsx)
        print(f"{color_green}Saved {len(all_data)} points to {outuput_xlsx}{color_reset}")
    else:
        print(f"{color_red}No data parsed. Check GPX files.{color_reset}")

def parse_kml(file_path):
    import xml.etree.ElementTree as ET

    ns = {
        'kml': 'http://www.opengis.net/kml/2.2',
        'gx': 'http://www.google.com/kml/ext/2.2'
    }

    tree = ET.parse(file_path)
    root = tree.getroot()
    placemarks = root.findall('.//kml:Placemark', ns)

    data = []

    for pm in placemarks:
        name = pm.findtext('kml:name', default='', namespaces=ns)
        description = pm.findtext('kml:description', default='', namespaces=ns)

        # Extract gx:Track if present
        track_elem = pm.find('.//gx:Track', ns)
        if track_elem is not None:
            whens = [w.text for w in track_elem.findall('gx:when', ns)]
            coords = [c.text.strip() for c in track_elem.findall('gx:coord', ns)]

            for i in range(min(len(whens), len(coords))):
                try:
                    lon, lat, *alt = map(float, coords[i].split())
                except:
                    continue

                row = {
                    '#': name,
                    'Time': whens[i],
                    'Latitude': lat,
                    'Longitude': lon,
                    'Altitude': alt[0] if alt else '',
                    'Direction': '',              # can be added if found in ExtendedData
                    'speed': '',
                    'AccuracyMeters': '',
                    'Name': '',
                    'Description': description,
                    'Source file information': os.path.basename(file_path)
                }
                data.append(row)
            continue  # done with this Placemark

        # Fallback for Point-based Placemarks
        coord_elem = pm.find('.//kml:Point/kml:coordinates', ns)
        if coord_elem is not None:
            try:
                lon, lat, *alt = map(float, coord_elem.text.strip().split(','))
            except:
                continue

            time = pm.findtext('.//kml:TimeStamp/kml:when', default='', namespaces=ns)

            ext_data = {}
            for data_tag in pm.findall('.//kml:ExtendedData//kml:Data', ns):
                key = data_tag.attrib.get('name', '').lower()
                value_elem = data_tag.find('kml:value', ns)
                value = value_elem.text if value_elem is not None else ''
                ext_data[key] = value

            row = {
                '#': name,
                'Time': time,
                'Latitude': lat,
                'Longitude': lon,
                'Altitude': alt[0] if alt else '',
                'speed': ext_data.get('speed', ''),
                'Direction': ext_data.get('heading') or ext_data.get('direction', ''),
                'AccuracyMeters': ext_data.get('accuracy', ''),
                'Name': ext_data.get('sensor') or ext_data.get('name', ''),
                'Description': description,
                'Source file information': os.path.basename(file_path),
            }

            data.append(row)

    return data


def write_xlsx(data, file_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sheet1'
    worksheet.freeze_panes = 'B2'

    # headers = [
        # "case", "Time", "Latitude", "Longitude", "Type", "City", "county",
        # "state", "zipcode", "Coordinate", "Time Original", "Icon", "original_file", "Altitude"
    # ]

    headers = [
        "#", "Time", "Latitude", "Longitude", "Address", "Group", "Subgroup"
        , "Description", "Type", "Source", "Deleted", "Tag"
        , "Source file information", "Service Identifier", "Carved", "Name"
        , "business", "number", "street", "city", "county", "state", "zipcode"
        , "country", "fulladdress", "query", "Sighting State", "Plate"
        , "Capture Time", "Capture Network", "Highway Name", "Coordinate"
        , "Capture Location Latitude", "Capture Location Longitude", "Container"
        , "Sighting Location", "Direction", "Time Local", "End time", "Category"
        , "Manually decoded", "Account", "PlusCode", "Time Original", "Timezone"
        , "Icon", "original_file", "case", "Origin Latitude", "Origin Longitude"
        , "Start Time", "Azimuth", "Radius", "Altitude", "Location"
        , "time_orig_start", "timezone_start", "Index", "speed", "parked", "MAC"
    ]

    # Write headers
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if header in ["Time", "Latitude", "Longitude"]:
            fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            cell.fill = fill

    # Write data
    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            worksheet.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name)

    workbook.save(file_path)

def usage():
    print(f"Usage: {sys.argv[0]} -G|-I file.gpx [-O gpx__output.xlsx]")
    print("Examples:")
    print(f"    {sys.argv[0]} -g single.gpx -O gpx__output.xlsx")
    print(f"    {sys.argv[0]} -G -O gpx_merged.xlsx")
    print(f"    {sys.argv[0]} -k single.kmz -O kmz__output.xlsx")

if __name__ == '__main__':
    main()
