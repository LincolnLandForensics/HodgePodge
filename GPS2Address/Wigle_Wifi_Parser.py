#!/usr/bin/python
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import re
import sys
import csv
import gzip
import shutil



import argparse  # for menu system
# import xlsxwriter
import openpyxl
from openpyxl import load_workbook, Workbook    # pip install openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "convert wigle .gz or .csv exports to gps2address.py locations format"
version = '1.0.2'

global headers
headers = [
    "#", "Time", "Latitude", "Longitude", "Address", "Group", "Subgroup"
    , "Description", "Type", "Source", "Deleted", "Tag"
    , "Source file information", "Service Identifier", "Carved", "Name"
    , "business", "number", "street", "city", "county", "state", "zipcode"
    , "country", "fulladdress", "query", "Sighting State", "Plate"
    , "Capture Time", "Capture Network", "Highway Name", "Coordinate"
    , "Capture Location Latitude", "Capture Location Longitude", "Container"
    , "Sighting Location", "Direction", "Time Local", "End time", "Category"
    , "Manually decoded", "Account", "PlusCode", "Time Original", "Timezone"
    , "Icon", "original_file", "case", "Origin Latitude", "Origin Longitude"
    , "Start Time", "Azimuth", "Radius", "Altitude", "Location"
    , "time_orig_start", "timezone_start", "Index", "speed", "parked"
    , "MAC", "SSID", "AuthMode", "Channel", "Frequency", "RSSI"
    , "AltitudeMeters", "AccuracyMeters", "RCOIs", "MfgrId", "CompanyName"
]


# Colorize section
global color_red
global color_green
global color_reset
color_green = ''
color_green = ''
color_red = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
  
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL
        

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global Row
    Row = 1  # defines arguments
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False)
    # parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-b','--blank', help='create blank sheet', required=False, action='store_true')    
    parser.add_argument('-w', '--wigleparse', help='parse wigle file csv', required=False, action='store_true')

    args = parser.parse_args()

    global input_file
    input_file = args.input if args.input else "WigleWifi_sample.csv"

    global output_xlsx
    # output_xlsx = args.output if args.output else "WigleWifi_sample.xlsx"

    if args.wigleparse:
        data = []
        process_wigle_file(input_file)
    elif args.blank:
        data = []
        write_xlsx(data)
        return 0 
        sys.exit()
    else:
        usage()
        
    # workbook.close()
    # Workbook.close()    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def format_function(bg_color='white'):
    global format
    format = workbook.add_format({
        'bg_color': bg_color
    })


def message_square(message, color):
    horizontal_line = f"+{'-' * (len(message) + 2)}+"
    empty_line = f"| {' ' * (len(message))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {message} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def MfgrId2Company(MfgrId):
    try:
        MfgrId = int(MfgrId)  # Ensure it's an integer
    except ValueError:
        return ""

    MfgrId = int(MfgrId)
    # print(f'MfgrId = type {type(MfgrId)}')  # temp
    
    mfgr_dict = {
        6: "Xerox Corporation",
        13: "XEROX CORPORATION",
        76: "Compaq Computer Corporation",
        87: "Hewlett-Packard Company",
        89: "International Business Machines",
        93: "VisiCorp",
        96: "AT&T",
        117: "Apple Computer Inc.",
        135: "Sun Microsystems",
        137: "Advanced Micro Devices",
        217: "Silicon Graphics, Inc.",
        263: "Cisco Systems, Inc.",
        283: "TOSHIBA CORPORATION",
        301: "Unisys",
        315: "Hewlett-Packard Company",
        369: "Intel Corporation",
        529: "Apple Computer Inc.",
        767: "Cisco Systems, Inc.",
        12849: "Cisco Systems, Inc.",
        1363: "Cisco Systems, Inc.",
        1494: "Cisco Systems, Inc.",
        1536: "Cisco Systems, Inc.",
        1704: "Cisco Systems, Inc.",
        1736: "Cisco Systems, Inc.",
        1744: "Cisco Systems, Inc.",
        1993: "Cisco Systems, Inc.",
        2409: "Cisco Systems, Inc.",
        2504: "Cisco Systems, Inc.",
        27475: "Cisco Systems, Inc.",
        29439: "Cisco Systems, Inc.",
        34817: "Cisco Systems, Inc.",
        34818: "Cisco Systems, Inc.",
        40857: "Cisco Systems, Inc.",
        42239: "Cisco Systems, Inc.",
        48872: "Cisco Systems, Inc.",
        61680: "Cisco Systems, Inc.",
        65024: "Cisco Systems, Inc.",
        65535: "Cisco Systems, Inc.",
    }
    return mfgr_dict.get(MfgrId, "")
    
    
def process_wigle_file(filename):
    if filename.endswith('.gz'):
        unzipped_filename = filename[:-3]  # Remove .gz extension
        # output_xlsx = (f'{unzipped_filename}.xlsx')

        try:
            with gzip.open(filename, 'rb') as f_in:
                with open(unzipped_filename, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            message = (f'Unzipped: {filename} -> {unzipped_filename}')
            message_square(message, color_green)
            filename = unzipped_filename
        except Exception as e:
            message = (f'Error unzipping file: {e}')
            message_square(message, color_red)            
            return
    
    if not os.path.isfile(filename):
        print(f"Error: File '{filename}' not found or is not a valid file.")
        message = (f"Error: File {filename} not found or is not a valid file.")
        message_square(message, color_red)        
        return
    
    if not filename.startswith('WigleWifi') or not filename.endswith('.csv'):
        message = (f"Invalid file: Filename must start with 'WigleWifi' and end with '.csv'")
        message_square(message, color_red)        
        return
    else:
        data = []
        csv_file = open(filename)
        source_file = filename
        row_count = 0
        for row in csv_file:
            row = row.split(',')
            row_data = {}
            description, group, subgroup, type_data, Name, Type = '', '', '', '', '', ''
            tag, CompanyName, country, source, Icon = '', '', '', 'Wigle', ''
       
            try:
                MAC = row[0] if len(row) > 0 else ''
                SSID = row[1] if len(row) > 1 else ''
                AuthMode = row[2] if len(row) > 2 else ''
                Time = row[3] if len(row) > 3 else ''
                Channel = row[4] if len(row) > 4 else ''
                Frequency = row[5] if len(row) > 5 else ''
                RSSI = row[6] if len(row) > 6 else ''
                latitude = row[7] if len(row) > 7 else ''
                longitude = row[8] if len(row) > 8 else ''
                AltitudeMeters = row[9] if len(row) > 9 else ''
                AccuracyMeters = row[10] if len(row) > 10 else ''
                RCOIs = row[11] if len(row) > 11 else ''
                MfgrId = row[12] if len(row) > 12 else ''
                subgroup = row[13] if len(row) > 13 else ''
                Type = row[13] if len(row) > 13 else ''
                
            except Exception as e:
                print(f"Error processing line : {e}")

            if MfgrId != '':
                CompanyName = MfgrId2Company(MfgrId)
            if latitude != '' and longitude != '':
                Coordinate = (f'{latitude},{longitude}')

            subgroup = subgroup.strip()    
            Type = Type.strip()
            # SSID = SSID.replace('\"', '')

            if AuthMode == 'LTE;us':
                Type = 'Tower'
                subgroup = 'LTE'            
                country = 'US'
                Icon = 'Tower'

            elif ' tv' in SSID.lower() or '(tv)' in SSID.lower():  # todo doesn't match (tv)
                subgroup = 'Display'
                Type = 'Display/Speaker'
                Icon = 'BT'

            elif 'LTE;' in AuthMode:
                subgroup = 'LTE' 
                Type = 'Tower'
                Icon = 'Tower'
            elif AuthMode == 'GSM':
                Type = 'Tower'
                subgroup = 'GSM' 
                Icon = 'Tower'
  
            elif "Desktop;" in AuthMode:
                subgroup = 'Desktop'
                Type = 'Desktop'
                Icon = 'BT'
            elif 'Display/Speaker' in AuthMode:
                Type = 'Display/Speaker'
                Icon = 'BT'
                if 'speaker' in SSID.lower():
                    subgroup = 'Speaker'
                elif ' tv ' in SSID.lower():
                    subgroup = 'Display'
                else:
                    subgroup = 'Display/Speaker'   
            elif 'speaker' in SSID.lower() or 'soundbar' in SSID.lower():
                Type = 'Display/Speaker'
                subgroup = 'Speaker'
                Icon = 'BT'




            elif "(oven)" in SSID.lower():
                subgroup = 'Oven'
                # Type = 'Display/Speaker'
                Icon = 'BT'


                    
            # elif Type == 'BT': # todo
            elif "BT" in Type or "BLE" in Type:
                subgroup = Type
                Type = 'BlueTooth'
                Icon = 'BT'
                if 'oven' in SSID.lower():
                    subgroup = 'Oven'
                elif 'qled' in SSID.lower():
                    subgroup = 'Display'                
                    Type = 'Display/Speaker'
                elif 'sound' in SSID.lower():
                    subgroup = 'Speaker'                
                    Type = 'Display/Speaker'
                elif 'officejet' in SSID.lower() or 'deskjet' in SSID.lower():
                    subgroup = 'Printer'                
                elif 'dryer' in SSID.lower():
                    subgroup = 'Dryer' 
                elif 'washer' in SSID.lower():
                    subgroup = 'Washer' 
                elif 'lamp' in SSID.lower() or ' light' in SSID.lower():
                    subgroup = 'Light' 
                elif 'tv' in SSID.lower():
                    Type = 'Display/Speaker'
                    subgroup = 'Display' 

            elif "WPA2" in AuthMode:
                subgroup = 'WPA2'
                Icon = 'WIFI'
                Type = 'WIFI'
            elif AuthMode == "[ESS]":
                Type = 'WIFI'
                subgroup = 'ESS'
                Icon = 'WIFI-open' 
                if 'officejet' in SSID.lower():
                    subgroup = 'Printer'
                if 'thermostat' in SSID.lower():
                    subgroup = 'Thermostat'                    
                    
                    
                    
            SSID = sanitize_string(SSID)
            type_data = Type
            
            if Icon == '':
                Icon = Type
            subgroup = subgroup.strip()
            group = group.strip()
            
            if SSID != '':
                description = (f'{description}\nSSID:{SSID}')            
            if MAC != '':
                description = (f'{description}\nMAC:{MAC}')             
            if AuthMode != '':
                description = (f'{description}\nAuthMode:{AuthMode}')             
            if Frequency != '':
                description = (f'{description}\nFrequency:{Frequency}')             
            if Channel != '':
                description = (f'{description}\nChannel:{Channel}')             
            if RSSI != '':
                description = (f'{description}\nReceived Signal Strength:{RSSI}')           
            if Type != '':
                description = (f'{description}\nType:{Type}')
            description = description.strip()   
            if group != '':
                description = (f'{description}\nGroup:{group}')            
            if subgroup != '':
                description = (f'{description}\nSubgroup:{subgroup}')                

            if CompanyName != '':
                description = (f'{description}\nCompanyName:{CompanyName}')

            # Apply sanitization
            description = sanitize_string(description) 
            description = description.strip()


 
            if Coordinate != 'Latitude,Longitude':
                row_count += 1
                row_data["#"] = SSID
                row_data["Time"] = Time
                row_data["Latitude"] = latitude
                row_data["Longitude"] = longitude 
                # row_data["Address"] = address
                row_data["Group"] = group
                row_data["Subgroup"] = subgroup
                row_data["Description"] = description
                row_data["Type"] = type_data
                # row_data["Tag"] = tag        
                row_data["Source"] = source
                row_data["Source file information"] = source_file
                row_data["Name"] = Name
                # row_data["business"] = business 
                # row_data["number"] = number 
                # row_data["street"] = street
                # row_data["city"] = city 
                # row_data["county"] = county 
                # row_data["state"] = state 
                # row_data["zipcode"] = zipcode
                row_data["country"] = country 
                # row_data["fulladdress"] = fulladdress
                # row_data["query"] = query
                # row_data["Plate"] = plate
                # row_data["Capture Time"] = capture_time
                # row_data["Highway Name"] = hwy
                row_data["Coordinate"] = Coordinate
                # row_data["Direction"] = direction
                # row_data["End time"] = end_time
                # row_data["Category"] = category
                # row_data["Time Original"] = time_orig
                # row_data["Timezone"] = timezone
                # row_data["PlusCode"] = PlusCode
                # row_data["Radius"] = Radius
                row_data["Icon"] = Icon
                row_data["original_file"] = filename # test


                row_data["MAC"] = MAC
                row_data["SSID"] = SSID
                row_data["AuthMode"] = AuthMode
                row_data["Channel"] = Channel  
                row_data["Frequency"] = Frequency        
                row_data["RSSI"] = RSSI
                row_data["AltitudeMeters"] = AltitudeMeters
                row_data["AccuracyMeters"] = AccuracyMeters
                row_data["RCOIs"] = RCOIs  
                row_data["filename"] = filename
                row_data["MfgrId"] = MfgrId 
                row_data["CompanyName"] = CompanyName 
                

                data.append(row_data)
        print(f'Processed {row_count} rows')
        
        # data = remove_duplicate_macs(data)
        global output_xlsx
        output_xlsx = filename
        output_xlsx = output_xlsx.replace('.csv', '.xlsx')  # task
        
        
        print(f'filename = {filename}   outputxlsx = {output_xlsx}')    # temp
        write_xlsx(data)


def remove_duplicate_macs(data):
    """
    Removes rows with duplicate MAC addresses and returns unique data.
    
    :param data: List of dictionaries containing the dataset
    :return: Filtered list with unique MAC addresses
    """
    unique_macs = set()
    filtered_data = []
    row_count = 0
    for row in data:
        mac_address = row.get("MAC", "").strip()

        if mac_address and mac_address not in unique_macs:
            unique_macs.add(mac_address)
            filtered_data.append(row)
            row_count += 1
    print(f'found {row_count} unique MACs')
    return filtered_data

        

def sanitize_string(text):
    # Remove control characters (non-printable ASCII)
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    
    # Optionally replace problematic characters like brackets
    text = text.replace("[", "(").replace("]", ")")
    
    return text.strip()


def write_xlsx(data):
    '''
    The write_locations() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''
    message = (f'Writing {output_xlsx}')
    message_square(message, color_green)

    try:
        data = sorted(data, key=lambda x: (x.get("SSID", ""), x.get("RSSI", ""), x.get("MAC", "")))
        print(f'Sorting by MAC with the strongest signal')
    except TypeError as error:
        print(f'{color_red}{error}{color_reset}')

    data = remove_duplicate_macs(data)

    try:
        data = sorted(data, key=lambda x: (x.get("Time", ""), x.get("MAC", "")))
        print(f'Sorting by Time')
    except TypeError as error:
        print(f'{color_red}{error}{color_reset}')    
    
    
    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Intel'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    log_headers = [
        "Date", "Subject", "Requesting Agency", "Requesting Agent", "Case"
        , "Summary of Findings", "Source", "Notes"
    ]


    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [3, 4, 5, 6, 49, 50]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [7,8, 13, 14, 15, 29, 30, 35, 36, 37, 38, 39, 40, 41, 42, 43]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill
        # elif col_index == 27:  # Red for column 27
            # fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
            # cell.fill = fill

    ## Excel column width

    worksheet.column_dimensions['A'].width = 25 # 
    worksheet.column_dimensions['B'].width = 20 # 
    worksheet.column_dimensions['C'].width = 20 # 
    worksheet.column_dimensions['D'].width = 25 # 
    worksheet.column_dimensions['E'].width = 1 # 
    worksheet.column_dimensions['F'].width = 10 # Group
    worksheet.column_dimensions['G'].width = 14 # Subgroup
    worksheet.column_dimensions['H'].width = 35 # Description
    worksheet.column_dimensions['I'].width = 15 # Type
    worksheet.column_dimensions['J'].width = 8 # Source
    worksheet.column_dimensions['K'].width = 1 # Deleted
    worksheet.column_dimensions['L'].width = 1 # Tag
    worksheet.column_dimensions['M'].width = 20 # Source file information
    worksheet.column_dimensions['N'].width = 1 # 
    worksheet.column_dimensions['O'].width = 1 #
    worksheet.column_dimensions['P'].width = 1 # 
    worksheet.column_dimensions['Q'].width = 1 # 
    worksheet.column_dimensions['R'].width = 1 # 
    worksheet.column_dimensions['S'].width = 1 # 
    worksheet.column_dimensions['T'].width = 1 #
    worksheet.column_dimensions['U'].width = 1 # 
    worksheet.column_dimensions['V'].width = 1 # 
    worksheet.column_dimensions['W'].width = 1 # 
    worksheet.column_dimensions['X'].width = 1 # 
    worksheet.column_dimensions['Y'].width = 1 # 
    worksheet.column_dimensions['Z'].width = 1 # 
    worksheet.column_dimensions['AA'].width = 1 # 
    worksheet.column_dimensions['AB'].width = 1 # 
    worksheet.column_dimensions['AC'].width = 1 # 
    worksheet.column_dimensions['AD'].width = 1 # 
    worksheet.column_dimensions['AE'].width = 1 # 
    worksheet.column_dimensions['AF'].width = 22 # Coordinate
    worksheet.column_dimensions['AG'].width = 1 # 
    worksheet.column_dimensions['AH'].width = 1 # 
    worksheet.column_dimensions['AI'].width = 1 # 
    worksheet.column_dimensions['AJ'].width = 1 # 
    worksheet.column_dimensions['AK'].width = 1 # 
    worksheet.column_dimensions['AL'].width = 1 # 
    worksheet.column_dimensions['AM'].width = 1 # 
    worksheet.column_dimensions['AN'].width = 1 # 
    worksheet.column_dimensions['AO'].width = 1 # 
    worksheet.column_dimensions['AP'].width = 1 # 
    worksheet.column_dimensions['AQ'].width = 1 # 
    worksheet.column_dimensions['AR'].width = 1 # 
    worksheet.column_dimensions['AS'].width = 1 # 
    worksheet.column_dimensions['AT'].width = 12 # icon
    worksheet.column_dimensions['AU'].width = 22 # original_file
    worksheet.column_dimensions['AV'].width = 1 # 
    worksheet.column_dimensions['AW'].width = 1 # 
    worksheet.column_dimensions['AX'].width = 1 # 
    worksheet.column_dimensions['AY'].width = 1 # 
    worksheet.column_dimensions['AZ'].width = 1 # 
    worksheet.column_dimensions['BA'].width = 1 # 
    worksheet.column_dimensions['BB'].width = 1 # 
    worksheet.column_dimensions['BC'].width = 1 # 
    worksheet.column_dimensions['BD'].width = 1 # 
    worksheet.column_dimensions['BE'].width = 1 # 
    worksheet.column_dimensions['BF'].width = 1 # 
    worksheet.column_dimensions['BG'].width = 1 # 
    worksheet.column_dimensions['BH'].width = 1 # 


    worksheet.column_dimensions['BI'].width = 20 # MAC
    worksheet.column_dimensions['BJ'].width = 15 # SSID
    
    worksheet.column_dimensions['BS'].width = 30 # companyName

    for i in range(len(data)):
        if data[i] is None:
            data[i] = ''


    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            try:
                cell_data = row_data.get(col_name)
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    # Create a new worksheet for color codes
    color_worksheet = workbook.create_sheet(title='ColorCode')
    color_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    color_worksheet.column_dimensions['A'].width = 14# Color
    color_worksheet.column_dimensions['B'].width = 20# Description


    # Excel row height
    color_worksheet.row_dimensions[2].height = 22  # Adjust the height as needed
    color_worksheet.row_dimensions[3].height = 22
    color_worksheet.row_dimensions[4].height = 23
    color_worksheet.row_dimensions[5].height = 23
    color_worksheet.row_dimensions[6].height = 40   # truck

    color_worksheet.cell(row=1, column=1).value = 'Color'
    color_worksheet.cell(row=1, column=2).value = 'description'
    color_worksheet.cell(row=2, column=1).value = 'Red'
    color_worksheet.cell(row=3, column=1).value = 'Orange'
    color_worksheet.cell(row=4, column=1).value = 'Green'
    color_worksheet.cell(row=5, column=1).value = 'Yellow'

    color_worksheet.cell(row=7, column=1).value = 'ABBREVIATIONS'
    color_worksheet.cell(row=8, column=1).value = 'AKA'
    color_worksheet.cell(row=9, column=1).value = 'DOB'
    color_worksheet.cell(row=10, column=1).value = 'VIS'
    color_worksheet.cell(row=11, column=1).value = 'VIN'
    color_worksheet.cell(row=12, column=1).value = 'VYR'
    color_worksheet.cell(row=13, column=1).value = 'VMA'
    color_worksheet.cell(row=14, column=1).value = 'LIC'
    color_worksheet.cell(row=15, column=1).value = 'LIY'
    color_worksheet.cell(row=16, column=1).value = 'DLN'
    color_worksheet.cell(row=17, column=1).value = 'DLS'

       
    color_worksheet.cell(row=2, column=2).value = 'Bad Intel or dead link'
    color_worksheet.cell(row=3, column=2).value = 'Research'
    color_worksheet.cell(row=4, column=2).value = 'Good Intel'
    color_worksheet.cell(row=5, column=2).value = 'Highlighted'

    color_worksheet.cell(row=8, column=2).value = 'Also Known As (Alias)'
    color_worksheet.cell(row=9, column=2).value = 'Date of Birth'
    color_worksheet.cell(row=10, column=2).value = 'Vehicle State'
    color_worksheet.cell(row=11, column=2).value = 'Vehicle Identification Number'
    color_worksheet.cell(row=12, column=2).value = 'Vehicle Year'
    color_worksheet.cell(row=13, column=2).value = 'Vehicle Make'
    color_worksheet.cell(row=14, column=2).value = 'License'
    color_worksheet.cell(row=15, column=2).value = 'License Year'
    color_worksheet.cell(row=16, column=2).value = 'Drivers License Number'
    color_worksheet.cell(row=17, column=2).value = 'Drivers License State'


    # colored fills
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    # Apply the orange fill to the cell in row 2, column 2
    color_worksheet.cell(row=2, column=2).fill = red_fill
    color_worksheet.cell(row=3, column=2).fill = orange_fill
    color_worksheet.cell(row=4, column=2).fill = green_fill
    color_worksheet.cell(row=5, column=2).fill = yellow_fill


    # Create a new worksheet for logs
    log_worksheet = workbook.create_sheet(title='Log')
    log_worksheet.freeze_panes = 'B2'  # Freeze cells

# Date, Subject, Requesting Agency, Requesting Agent, Case, Summary of Findings, Source, Notes, Requestor

    # Excel column width
    log_worksheet.column_dimensions['A'].width = 14# Date
    log_worksheet.column_dimensions['B'].width = 20# Subject
    log_worksheet.column_dimensions['C'].width = 24# Requesting Agency
    log_worksheet.column_dimensions['D'].width = 20# Requesting Agent
    log_worksheet.column_dimensions['E'].width = 14# Case
    log_worksheet.column_dimensions['F'].width = 20# Summary of Findings
    log_worksheet.column_dimensions['G'].width = 14# Source
    log_worksheet.column_dimensions['H'].width = 25# Notes

    log_worksheet.cell(row=1, column=1).value = 'Date'
    log_worksheet.cell(row=1, column=2).value = 'Subject'
    log_worksheet.cell(row=1, column=3).value = 'Requesting Agency'
    log_worksheet.cell(row=1, column=4).value = 'Requesting Agent'
    log_worksheet.cell(row=1, column=5).value = 'Case'
    log_worksheet.cell(row=1, column=6).value = 'Summary of Findings'
    log_worksheet.cell(row=1, column=7).value = 'Notes'



    workbook.save(output_xlsx)
    workbook.close()

    
def usage():
    file = sys.argv[0].split('\\')[-1]
    print("\nDescription: " + description)
    print(file + " Version: %s by %s" % (version, author))
    print("\nExample:")
    print("\t" + file + " -w -I WigleWifi_sample.csv")


if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
1.0.0 - removes dulicate MAC's and keeps the stongest signal
0.0.1 - convert MfgrId to a real company
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

keep the .gz filename

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""



"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
