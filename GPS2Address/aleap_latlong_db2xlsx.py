"""
LatLong DB to XLSX Converter
Converts SQLite _latlong.db data into Aleap_latlong.xlsx
Provides both a GUI mode and a command-line mode (-c).
"""

import os
import sys
import argparse
import sqlite3
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DEFAULT_OUTPUT = "Aleap_latlong.xlsx"

print(f'hello world')
# -----------------------------------------------------------
#  UTILITY: write colored text into Tkinter Text widget
# -----------------------------------------------------------
def log(msg, box, color="black"):
    box.configure(state="normal")
    box.insert(tk.END, msg + "\n", color)
    box.configure(state="disabled")
    box.see(tk.END)


# -----------------------------------------------------------
#  CORE CONVERTER FUNCTION
# -----------------------------------------------------------
def convert_db_to_xlsx(db_path, xlsx_path, message_box=None):
    use_gui = message_box is not None

    def gui_log(msg, color="black"):
        if use_gui:
            log(msg, message_box, color)
        else:
            print(msg)

    # Use defaults if empty
    if not db_path:
        db_path = os.path.join(os.getcwd(), "_latlong.db")
    if not xlsx_path:
        xlsx_path = os.path.join(os.getcwd(), DEFAULT_OUTPUT)

    gui_log(f"📁 Converting {db_path} → {xlsx_path}", color="blue")

    # Connect to SQLite
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT timestamp AS time,
                   latitude AS Latitude,
                   longitude AS Longitude,
                   activity AS Type,
                   timestamp AS 'Time Original'
            FROM data;
        """)
        rows = cursor.fetchall()
    except Exception as e:
        gui_log(f"❌ Failed to read SQLite DB: {e}", color="red")
        return False

    headers = ["time", "Latitude", "Longitude", "Type", "Time Original",
               "original_file", "Icon", "group", "Subgroup"]

    wb = Workbook()
    ws = wb.active
    ws.title = "LatLong"
    ws.freeze_panes = "B2"

    # Write headers
    ws.append(headers)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", horizontal="left")

    # Process rows
    for row in rows:
        time, lat, lon, Type, time_original = row

        if isinstance(time, str) and "." in time:
            time = time.split('.')[0]

        original_file = "_latlong.db from ALEAPP"
        Icon = ""
        group = ""
        Subgroup = ""

        # Classification logic
        if "Google Photos" in (Type or ""):
            Icon = "Images"
            group = "Media remote" if "Remote" in Type else "Media"
            Subgroup = "Media"
        elif "Searched" in (Type or "") or "Searches" in (Type or ""):
            Icon = "Searched"
            Subgroup = "SearchedPlaces"
        elif "Google Maps Last Trip" in (Type or ""):
            Icon = "Car"

        extended_row = [time, lat, lon, Type, time_original,
                        original_file, Icon, group, Subgroup]

        ws.append(extended_row)
        for col in range(1, len(extended_row) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical="top", horizontal="left")

    # Column widths
    widths = [30, 20, 20, 30, 30, 30, 15, 15, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Save
    try:
        wb.save(xlsx_path)
        gui_log(f"✅ Saved output to {xlsx_path}", color="green")
        return True
    except Exception as e:
        gui_log(f"❌ Failed to save XLSX: {e}", color="red")
        return False


# -----------------------------------------------------------
#  COMMAND-LINE MODE (-c)
# -----------------------------------------------------------
def cli_mode(input_db, output_xlsx):
    ok = convert_db_to_xlsx(input_db, output_xlsx, message_box=None)
    sys.exit(0 if ok else 1)


# -----------------------------------------------------------
#  GUI MODE
# -----------------------------------------------------------
def start_conversion():
    db_path = input_entry.get().strip()
    xlsx_path = output_entry.get().strip()
    message_box.configure(state="normal")
    message_box.delete("1.0", tk.END)
    message_box.configure(state="disabled")
    convert_db_to_xlsx(db_path, xlsx_path, message_box)


def build_gui():
    global input_entry, output_entry, message_box

    root = tk.Tk()
    root.title("LatLong DB to XLSX Converter")

    tk.Label(root, text="📎 Convert _latlong.db → Aleap_latlong.xlsx",
             font=("Arial", 10, "bold")).pack(pady=(5, 0))

    frame = tk.Frame(root)
    frame.pack(pady=5)

    tk.Label(frame, text="Input .db File:").grid(row=0, column=0, sticky="e")
    input_entry = tk.Entry(frame, width=50)
    input_entry.insert(0, os.path.join(os.getcwd(), "_latlong.db"))
    input_entry.grid(row=0, column=1)
    tk.Button(frame, text="Browse",
              command=lambda: input_entry.delete(0, tk.END) or
                              input_entry.insert(0, filedialog.askopenfilename(filetypes=[("SQLite DB", "*.db")]))
              ).grid(row=0, column=2)

    tk.Label(frame, text="Output .xlsx File:").grid(row=1, column=0, sticky="e")
    output_entry = tk.Entry(frame, width=50)
    output_entry.insert(0, os.path.join(os.getcwd(), DEFAULT_OUTPUT))
    output_entry.grid(row=1, column=1)
    tk.Button(frame, text="Browse",
              command=lambda: output_entry.delete(0, tk.END) or
                              output_entry.insert(0, filedialog.asksaveasfilename(
                                  defaultextension=".xlsx",
                                  filetypes=[("Excel File", "*.xlsx")]))
              ).grid(row=1, column=2)

    tk.Button(root, text="Convert", command=start_conversion, bg="lightblue").pack(pady=10)

    # Colored Text widget
    message_box = tk.Text(root, width=90, height=12, state="disabled")
    message_box.tag_config("red", foreground="red")
    message_box.tag_config("green", foreground="green")
    message_box.tag_config("blue", foreground="blue")
    message_box.tag_config("black", foreground="black")
    message_box.pack(pady=5)

    root.mainloop()


# -----------------------------------------------------------
#  MAIN ENTRY POINT
# -----------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LatLong DB → XLSX Converter")
    parser.add_argument("-c", "--convert", nargs="*", help="Run in command-line mode. "
                                                           "Usage: -c [input.db output.xlsx]")

    args = parser.parse_args()

    if args.convert is not None:
        # No args → use defaults
        if len(args.convert) == 0:
            cli_mode("_latlong.db", DEFAULT_OUTPUT)
        elif len(args.convert) == 1:
            cli_mode(args.convert[0], DEFAULT_OUTPUT)
        else:
            cli_mode(args.convert[0], args.convert[1])

    # If no -c provided → launch GUI
    build_gui()
