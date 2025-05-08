#!/usr/bin/env python3
# coding: utf-8
'''
read GPS coordinates and convert them to addresses
or
read addresses and convert them to coordinates
and
read GPS coordinates (xlsx) and convert them to KML
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>
# from functools import cache

from fastkml import kml  # pip install fastkml
from shapely.geometry import Point, LineString, Polygon  # pip install fastkml shapely openpyxl

import io
import requests    # pip install requests
from openpyxl.drawing.image import Image

import os
import re
import sys
import time
import random
import openpyxl
import simplekml    # pip install simplekml
import geohash2    # pip install geohash2
from datetime import datetime

from geopy.distance import distance # test
from math import radians, cos, sin  # test

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment



import argparse  # for menu system
from openpyxl import load_workbook, Workbook

from geopy.geocoders import Nominatim   # pip install geopy
geolocator = Nominatim(user_agent="GeoTraxer")

# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = "convert GPS coordinates to addresses or visa versa & create a KML file"
version = '1.4.1'

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>
# @cache
def main():
    
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-H','--howto', help='help module', required=False, action='store_true')
    parser.add_argument('-c', '--create', help='create blank input sheet', required=False, action='store_true')
    parser.add_argument('-i', '--intel', help='convert intel sheet to locations', required=False, action='store_true')
    parser.add_argument('-k', '--kml', help='xlsx to kml with nothing else', required=False, action='store_true')
    parser.add_argument('-K', '--kml2xlsx', help='kml to xlsx', required=False, action='store_true')

    parser.add_argument('-R', '--read', help='read xlsx', required=False, action='store_true')
    parser.add_argument('-r', '--read_basic', help='read basic xlsx', required=False, action='store_true')

    args = parser.parse_args()

    if args.howto:  # this section might be redundant
        parser.print_help()
        usage()
        return 0
        sys.exit()

    # global input_xlsx
    global output_xlsx
    global output_kml

    global input_kml
    # input_kml = 'Locations.kml'
    input_kml = ''
    
    if not args.input: 
        input_xlsx = "locations.xlsx"        
    elif args.input.lower().endswith('.kml'):
        # input_xlsx = args.input
        input_kml = args.input
    else:
        input_xlsx = args.input

    global datatype
    datatype = ''
    if args.input is not None:
        if args.input.lower().endswith('.xlsx'):
            datatype = input_xlsx
        elif args.input.lower().endswith('.kml'):
            datatype = args.input
        elif args.input.lower().endswith('.xls'):
            print(f'Convert {args.input} to .xlsx format first')
            exit(1)  # Exit with a nonzero status to indicate an error


    else:
        datatype = ''
     
    datatype = datatype.replace('.xlsx', '').replace('.kml', '')

    if '\\' in datatype:
        datatype = datatype.split('\\')[-1]

    output_kml = (f'GPS_{datatype}.kml')

    if not args.output: 
        output_xlsx = (f'Locations_{datatype}.xlsx') 
      
    else:
        output_xlsx = args.output

    if args.create:
        data = []
        msg_blurb = (f'Writing to {output_xlsx}')
        msg_blurb_square(msg_blurb, color_green)
        write_locations(data)

    elif args.kml:
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            msg_blurb = (f'Reading {input_xlsx}')
            msg_blurb_square(msg_blurb, color_green)

            # data = read_xlsx(input_xlsx)
            (data, coordinates) = read_locations(input_xlsx)

            # create kml file
            write_locations(data)   # ??
            write_kml(data)
            travel_path_kml(coordinates)
            
            workbook.close()
            # print(f'{color_green}Writing to {output_xlsx} {color_reset}')
            msg_blurb = (f'Writing to {output_kml}')
            msg_blurb_square(msg_blurb, color_blue)

            print(f'''\n\n{color_yellow}
            visit https://earth.google.com/
            <file><Import KML> select gps.kml <open>
            {color_reset}\n''')
        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()
    elif args.kml2xlsx:
        data = []
        file_exists = os.path.exists(input_kml)
        if file_exists == True:
            msg_blurb = (f'Reading {input_kml}')
            msg_blurb_square(msg_blurb, color_green)

            # (data, coordinates) = kml_to_excel(input_kml, output_xlsx)
            (data, coordinates) = kml_to_excel(data)

            # create kml file
            write_locations(data)   # ??
            # write_kml(data)
            # travel_path_kml(coordinates)
            
            workbook.close()
            msg_blurb = (f'Writing to {output_xlsx}')
            msg_blurb_square(msg_blurb, color_blue)

            # print(f'''\n\n{color_yellow}
            # visit https://earth.google.com/
            # <file><Import KML> select gps.kml <open>
            # {color_reset}\n''')
        else:
            print(f'{color_red}{input_kml} does not exist{color_reset}')
            exit()

    elif args.intel:
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            msg_blurb = (f'Reading {input_xlsx}')
            msg_blurb_square(msg_blurb, color_green)

            data = read_intel(input_xlsx)
            # write_kml(data)
            write_locations(data)
            msg_blurb = (f'Writing to {output_xlsx}')
            msg_blurb_square(msg_blurb, color_green)

        else:
            msg_blurb = (f'{input_xlsx} does not exist')
            msg_blurb_square(msg_blurb, color_red)
            # print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()
            
    elif args.read:
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            msg_blurb = (f'Reading {input_xlsx}')
            msg_blurb_square(msg_blurb, color_green)
            
            # data = read_xlsx(input_xlsx)
            (data, coordinates) = read_locations(input_xlsx)
            data = read_gps(data)
            write_locations(data)
            write_kml(data)
            
            travel_path_kml(coordinates)
            workbook.close()
            msg_blurb = (f'Writing to {output_xlsx}')
            msg_blurb_square(msg_blurb, color_green)

            print(f'''\n\n{color_yellow}
            visit https://earth.google.com/
            <file><Import KML> select gps.kml <open>
            {color_reset}\n''')
        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()

    elif args.read_basic:
        
        data = []
        file_exists = os.path.exists(input_xlsx)
        if file_exists == True:
            msg_blurb = (f'Reading basic {input_xlsx} no kml')
            msg_blurb_square(msg_blurb, color_green)
            if input_kml.lower().endswith('.kml'):
                data = kml_to_xlsx(input_kml, output_xlsx)
                write_locations(data)
            else:    
                # data = read_xlsx(input_xlsx)
                (data, coordinates) = read_locations(input_xlsx)
                # print(f'{coordinates}') # temp
                # data = read_gps(data)
                write_locations(data)
                write_kml(data)
                travel_path_kml(coordinates)
                workbook.close()
                msg_blurb = (f'Writing to {output_xlsx}')
                msg_blurb_square(msg_blurb, color_green)

                print(f'''\n\n{color_yellow}
                visit https://earth.google.com/
                <file><Import KML> select gps.kml <open>
                {color_reset}\n''')
        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()


    else:
        usage()
    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def case_number_prompt():
    # Prompt the user to enter the case number
    case_number = input("Please enter the Case Number: ")
    # Assign the entered value to Case
    case_prompt = case_number
    return case_prompt

def convert_timestamp(timestamp, time_orig, timezone):
    if timezone is None:
        timezone = ''
    if time_orig is None:
        time_orig = ''

    timestamp = str(timestamp)

    if re.match(r'\d{1,2}/\d{1,2}/\d{4} \d{1,2}:\d{2}:\d{2}\.\d{3} (AM|PM)', timestamp):
        # Define the expected format
        expected_format = "%m/%d/%Y %I:%M:%S.%f %p"

        # Parse the string into a datetime object
        dt_obj = datetime.strptime(timestamp, expected_format)

        # Remove microseconds
        dt_obj = dt_obj.replace(microsecond=0)

        # Format the datetime object back into a string with the specified format
        # timestamp = dt_obj.strftime("%Y/%m/%d %I:%M:%S %p")
        timestamp = dt_obj.strftime("%Y-%m-%d %I:%M:%S %p")


    # Regular expression to find all timezones
    timezone_pattern = r"([A-Za-z ]+)$"
    matches = re.findall(timezone_pattern, timestamp)

    # Extract the last timezone match
    if matches:
        timezone = matches[-1]
        timestamp = timestamp.replace(timezone, "").strip()
    else:
        timezone = ''
        
    if time_orig == "":
        time_orig = timestamp
    else:
        timezone = ''


    # timestamp = timestamp.replace(' at ', ' ')
    if "(" in timestamp:
        timestamp = timestamp.split('(')
        timezone = timestamp[1].replace(")", '')
        timestamp = timestamp[0]
    elif " CDT" in timestamp:
        timezone = "CDT"
        timestamp = timestamp.replace(" CDT", "")
    elif " CST" in timestamp:
        timezone = "CST"
        timestamp = timestamp.replace(" CST", "")

    formats = [
        "%B %d, %Y, %I:%M:%S %p %Z",    # June 13, 2022, 9:41:33 PM CDT (Flock)
        "%B %d, %Y at %I:%M:%S %p %Z",  # June 17, 2024 at 3:41:53 PM CDT (flock?)
        "%Y:%m:%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",    # 2005-10-18 10:58:29
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M:%S.%f %p", # '3/8/2024 11:06:47.358 AM  # task
        "%m/%d/%Y %I:%M %p",  # timestamps without seconds
        "%m/%d/%Y %H:%M:%S",  # timestamps in military time without seconds
        "%m-%d-%y at %I:%M:%S %p %Z", # test 09-10-23 at 4:29:12 PM CDT
        "%m-%d-%y %I:%M:%S %p",
        "%B %d, %Y at %I:%M:%S %p",
        "%B %d, %Y %I:%M:%S %p %Z",
        "%B %d, %Y %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p %Z",
        "%Y-%m-%dT%H:%M:%SZ",  # ISO 8601 format with UTC timezone
        "%Y/%m/%d %H:%M:%S",  # 2022/06/13 21:41:33
        "%d-%m-%Y %I:%M:%S %p",  # 13-06-2022 9:41:33 PM
        "%d/%m/%Y %H:%M:%S",  # 13/06/2022 21:41:33
        "%Y-%m-%d %I:%M:%S %p",  # 2022-06-13 9:41:33 PM
        "%Y%m%d%H%M%S",  # 20220613214133
        "%Y%m%d %H%M%S",  # 20220613 214133
        "%m/%d/%y %H:%M:%S",  # 06/13/22 21:41:33
        "%d-%b-%Y %I:%M:%S %p",  # 13-Jun-2022 9:41:33 PM
        "%d/%b/%Y %H:%M:%S",  # 13/Jun/2022 21:41:33
        "%Y/%b/%d %I:%M:%S %p",  # 2022/Jun/13 9:41:33 PM
        "%d %b %Y %H:%M:%S",  # 13 Jun 2022 21:41:33
        "%A, %B %d, %Y %I:%M:%S %p %Z",  # Monday, June 13, 2022 9:41:33 PM CDT ?
        "%m/%d/%Y %H:%M",               # Month/Day/Year Hour:Minute (new format)    
        "%Y-%m-%d %H:%M:%S%z",  #         2009-04-11 19:37:36-05:00
        "%m/%d/%Y",  # date only
        "%m/%d/%y",  # date only
        "%m-%d-%y",  # date only, short year
        "%m-%d-%Y",  # date only, long year        
        "%Y-%m-%d",  # date only in ISO format        
        "%Y/%m/%d",  # date only      

        "%A, %B %d, %Y %I:%M:%S %p"     # Monday, June 13, 2022 9:41:33 PM CDT
    ]

    for fmt in formats:
        try:
            dt_obj = datetime.strptime(timestamp.strip(), fmt)
            timestamp = dt_obj
            return timestamp, time_orig, timezone
                        
        except ValueError:
            pass

    raise ValueError(f"{time_orig} Timestamp format not recognized")

def direction_convert(direction):
    # If the input is a valid direction letter (N, E, S, W), return it directly
    if isinstance(direction, str) and direction in ["N", "E", "S", "W"]:
        return direction
    
    # If the input is not a number or outside the valid range, return an error
    if not isinstance(direction, (int, float)) or direction < 0 or direction > 360:
        return "Invalid input. Must be a degree between 0 and 360, or a valid direction letter."
    
    # Simplified cardinal directions with their corresponding ranges
    directions = [
        ("N", 0, 45),   # North (0° to 45° or 360°)
        ("E", 45, 135),  # East (45° to 135°)
        ("S", 135, 225), # South (135° to 225°)
        ("W", 225, 315), # West (225° to 315°)
        ("N", 315, 360)  # North again (315° to 360°)
    ]
    
    # Iterate through each direction and match the degree
    for dir_label, min_deg, max_deg in directions:
        if min_deg <= direction < max_deg:
            return dir_label
            
            
def gps_cleanup(latitude, longitude):
    latitude = latitude.replace("'", '').replace("\"", '').replace("°", '.')   # don't replace .
    longitude = longitude.replace("'", '').replace("\"", '').replace("°", '.')    

    latitude = re.sub(r'[a-zA-Z]', '', latitude)
    longitude = re.sub(r'[a-zA-Z]', '', longitude)

    try:
        # Convert the cleaned string to a float
        latitude = float(latitude)
        longitude = float(longitude)
    except ValueError:
        # print(f'Error: {ValueError}') 
        # print(f'latitude type = {type(latitude)} = {latitude}') # temp
        pass
    return (latitude, longitude)

def get_coordinates(geometry):
    """Extract latitude, longitude, and formatted coordinates from different KML geometries."""
    latitude, longitude, coordinate = '', '', ''

    if isinstance(geometry, Point):
        latitude, longitude = geometry.y, geometry.x
        coordinate = f'{latitude},{longitude}'

    elif isinstance(geometry, LineString):
        coordinates = [(point[1], point[0]) for point in geometry.coords]
        latitude, longitude = coordinates[0]
        coordinate = ', '.join([f'{lat},{lon}' for lat, lon in coordinates])

    elif isinstance(geometry, Polygon):
        coordinates = [(point[1], point[0]) for point in geometry.exterior.coords]
        latitude, longitude = coordinates[0]
        coordinate = ', '.join([f'{lat},{lon}' for lat, lon in coordinates])

    elif latitude == '':
        coordinate = str(geometry)
        coordinate = coordinate.split('(')[1].replace(')', '')
        coordinate = coordinate.split(' ')
        longitude = coordinate[0]
        latitude = coordinate[1]

        # coordinate = (f'{latitude},{longitude}')
        coordinate = f'{longitude},{latitude}'
    return latitude, longitude, coordinate

        
def haversine(lat1, lon1, lat2, lon2):
    from math import radians, cos, sin, sqrt, atan2
    distance = ''
    R = 6371.0  # Radius of the Earth in km
    try:
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])

        dlat = lat2 - lat1
        dlon = lon2 - lon1

        a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))

        distance = R * c
    except ValueError as e:
        print(f'{e} error ') # temp
    return distance

def kml_to_excel(data):
# def kml_to_excel(input_kml, output_xlsx):
    coordinates = []
    from fastkml import kml  # pip install fastkml
    from shapely.geometry import Point, LineString, Polygon  # pip install fastkml shapely openpyxl

    with open(input_kml, 'rb') as file:  # Open the KML file as binary
        doc = file.read()

    k = kml.KML()
    k.from_string(doc)
    placemarks = []

    # Traverse through features in the KML
    for feature in k.features():
        if isinstance(feature, kml.Document) or isinstance(feature, kml.Folder):
            for placemark in feature.features():

                if isinstance(placemark, kml.Placemark):
                    (tag, type_data, styleUrl, Altitude, source_file, business) = ('', '', '', '', input_kml, '')
                    (fulladdress, IconStyle, latitude, longitude, coordinate) = ('', '', '' ,'', '')

                    name = placemark.name or ""
                    if "/" in name:
                        name = ''
                    elif ".MOV" in name or ".mov" in name:
                        type_data = "Videos"
                        description = name
                        IconStyle = 'Videos'                        
                    
                    description = placemark.description or ""
                    description = description.replace('<p>', '').replace('</p>', '')
                    if "/" in description:
                        description = ''
                    if ".JPG" in name or ".jpg" in name or ".jpeg" in name:
                        type_data = "Images"
                        description = name
                        IconStyle = 'Images'
                        
                        
                    # Extract timestamp if present
                    Time = ""
                    if placemark.timeStamp:
                        try:
                            Time = str(placemark.timeStamp).split('.')[0]
                        except ValueError:
                            Time = ""  # Handle invalid timestamp gracefully

                    # Extract coordinate
                    latitude, longitude, coordinate = get_coordinates(placemark.geometry)
                    print(f'{Time}  {coordinate}')  # temp

                    data.append({
                        '#': '',
                        'Name': name,
                        'Time': Time,
                        'Latitude': latitude,
                        'Longitude': longitude,
                        'Coordinate': coordinate,            
                        'Description': description,
                        'Tag': tag,            
                        'Type': type_data,            
                        'styleUrl': styleUrl,
                        'Altitude': Altitude,
                        'Source file information': source_file,
                        'business': business,    
                        'fulladdress': fulladdress ,    
                        'Icon': IconStyle
                    })

    # for idx, placemark in enumerate(placemarks, start=1):
        # ws.append([idx] + list(placemark))

    # Auto-adjust column widths
    # for col in ws.columns:
        # max_length = 0
        # col_letter = col[0].column_letter
        # for cell in col:
            # try:
                # if len(str(cell.value)) > max_length:
                    # max_length = len(cell.value)
            # except:
                # pass
        # ws.column_dimensions[col_letter].width = max(max_length, 10)

    # wb.save(output_xlsx)

    return data, coordinates

    
def random_color():
    return simplekml.Color.rgb(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
    
def travel_path_kml(coordinates):
    '''
    This reads latitude and longitude coordinates and generates a KML file with lines 
    connecting the points and no icons displayed.
    '''
    
    kml2 = simplekml.Kml()
    heat = kml2.newfolder(name="travel_path")

    # output_file = f'travel_path_example.kml' 
    output_file = (f'travel_path_{datatype}.kml') 
    # for coord in coordinates:
        # lat, long, timestamp = coord
        # print(f"Latitude: {lat}, Longitude: {long}, Timestamp: {timestamp}")
        
    for i in range(len(coordinates) - 1):
        new_color = simplekml.Color.red    # test
        # (new_color, linestring) = ('ff2f41ac', '')
        coord1 = coordinates[i]
        coord2 = coordinates[i + 1]
        lat1, lon1, time1 = coord1
        lat2, lon2, time2 = coord2

        lat_long1 = (f'{lat1},{lon1}')
        lat_long2 = (f'{lat2},{lon2}')  

        if i == len(coordinates) - 1:
            print(f"skipping last set")
        # else:
            # print(f"Latitude: {lat}, Longitude: {long}, Timestamp: {timestamp}")

        
        same_day, time_diff = time_compare(time1, time2)

        try:
            distance = haversine(lat1, lon1, lat2, lon2)
        except:
            distance = 0
        # print(f'distance = {distance} km/h') # temp
        # Assuming an average speed limit of 100 km/h to determine feasibility
        max_distance_possible = time_diff * 100
        if 0.001 < time_diff < 1400 and lat_long1 != lat_long2:
        # if 0.001 < time_diff < 1400 and distance < max_distance_possible:
            linestring = heat.newlinestring(coords=[(lon1, lat1), (lon2, lat2)])
            # linestring.style.linestyle.color = random_color()
            linestring.style.linestyle.color = new_color
            linestring.style.linestyle.width = 2
        else:
            new_color = random_color()
            # linestring.style.linestyle.color = random_color()
            # print(f'new_color = {new_color}')   # temp
            
            
    # Ensure no icons are visible
    heat.stylemap.normalstyle.iconstyle.scale = 0
    heat.stylemap.normalstyle.iconstyle.icon.href = ''

    kml2.save(output_file)

    msg_blurb = (f'extra travel_path KML file {output_file} created successfully!')
    msg_blurb_square(msg_blurb, color_blue) 
    
def travel_path_kml_old(coordinates):
    '''
    This reads latitude and longitude coordinates and generates a KML file with lines 
    connecting the points and no icons displayed.
    integrate this with (same_day, time_dif) = time_compare(Time_previous, Time)
    calculate distance between lat1/long1 and lat2/long2. calclulate time to see if this trip was possible
    randomize line color for different trips
    '''

    kml2 = simplekml.Kml()
    heat = kml2.newfolder(name="travel_path")

    output_file = (f'travel_path_{datatype}.kml') 
    # Create lines between consecutive points
    for i in range(len(coordinates) - 1):
        (same_day, time_diff) = ('', '')
        coord1 = coordinates[i]
        coord2 = coordinates[i + 1]
        try:
            # Access individual components
            lat1, long1, time1 = coord1
            lat2, long2, time2 = coord2            

            long_lat1 = (f'{long1},{lat1}')
            long_lat2 = (f'{long2},{lat2}')            
            lat_long1 = (f'{lat1},{long1}')
            lat_long2 = (f'{lat2},{long2}')            


        except ValueError as e:
            print(f'{e} Time error {Time} {coordinates} ') # temp
            # print(f'Error: {e}, coord1: {coord1}, coord2: {coord2}')
        
        (same_day, time_diff) = time_compare(time1, time2)
        # if same_day and time_diff != '0.0':
        if 0.001 < time_diff < 1400 and lat_long1 != lat_long2:
        
        # if 0.001 < time_diff < 1400 and long_lat1 != long_lat2:
        # if same_day and time_diff != '0.0' and long_lat1 != long_lat2:
            # print(f'less than 1 day with different location {long_lat1} - {long_lat2}')
            print(f'less than 1 day with different location {lat_long1} - {lat_long2}')

            # heat.newlinestring(coords=[long_lat1, long_lat2])
            heat.newlinestring(coords=[lat_long1, lat_long2])

    # Ensure no icons are visible
    heat.stylemap.normalstyle.iconstyle.scale = 0
    heat.stylemap.normalstyle.iconstyle.icon.href = ''

    # Ensure lines are displayed in color
    # heat.stylemap.normalstyle.linestyle.color = simplekml.Color.blue # only does white
    heat.style.linestyle.color = random_color()
    # heat.style.linestyle.color = simplekml.Color.red  # Make the line red
    # heat.style.linestyle.color = simplekml.Color.orange  # Make the line orange
    # heat.style.linestyle.color = simplekml.Color.yellow  # Make the line yellow
    # heat.style.linestyle.color = simplekml.Color.green  # Make the line green
    # heat.style.linestyle.color = simplekml.Color.blue  # Make the line blue
    # heat.style.linestyle.color = simplekml.Color.purple  # Make the line purple
    # heat.style.linestyle.color = simplekml.Color.white  # Make the line white
    # heat.style.linestyle.color = simplekml.Color.black  # Make the line black

    heat.style.linestyle.width = 2  # Set line width

    kml2.save(output_file)

    msg_blurb = (f'extra travel_path KML file {output_file} created successfully!')
    msg_blurb_square(msg_blurb, color_blue)    
    
def kml_to_xlsx(kml_file, xlsx_file):
    data = []
    from pykml import parser    # pip install pykml
    with open(kml_file, 'r') as f:
        doc = parser.parse(f)

    for placemark in doc.getroot().Document.Placemark:
        (name, coords, latitude, longitude, description, styleUrl, IconStyle) = ('', '', '', '', '', '', '')
        (Altitude, Time, type_data, source_file, business, tag) = ('', '', '', '', '', '')
        (fulladdress) = ('')
        name = placemark.name.text
        coordinate = placemark.Point.coordinates.text
        coords = placemark.Point.coordinates.text.split(',')
        latitude = coords[1]
        longitude = coords[0]
        try:
            Altitude = coords[2]
        except:pass
        
        coordinate = coordinate.rstrip(',0.0')
        description = placemark.description.text
        pattern = r'([A-Z]+):([^:\n]+)'
        time_pattern = r'([A-Z]+):([^\n]+)'
        matches = re.findall(pattern, description)
        matches2 = re.findall(time_pattern, description)

        for key2, value2 in matches2:
            if key2 == 'TIME':
                Time = value2.strip()

        key_value_pairs = {}

        for key, value in matches:
            if key == 'Business':
                business = value.strip()
            elif key == 'TYPE':
                type_data = value.strip()
            elif key == 'SOURCE':
                source_file = value.strip()
            elif key == 'TAG':
                tag = value.strip()
            elif key == 'ADDRESS':
                fulladdress = value.strip()     
             
        
        
        styleUrl = placemark.styleUrl.text
        # IconStyle = placemark.IconStyle.text

        # Append extracted data to the list
        data.append({
            '#': name,
            'Name': name,
            'Time': Time,
            'Latitude': latitude,
            'Longitude': longitude,
            'Coordinate': coordinate,            
            'Description': description,
            'Tag': tag,            
            'Type': type_data,            
            'styleUrl': styleUrl,
            'Altitude': Altitude,
            'Type': type_data,
            'Source file information': source_file,
            'business': business,    
            'fulladdress': fulladdress ,    
            'Icon': IconStyle
        })

    return data

def long_lat_flip(latitude, longitude, coordinate, Altitude):
    coordinate = coordinate.replace('(', '').replace(')', '')
    if latitude == ''  and longitude == '' and coordinate != '':
        if ',' in coordinate:
            coordinate_split = coordinate.split(',')
            latitude = coordinate_split[0].strip()
            longitude = coordinate_split[1].strip()
            try:
                if Altitude == '':
                    Altitude = coordinate_split[2].strip()
            except:pass
                
            if Altitude == '':
                # coordinate = f'{latitude},{longitude}'
                coordinate = f'{longitude},{latitude}'
            elif Altitude != '':
                # coordinate = f'{latitude},{longitude},{Altitude}' 
                coordinate = f'{longitude},{latitude},{Altitude}' 
    return (latitude, longitude, coordinate, Altitude)
 
def msg_blurb_square(msg_blurb, color):
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')

def radius_azimuth(kml, no, description, latitude, longitude, Azimuth, Radius, Altitude, point_icon):
    try:
        Radius = float(Radius)
    except:
        Radius = 0
        
    if Azimuth == '':
        Azimuth = '.1'
    if Altitude == '':
        Altitude = 0

    point = kml.newpoint(
        name=f"{no}",
        description=f"{description}",
        coords=[(longitude, latitude, Altitude)]
    )
    point.style.iconstyle.icon.href = point_icon
    point.altitudemode = simplekml.AltitudeMode.relativetoground

    # Function to calculate destination point given start point, bearing, and distance
    def destination_point(lat, lon, bearing, distance_km):
        R = '6371.0'  # Radius of the Earth in kilometers
        bearing = radians(bearing)

        lat1 = radians(lat)
        lon1 = radians(lon)

        lat2 = sin(lat1) * cos(distance_km / R) + cos(lat1) * sin(distance_km / R) * cos(bearing)
        lat2 = degrees(atan2(sin(bearing) * sin(distance_km / R) * cos(lat1), cos(distance_km / R) - sin(lat1) * sin(lat2)))
        lon2 = lon1 + atan2(sin(bearing) * sin(distance_km / R) * cos(lat1), cos(distance_km / R) - sin(lat1) * sin(lat2))
        lon2 = degrees(lon2)

        return lat2, lon2

    # Calculate circle points
    num_points = 360  # number of points in the circle
    circle_points = []
    for i in range(num_points):
        try:
            angle = i
            destination = distance(kilometers=Radius / 1000.0).destination((latitude, longitude), angle)
            circle_points.append((destination.longitude, destination.latitude, Altitude))
        except ValueError:
            print(f'Radius Error {ValueError} {latitude} {angle}') # temp
            pass
    # Add the circle to the KML
    pol = kml.newpolygon(name="Circle", outerboundaryis=circle_points)
    pol.altitudemode = simplekml.AltitudeMode.relativetoground
    pol.style.linestyle.color = simplekml.Color.red  # Circle outline color
    pol.style.linestyle.width = 2  # Circle outline width
    pol.style.polystyle.color = simplekml.Color.changealphaint(100, simplekml.Color.red)  # Circle fill color with transparency

    try:
        # Calculate the azimuth point
        azimuth_distance = Radius / 1000.0  # convert radius to kilometers
        destination = distance(kilometers=azimuth_distance).destination((latitude, longitude), Azimuth)
        azimuth_point = (destination.longitude, destination.latitude, Altitude)

        # Add the azimuth line to the KML
        line = kml.newlinestring(name="Azimuth Line", coords=[(longitude, latitude, Altitude), azimuth_point])
        line.altitudemode = simplekml.AltitudeMode.relativetoground


        # line.style.linestyle.color = simplekml.Color.red  # Make the line red
        line.style.linestyle.color = simplekml.Color.orange  # Make the line orange
        # line.style.linestyle.color = simplekml.Color.yellow  # Make the line yellow
        # line.style.linestyle.color = simplekml.Color.green  # Make the line green
        # line.style.linestyle.color = simplekml.Color.blue  # Make the line blue
        # line.style.linestyle.color = simplekml.Color.purple  # Make the line purple
        # line.style.linestyle.color = simplekml.Color.white  # Make the line white
        # line.style.linestyle.color = simplekml.Color.black  # Make the line black
        # line.style.linestyle.color = simplekml.Color.blue  # Line color
        line.style.linestyle.width = 2  # Line width

        # Save the KML object to a file
        # kml.save("my_point_with_circle_and_azimuth.kml")
    except:
        pass
    return (kml, point)
    
def read_gps(data): 

    """Read data and return as a list of dictionaries.
    It extracts headers from the 
    first row and then iterates through the data rows, creating dictionaries 
    for each row with headers as keys and cell values as values.
    
    """


    coordinates_list = [] # a list of coordinate and fulladdress   

    for row_index, row_data in enumerate(data):
        print(f'\n{row_index + 2} ______________________\n')
        (street, fulladdress) = ('', '')
        
        (zipcode, business, number, street, city, county) = ('', '', '', '', '', '')
        (state, latitude, longitude, query, coordinate) = ('', '', '', '', '')
        (Index, country, latitude, longitude, PlusCode) = ('', '', '', '', '')
        (county, query, type_data) = ('', '', '')
        (location, skip, address) = ('', '', '')
        (type_data, plate, country_code, state_ftk, city_ftk, hwy) = ('', '', '', '', '', '')
        (direction, name_data, no, account, container, time_local) = ('', '', '', '', '', '') 
        (deleted, service_id, carved, sighting_state, sighting_location, manually_decoded) = ('', '', '', '', '', '')

        fulladdress = row_data.get("fulladdress")
        if fulladdress is None:
            fulladdress = ''  

        name_data = row_data.get("Name")
        latitude = row_data.get("Latitude") # works
        if latitude is None:
            latitude = ''
        latitude = str(latitude)
        
        longitude = row_data.get("Longitude") # works
        if longitude is None:
            longitude = ''  
        longitude = str(longitude)
      
        address = row_data.get("Address") # works
        type_data = row_data.get("Type") # works  
        business = row_data.get("business")
        number = row_data.get("number")
        street = row_data.get("street")
        city = row_data.get("city")
        Time = row_data.get("Time")
        county = row_data.get("county")
        state = row_data.get("state")
        zipcode = row_data.get("zipcode")
        query = row_data.get("query")
        country = row_data.get("country")
        Subgroup = row_data.get("Subgroup")
        
        coordinate = row_data.get("Coordinate")
        row_data["Index"] = (row_index + 2)
        PlusCode = row_data.get("PlusCode")
        Radius = row_data.get("Radius") 
        Azimuth = row_data.get("Azimuth") # test
        speed = row_data.get("speed") # test
        parked = row_data.get("parked") # test

        if Radius is None:
            Radius == ''
        else:
            try:
                Radius = float(Radius)
            except ValueError:
                Radius == ''

        
# add uniq coordinates to coordinates_list
        
        # Iterate over the list of coordinate and fulladdress
        # for coordinate, fulladdress in coordinates_list:
            # print(f'test')  # temp
            # coordinate = create_coordinate(latitude, longitude)
            # check_full_address(coordinates_to_addresses, coordinate)

        # Function to check if full address exists for a coordinate
        # def check_full_address(coordinates_list, coordinate):
            # if coordinate in coordinates_list:
                # print("Full address exists.")
        

# skip lines with fulladdress
        # if fulladdress != '' or fulladdress == 'Soul Buoy':  
        if "," in fulladdress:        
            print(f'skipping fulladdress')    # temp
        # if len(fulladdress) > 2 or fulladdress == 'Soul Buoy':        
            if fulladdress == 'Soul Buoy':
                fulladdress = ''
        # if len(fulladdress) < 2 or fulladdress == 'Soul Buoy':        
            # fulladdress = ''
            skip == 'skip'
# GPS to fulladdress
        elif latitude != '' and isinstance(latitude, str) and len(latitude) > 5:
            print(f'trying to get full address')    # temp
            if latitude != '' and isinstance(latitude, str) and len(latitude) > 3:

                ##  if no fulladdress
                if len(fulladdress) < 2:
                    query = (f'{latitude}, {longitude}') # backwards

                    try:
                        # location = geolocator.reverse((longitude, latitude), language='en')

                        location = geolocator.reverse((latitude, longitude), language='en') #lat/long
                    except Exception as e:
                        print(f"{color_red}Error : {str(e)}{color_reset}") 
                    try:
                        fulladdress = location.address

                    except Exception as e:
                        print(f"{color_red}Error : {str(e)}{color_reset}") 

                    if fulladdress == 'Soul Buoy':        
                        fulladdress = ''
  

                    time.sleep(8)   ## Sleep x seconds

# address to gps / Full address
        elif address != '':
            if len(address) > 3:

                try:
                    print(f'trying address resolution')
                    location = geolocator.geocode(address)

                    if location:
                        latitude = location.latitude
                        longitude = location.longitude
                        query = (f'{address}')
                    else:
                        print(f"Location not found for address: {address}")

                except Exception as e:
                    print(f"Error: {str(e)}")

            try:
                location = geolocator.reverse((latitude, longitude), language='en')
                fulladdress = location.address

            except Exception as e:
                print(f"{color_red}Error: {str(e)}{color_reset}")

            if fulladdress == 'Soul Buoy':        
                fulladdress = ''
                skip == 'skip'

# coordinate        
            if latitude != '' and longitude != '' and coordinate == '':
                coordinate = (f'{latitude},{longitude}')
                # coordinate = (f'{longitude},{latitude}')

                time.sleep(8)   # Sleep for x seconds

        elif PlusCode != '':
            print(f'testing PlusCode {PlusCode}')
            
            try:
                decoded_location = geohash2.decode(PlusCode)
                latitude = decoded_location[0]
                longitude = decoded_location[1]

            
                print(f"Coordinates for Geohash '{PlusCode}': {decoded_location[0]}, {decoded_location[1]}")
                
                
                
                # location = geolocator.geocode(PlusCode)
                if decoded_location is not None:
                    # latitude, longitude = location.latitude, location.longitude
                    print(f'PlusCode {PlusCode} = {latitude}, {longitude}')
                    fulladdress = location.address
                else:
                    print(f"Unable to find coordinates for address: {PlusCode}")
            except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} PlusCode = <{PlusCode}>")  

        else:
            print(f'{color_red}none of the above{color_reset}')

# parse fulladdress
        if len(fulladdress) > 5 and skip != 'skip' and "," in fulladdress:
            try:
                address_parts = fulladdress.split(', ')
                if country != '' and len(address_parts) >= 1:
                    country = address_parts[-1]
                if zipcode != '' and len(address_parts) >= 2:   # task
                    zipcode = address_parts[-2]
                if len(state) <= 2 and len(address_parts) >= 3: # test
                # if state == '' and len(address_parts) >= 3:
                    state = address_parts[-3]
                if len(country) <= 2 and len(address_parts) >= 4:
                    county = address_parts[-4]
                if len(address_parts) >= 6:
                    if "Township" in address_parts[-5]:
                        city = address_parts[-6]
                    else:    
                        county = address_parts[-5]
                        city = address_parts[-6]
                
                # if len(address_parts) == 7:
                    # print(f'there are 7 parts') # temp
                if fulladdress.count(',') == 7:    # Check if there are exactly 7 commas
                    # print(f'there are also 7 parts') # temp
                    if 'Township' in address_parts[3]:    # task
                             
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[0]
                        street = address_parts[1]

                    else:
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[1]
                        street = address_parts[2]
    
                elif fulladdress.count(',') == 5:
                    if 'Township' in address_parts[1]:    # task
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[1]
                        street = address_parts[2]

                    else:
                        address_parts = fulladdress.split(', ')    # Split the address by commas
                        business = address_parts[0]
                        number = address_parts[1]
                        street = address_parts[2]

                if "United States" in country:
                    country = "United States" 

                if street.endswith(" Township"):
                    street == ''                

                try:
                    if business == '' and address_parts[1].isdigit():
                        business = address_parts[0]
                except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} Business = <{business}>")  


                try:
                    if business.isdigit():
                        business = ''
                    elif business.endswith(" Street") or business.endswith(" Road") or business.endswith(" Tollway") or business.endswith(" Avenue"): 
                        business = ''
                except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} Business2 = <{business}>")  


                if address_parts[0].isdigit():
                    number = address_parts[0]
                    if street == '':
                        street = address_parts[1]
                if number is None:
                    number = ''
                    
                try:        
                    number = number if number.isdigit() else ''
                except Exception as e:
                    print(f"{color_red}Error : {str(e)}{color_reset} number = <{number}>")  


            except Exception as e:
                print(f"{color_red}Error : {str(e)}{color_reset} Business = <{business}> Full address =<{address_parts}>")  

        if type_data == '':
            type_data = active_sheet_title
        print(f'active_sheet_title = {active_sheet_title}   type_data = {type_data}') # temp

# Icon    
        Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''

        if Icon != "":
            Icon = Icon        
        elif type_data == "Calendar":
            Icon = "Calendar"
        elif type_data == "LPR":
            Icon = "LPR"
        elif type_data == "Images":
            Icon = "Images"
        elif type_data == "Intel":
            Icon = "Intel"
        elif type_data == "Locations":
            Icon = "Locations"
            if Subgroup == "SearchedPlaces":
                Icon = "Searched"
            elif Subgroup == "Shared":
                Icon = "Shared"   
            elif Subgroup == "Mentioned":
                Icon = "Locations"  # task
        elif type_data == "Searched Items":
            Icon = "Searched"
        elif type_data == "Toll":
            Icon = "Toll"
        elif type_data == "Videos":
            Icon = "Videos"




# write rows to data
        row_data["Latitude"] = latitude
        row_data["Longitude"] = longitude 
        row_data["Address"] = address
        row_data["business"] = business 
        row_data["number"] = number 
        row_data["street"] = street
        row_data["city"] = city 
        row_data["county"] = county 
        row_data["state"] = state 
        row_data["zipcode"] = zipcode
        row_data["country"] = country 
        row_data["fulladdress"] = fulladdress
        row_data["query"] = query
        row_data["Coordinate"] = coordinate
        row_data["PlusCode"] = PlusCode
        row_data["Icon"] = Icon
        row_data["Radius"] = Radius
        
        print(f'\nName: {name_data}\nCoordinate: {coordinate}\naddress = {address}\nbusiness = {business}\nfulladdress = {fulladdress}\n')

    return data

def read_intel(input_xlsx):
    """Read intel_.xlsx sheet and convert it to locations format.    
    """

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []

# active sheet (current sheet)
    active_sheet = wb.active
    global active_sheet_title
    active_sheet_title = active_sheet.title   


    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        print(f"{color_red}No data found in the Excel file.{color_reset}")
        return None

    for row_index, row_data in enumerate(data):

        (fullname, phone, business, fulladdress, city, state) = ('', '', '', '', '', '')
        (note, sosagent, Time, Latitude, Longitude, Coordinate) = ('', '', '', '', '', '')
        (Source, description, type_data, Name, tag, source) = ('', '', 'Intel', '', '', '')
        (group, subgroup, number, street, county, zipcode) = ('', '', '', '', '', '')
        (country, query, plate, capture_time, hwy, direction) = ('', '', '', '', '', '')
        (end_time, category, time_orig, timezone, PlusCode, Icon) = ('', '', '', '', '', '')
        (Radius) = ('')
# Time
        # Time = ''
        Time = row_data.get("Time")
        if Time is None:
            Time = ''

# Latitude
        # latitude = ''
        latitude = row_data.get("Latitude")
        latitude = str(latitude)
        if latitude is None or latitude == 'None':
            latitude = ''        

# Latitude
        # longitude = ''
        longitude = row_data.get("Longitude")
        longitude = str(longitude)
        if longitude is None or longitude == 'None':
            longitude = ''        

# Coordinate    
        Coordinate = ''
        Coordinate = row_data.get("Coordinate")
        if Coordinate is None:
            Coordinate = ''
            
# address
        address = row_data.get("fulladdress")
        if address is None:
            address = ''    
            
# Name    
        Name = ''
        Name = row_data.get("fullname")
        if Name is None:
            Name = ''

# source
        source = row_data.get("Source")
        if source is None:
            source = ''        
            
# source file
        source_file = row_data.get("Source file information")
        if source_file is None:
            source_file = ''

# original_file
        original_file = row_data.get("original_file")
        if original_file is None or original_file == "":
            original_file = input_xlsx
            
# sosagent    
        sosagent = ''
        sosagent = row_data.get("sosagent")
        if sosagent is None:
            sosagent = ''
            
# phone    
        phone = ''
        phone = row_data.get("phone")
        if phone is None:
            phone = ''
        else:
            description = (f'{description}\nPhone:{phone}')

# business    
        business = ''
        business = row_data.get("business")
        if business is None:
            business = ''
        else:
            description = (f'{description}\nBusiness:{business}')

        description = description.strip()

# city    
        city = ''
        city = row_data.get("city")
        if city is None:
            city = ''

# country    
        country = ''
        country = row_data.get("Country")
        if country is None:
            country = ''


# state    
        state = row_data.get("state")
        if state is None:
            state = ''

# state    
        if state == '':
            state = row_data.get("State/Province")
            if state is None:
                state = ''


# Icon    
        Icon = ''
        Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''

      
# write rows to data
        row_data["Time"] = Time
        row_data["Latitude"] = latitude
        row_data["Longitude"] = longitude 
        row_data["Address"] = address
        # row_data["Group"] = group
        row_data["Subgroup"] = subgroup
        row_data["Description"] = description
        row_data["Type"] = type_data
        row_data["Tag"] = tag        
        row_data["Source"] = source
        row_data["Source file information"] = source_file
        row_data["Name"] = Name
        row_data["business"] = business 
        # row_data["number"] = number 
        # row_data["street"] = street
        row_data["city"] = city 
        # row_data["county"] = county 
        row_data["state"] = state 
        # row_data["zipcode"] = zipcode
        # row_data["country"] = country 
        row_data["fulladdress"] = fulladdress
        # row_data["query"] = query
        # row_data["Plate"] = plate
        # row_data["Capture Time"] = capture_time
        # row_data["Highway Name"] = hwy
        row_data["Coordinate"] = Coordinate
        # row_data["Direction"] = direction
        # row_data["End time"] = end_time
        # row_data["Category"] = category
        # row_data["Time Original"] = time_orig
        # row_data["Timezone"] = timezone
        # row_data["PlusCode"] = PlusCode
        # row_data["Radius"] = Radius
        row_data["Icon"] = Icon
        row_data["original_file"] = original_file # test
        # index
     
     
    return data

def read_locations(input_xlsx):

    """Read data from an xlsx file and return as a list of dictionaries.
    Read XLSX Function: The read_xlsx() function reads data from the input 
    Excel file using the openpyxl library. It extracts headers from the 
    first row and then iterates through the data rows.    
    """

    try:
        wb = openpyxl.load_workbook(input_xlsx)
        ws = wb.active
    except ValueError as e:
        print(f"Error reading : {input_xlsx}")
    data = []
    coordinates = []
    
# active sheet (current sheet)
    active_sheet = wb.active
    global active_sheet_title
    active_sheet_title = active_sheet.title   

    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        print(f"{color_red}No data found in the Excel file.{color_reset}")
        return None

# active sheet (current sheet)
    active_sheet = wb.active
    active_sheet_title = active_sheet.title   


    case_prompt = case_number_prompt()


    for row_index, row_data in enumerate(data):
        (zipcode, business, number, street, city, county) = ('', '', '', '', '', '')
        (state, fulladdress, Latitude, Longitude, query, Coordinate) = ('', '', '', '', '', '')
        (Index, country, capture_time, PlusCode, time_orig, Icon) = ('', '', '', '', '', '')
        (description, group, subgroup, source, source_file, tag) = ('', '', '', '', '', '')
        (Subgroup) = ('')
        (Time, capture_date, timezone, original_file) = ('', '', '', '')  
        (end_time, category, latitude, longitude, coordinate, address) = ('', '', '', '', '', '')
        (from_point, to_point, case) = ('', '', '')
        (type_data, plate, country_code, state_ftk, city_ftk, hwy) = ('', '', '', '', '', '')
        (direction, name_data, no, account, container, time_local) = ('', '', '', '', '', '') 
        (deleted, service_id, carved, sighting_state, sighting_location, manually_decoded) = ('', '', '', '', '', '')
        (origin_latitude, origin_longitude, start_time, location) = ('', '', '', '')
        (Azimuth, Radius, Altitude, time_orig_start, timezone_start) = ('', '', '', '', '')
        (speed, parked) = ('', '')

        if Icon == '':  # there is an icon section at the bottom
            Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''

# Name
        name_data = row_data.get("Name")
        if name_data is None:
            name_data = ''
        if name_data == '':
            name_data = row_data.get("File Name")   # Axiom Pictures
            if name_data is None:
                name_data = ''            
        if name_data == '':
            name_data = row_data.get("Network Name (SSID)")   # iOS Wifi
            if name_data is None:
                name_data = ''   

# no 
        no = row_data.get("#")
        if no is None:
            no = ''
        # add row number as pin name
        # if no == '':
            # no = row_index + 2

# Description    
        description = row_data.get("Description")
        if description is None:
            description = ''

# Time
        Time = row_data.get("Time")
        if Time is None:
            Time = ''

        if Time == '':
            Time = row_data.get("Vicinity Exit Date/Time")
            if Time is None:
                Time = ''

        if Time == '':
            Time = row_data.get("Timestamp Date/Time - UTC+00:00 (M/d/yyyy)")   # Find my locations (Axiom)
            Timezone = 'GMT'
            if Time is None:
                Time = ''
        if Time == '':
            Time = row_data.get("END_DATE")   # Sprint CDMA Cell dump
            if Time is None:
                Time = ''
        if Time == '':
            Time = row_data.get("Created Date/Time - UTC+00:00 (M/d/yyyy)")   # Axiom Apple Maps Searches
            timezone = 'GMT'    # test
            if Time is None:
                Time = ''
        if Time == '':
            Time = row_data.get("Last Joined Date/Time - UTC-06:00 (M/d/yyyy)[DST]")   # Axiom iOS wifi
            if Time is None:
                Time = ''

        if Time == '':
            Time = row_data.get("Created Date/Time - UTC-06:00 (M/d/yyyy)[DST]")   # Axiom Significant locations
            if Time is None:
                Time = ''
                

# Start Time
        start_time = row_data.get("Start Date/Time - UTC-06:00 (M/d/yyyy)[DST]")
        if start_time is None:
            start_time = ''
        if start_time == '':
            start_time = row_data.get("Start Date/Time - UTC+00:00 (M/d/yyyy)")
            if start_time is None:
                start_time = ''

        if start_time == '':
            start_time = row_data.get("Vicinity Entry Date/Time")
            if start_time is None:
                start_time = ''
        if Time == '':
            Time = row_data.get("START_DATE")   # Sprint CDMA Cell dump
            if Time is None:
                Time = ''


# Start Time original
        time_orig_start = row_data.get("time_orig_start")
        if time_orig_start is None:
            time_orig_start = ''

# timezone_start
        timezone_start = row_data.get("timezone_start")
        if timezone_start is None:
            timezone_start = ''
            
# Time
        if Time == '':
            Time = row_data.get("End Date/Time - UTC-06:00 (M/d/yyyy)[DST]")
            if Time is None:
                Time = ''


# time_orig
        time_orig = row_data.get("Time Original")
        if time_orig is None:
            time_orig = ''

# Capture Time
        capture_time  = row_data.get("Capture Time") 
        if capture_time is None:
            capture_time = ''     

        if Time == '':
            if capture_time != '':
                Time = capture_time

# Capture Date
        capture_date  = row_data.get("Capture Date") 
        if capture_date is None:
            capture_date = ''     

        if Time == '':
            if capture_date != '':
                Time = capture_date

# Apple Maps time
        apple_maps_date  = row_data.get("Date/Time - UTC-06:00 (M/d/yyyy)[DST]") 
        if apple_maps_date is None:
            apple_maps_date = ''     

        if Time == '':
            if apple_maps_date != '':
                Time = apple_maps_date
                
# Time (LOCAL)
        time_local  = row_data.get("Time (LOCAL)") 
        if time_local is None:
            time_local = ''     

        if Time == '':
            if time_local != '':
                Time = time_local
                capture_date = time_local
        if Time == '':
            Time  = row_data.get("Start Date/Time - UTC+00:00 (M/d/yyyy)") 
            if Time is None:
                Time = ''               
# GPS tracker
        if Time == '':
            Time  = row_data.get("report_time (GMT)") 
            if Time is None:
                Time = ''    
            else:
                timezone = 'GMT'

# GPS tracker 2
        if Time == '':
            Time  = row_data.get("Date Time (CT)") 
            if Time is None:
                Time = ''    
            else:
                timezone = 'CT'
                
# timezone
        if timezone == '':
            timezone  = row_data.get("Timezone")
            if timezone is None:
                timezone = ''  

# time Gmail warrant retuirn
        time2  = row_data.get("Timestamp (UTC)")
        if time2 is None:
            time2 = ''  
        if Time == "" and time2 != "":
            Time = time2
            timezone = "UTC"
            

# convert time
        # output_format = "%m/%d/%Y %H:%M:%S"  # Changed to military time
        # output_format = "%Y/%m/%d %H:%M:%S"  # Changed to ISO military time
        output_format = "%Y-%m-%d %H:%M:%S"  # Changed to ISO 8601 military time
        # output_format = "%Y-%m-%dT%H:%M:%SZ"    # Google Earth format

        # pattern = r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$'
        pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'  # ISO 8601 ? military time

        if time_orig == '' and Time != '': # copy the original time
            time_orig = Time

        if Time != '':
            try:
                (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
                Time = Time.strftime(output_format)
                
                # print(f'            Time = {Time}   time_orig = {time_orig}')   # temp
                if Time is None:
                    Time = ''              
                
            except ValueError as e:
                print(f"Error time2: {e} - {Time}")
                Time = ''    # temp rem of this
                pass

# start time convert
        if time_orig_start == '' and start_time != '': # copy the original time
            time_orig_start = start_time

        if start_time != '':
            try:
                (start_time, time_orig_start, timezone_start) = convert_timestamp(start_time, time_orig_start, timezone_start)
                start_time = start_time.strftime(output_format)

                if start_time is None:
                    start_time = ''              
                
            except ValueError as e:
                # print(f"Error time3: {e} - {start_time}")
                start_time = '' # temp rem of this
                pass

# End time
        end_time = row_data.get("End time")
        if end_time is None:
            end_time = ''        
        if end_time == '':
            end_time = row_data.get("End Date/Time - UTC+00:00 (M/d/yyyy)")
            if end_time is None:
                end_time = '' 

# Category
        category = row_data.get("Category")
        if category is None:
            category = ''        

# Altitude
        Altitude = row_data.get("Altitude")
        if Altitude is None:
            Altitude = ''  

        if Altitude == '':
            Altitude = row_data.get("Altitude (meters)")    # test Axiom Live Photos
            if Altitude is None:
                Altitude = ''  
        # if Altitude == '':
            # Altitude = row_data.get("altitude")    # GPS tracker
            # if Altitude is None:
                # Altitude = ''  


# gps

        latitude = row_data.get("Latitude")
        latitude = str(latitude)
        if latitude is None or latitude == 'None':
            latitude = ''        
        if latitude == '':
            latitude = row_data.get("latitude")
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = ''             

        longitude = row_data.get("Longitude")
        longitude = str(longitude)
        if longitude is None or longitude == 'None':
            longitude = ''        
        if longitude == '':
            longitude = row_data.get("longitude")
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = ''   



        if latitude == '0.0' or latitude == '0':
            latitude = '' 
        if longitude == '0.0' or longitude == '0':
            longitude = '' 

        # if len(longitude) <3:
            # print(f'blah t{longitude}t')   # temp
            # longitude = ''

        if latitude == '': 
            latitude = row_data.get("Capture Location Latitude")
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = ''        

        if latitude == '': 
            latitude = row_data.get("LAT")
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = ''  
                
      

        if latitude == '': 
            latitude = row_data.get("Destination Latitude")
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = ''   
        if latitude == '': 
            latitude = row_data.get("GPS Latitude")
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = '' 
        if latitude == '': 
            latitude = row_data.get("Latitude (On App Startup)")    # Axiom Uber Accounts
            latitude = str(latitude)
            if latitude is None or latitude == 'None':
                latitude = '' 



                
        if longitude == '':
            longitude = row_data.get("Capture Location Longitude")
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = ''  
                
        if longitude == '': 
            longitude = row_data.get("Destination Longitude")
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = '' 

        if longitude == '': 
            longitude = row_data.get("LONG")
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = '' 

        if longitude == '': 
            longitude = row_data.get("GPS Longitude")
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = '' 
            else:
                longitude_ref = row_data.get("GPS Longitude Reference")
                
                if longitude_ref == 'West':
                    longitude = '-' + longitude

        if longitude == '': 
            longitude = row_data.get("Longitude (On App Startup)")  # Axiom Uber Accounts
            longitude = str(longitude)
            if longitude is None or longitude == 'None':
                longitude = '' 


        (latitude, longitude) = gps_cleanup(latitude, longitude)

        origin_latitude = row_data.get("Origin Latitude")
        origin_latitude = str(origin_latitude)
        if origin_latitude is None or origin_latitude == 'None':
            origin_latitude = ''  

        origin_longitude = row_data.get("Origin Longitude")
        origin_longitude = str(origin_longitude)
        if origin_longitude is None or origin_longitude == 'None':
            origin_longitude = ''  

        if origin_longitude == '' and start_time == '': # cellebrite journey
            to_point = row_data.get("To point")
            if to_point is None or to_point == 'None':
                to_point = '' 
            from_point = row_data.get("From point")
            if from_point is None or from_point == 'None':
                from_point = ''             
            
            if '): (' in to_point:
                to_point = to_point.split('): (')
                if Time == '':
                    if '(' in to_point[0]:
                        Time = to_point[0].split('(')[0]
                        if Time != "":
                            (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
                        
                        timezone = to_point[0].split('(')[1]    # test
                        if latitude == '':
                            coordinate = to_point[1].split('), ')[0]
                            address = to_point[1].split('), ')[1]

                if start_time == '':
                    if '): (' in from_point:
                        start_time = from_point.split('): (')[0]
                        start_time = start_time.split('(')[0]
                        from_point = from_point.split('): (')[1]    # .replace('),','')
                        if start_time != "":
                            (start_time, time_orig_start, timezone_start) = convert_timestamp(start_time, timezone_start, timezone_start)   # test
                        timezone_start = ''
                        
                        if origin_latitude == '':
                            
                            coordinate_start = from_point[1].split('),')[0]
                            
                            from_point = from_point.split(', ')
                            origin_latitude = from_point[0]
                            origin_longitude = from_point[1]    # .split('),')[0]
                            origin_longitude = origin_longitude.replace(')', '')    # .strip()



# coordinate        

        if latitude != '' and longitude != '' and Altitude != '':
            coordinate = (f'{latitude},{longitude},{Altitude}')
            # coordinate = (f'{longitude},{latitude},{Altitude}')
        elif latitude != '' and longitude != '':
            coordinate = (f'{latitude},{longitude}')
            # coordinate = (f'{longitude},{latitude}')
   
        elif row_data.get("Coordinate") != None:
            coordinate = row_data.get("Coordinate")
            # (latitude, longitude, coordinate, Altitude) = long_lat_flip(latitude, longitude, coordinate, Altitude)
        elif row_data.get("Capture Location") != None:
            coordinate = row_data.get("Capture Location")
            # (latitude, longitude, coordinate, Altitude) = long_lat_flip(latitude, longitude, coordinate, Altitude)
        elif row_data.get("Capture Location (Latitude,Longitude)") != None:
            coordinate = row_data.get("Capture Location (Latitude,Longitude)")
            # (latitude, longitude, coordinate, Altitude) = long_lat_flip(latitude, longitude, coordinate, Altitude)
        elif row_data.get("Coordinate(Lat., Long)") != None:
            coordinate = row_data.get("Coordinate(Lat., Long)")
            # (latitude, longitude, coordinate, Altitude) = long_lat_flip(latitude, longitude, coordinate, Altitude)

        # if len(coordinate) > 6:
        coordinate = coordinate.replace('(', '').replace(')', '')

        if latitude == ''  and longitude == '' and coordinate != '':
            if ',' in coordinate:
                coordinate_split = coordinate.split(',')
                latitude = coordinate_split[0].strip()
                longitude = coordinate_split[1].strip()
                try:
                    if Altitude == '':
                        Altitude = coordinate_split[2].strip()
                except:pass
                    
                if Altitude == '':
                    coordinate = f'{latitude},{longitude}'
                elif Altitude != '':
                    coordinate = f'{latitude},{longitude},{Altitude}'

        if latitude != '' and longitude != '' and Time != '':
        # if latitude != '' and longitude != '': # test
            coordinates.append((latitude,longitude, Time))

            # coordinates.append((longitude, latitude))
            # (f"{row['Latitude']}, {row['Longitude']}", row['Time'])
# address
        
        address = row_data.get("Address")
        if address is None:
            address = ''        
        if address == '':
            address = row_data.get("TOWER ADDR")
            if address is None:
                address = ''    
        if address == '':
            address = row_data.get("Estimated Address")
            if address is None:
                address = ''    


# group
        group = row_data.get("Group")
        if group is None:
            group = ''

# subgroup
        subgroup = row_data.get("Subgroup")
        if subgroup is None or subgroup == 'None':
            subgroup = ''
        
        if subgroup == '':
            subgroup = row_data.get("Location Type")    # Axiom find my location / Significant Locations 
            if subgroup is None or subgroup == 'None':
                subgroup = ''
            if subgroup == "Safe Location" or subgroup == "Home":
                tag = 'Of interest'
            
            if subgroup == "Home":
                # tag = 'Of interest'
                if Icon == '' :
                        Icon = 'Home'   # task
                  
# type_data


        if type_data == '':
            type_data = row_data.get("Type")
        if type_data is None:
            type_data = ''  
        if type_data == '':
            type_data = row_data.get("type")
        if type_data is None:
            type_data = ''  

        if type_data == '':
            type_data = active_sheet_title
        # print(f'active_sheet_title = {active_sheet_title}   type_data = {type_data}') # temp

        # Icon = row_data.get("Icon")
        if type_data is None:
            type_data = ''

        if type_data == '':
            if subgroup == '':
                type_data = subgroup
            elif "Search" in original_file:
                type_data = "Searched"
                if Icon == '':
                    Icon = "Searched" 
            elif "Chats" in original_file:
                type_data = "Chats"

            elif "Search" in active_sheet_title:    # test
                # type_data = "Searched"
                if Icon == '':
                    Icon = "Searched" 
            elif active_sheet_title == 'Apple Maps Trips':
                type_data = "Apple Maps Trips"
                Icon = "Map" 

            elif active_sheet_title == 'Searched Items':
                type_data = "Searched"
                if Icon == '':
                    Icon = "Searched" 
            elif active_sheet_title == 'Parked Car Locations':
                type_data = "Parked Car Locations"
                if Icon == '':
                    Icon = "Car4" 
        elif type_data != '' and Icon == '':
            # if type_data == 'gps':
            if type_data == 'gps' or type_data == 'cell':

                Icon = 'Car4'


        # Apple Maps Trips = "Map"
        # Google Maps https://banner2.cleanpng.com/lnd/20240523/jeo/axziljwt9.webp = "Map"
        # Lyft Last Known Location
        # Lyft Location Shortcuts
        # Parked Car Locations
        # PlaceCardTap
        # Apple Maps = "Map"
            # Show
        # Significant Locations
        # Significant Locations Visits
        # Uber Cached Locations




# tag
        if tag == '':
            tag = row_data.get("Tag")
            if tag is None:
                tag = ''
        
        if tag == '':
            tag = row_data.get("Tags")
            if tag is None:
                tag = ''        

# tag from label
        label  = row_data.get("Label") 
        if label is None:
            label = ''     

        if tag == '':
            if label != '':
                tag = label



# source
        if source == '':
            source = row_data.get("Source")
        if source is None:
            source = ''
        if type_data == 'Unknown' or type_data == 'Locations':  # cellebrite Journeys
            type_data = source
        if source == '':
            source = active_sheet_title # test



# carved
        carved = row_data.get("Source file information")
        if carved is None or (isinstance(carved, str) and 'xlsx' in carved):
        # if carved is None or 'xlsx' in carved:
            carved = ''

# source file
        source_file = row_data.get("Source file information")
        if source_file is None:
            source_file = ''

# original_file
        original_file = row_data.get("original_file")
        if original_file is None or original_file == "":
            original_file = input_xlsx


        if 'Apple Maps' in original_file:
            # source = 'Apple Maps'
            type_data = 'Apple Maps'
            icon = "Map"
            if type_data == '':
                type_data = 'Locations'
        elif 'Google Maps' in original_file:
            source = 'Google Maps'
            icon = "Map"
            if type_data == '':
                type_data = 'Locations'
            
# business  
        business = row_data.get("business")
        if business is None:
            business = ''

# fulladdress
        fulladdress  = row_data.get("fulladdress")
        if fulladdress is None:
            fulladdress = ''       
        if fulladdress == '':
            fulladdress  = row_data.get("Location Address") # Find My Locations (Axiom)
            if fulladdress is None:
                fulladdress = ''         
        
        


# query
        query  = row_data.get("query")
        if query is None:
            query = ''     

        if query == '':
            query  = row_data.get("Search Query")   # google maps
            if query is None:
                query = ''  



# Plate
        plate  = row_data.get("Plate")         # red
        if plate is None:
            plate = ''     

        if plate != '' and type_data == '':
            type_data = 'LPR'

# country
        country = row_data.get("country")
        if country is None:
            country = ''     

        if country == '':
            country = row_data.get("Country")
            if country is None:
                country = ''   

# Country Code
        country_code  = row_data.get("Country Code") 
        if country_code is None:
            country_code = ''     

        if country == '':
            if country_code != '':
                country = country_code

# county  
        county = row_data.get("county")
        if county is None:
            county = ''

# state  
        state = row_data.get("state")
        if state is None:
            state = ''
        
        if state == '':
            state = row_data.get("State/Province")
            if state is None:
                state = ''        

# state_ftk
        state_ftk  = row_data.get("Region") 
        if state_ftk is None:
            state_ftk = ''     
        if state == '':
            if state_ftk != '':
                state = state_ftk

        if state == '' and latitude != '' and longitude != '':
            try:
                if 36.970298 <= float(latitude) <= 42.508337 and -91.516837 <= float(longitude) <= -87.3215:
                    state = 'Il'
                elif 40.61364 <= float(latitude) <= 40.61364 and -95.774704 <= float(longitude) <= -89.1745:
                    state = 'Mo'
                elif 41.761089 <= float(latitude) <= 41.761089 and --87.3140 <= float(longitude) <= -84.784579:
                    state = 'In'
            except:
                pass
        if country == '' and latitude != '' and longitude != '':
            try:
                if 24.396308 <= float(latitude) <= 49.384358 and -125.001402 <= float(longitude) <= -66.93457:
                    country = 'US'
            except:pass    

# zipcode  
        zipcode = row_data.get("zipcode")
        if zipcode is None:
            zipcode = ''

        if zipcode == '':
            zipcode = row_data.get("ZIP/Postal Code")
            if zipcode is None:
                zipcode = ''        

        if zipcode == '':
            zipcode = row_data.get("ZIP")
            if zipcode is None:
                zipcode = ''    

# number  
        number = row_data.get("number")
        if number is None:
            number = ''
            
# street  
        street = row_data.get("street")
        if street is None:
            street = ''


# city  
        city = row_data.get("city")
        if city is None:
            city = ''

# City
        city_ftk  = row_data.get("City") 
        if city_ftk is None:
            city_ftk = ''     

        if city == '':
            if city_ftk != '':
                city = city_ftk 
 
        if city == '':
            city = row_data.get("CITY")
            if city is None:
                city = ''
            
# hwy
        hwy  = row_data.get("Highway Name")
        if hwy is None:
            hwy = ''           

        if hwy == '':
            hwy  = row_data.get("Capture Camera")
            if hwy is None:
                hwy = ''           

# Direction
        direction  = row_data.get("Direction")
        if direction is None:
            direction = ''    
        if direction == '':
            direction  = row_data.get("direction")
            if direction is None:
                direction = ''  
        if direction == '':
            direction  = row_data.get("Heading")
            if direction is None:
                direction = ''  

        if direction == '':
            match = re.search(r'\((\w+)\)',hwy)
            if ' NB' in hwy:
                direction = 'N'
            elif ' EB' in hwy:
                direction = 'E'
            elif ' SB' in hwy:
                direction = 'S'
            elif ' WB' in hwy:
                direction = 'W'
            elif match:
                direction = match.group(1).replace('B','')
        
        try:
            direction = direction.replace('Northbound', 'N')
            direction = direction.replace('Eastbound', 'E')
            direction = direction.replace('Southbound', 'S')
            direction = direction.replace('Westbound', 'W')
            direction = direction.replace('Unknown', '')

            direction = direction.replace('North', 'N')
            direction = direction.replace('East', 'E')
            direction = direction.replace('South', 'S')
            direction = direction.replace('West', 'W')
          
            
        except: pass    

        if isinstance(direction, (int, float)) and 0 <= direction <= 360:
            # add original direction to note:
            direction = direction_convert(direction)
            
# PlusCode
        PlusCode  = row_data.get("PlusCode")
        if PlusCode is None:
            PlusCode = ''  

# Icon    
        if Icon == '':
            Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''
        if Icon != "":
            Icon = Icon[0].upper() + Icon[1:].lower()

        if Icon != "":
            Icon = Icon        
        elif type_data == "Calendar":
            Icon = "Calendar"
        elif type_data == "LPR":
            Icon = "LPR"
        elif type_data == "Images":
            Icon = "Images"
        elif type_data == "Picture":    # axiom
            type_data == "Images"   # task
            Icon = "Images"
        elif type_data == "Pictures":    # axiom Pictures
            type_data == "Images"   # task
            Icon = "Images"
        elif type_data == "Live Photos":    # axiom Live Photos
            type_data == "Images"   # task
            if subgroup == '':
                subgroup = 'Live Photos'
            Icon = "Images"
        elif type_data == "Intel":
            Icon = "Intel"
        elif type_data == "Find My Locations":    # axiom
            # type_data == "Images"   # task
            Icon = "Yellow"
        # elif type_data            
            # if subgroup == "SearchedPlaces":
                # Icon = "Searched"
            # elif subgroup == "Shared":
                # Icon = "Shared"   
            # elif subgroup == "Mentioned":
                # Icon = "Locations"  # task
            # elif subgroup == "HarvestedCellTower":
                # Icon = "Locations"   # task add a phone icon or tower
            # elif subgroup == "MediaProbablyCaptured":
                # Icon = "Images"  
            # elif subgroup == "MobilePayment":
                # Icon = "Payment"  # add $ icon
            # elif subgroup == "VehicleParked":
                # Icon = "Car"  ations"


        elif type_data == "Cell Towers":
            Icon = "Locations"  # task add a phone icon or tower

        elif type_data == "Searched Items":
            Icon = "Searched"
        elif type_data == "Toll":
            Icon = "Toll"
        elif type_data == "Videos":
            Icon = "Videos"

        elif subgroup == "SearchedPlaces":  # cellebrite poi searches 
            Icon = "Searched"
        elif subgroup == "Shared":
            Icon = "Shared"   
        elif subgroup == "Mentioned":
            Icon = "Locations"  # task
        elif subgroup == "HarvestedCellTower":
            Icon = "Tower"
        elif subgroup == "MediaProbablyCaptured":
            Icon = "Images"  
        elif subgroup == "MobilePayment":
            Icon = "Payment"  # add $ icon
        elif subgroup == "VehicleParked":
            Icon = "Car"  
        elif tag != "":
            Icon = "Red"  


        '''
        Apple Maps Trips
        Google Maps https://banner2.cleanpng.com/lnd/20240523/jeo/axziljwt9.webp
        Lyft Last Known Location
        Lyft Location Shortcuts
        Parked Car Locations
        PlaceCardTap
        Apple Maps
            Show
        Significant Locations
        Significant Locations Visits
        Uber Cached Locations


        '''


# case    
        if case == '':
            case = row_data.get("case")
        if case is None:
            case = ''
        if case == '':
            case = case_prompt
            
 # case 2   
        # if case == '':
            # case = row_data.get("Evidence number")  # breaks Axiom "Find my locations"
            # if case is None:
                # case = ''           

# location
        location  = row_data.get("Location")
        if location is None:
            location = ''  

# Azimuth
        Azimuth  = row_data.get("Azimuth")
        if Azimuth is None:
            Azimuth = '' 

# Radius
        if Radius == '':
            Radius  = row_data.get("Radius")
            if Radius is None:
                Radius = ''  
        if Radius == '':
            Radius  = row_data.get("Accuracy (meters)")
            if Radius is None:
                Radius = ''  
        if Radius == '':
            Radius  = row_data.get("Accuracy")
            if Radius is None:
                Radius = ''  

        if Radius is None:
            Radius == ''
        else:
            try:
                Radius = float(Radius)
            except ValueError:
                Radius == ''

# speed
        if speed == '':
            speed  = row_data.get("speed")
            if speed is None:
                speed = '' 
        if speed == '':
            speed  = row_data.get("Speed (mph)")
            if speed is None:
                speed = '' 

# parked
        if parked == '':
            parked  = row_data.get("parked")
            if parked is None:
                parked = '' 
        if parked == '':
            parked  = row_data.get("Park Time")
            if parked is None:
                parked = '' 
                
# write rows to data
        row_data["#"] = no
        row_data["Time"] = Time
        row_data["Latitude"] = latitude
        row_data["Longitude"] = longitude 
        row_data["Address"] = address
        row_data["Group"] = group
        row_data["Subgroup"] = subgroup
        row_data["Description"] = description
        row_data["Type"] = type_data
        row_data["Source"] = source
        # row_data["Deleted"] = deleted
        row_data["Tag"] = tag        
        row_data["Source file information"] = source_file
        # row_data["Service Identifier"] = service_id   
        # row_data["Carved"] = carved        
        row_data["Name"] = name_data
        row_data["business"] = business 
        row_data["number"] = number 
        row_data["street"] = street
        row_data["city"] = city 
        row_data["county"] = county 
        row_data["state"] = state 
        row_data["zipcode"] = zipcode
        row_data["country"] = country 
        row_data["fulladdress"] = fulladdress
        row_data["query"] = query
        # row_data["Sighting State"] = sighting_state
        row_data["Plate"] = plate
        row_data["Capture Time"] = capture_time
        row_data["Highway Name"] = hwy
        row_data["Coordinate"] = coordinate
        row_data["Capture Location Latitude"] = latitude        
        row_data["Capture Location Longitude"] = longitude       
        # row_data["Container"] = container       
        # row_data["Sighting Location"] = sighting_location        
        row_data["Direction"] = direction
        # row_data["Time Local"] = time_local
        row_data["End time"] = end_time
        row_data["Category"] = category
        # row_data["Manually decoded"] = manually_decoded        
        # row_data["Account"] = account
        row_data["PlusCode"] = PlusCode        
        row_data["Time Original"] = time_orig
        row_data["Timezone"] = timezone
        row_data["Icon"] = Icon        
        row_data["original_file"] = original_file
        row_data["case"] = case 
        row_data["Origin Latitude"] = origin_latitude
        row_data["Origin Longitude"] = origin_longitude
        row_data["Start Time"] = start_time
        row_data["Location"] = location
        row_data["Azimuth"] = Azimuth
        row_data["Radius"] = Radius
        row_data["Altitude"] = Altitude
        row_data["time_orig_start"] = time_orig_start
        row_data["timezone_start"] = timezone_start
        row_data["speed"] = speed
        row_data["parked"] = parked



    return (data, coordinates)

def point_icon_maker(Icon):        
    # Define different default icons
    # square_icon = 'http://maps.google.com/mapfiles/kml/shapes/square.png'
    # triangle_icon = 'http://maps.google.com/mapfiles/kml/shapes/triangle.png'
    # star_icon = 'http://maps.google.com/mapfiles/kml/shapes/star.png'
    # polygon_icon = 'http://maps.google.com/mapfiles/kml/shapes/polygon.png'
    # circle_icon = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'
    # yellow_circle_icon = 'http://maps.google.com/mapfiles/kml/paddle/ylw-circle.png'
    # red_circle_icon = 'http://maps.google.com/mapfiles/kml/paddle/red-circle.png'
    # white_circle_icon = 'http://maps.google.com/mapfiles/kml/paddle/wht-circle.png'

    ap_unlocked_icon = 'http://maps.google.com/mapfiles/kml/pal4/icon51.png'   # red star
    ap_locked_icon = 'http://maps.google.com/mapfiles/kml/pal4/icon59.png'   # white star
    bluetooth_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon57.png'   # white circle        
    calendar_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon23.png' # paper
    car_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon15.png'   # red car
    car_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon15.png'   # red car
    car2_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon47.png'  # yellow car
    car3_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon54.png'  # green car with circle
    car4_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon7.png'  # red car with circle
    chat_icon = 'https://maps.google.com/mapfiles/kml/shapes/post_office.png' # email
    default_icon = 'https://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png'   # yellow pin
    display_icon = 'http://maps.google.com/mapfiles/kml/pal3/icon61.png'   # small black square      
    green_icon = 'https://maps.google.com/mapfiles/kml/pushpin/grn-pushpin.png'
    home_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon56.png'
    images_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon46.png'
    intel_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon44.png'
    locations_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon28.png'    # yellow paddle
    office_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon21.png'
    orange_icon = 'https://maps.google.com/mapfiles/kml/paddle/orange-blank.png'
    payment_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon50.png'
    red_icon = 'https://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png'
    searched_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon0.png'  #  
    shared_icon = 'https://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png'    # need new pin
    toll_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-none.png'
    tower_icon = 'http://maps.google.com/mapfiles/kml/shapes/target.png' # Bullseye
    truck_icon = 'https://maps.google.com/mapfiles/kml/shapes/truck.png'    # blue truck
    videos_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon30.png'
    yellow_icon = 'https://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png'
 
    n_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-0.png'
    e_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-4.png'
    s_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-8.png'
    w_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-12.png'

   
    if Icon == "Lpr" or Icon == "Car":
        point_icon = car_icon
  
    elif Icon.lower() == "wifi":
        point_icon = ap_locked_icon
    elif Icon.lower() == "wifi-open":
        point_icon = ap_unlocked_icon
    elif Icon.lower() == "bt":
        Icon = 'BT'
        point_icon = bluetooth_icon
    elif Icon.lower() == "display":
        Icon = 'Display'
        point_icon = display_icon
    elif Icon == "Car2":
        point_icon = car2_icon
    elif Icon == "Car3":
        point_icon = car3_icon
    elif Icon == "Car4":
        point_icon = car4_icon
    elif Icon == "Truck":
        point_icon = truck_icon
    elif Icon == "Calendar":
        point_icon = calendar_icon
    elif Icon == "Chat":
        point_icon = chat_icon
    elif Icon == "Home":
        point_icon = home_icon
    elif Icon == "Images":
        point_icon = images_icon
    elif Icon == "Intel":
        point_icon = intel_icon
    elif Icon == "Map":
        point_icon = map_icon
    elif Icon == "Office":
        point_icon = office_icon
    elif Icon == "Payment":
        point_icon = payment_icon
    elif Icon == "Searched":
        point_icon = searched_icon
    elif Icon == "Shared":
        point_icon = shared_icon
    elif Icon == "Videos":
        point_icon = videos_icon
    elif Icon == "Locations":
        point_icon = locations_icon
    elif Icon == "Toll":
        point_icon = toll_icon
    elif Icon.lower() == "tower":
        point_icon = tower_icon
        # print(f'found a tower Icon {Icon}')    # temp
    elif Icon == "Lte":
        point_icon = tower_icon
        # print(f'found a tower Icon {Icon}')    # temp
    elif Icon == "Gsm":
        point_icon = tower_icon
        print(f'found a tower Icon {Icon}')    # temp
        # (kml, point) = radius_azimuth(kml, no, description, latitude, longitude, Azimuth, Radius, Altitude, point_icon)
    elif Icon == "Yellow":
        point_icon = yellow_icon
    elif Icon == "Red":
        point_icon = red_icon
    elif Icon == "Green":
        point_icon = green_icon
    elif Icon == "Orange":
        point_icon = orange_icon

    elif Icon == "N":
        point_icon = n_icon
    elif Icon == "E":
        point_icon = e_icon
    elif Icon == "S":
        point_icon = s_icon
    elif Icon == "W":
        point_icon = w_icon

    else:
        point_icon = default_icon

    return point_icon

def time_compare(time1, time2):
    time_format = "%Y-%m-%d %H:%M:%S"

    # Convert to datetime only if needed
    if isinstance(time1, str):
        time1 = datetime.strptime(time1, time_format)
    if isinstance(time2, str):
        time2 = datetime.strptime(time2, time_format)

    # Compare the dates
    same_day = time1.date() == time2.date()

    # Calculate the difference in minutes
    try:
        time_diff = abs((time2 - time1).total_seconds() / 60.0)
    except Exception:
        time_diff = ''

    return (same_day, time_diff)
    
def write_kml(data):
    '''
    The write_kml() function receives the processed data as a list of 
    dictionaries and writes it to a kml using simplekml. 
    '''

    # Create KML object
    kml = simplekml.Kml()

    for row_index, row_data in enumerate(data):
        (description_start) = ('')
        index_data = row_index + 2  # excel row starts at 2, not 0
        
        no = row_data.get("#") #
        if no is None:
            no = ''
        # if no is None or no == '':
            # no = index_data

        Time = row_data.get("Time") #
        latitude = row_data.get("Latitude") #
        longitude = row_data.get("Longitude") #
        address = row_data.get("Address") #
        group_data = row_data.get("Group") 
        subgroup = row_data.get("Subgroup")   
        description = row_data.get("Description")
        type_data = row_data.get("Type")#
        tag = row_data.get("Tag")
        source_file = row_data.get("Source file information") #
        name_data = row_data.get("Name")
        business = row_data.get("business")
        fulladdress  = row_data.get("fulladdress")
        plate  = row_data.get("Plate")   
        hwy  = row_data.get("Highway Name") #
        coordinate = row_data.get("Coordinate")
        direction  = row_data.get("Direction")
        end_time = row_data.get("End time")
        category = row_data.get("Category")
        Icon = row_data.get("Icon")
        original_file = row_data.get("original_file")
        case = row_data.get("case") # task

        start_time = row_data.get("Start Time") # test
        origin_latitude = row_data.get("Origin Latitude") # test
        origin_longitude = row_data.get("Origin Longitude") # test

        # print(f'2origin_longitude = {origin_longitude}   origin_latitude = {origin_latitude}')   # temp

        Azimuth = row_data.get("Azimuth") # test
        Radius = row_data.get("Radius") # test
        Altitude = row_data.get("Altitude") # test
        time_orig_start = row_data.get("time_orig_start") # test
        timezone_start = row_data.get("timezone_start") # test
        speed = row_data.get("speed") # test
        parked = row_data.get("parked") # test

# Radius
        if Radius is None:
            Radius = ''
        else:
            try:
                Radius = float(Radius)
                # Radius = float_radius
            except ValueError:
                Radius = ''

        if Altitude == '' or Altitude is None:
            Altitude == ''
        
        if name_data != '':
            (description) = (f'{description}\nNAME: {name_data}')

        if coordinate != '':
            (description) = (f'{description}\nCOORDINATE: {coordinate}')    # test
        
        if Time != '':
            (description) = (f'{description}\nTIME: {Time}')

        if end_time != '':
            (description) = (f'{description}\nEnd Time: {end_time}')

        if address != '':
            (description) = (f'{description}\n{address}')

        elif fulladdress != '':
            (description) = (f'{description}\nADDRESS: {fulladdress}')

        if hwy != '':
            (description) = (f'{description}\nHWY NAME: {hwy}')
            
        if direction != '':
            (description) = (f'{description}\nDIRECTION: {direction}')

        # if source_file != '' and source_file != None:
            # (description) = (f'{description}\nSOURCE: {source_file}')

        if tag != '':
            (description) = (f'{description}\nTAG: {tag}')    # test

        if type_data != '':
            (description) = (f'{description}\nTYPE: {type_data}')

        if group_data != '':
            (description) = (f'{description} / {group_data}')

        if subgroup != '' and subgroup != 'UNKNOWN':
            (description) = (f'{description} / {subgroup}')

        if business != '':
            (description) = (f'{description}\nBUSINESS: {business}')
            
        if plate != '':
            (description) = (f'{description}\nPLATE: {plate}')
        if case != '' and case is not None:
            (description) = (f'{description}\nCASE: {case}')
        if Altitude != '' and Altitude != '0' :
            (description) = (f'{description}\nALTITUDE: {Altitude}')
        if Azimuth != '' and Azimuth != '0' :
            (description) = (f'{description}\nAZIMUTH: {Azimuth}')
        if Radius != '' and Radius != '0' :
            (description) = (f'{description}\nRadius: {Radius}')
        if original_file != '':
            (description) = (f'{description}\nSOURCE: {original_file}')
        if speed != '':
            (description) = (f'{description}\nSPEED: {speed}')
        if parked != '':
            (description) = (f'{description}\nPARKED: {parked}')

# row # in description
        (description) = (f'{description}\nROW#: {row_index + 2}') # test


# description_start
        if start_time != '' and start_time is not None:
            (description_start) = (f'{description_start}START TIME: {start_time}')

        point = ''  # Initialize point variable outside the block
        point_icon = point_icon_maker(Icon)
        point_start = ''
        
        if latitude == '' or longitude == '' or latitude == None or longitude == None:
            # print(f'skipping row {index_data} - No GPS')
            skip = 'skip'   # useless variable

        elif isinstance(Radius, float):
            (kml, point) = radius_azimuth(kml, no, description, latitude, longitude, Azimuth, Radius, Altitude, point_icon)

        elif point_icon != '':
            point = kml.newpoint(
                name=f"{no}",                
                description=f"{description}",
                coords=[(longitude, latitude, Altitude)]
            )
            point.style.iconstyle.icon.href = point_icon
            point.altitudemode = simplekml.AltitudeMode.relativetoground

        if tag != '':   # mark label yellow if tag is not blank
            try:
                point.style.labelstyle.color = simplekml.Color.yellow  # Set label text color
                # point.style.labelstyle.scale = 1.2  # Adjust label scale if needed    # task
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

        if origin_latitude != '' and origin_longitude != '':
            point_start = kml.newpoint(
                name=f"{no}_start",                
                description=f"{description_start}",
                coords=[(origin_longitude, origin_latitude)]
              
            )
# Create line and specify style
            line = kml.newlinestring(name="trip line", coords=[(longitude, latitude), (origin_longitude, origin_latitude)])

            # line.style.linestyle.color = simplekml.Color.red  # Make the line red
            # line.style.linestyle.color = simplekml.Color.orange  # Make the line orange
            # line.style.linestyle.color = simplekml.Color.yellow  # Make the line yellow
            # line.style.linestyle.color = simplekml.Color.green  # Make the line green
            line.style.linestyle.color = simplekml.Color.blue  # Make the line blue
            # line.style.linestyle.color = simplekml.Color.purple  # Make the line purple
            # line.style.linestyle.color = simplekml.Color.white  # Make the line white
            # line.style.linestyle.color = simplekml.Color.black  # Make the line black

            line.style.linestyle.width = 2  # Set line width

    kml.save(output_kml)    # Save the KML document to the specified output file

    msg_blurb = (f'KML file {output_kml} created successfully!')
    msg_blurb_square(msg_blurb, color_blue)

def write_locations(data):
    '''
    The write_locations() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''
    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    # print(f'Reading {active_sheet_title} sheet\n')
    try:
        worksheet.title = active_sheet_title
    except Exception as e:
        print(f"{color_red}Error : {str(e)}{color_reset}")

    # worksheet.title = 'Locations'
    # header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'C2'  # Freeze cells
    worksheet.selection = 'B2'

    # Apply default style to the header row (bold font)
    for cell in worksheet[1]:   # test
        cell.font = Font(bold=True)


    headers = [
        "#", "Time", "Latitude", "Longitude", "Address", "Group", "Subgroup"
        , "Description", "Type", "Source", "Deleted", "Tag"
        , "Source file information", "Service Identifier", "Carved", "Name"
        , "business", "number", "street", "city", "county", "state", "zipcode"
        , "country", "fulladdress", "query", "Sighting State", "Plate"
        , "Capture Time", "Capture Network", "Highway Name", "Coordinate"
        , "Capture Location Latitude", "Capture Location Longitude", "Container"
        , "Sighting Location", "Direction", "Time Local", "End time", "Category"
        , "Manually decoded", "Account", "PlusCode", "Time Original", "Timezone"
        , "Icon", "original_file", "case", "Origin Latitude", "Origin Longitude"
        , "Start Time", "Azimuth", "Radius", "Altitude", "Location"
        , "time_orig_start", "timezone_start", "Index", "speed", "parked", "MAC"
    ]

    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [2, 3, 4]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [1, 5, 6, 7, 8, 9, 15, 16, 24, 30, 31, 36, 38]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill
        elif col_index == 27:  # Red for column 27
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
            cell.fill = fill

    ## Excel column width
    worksheet.column_dimensions['A'].width = 8# #
    worksheet.column_dimensions['B'].width = 19# Time
    worksheet.column_dimensions['C'].width = 18# Latitude
    worksheet.column_dimensions['D'].width = 18# Longitude
    worksheet.column_dimensions['E'].width = 45# Address
    worksheet.column_dimensions['F'].width = 14# Group
    worksheet.column_dimensions['G'].width = 13# Subgroup
    worksheet.column_dimensions['H'].width = 17# Description
    worksheet.column_dimensions['I'].width = 9# Type
    worksheet.column_dimensions['J'].width = 10# Source
    worksheet.column_dimensions['K'].width = 10# Deleted
    worksheet.column_dimensions['L'].width = 11# Tag
    worksheet.column_dimensions['M'].width = 20# Source file information
    worksheet.column_dimensions['N'].width = 15# Service Identifier
    worksheet.column_dimensions['O'].width = 7# Carved
    worksheet.column_dimensions['P'].width = 15# Name
    
    ## bonus
    worksheet.column_dimensions['Q'].width = 20# business 
    worksheet.column_dimensions['R'].width = 10# number
    worksheet.column_dimensions['S'].width = 20# street 
    worksheet.column_dimensions['T'].width = 15# city   
    worksheet.column_dimensions['Y'].width = 25# county    
    worksheet.column_dimensions['V'].width = 12# state   
    worksheet.column_dimensions['W'].width = 8# zipcode     
    worksheet.column_dimensions['X'].width = 6# country    
    worksheet.column_dimensions['Y'].width = 26# FullAddress   
    worksheet.column_dimensions['Z'].width = 26# query

    ##  Flock
    worksheet.column_dimensions['AA'].width = 11# Sighting State
    worksheet.column_dimensions['AB'].width = 11# Plate
    worksheet.column_dimensions['AC'].width = 22# Capture Time
    worksheet.column_dimensions['AD'].width = 15# Capture Network
    worksheet.column_dimensions['AE'].width = 21# Highway Name
    worksheet.column_dimensions['AF'].width = 30# Coordinate
    worksheet.column_dimensions['AG'].width = 20# Capture Location Latitude
    worksheet.column_dimensions['AH'].width = 20# Capture Location Longitude

    ##
    worksheet.column_dimensions['AI'].width = 10# Container
    worksheet.column_dimensions['AJ'].width = 14# Sighting Location
    worksheet.column_dimensions['AK'].width = 10# Direction
    worksheet.column_dimensions['AL'].width = 11# Time Local
    worksheet.column_dimensions['AM'].width = 25# End time
    worksheet.column_dimensions['AN'].width = 10# Category
    worksheet.column_dimensions['AO'].width = 18# Manually decoded
    worksheet.column_dimensions['AP'].width = 10# Account
    worksheet.column_dimensions['AQ'].width = 25 # PlusCode
    worksheet.column_dimensions['AR'].width = 21 # Time Original
    worksheet.column_dimensions['AS'].width = 9 # Timezone
    worksheet.column_dimensions['AT'].width = 10 # Icon   
    worksheet.column_dimensions['AU'].width = 20 # original_file
    worksheet.column_dimensions['AV'].width = 10 # case
    worksheet.column_dimensions['AW'].width = 20 # Origin Latitude
    worksheet.column_dimensions['AX'].width = 20 # Origin Longitude
    worksheet.column_dimensions['AY'].width = 25 # Start Time
    worksheet.column_dimensions['AZ'].width = 20 # Azimuth
    worksheet.column_dimensions['BA'].width = 6 # Radius
    worksheet.column_dimensions['BB'].width = 20 # Altitude
    worksheet.column_dimensions['BC'].width = 20 # Location
    worksheet.column_dimensions['BD'].width = 25 # time_orig_start
    worksheet.column_dimensions['BE'].width = 17 # timezone_start
    worksheet.column_dimensions['BF'].width = 6 # Index

    
    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")


    # Create a new worksheet for color codes
    color_worksheet = workbook.create_sheet(title='Icons')
    color_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    color_worksheet.column_dimensions['A'].width = 11# Icon sample
    color_worksheet.column_dimensions['B'].width = 11# Name
    color_worksheet.column_dimensions['C'].width = 65# Description

    # Excel row height
    color_worksheet.row_dimensions[2].height = 22  # Adjust the height as needed
    color_worksheet.row_dimensions[3].height = 22
    color_worksheet.row_dimensions[4].height = 23
    color_worksheet.row_dimensions[5].height = 23
    color_worksheet.row_dimensions[6].height = 40   # truck
    color_worksheet.row_dimensions[7].height = 6
    color_worksheet.row_dimensions[8].height = 24
    color_worksheet.row_dimensions[9].height = 22
    color_worksheet.row_dimensions[10].height = 22
    color_worksheet.row_dimensions[11].height = 22
    color_worksheet.row_dimensions[12].height = 23
    color_worksheet.row_dimensions[13].height = 55  # Yellow pin
    color_worksheet.row_dimensions[14].height = 25
    color_worksheet.row_dimensions[15].height = 25
    color_worksheet.row_dimensions[16].height = 23
    color_worksheet.row_dimensions[17].height = 6
    color_worksheet.row_dimensions[18].height = 38
    color_worksheet.row_dimensions[19].height = 38
    color_worksheet.row_dimensions[20].height = 38
    color_worksheet.row_dimensions[21].height = 38
    color_worksheet.row_dimensions[22].height = 38
    color_worksheet.row_dimensions[23].height = 6
    color_worksheet.row_dimensions[24].height = 15 # yellow font
    color_worksheet.row_dimensions[25].height = 55  # red_icon
    color_worksheet.row_dimensions[26].height = 50  # chats

    color_worksheet.row_dimensions[27].height = 28 # payment
    color_worksheet.row_dimensions[28].height = 50  # Tower
    color_worksheet.row_dimensions[29].height = 15  # bluetooth
    color_worksheet.row_dimensions[30].height = 22  # WIFI-open
    color_worksheet.row_dimensions[31].height = 22  # WIFI
    color_worksheet.row_dimensions[33].height = 15
    color_worksheet.row_dimensions[33].height = 15


    
    # Define color codes
    color_worksheet['A1'] = ' '
    color_worksheet['B1'] = 'Icon'
    color_worksheet['C1'] = 'Icon Description'

    icon_data = [

        ('', 'Car', 'LPR red car (License Plate Reader)'),
        ('', 'Car2', 'LPR yellow car'),
        ('', 'Car3', 'LPR green car with circle'),
        ('', 'Car4', 'LPR red car with circle'),
        ('', 'Truck', 'LPR truck'),         
        ('', '', ''),
        ('', 'Calendar', 'Calendar'), 
        ('', 'Home', 'Home'),                
        ('', 'Images', 'Photo'),
        ('', 'Intel', 'blue I for Intel'),  
        ('', 'Locations', 'Reticle'),  
        ('', 'Yellow', 'Yellow pin (default)'),  
        ('', 'Office', 'Office'),         
        ('', 'Searched', 'Searched Item'),          
        ('', 'Videos', 'Video clip'),        
        ('', '', ''),
        ('', 'Toll', 'Blue square'), 
        ('', 'N', 'Northbound blue arrow'),
        ('', 'E', 'Eastbound blue arrow'),
        ('', 'S', 'Southbound blue arrow'),
        ('', 'W', 'Westbound blue arrow'),
        ('', '', ''),
        ('', 'Yellow font', 'Tagged'),
        ('', 'Red', 'Red pin for Tagged items'),        
        ('', 'Chats', 'Chats'),   # 
        ('', 'Payment', 'Payment'),        
        ('', 'Tower', 'Bullseye'),
        ('', 'Bluetooth', 'white circle'),
        ('', 'WIFI-open', 'red star'),
        ('', 'WIFI', 'white star'),
        ('', 'Display/Sound', 'white square'),        
        ('', '', ''),
        ('', 'blue lines', 'trips with a start and end'),
        ('', 'red lines', 'coordinates with timestamps within a short period of time (like same day)'),        
        ('', 'red circles', 'indicate radius of the signal and/or accuracy of the point'),
        ('', 'NOTE', 'visit https://earth.google.com/ <file><Import KML> select gps.kml <open>'),
        ('', '', 'Timestamps are in ISO 8601 military time (Year-Month-Day Hour:Minute:Seconds)')
    ]

    for row_index, (icon, tag, description) in enumerate(icon_data):
        color_worksheet.cell(row=row_index + 2, column=1).value = icon
        color_worksheet.cell(row=row_index + 2, column=2).value = tag
        color_worksheet.cell(row=row_index + 2, column=3).value = description
    ap_unlocked_icon = 'http://maps.google.com/mapfiles/kml/pal4/icon51.png'   # red star
    ap_locked_icon = 'http://maps.google.com/mapfiles/kml/pal4/icon59.png'   # white star
    bluetooth_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon57.png'   # white circle        
    calendar_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon23.png' # paper
    car_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon15.png'   # red car
    car_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon15.png'   # red car
    car2_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon47.png'  # yellow car
    car3_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon54.png'  # green car with circle
    car4_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon7.png'  # red car with circle
    chat_icon = 'https://maps.google.com/mapfiles/kml/shapes/post_office.png' # email
    default_icon = 'https://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png'   # yellow pin
    display_icon = 'http://maps.google.com/mapfiles/kml/pal3/icon61.png'   # small black square    
    green_icon = 'https://maps.google.com/mapfiles/kml/pushpin/grn-pushpin.png'
    home_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon56.png'
    images_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon46.png'
    intel_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon44.png'
    locations_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon28.png'    # yellow paddle
    office_icon = 'https://maps.google.com/mapfiles/kml/pal3/icon21.png'
    orange_icon = 'https://maps.google.com/mapfiles/kml/paddle/orange-blank.png'
    payment_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon50.png'
    red_icon = 'https://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png'
    searched_icon = 'https://maps.google.com/mapfiles/kml/pal4/icon0.png'  #  
    toll_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-none.png'
    tower_icon = 'http://maps.google.com/mapfiles/kml/shapes/target.png' # Bullseye
    truck_icon = 'https://maps.google.com/mapfiles/kml/shapes/truck.png'    # blue truck
    videos_icon = 'https://maps.google.com/mapfiles/kml/pal2/icon30.png'
    yellow_icon = 'https://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png'
 
    n_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-0.png'
    e_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-4.png'
    s_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-8.png'
    w_icon = 'https://earth.google.com/images/kml-icons/track-directional/track-12.png'
    
    try:
        # Insert graphic from URL into cell of color_worksheet

        response = requests.get(car_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A2')

        response = requests.get(car2_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A3')

        response = requests.get(car3_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A4')

        response = requests.get(car4_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A5')

        response = requests.get(truck_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A6')

        response = requests.get(calendar_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A8')
        
        response = requests.get(home_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A9')

        response = requests.get(images_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A10')

        response = requests.get(intel_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A11')

        response = requests.get(locations_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A12')

        response = requests.get(default_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A13')

        response = requests.get(office_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A14')        

        response = requests.get(searched_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A15')

        response = requests.get(videos_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A16')
        
        response = requests.get(toll_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A18')

        response = requests.get(n_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A19')

        response = requests.get(e_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A20')

        response = requests.get(s_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A21')

        response = requests.get(w_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A22')

        response = requests.get(red_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A25')


        response = requests.get(chat_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A26')     

        response = requests.get(payment_icon)   # task
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A27')  

        response = requests.get(tower_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A28')
 
        response = requests.get(bluetooth_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A29')

        response = requests.get(ap_unlocked_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A30')

        response = requests.get(ap_locked_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A31')

        response = requests.get(display_icon)
        img = Image(io.BytesIO(response.content))
        color_worksheet.add_image(img, 'A32')
        
    except:
        pass

    
    workbook.save(output_xlsx)

def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description2}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}insert your input into locations.xlsx')
    print(f'\nExample:')
    print(f'    {file} -c -O input_blank.xlsx') 
    print(f'    {file} -k -I locations.xlsx  # xlsx 2 kml with no internet processing')     
    print(f'    {file} -K -I Locations.kml  # kml 2 xlsx with no internet processing')     

    print(f'    {file} -r -I locations.xlsx -O Locations_.xlsx')
    print(f'    {file} -R')
    print(f'    {file} -R -I locations.xlsx -O Locations_.xlsx') 
    print(f'    {file} -R -I locations_FTK.xlsx -O Locations_.xlsx') 
    print(f'    {file} -R -I Flock.xlsx -O Locations_Flock.xlsx')   
    print(f'    {file} -R -I Journeys.xlsx -O Locations_Journeys.xlsx   # beta')  
    print(f'    {file} -R -I MediaLocations_.xlsx')  
    print(f'    {file} -R -I PointsOfInterest.xlsx -O Locations_PointsOfInterest.xlsx') 
    print(f'    {file} -R -I Tolls.xlsx -O Locations_Tolls.xlsx')     
    print(f'    {file} -i -I intel_.xlsx -O Locations_Intel_.xlsx')  
    print(f'    {file} -i -I intel_SearchedItems_.xlsx')  
    print(f'    {file} -i -I intel_Chats_.xlsx')  

    
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
1.3.7 - added Bluetooth, cell tower and AP points
1.3.1 - Trip path (default blue line, but can change color manually)
1.3.0 - Injest Celltower dumps (Azimuth, Radius)
1.2.1 - can populate # with custom display or if it's blank it will put in the column number
1.2.0 - added icons
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
line 2610 errors sometimes point_icon = shared_icon


test https://github.com/AXYS-Cyber/iCatch.git
add a 10,000 line splitter to create seperate kml's when it's over 10,000



fix the bluetooth_icon- it's very big. maybe pick a different one
https://github.com/UMIT2Lab/iGem

batchgeo.com is an online alternative (similar)
research Gx:trax module

compare timestamps and create trips based on timestamp (same day or within so many minutes.)
(same_day, time_diff) = time_compare(Time_previous, Time)
placeid (google)
travel path only does white lines (fix that) randomize trip line colors for different trips
try Castviz
integrate altitude & altitudeMode 
convert .kml to xlsx
Output for CellHawk Generic Cell tower output

bearing (=direction), speed

create a list of coordinates[Coordinate:full address] if the contacts already exists, don't check for the full address
if the internet isn't connected, don't try to check full address

if address / fulldata only, get lat/long
if it's less than 3000 skip the sleep timer

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
avoid processing Google Maps Tiles, Apple Maps - Biome App Intents.xlsx, iOS maps
connection timeout after about 4000 attempts
with the sleep timer set to 10 (sec) it doesn't crap out.

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

