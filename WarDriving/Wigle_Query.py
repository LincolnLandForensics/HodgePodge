 #!/usr/bin/env python3
# coding: utf-8
'''
read xlsx, write xlsx with only openpyxl
Pandas is too big
send data all at once so it can be sorted if needed
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>> 

import json
import requests
from requests.auth import HTTPBasicAuth

 
import os
import re
import sys
import time
import openpyxl
import simplekml    # pip install simplekml
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import argparse  # for menu system
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build
    if major_version >= 2 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = "Read WIFI & BT MAC addresses and query Wigle for sightings"
version = '0.1.2'

global USERNAME
# must sign up for a wigle api key, similar to this username:password
USERNAME = ""    # AID19691969196919691969196919691969
global PASSWORD
PASSWORD = ""   # b77e1969196919691969196919691969



global headers
# update the headers if you don't want all of them.
headers = [
    "#", "Time", "Latitude", "Longitude", "Address", "Group", "Subgroup"
    , "Description", "Type", "Source", "Deleted", "Tag"
    , "Source file information", "Service Identifier", "Carved", "Name"
    , "business", "number", "street", "city", "county", "state", "zipcode"
    , "country", "fulladdress", "query", "Sighting State", "Plate"
    , "Capture Time", "Capture Network", "Highway Name", "Coordinate"
    , "Capture Location Latitude", "Capture Location Longitude", "Container"
    , "Sighting Location", "Direction", "Time Local", "End time", "Category"
    , "Manually decoded", "Account", "PlusCode", "Time Original", "Timezone"
    , "Icon", "original_file", "case", "Origin Latitude", "Origin Longitude"
    , "Start Time", "Azimuth", "Radius", "Altitude", "Location"
    , "time_orig_start", "timezone_start", "Index", "speed", "parked"
    , "MAC", "SSID", "AuthMode", "Channel", "Frequency", "dB"
    , "AltitudeMeters", "AccuracyMeters", "RCOIs", "MfgrId", "CompanyName"
    , "Data", "Packet Type", "Hits", "length"
]


global companies
companies = {
        "74:EC:B2": "Amazon Technologies Inc.",
        "E4:F0:42": "Google, Inc.",
        "A0:D7:F3": "Samsung Electronics Co.,Ltd",
        "D4:3A:2C": "Google, Inc."
    }
    
    

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-b', '--blank', help='create blank input sheet', required=False, action='store_true')
    parser.add_argument('-Q', '--query', help='wigle query', required=False, action='store_true')

    args = parser.parse_args()
    data = []

    global input_xlsx
    input_xlsx = args.input if args.input else "input_.xlsx"

    global output_xlsx
    output_xlsx = args.output if args.output else "output_.xlsx"

    if args.blank:
        data = []
        print(f'{color_green}Writing to {output_xlsx} {color_reset}')
        write_xlsx(data,output_xlsx)
        
    elif args.query:
            
        file_exists = os.path.exists(input_xlsx)

        datatype = input_xlsx
        datatype = datatype.replace('.xlsx', '')

        if not args.output: 
            output_xlsx = (f'intel_{datatype}.xlsx') 
  
        else:
            output_xlsx = args.output


        if file_exists == True:
            msg_blurb = (f'Reading {input_xlsx}')
            msg_blurb_square(msg_blurb, color_green)

            data = read_xlsx(input_xlsx)
            # write_xlsx(data
            write_xlsx(data,output_xlsx)

            workbook.close()
            msg_blurb = (f'Writing to {output_xlsx}')
            msg_blurb_square(msg_blurb, color_green)            
            # print(f'{color_green}Writing to {output_xlsx} {color_reset}')

        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()

    else:
        usage()
    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>


def companies_read():
    # Check if the CSV file exists
    csv_file = 'mac-vendors-export.csv'
    if os.path.exists(csv_file):
        # Read the CSV file and update the companies dictionary
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                mac_prefix = row[0]
                vendor_name = row[1]
                companies[mac_prefix] = vendor_name


def company_lookup(MAC):
    company = ''

    # Extract the first 8 characters of MAC
    mac_prefix = MAC[:8]  # Get the first 8 characters

    # Return the company name if found, else an empty string
    return companies.get(mac_prefix, "") 

    
def convert_timestamp(timestamp, time_orig, timezone):
    if timezone is None:
        timezone = ''
    if time_orig is None:
        time_orig = ''

    timestamp = str(timestamp)

    # Regular expression to find all timezones
    timezone_pattern = r"([A-Za-z ]+)$"
    matches = re.findall(timezone_pattern, timestamp)

    # Extract the last timezone match
    if matches:
        timezone = matches[-1]
        timestamp = timestamp.replace(timezone, "").strip()
    else:
        timezone = ''
        
    if time_orig == "":
        time_orig = timestamp
    else:
        timezone = ''


    # timestamp = timestamp.replace(' at ', ' ')
    if "(" in timestamp:
        timestamp = timestamp.split('(')
        timezone = timestamp[1].replace(")", '')
        timestamp = timestamp[0]
    elif " CDT" in timestamp:
        timezone = "CDT"
        timestamp = timestamp.replace(" CDT", "")
    elif " CST" in timestamp:
        timezone = "CST"
        timestamp = timestamp.replace(" CST", "")




    formats = [
        "%B %d, %Y, %I:%M:%S %p %Z",    # June 13, 2022, 9:41:33 PM CDT (Flock)
        "%Y:%m:%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",  # timestamps without seconds
        "%m/%d/%Y %H:%M:%S",  # timestamps in military time without seconds
        "%m-%d-%y at %I:%M:%S %p %Z", # test 09-10-23 at 4:29:12 PM CDT
        "%m-%d-%y %I:%M:%S %p",
        "%B %d, %Y at %I:%M:%S %p %Z",
        "%B %d, %Y at %I:%M:%S %p",
        "%B %d, %Y %I:%M:%S %p %Z",
        "%B %d, %Y %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p %Z",
        "%Y-%m-%dT%H:%M:%SZ",  # ISO 8601 format with UTC timezone
        "%Y/%m/%d %H:%M:%S",  # 2022/06/13 21:41:33
        "%d-%m-%Y %I:%M:%S %p",  # 13-06-2022 9:41:33 PM
        "%d/%m/%Y %H:%M:%S",  # 13/06/2022 21:41:33
        "%Y-%m-%d %I:%M:%S %p",  # 2022-06-13 9:41:33 PM
        "%Y%m%d%H%M%S",  # 20220613214133
        "%Y%m%d %H%M%S",  # 20220613 214133
        "%m/%d/%y %H:%M:%S",  # 06/13/22 21:41:33
        "%d-%b-%Y %I:%M:%S %p",  # 13-Jun-2022 9:41:33 PM
        "%d/%b/%Y %H:%M:%S",  # 13/Jun/2022 21:41:33
        "%Y/%b/%d %I:%M:%S %p",  # 2022/Jun/13 9:41:33 PM
        "%d %b %Y %H:%M:%S",  # 13 Jun 2022 21:41:33
        "%A, %B %d, %Y %I:%M:%S %p %Z",  # Monday, June 13, 2022 9:41:33 PM CDT ?
        "%A, %B %d, %Y %I:%M:%S %p"     # Monday, June 13, 2022 9:41:33 PM CDT
    ]

    for fmt in formats:
        try:
            dt_obj = datetime.strptime(timestamp.strip(), fmt)
            timestamp = dt_obj
            return timestamp, time_orig, timezone
                        
        except ValueError:
            pass

    raise ValueError(f"{time_orig} Timestamp format not recognized")



def identify_device_type(CompanyName, service_uuids):
    """Determines the most likely device type based on UUIDs and manufacturer"""
    if CompanyName in DEVICE_TYPES:
        for device_type, uuid_list in DEVICE_TYPES[CompanyName].items():
            if any(uuid in service_uuids for uuid in uuid_list):
                return device_type
    return "Unknown Device Type"    
    
def msg_blurb_square(msg_blurb, color):
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"
    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def read_xlsx(input_xlsx):

    """Read data from an xlsx file and return as a list of dictionaries.
    Read XLSX Function: The read_xlsx() function reads data from the input 
    Excel file using the openpyxl library. It extracts headers from the 
    first row and then iterates through the data rows.    
    """

    if USERNAME == '':
        msg_blurb = (f"you need to enter your Wigle API USERNAME key, or this won't work")
        msg_blurb_square(msg_blurb, color_red)        
        exit()
    # if PASSWORD == '':
    if len(PASSWORD) < 15:
        msg_blurb = (f'you need to enter your Wigle API PASSWORD key')
        msg_blurb_square(msg_blurb, color_red)
            
        
        
        exit()
    # USERNAME = "AID3eb88c99ae363e3385bb9058dd262e5a"
    # PASSWORD = "b77efe08e11e816fb77270c9c2938d97"

    
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []
    datatype = input_xlsx
    datatype = datatype.replace('.xlsx', '')
    
    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        msg_blurb = (f'No data found in the Excel file: {input_xlsx}')
        msg_blurb_square(msg_blurb, color_red)
    
        exit()
        return None

# active sheet (current sheet)
    active_sheet = wb.active
    global active_sheet_title
    active_sheet_title = active_sheet.title    


    for row_index, row_data in enumerate(data):
        (Time, time_orig, timezone, response, tag, Description) = ('', '', '', '', '', '')
        (Latitude, Longitude, Coordinate, raw_data, Time) = ('', '', '', '', '')
        (city, state, country, Altitude, note, MAC) = ('', '', '', '', '', '')
        (source, source_file, original_file, Icon, zipcode, Channel) = ('', '', '', '', '', '')
        (number, street, Subgroup, fulladdress, SSID, CompanyName) = ('', '', '', '', '', 'temp')
# type_data
        Description = row_data.get("Description")
        if Description is None:
            Description = ''          
        
# type_data
        type_data = row_data.get("Type")
        if type_data is None:
            type_data = ''        

# Data
        raw_data = row_data.get("Data")
        if raw_data is None:
            raw_data = ''               
       
# MAC
        MAC = row_data.get("MAC")
        if MAC is None:
            MAC = ''         
        if MAC != '':
            print(f'{color_blue}MAC={MAC}{color_reset}' )
            if type_data == 'WIFI' or type_data == 'BT' or type_data == 'BLE':
                print(f'{color_yellow}sleeping for 40 seconds{color_reset}' )
                time.sleep(40)            
                result = query_wigle_wifi_api(USERNAME, PASSWORD, type_data, MAC)
   
                response = json.dumps(result)  # Convert dictionary to JSON string
                parsed_response = response_parse_wifi(response)

                Description = (f'{Description}{parsed_response}')
                Data = response

                
                print(parsed_response)

                if parsed_response:
                    Latitude = parsed_response["Latitude"]
                    Longitude = parsed_response["Longitude"]
                    SSID = parsed_response["SSID"]
                    Time = parsed_response["Time"]
                    number = parsed_response["number"]
                    street = parsed_response["street"]
                    city = parsed_response["city"]
                    state = parsed_response["state"]
                    country = parsed_response["country"]
                    zipcode = parsed_response["zipcode"]
                    MAC = parsed_response["MAC"]
                    Subgroup = parsed_response["Subgroup"]
                    Channel = parsed_response["Channel"]
                    SSID = parsed_response["SSID"]
                    Description = (f'Latitude: {Latitude}\nLongitude: {Longitude}\nTime: {Time}\nNumber: {number}\nstreet: {street}\ncity: {city}\nState: {state}\ncountry: {country}\nZipcode: {zipcode}\nMAC: {MAC}\nSubgroup: {Subgroup}\nChannel: {Channel})')
                
        if Latitude is None or Latitude == '':
            Latitude = row_data.get("Latitude")
            Longitude = row_data.get("Longitude")
            if Latitude is None:
                Latitude == ''
                Longitude == ''

# state
        if state == '':
            state = row_data.get("state")
        if state is None:
            state = ''

# city
        if city == '':
            city = row_data.get("city")
        if city is None:
            city = ''

# country
        if country == '':
            country = row_data.get("country")
        if country is None:
            country = ''


# fulladdress
        if fulladdress == '':
        # if number == '' and street == '' and city == '' and state == '' and zipcode == '':
            fulladdress = (f'{number} {street}, {city}, {state} {zipcode}')
        if fulladdress == ' , ,  ':
            fulladdress = ''
            
            
# source
        if source == '':
            source = row_data.get("Source")
        if source is None:
            source = ''


# source file
        if source_file == '':
            source_file = row_data.get("Source file information")
        if source_file is None:
            source_file = ''

# original_file
        if original_file == '':
            original_file = row_data.get("original_file")
        if original_file is None or original_file == "":
            original_file = input_xlsx
          
# Icon    
        if Icon == '':
            Icon = row_data.get("Icon")
        # Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''
        if Icon == "":
            if "Searched" in original_file:
               Icon = "Searched"            
            elif "Chats" in original_file:
               Icon = "Chats"  


# Time
        if Time == '':        
            Time = row_data.get("Time")
        if Time is None:
            Time = ''

        if (Coordinate == '' or Coordinate is None) and Altitude == '':
            if Latitude is None:
                Latitude == ''
            if Longitude is None:
                Longitude == ''                
            else:    
                Coordinate = (f'{Latitude},{Longitude}')
            if 'None' in Coordinate:
                Coordinate == ''

# convert time
        output_format = "%Y-%m-%d %H:%M:%S "    # ISO 8601
        # output_format = "%Y-%m-%dT%H:%M:%SZ"    # Google Earth format
        # pattern = r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$'
        pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'  # ISO military time

        if time_orig == '' and Time != '': # copy the original time
            time_orig = Time
        # try:
            # (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
            # Time = Time.strftime(output_format)
            # if Time is None:
                # Time = ''              
            
        # except ValueError as e:
            # if Time != "":
                # print(f"Error time2: {e} - {Time}")
                # Time = ''    # temp rem of this

        if Coordinate == 'None,None':
            Coordinate == ''

# CompanyName
        CompanyName = company_lookup(MAC) 
        print(f'CompanyName = {CompanyName}')   # temp
    
                
# write rows to data

        row_data["country"] = country        
        row_data["Description"] = Description  
        row_data["note"] = note
        row_data["MAC"] = MAC    
        row_data["number"] = number   
        row_data["street"] = street  
        row_data["Subgroup"] = Subgroup          
        row_data["city"] = city 
        row_data["zipcode"] = zipcode
        row_data["state"] = state 
        row_data["fulladdress"] = fulladdress 
        row_data["Time"] = Time
        row_data["Latitude"] = Latitude  
        row_data["Longitude"] = Longitude  
        row_data["Coordinate"] = Coordinate  
        row_data["Source file information"] = source_file     
        row_data["original_file"] = original_file     
        # row_data["Tag"] = tag     
        row_data["Type"] = type_data     
        row_data["Channel"] = Channel 
        row_data["Icon"] = Icon 
        row_data["Data"] = raw_data 
        row_data["SSID"] = SSID         
        row_data["#"] = SSID
        row_data["Manually decoded"] = str(parsed_response)
        row_data["CompanyName"] = CompanyName        

    return data


def write_xlsx(data,output_xlsx):
    '''
    The write_locations() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''

    msg_blurb = (f'Writing {output_xlsx}')
    msg_blurb_square(msg_blurb, color_green)
    
    
    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Locations'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    log_headers = [
        "Date", "Subject", "Requesting Agency", "Requesting Agent", "Case"
        , "Summary of Findings", "Source", "Notes"
    ]


    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [3, 4, 5, 6, 49, 50]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [7,8, 13, 14, 15, 29, 30, 35, 36, 37, 38, 39, 40, 41, 42, 43]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill
        # elif col_index == 27:  # Red for column 27
            # fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
            # cell.fill = fill

    ## Excel column width

    worksheet.column_dimensions['A'].width = 25 # 
    worksheet.column_dimensions['B'].width = 20 # 
    worksheet.column_dimensions['C'].width = 20 # 
    worksheet.column_dimensions['D'].width = 25 # 
    worksheet.column_dimensions['E'].width = 1 # 
    worksheet.column_dimensions['F'].width = 10 # Group
    worksheet.column_dimensions['G'].width = 14 # Subgroup
    worksheet.column_dimensions['H'].width = 35 # Description
    worksheet.column_dimensions['I'].width = 15 # Type
    worksheet.column_dimensions['J'].width = 8 # Source
    worksheet.column_dimensions['K'].width = 1 # Deleted
    worksheet.column_dimensions['L'].width = 5 # Tag
    worksheet.column_dimensions['M'].width = 20 # Source file information
    worksheet.column_dimensions['N'].width = 1 # 
    worksheet.column_dimensions['O'].width = 1 #
    worksheet.column_dimensions['P'].width = 10 # 
    worksheet.column_dimensions['Q'].width = 1 # 
    worksheet.column_dimensions['R'].width = 1 # 
    worksheet.column_dimensions['S'].width = 1 # 
    worksheet.column_dimensions['T'].width = 1 #
    worksheet.column_dimensions['U'].width = 1 # 
    worksheet.column_dimensions['V'].width = 1 # 
    worksheet.column_dimensions['W'].width = 1 # 
    worksheet.column_dimensions['X'].width = 1 # 
    worksheet.column_dimensions['Y'].width = 32 # 
    worksheet.column_dimensions['Z'].width = 1 # 
    worksheet.column_dimensions['AA'].width = 1 # 
    worksheet.column_dimensions['AB'].width = 1 # 
    worksheet.column_dimensions['AC'].width = 1 # 
    worksheet.column_dimensions['AD'].width = 1 # 
    worksheet.column_dimensions['AE'].width = 1 # 
    worksheet.column_dimensions['AF'].width = 22 # Coordinate
    worksheet.column_dimensions['AG'].width = 1 # 
    worksheet.column_dimensions['AH'].width = 1 # 
    worksheet.column_dimensions['AI'].width = 1 # 
    worksheet.column_dimensions['AJ'].width = 1 # 
    worksheet.column_dimensions['AK'].width = 1 # 
    worksheet.column_dimensions['AL'].width = 1 # 
    worksheet.column_dimensions['AM'].width = 1 # 
    worksheet.column_dimensions['AN'].width = 1 # 
    worksheet.column_dimensions['AO'].width = 12 # 
    worksheet.column_dimensions['AP'].width = 1 # 
    worksheet.column_dimensions['AQ'].width = 1 # 
    worksheet.column_dimensions['AR'].width = 1 # 
    worksheet.column_dimensions['AS'].width = 1 # 
    worksheet.column_dimensions['AT'].width = 12 # icon
    worksheet.column_dimensions['AU'].width = 23 # original_file
    worksheet.column_dimensions['AV'].width = 9 # case
    worksheet.column_dimensions['AW'].width = 1 # 
    worksheet.column_dimensions['AX'].width = 1 # 
    worksheet.column_dimensions['AY'].width = 1 # 
    worksheet.column_dimensions['AZ'].width = 1 # 
    worksheet.column_dimensions['BA'].width = 1 # 
    worksheet.column_dimensions['BB'].width = 1 # 
    worksheet.column_dimensions['BC'].width = 1 # 
    worksheet.column_dimensions['BD'].width = 1 # 
    worksheet.column_dimensions['BE'].width = 1 # 
    worksheet.column_dimensions['BF'].width = 1 # 
    worksheet.column_dimensions['BG'].width = 1 # 
    worksheet.column_dimensions['BH'].width = 1 # 


    worksheet.column_dimensions['BI'].width = 20 # MAC
    worksheet.column_dimensions['BJ'].width = 15 # SSID
    
    worksheet.column_dimensions['BS'].width = 30 # companyName

    if not data:
        print("data is empty.") 
        exit()
    else:
        for i in range(len(data)):
            if data[i] is None:
                data[i] = ''
        

    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            try:
                cell_data = row_data.get(col_name)
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    # Create a new worksheet for color codes
    color_worksheet = workbook.create_sheet(title='ColorCode')
    color_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    color_worksheet.column_dimensions['A'].width = 14# Color
    color_worksheet.column_dimensions['B'].width = 20# Description


    # Excel row height
    color_worksheet.row_dimensions[2].height = 22  # Adjust the height as needed
    color_worksheet.row_dimensions[3].height = 22
    color_worksheet.row_dimensions[4].height = 23
    color_worksheet.row_dimensions[5].height = 23
    color_worksheet.row_dimensions[6].height = 40   # truck

    color_worksheet.cell(row=1, column=1).value = 'Color'
    color_worksheet.cell(row=1, column=2).value = 'description'
    color_worksheet.cell(row=2, column=1).value = 'Red'
    color_worksheet.cell(row=3, column=1).value = 'Orange'
    color_worksheet.cell(row=4, column=1).value = 'Green'
    color_worksheet.cell(row=5, column=1).value = 'Yellow'

    color_worksheet.cell(row=7, column=1).value = 'ABBREVIATIONS'
    color_worksheet.cell(row=8, column=1).value = 'AKA'
    color_worksheet.cell(row=9, column=1).value = 'DOB'
    color_worksheet.cell(row=10, column=1).value = 'VIS'
    color_worksheet.cell(row=11, column=1).value = 'VIN'
    color_worksheet.cell(row=12, column=1).value = 'VYR'
    color_worksheet.cell(row=13, column=1).value = 'VMA'
    color_worksheet.cell(row=14, column=1).value = 'LIC'
    color_worksheet.cell(row=15, column=1).value = 'LIY'
    color_worksheet.cell(row=16, column=1).value = 'DLN'
    color_worksheet.cell(row=17, column=1).value = 'DLS'

       
    color_worksheet.cell(row=2, column=2).value = 'Bad Intel or dead link'
    color_worksheet.cell(row=3, column=2).value = 'Research'
    color_worksheet.cell(row=4, column=2).value = 'Good Intel'
    color_worksheet.cell(row=5, column=2).value = 'Highlighted'

    color_worksheet.cell(row=8, column=2).value = 'Also Known As (Alias)'
    color_worksheet.cell(row=9, column=2).value = 'Date of Birth'
    color_worksheet.cell(row=10, column=2).value = 'Vehicle State'
    color_worksheet.cell(row=11, column=2).value = 'Vehicle Identification Number'
    color_worksheet.cell(row=12, column=2).value = 'Vehicle Year'
    color_worksheet.cell(row=13, column=2).value = 'Vehicle Make'
    color_worksheet.cell(row=14, column=2).value = 'License'
    color_worksheet.cell(row=15, column=2).value = 'License Year'
    color_worksheet.cell(row=16, column=2).value = 'Drivers License Number'
    color_worksheet.cell(row=17, column=2).value = 'Drivers License State'


    # colored fills
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    # Apply the orange fill to the cell in row 2, column 2
    color_worksheet.cell(row=2, column=2).fill = red_fill
    color_worksheet.cell(row=3, column=2).fill = orange_fill
    color_worksheet.cell(row=4, column=2).fill = green_fill
    color_worksheet.cell(row=5, column=2).fill = yellow_fill


    # Create a new worksheet for logs
    log_worksheet = workbook.create_sheet(title='Log')
    log_worksheet.freeze_panes = 'B2'  # Freeze cells

# Date, Subject, Requesting Agency, Requesting Agent, Case, Summary of Findings, Source, Notes, Requestor

    # Excel column width
    log_worksheet.column_dimensions['A'].width = 14# Date
    log_worksheet.column_dimensions['B'].width = 20# Subject
    log_worksheet.column_dimensions['C'].width = 24# Requesting Agency
    log_worksheet.column_dimensions['D'].width = 20# Requesting Agent
    log_worksheet.column_dimensions['E'].width = 14# Case
    log_worksheet.column_dimensions['F'].width = 20# Summary of Findings
    log_worksheet.column_dimensions['G'].width = 14# Source
    log_worksheet.column_dimensions['H'].width = 25# Notes

    log_worksheet.cell(row=1, column=1).value = 'Date'
    log_worksheet.cell(row=1, column=2).value = 'Subject'
    log_worksheet.cell(row=1, column=3).value = 'Requesting Agency'
    log_worksheet.cell(row=1, column=4).value = 'Requesting Agent'
    log_worksheet.cell(row=1, column=5).value = 'Case'
    log_worksheet.cell(row=1, column=6).value = 'Summary of Findings'
    log_worksheet.cell(row=1, column=7).value = 'Notes'



    workbook.save(output_xlsx)
    workbook.close()

def query_wigle_api(username: str, password: str):
    url = "https://api.wigle.net/api/v2/profile/user"
    headers = {"Accept": "application/json"}
    
    response = requests.get(url, headers=headers, auth=HTTPBasicAuth(username, password))
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error_wigle: {response.status_code} - {response.text}")
        return None

def query_wigle_wifi_api(username: str, password: str, type_data, MAC):
    if ":" in MAC:
        MAC = MAC.replace(":","%3A")
    if type_data.lower() == "wifi":
        url = (f"https://api.wigle.net/api/v3/detail/wifi/{MAC}")
    elif type_data.lower() == "bt" or type_data.lower() == "ble":
        url = (f"https://api.wigle.net/api/v3/detail/bt/{MAC}")
    # else:
        # url = "https://api.wigle.net/api/v2/profile/user"
        # print(f'fialed')
    # print(f'url =  {url}')

    headers = {"Accept": "application/json"}
    
    response = requests.get(url, headers=headers, auth=HTTPBasicAuth(username, password))
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error_wigle: {response.status_code} - {response.text}")
        # return None
        return response.text
        
        
        
def response_parse_wifi(response):
    parsed_data = {}
    try:
        data_wigle = json.loads(response)
        if "too many DETAIL queries today." in data_wigle:
            # print(f'too many DETAIL queries today.')
            msg_blurb = (f'too many DETAIL queries today. Sleeping for 2 minutes.')
            msg_blurb_square(msg_blurb, color_red)
            # exit()
            time.sleep(120)
        elif not isinstance(data_wigle, dict):  # Ensure data is a dictionary
            print("Error: Parsed data is not a dictionary.")
            return ''        
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
        return ''

    try:
        parsed_data = {
            "Latitude": data_wigle.get("trilateratedLatitude") if data_wigle.get("trilateratedLatitude") else '',
            "Longitude": data_wigle.get("trilateratedLongitude") if data_wigle.get("trilateratedLongitude") else '',
            "SSID": data_wigle.get("name") if data_wigle.get("name") else '',        
            "Time": data_wigle.get("lastUpdate") if data_wigle.get("lastUpdate") else '',
            "number": data_wigle.get("streetAddress", {}).get("housenumber") if data_wigle.get("streetAddress") else '',
            "street": data_wigle.get("streetAddress", {}).get("road") if data_wigle.get("streetAddress") else '',
            "city": data_wigle.get("streetAddress", {}).get("city") if data_wigle.get("streetAddress") else '',
            "state": data_wigle.get("streetAddress", {}).get("region") if data_wigle.get("streetAddress") else '',
            "country": data_wigle.get("streetAddress", {}).get("country") if data_wigle.get("streetAddress") else '',
            "zipcode": data_wigle.get("streetAddress", {}).get("postalcode") if data_wigle.get("streetAddress") else '',
            "MAC": data_wigle.get("networkId")  if data_wigle.get("networkId") else '',        
            "Subgroup": data_wigle.get("encryption")  if data_wigle.get("encryption") else '',
            "Channel": data_wigle.get("channel")  if data_wigle.get("channel") else '',
            "SSID": data_wigle.get("locationClusters", [{}])[0].get("clusterSsid", "") if data_wigle.get("locationClusters") else 'test',       
        }

    except Exception as e:
        print(f"{color_red}Error: {str(e)}{color_reset}")

    
    return parsed_data 
    

def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description2}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}export from Cellebrite categories')
    print(f'\nExample:')
   
    print(f'    {file} -b -O input_blank.xlsx') 
    print(f'    {file} -Q -I Calls.xlsx  ')      

                
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
-u create a wigle username lookup

if instagram id found create a url
Icon, type_data and origin update
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
create a blank sheet -b
insert mac addresses into MAC
Make sure to put WIFI, BT or BLE in type, or it will skip it.


Error_wigle: 400 - {"errors":["path param btNetworkId must match \"^([0-9a-fA-F]{2}:){5}[0-9a-fA-F]{2}$\""]}

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

