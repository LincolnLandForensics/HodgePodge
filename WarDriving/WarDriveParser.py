#!/usr/bin/python
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import re
import sys
import csv
import glob  # Import the glob module for pattern matching
import gzip
import time
import shutil
import struct
import string
import binascii
import argparse  # for menu system
import openpyxl # pip install openpyxl
import unicodedata
from openpyxl import load_workbook, Workbook 
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


import json
import requests
from requests.auth import HTTPBasicAuth



# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "Convert wigle .gz or .csv exports to gps2address.py locations format or convert HackRf logs. Convert MAC to company name."
version = '1.2.2'


global USERNAME
USERNAME = "AID3eb88c99ae363e3385bb9058dd262e5a"
global PASSWORD
PASSWORD = "b77efe08e11e816fb77270c9c2938d97"

global hackRF_drive
hackRF_drive= 'H'

global logfolder
logfolder=r'\Logs'  # D:\Forensics\scripts\python\Logs

global headers
# update the headers if you don't want all of them.
headers = [
    "#", "Time", "Latitude", "Longitude", "Address", "Group", "Subgroup"
    , "Description", "Type", "Source", "Deleted", "Tag"
    , "Source file information", "Service Identifier", "Carved", "Name"
    , "business", "number", "street", "city", "county", "state", "zipcode"
    , "country", "fulladdress", "query", "Sighting State", "Plate"
    , "Capture Time", "Capture Network", "Highway Name", "Coordinate"
    , "Capture Location Latitude", "Capture Location Longitude", "Container"
    , "Sighting Location", "Direction", "Time Local", "End time", "Category"
    , "Manually decoded", "Account", "PlusCode", "Time Original", "Timezone"
    , "Icon", "original_file", "case", "Origin Latitude", "Origin Longitude"
    , "Start Time", "Azimuth", "Radius", "Altitude", "Location"
    , "time_orig_start", "timezone_start", "Index", "speed", "parked"
    , "MAC", "SSID", "AuthMode", "Channel", "Frequency", "dB"
    , "AltitudeMeters", "AccuracyMeters", "RCOIs", "MfgrId", "CompanyName"
    , "Data", "Packet Type", "Hits", "length"
]

global companies
companies = {
        "74:EC:B2": "Amazon Technologies Inc.",
        "E4:F0:42": "Google, Inc.",
        "A0:D7:F3": "Samsung Electronics Co.,Ltd",
        "D4:3A:2C": "Google, Inc."
    }


global COMPANY_IDS
# Expanded Bluetooth company identifiers
COMPANY_IDS = {
    0x0006: "Apple, Inc.",
    0x004C: "Apple, Inc. (iBeacon)",
    0x0075: "Samsung Electronics Co. Ltd.",
    0x00E0: "Google, Inc.",
    0x0059: "Nordic Semiconductor ASA",
    0x000D: "Texas Instruments Inc.",
    0x0033: "Microsoft Corporation",
    0x01D6: "Fitbit Inc.",
    0x004D: "Sony Corporation",
    0x03DA: "Garmin International",
    0x0157: "Huawei Technologies Co., Ltd.",
    0x0171: "Xiaomi Inc.",
    0x02E5: "Oppo Mobile Telecommunications Corp., Ltd.",
    0x0221: "OnePlus Technology (Shenzhen) Co., Ltd.",
}

global AD_TYPE_MAP
# Advertisement Data Types
AD_TYPE_MAP = {
    0x01: "Flags",
    0x02: "Incomplete List of 16-bit Service Class UUIDs",
    0x03: "Complete List of 16-bit Service Class UUIDs",
    0x04: "Incomplete List of 32-bit Service Class UUIDs",
    0x05: "Complete List of 32-bit Service Class UUIDs",
    0x06: "Incomplete List of 128-bit Service Class UUIDs",
    0x07: "Complete List of 128-bit Service Class UUIDs",
    0x08: "Shortened Local Name",
    0x09: "Complete Local Name",
    0x0A: "TX Power Level",
    0xFF: "Manufacturer Specific Data",
}

global DEVICE_TYPES
# Expanded Known Device Type UUIDs
DEVICE_TYPES = {
    "Apple, Inc.": {
        "iPhone": ["180A"],
        "Apple Watch": ["181D", "180D"],
        "AirPods": ["FDCA", "FDAF"],
        "iPad": ["180A"],  # Same as iPhone
        "MacBook": ["180A"],  # Generic Apple device UUID
    },
    "Samsung Electronics Co. Ltd.": {
        "Galaxy Phone": ["180A"],
        "Galaxy Watch": ["181D"],
        "Galaxy Buds": ["FD08"],
    },
    "Google, Inc.": {
        "Pixel Phone": ["180A"],
        "Pixel Buds": ["FE9F"],
    },
    "Fitbit Inc.": {
        "Fitbit Tracker": ["181D", "180D"],
    },
    "Garmin International": {
        "Garmin Watch": ["181D"],
    },
    "Huawei Technologies Co., Ltd.": {
        "Huawei Watch": ["181D"],
        "Huawei Phone": ["180A"],
    },
    "Xiaomi Inc.": {
        "Xiaomi Mi Band": ["181D"],
        "Xiaomi Phone": ["180A"],
    },
    "Sony Corporation": {
        "Sony Headphones": ["FDCE"],
        "Sony Smartwatch": ["181D"],
    },
    "Microsoft Corporation": {
        "Surface": ["180A"],
        "Xbox Controller": ["1812"],  # HID Game Controller UUID
    },
}

    
# Colorize section
global color_red
global color_green
global color_reset
color_green = ''
color_green = ''
color_red = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
  
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL
        

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global Row
    Row = 1  # defines arguments
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-b','--blank', help='create blank sheet', required=False, action='store_true')  
    parser.add_argument('-C', '--clear', help='clear logs off the HackRF', required=False, action='store_true')      
    parser.add_argument('-L', '--logs', help='log grabber (HackRF)', required=False, action='store_true')
    parser.add_argument('-p', '--parseHackRF', help='parse HackRF text', required=False, action='store_true')    
    parser.add_argument('-w', '--wigleparse', help='parse wigle file csv', required=False, action='store_true')
    parser.add_argument('-x', '--xlsxparse', help='parse xlsx sheet', required=False, action='store_true')

    args = parser.parse_args()

    global input_folder
    input_folder = args.input if args.input else "\logs"  

    global input_file
    input_file = args.input if args.input else "WigleWifi_sample.csv"

    global output_xlsx
    output_xlsx = args.output if args.output else "WarDrive_.xlsx"
    data = []
    if args.clear:
        clear_logs()
    elif args.logs:
        log_grab() 
    elif args.parseHackRF:
        companies_read()

        if not os.path.exists(input_folder):
            print(f"Error: The directory '{input_folder}' does not exist.")
            # return
        else:
            message = (f"Reading files in {input_folder}")
            message_square(message, color_green)             
        parse_hackRF(input_folder, output_xlsx, data)        
    elif args.wigleparse:
        companies_read()
        process_wigle_file(input_file, data)

        
    elif args.xlsxparse:
        companies_read()    
        input_xlsx = input_file
        # process_xlsx_file(input_file, data)

        file_exists = os.path.exists(input_file)

        datatype = input_file
        datatype = datatype.replace('.xlsx', '')

        if not args.output: 
            output_xlsx = (f'intel_{datatype}.xlsx') 
  
        else:
            output_xlsx = args.output


        if file_exists == True:
            msg_blurb = (f'Reading {input_xlsx}')
            message_square(msg_blurb, color_green)

            data = read_xlsx(input_xlsx)
            # write_xlsx(data
            write_xlsx(data,output_xlsx)

            workbook.close()
            msg_blurb = (f'Writing to {output_xlsx}')
            message_square(msg_blurb, color_green)            
            # print(f'{color_green}Writing to {output_xlsx} {color_reset}')


    elif args.blank:
        write_xlsx(data,output_xlsx)
        return 0 
        sys.exit()
    else:
        usage()
        
    # workbook.close()
    # Workbook.close()    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def watchList_check(MAC, Name):
    # Check if the CSV file exists
    Tag =  ''
    watchList_file = 'watchList.csv'
    if os.path.exists(watchList_file):
        # Read the CSV file and update the companies dictionary
        with open(watchList_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                mac = row[0]
                if MAC == mac:
                    Tag = 'watchList'
                    if Name == '':
                        Name = row[1]
    return Tag, Name


def BLE_Data_Translate(data):
    length, ad_type, company_id, raw_data = ('', '', '', '')
    
    try:
        # Fix odd-length hex strings by padding (if needed)
        if len(data) % 2 != 0:
            data = "0" + data  # Add leading zero   
            
        # Convert hex string to raw bytes (binary data)
        raw_bytes = binascii.unhexlify(data)

        # Extract key parts of the packet
        length = raw_bytes[0]  # First byte: Length
        ad_type = raw_bytes[1]  # Second byte: Advertising type (should be 0xFF for Manufacturer Data)
        company_id = int.from_bytes(raw_bytes[2:4], byteorder="little")  # Convert 2-byte ID to int

        payload = raw_bytes[4:]  # Remaining data (MAC, UUID, or sensor values)
        raw_data = payload

        # print(f"Raw Payload: {payload.hex().upper()}")  # Convert binary payload to hex string
        raw_data = payload.hex().upper()   # temp
    except Exception as e:
        # print(f"Error processing data '{data}': {e}")
        print(f'{e}')
    
    return length, ad_type, company_id, raw_data
    
    
def clear_logs():
    # Prompt user for drive letter, use default if left blank
    drive = input(f"\nIf you continue, you will delete HackRF logs!\n\nEnter the drive letter where HackRF is plugged in (default is {hackRF_drive}): \n").strip().lower()
    if not drive:
        drive = hackRF_drive.lower()

    # Paths for .csv and .txt files
    source_csv = f"{drive.upper()}:/BLERX/Lists/*.csv"
    source_txt_1 = f"{drive.upper()}:/BLERX/LOGS/*.TXT"
    source_txt_2 = f"{drive.upper()}:/LOGS/*.TXT"
    
    # Delete .csv files
    for file in glob.glob(source_csv):
        os.remove(file)  # Remove the file
        print(f"Deleted: {file}")

    # Delete .txt files from BLERX/LOGS
    for file in glob.glob(source_txt_1):
        os.remove(file)  # Remove the file
        print(f"Deleted: {file}")
        
    # Delete .txt files from LOGS
    for file in glob.glob(source_txt_2):
        os.remove(file)  # Remove the file
        print(f"Deleted: {file}")

    # Optional: Display a message indicating the logs have been deleted
    message = (f'HackRF logs have been deleted from drive {drive.upper()}')
    message_square(message, color_green)


def companies_read():
    # Check if the CSV file exists
    csv_file = 'mac-vendors-export.csv'
    if os.path.exists(csv_file):
        # Read the CSV file and update the companies dictionary
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                mac_prefix = row[0]
                vendor_name = row[1]
                companies[mac_prefix] = vendor_name


def company_lookup(MAC):
    company = ''

    # Extract the first 8 characters of MAC
    mac_prefix = MAC[:8]  # Get the first 8 characters

    # Return the company name if found, else an empty string
    return companies.get(mac_prefix, "") 
    
    
def format_function(bg_color='white'):
    global format
    format = workbook.add_format({
        'bg_color': bg_color
    })


def identify_device_type(CompanyName, service_uuids):
    """Determines the most likely device type based on UUIDs and manufacturer"""
    if CompanyName in DEVICE_TYPES:
        for device_type, uuid_list in DEVICE_TYPES[CompanyName].items():
            if any(uuid in service_uuids for uuid in uuid_list):
                return device_type
    return "Unknown Device Type"
    
                
def iso8601_timestamp(timestamp):
    # Convert string to datetime object
    try:
        dt = datetime.strptime(timestamp, "%Y%m%d%H%M%S")

        # Convert to ISO 8601 format
        iso_timestamp = dt.isoformat()
        return iso_timestamp
    except Exception as e:
        # print(f"Error processing timestamp {timestamp}: {e}")
        return timestamp


def log_grab():
    # Prompt user for drive letter, use default if left blank
    drive = input(f"\nEnter the drive letter the HackRF is plugged into (default is {hackRF_drive}): \n").strip().lower()
    if not drive:
        drive = hackRF_drive.lower()

    # Paths for .csv and .txt files
    source_csv = f"{drive.upper()}:\BLERX\Lists\*.csv"
    source_txt_1 = f"{drive.upper()}:\BLERX\LOGS\*.TXT"
    source_txt_2 = f"{drive.upper()}:\LOGS\*.TXT"
    
    # Create the log folder if it doesn't exist
    if not os.path.exists(logfolder):
        os.makedirs(logfolder)

    # Copy .csv files to log folder
    for file in glob.glob(source_csv):
        shutil.copy(file, logfolder)
        print(f"Copied: {file}")

    # Copy .txt files from BLERX/LOGS to log folder
    for file in glob.glob(source_txt_1):
        shutil.copy(file, logfolder)
        print(f"Copied: {file}")
        
    # Copy .txt files from LOGS to log folder
    for file in glob.glob(source_txt_2):
        shutil.copy(file, logfolder)
        print(f"Copied: {file}")

    message = (f'HackRF logs copied to {logfolder}')
    message_square(message, color_green) 
    
    
def message_square(message, color):
    horizontal_line = f"+{'-' * (len(message) + 2)}+"
    empty_line = f"| {' ' * (len(message))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {message} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def MfgrId2Company(MfgrId):
    try:
        MfgrId = int(MfgrId)  # Ensure it's an integer
    except ValueError:
        return ""

    MfgrId = int(MfgrId)
    # print(f'MfgrId = type {type(MfgrId)}')  # temp
    
    mfgr_dict = {
        6: "Xerox Corporation",
        13: "XEROX CORPORATION",
        76: "Compaq Computer Corporation",
        87: "Hewlett-Packard Company",
        89: "International Business Machines",
        93: "VisiCorp",
        96: "AT&T",
        117: "Apple Computer Inc.",
        135: "Sun Microsystems",
        137: "Advanced Micro Devices",
        217: "Silicon Graphics, Inc.",
        263: "Cisco Systems, Inc.",
        283: "TOSHIBA CORPORATION",
        301: "Unisys",
        315: "Hewlett-Packard Company",
        369: "Intel Corporation",
        529: "Apple Computer Inc.",
        767: "Cisco Systems, Inc.",
        12849: "Cisco Systems, Inc.",
        1363: "Cisco Systems, Inc.",
        1494: "Cisco Systems, Inc.",
        1536: "Cisco Systems, Inc.",
        1704: "Cisco Systems, Inc.",
        1736: "Cisco Systems, Inc.",
        1744: "Cisco Systems, Inc.",
        1993: "Cisco Systems, Inc.",
        2409: "Cisco Systems, Inc.",
        2504: "Cisco Systems, Inc.",
        27475: "Cisco Systems, Inc.",
        29439: "Cisco Systems, Inc.",
        34817: "Cisco Systems, Inc.",
        34818: "Cisco Systems, Inc.",
        40857: "Cisco Systems, Inc.",
        42239: "Cisco Systems, Inc.",
        48872: "Cisco Systems, Inc.",
        61680: "Cisco Systems, Inc.",
        65024: "Cisco Systems, Inc.",
        65535: "Cisco Systems, Inc.",
    }
    return mfgr_dict.get(MfgrId, "")


def packet_lookup(msg_type):
    # Dictionary of message types and corresponding descriptions
    packet_details = {
        "ADV_IND": "Scannable advertising indication for all devices",
        "ADV_DIRECT_IND": "Directed advertising to a specific device indicating that only that device can connect",
        "ADV_NONCONN_IND": "Advertising indication but not accepting connections or scans",
        "SCAN_REQ": "Sent by a device to receive more info from scannable advertisers",
        "SCAN_RSP": "Response to 03 containing more info",
        "CONNECT_REQ": "Sent to an advertiser to initiate a connection",
        "ADV_SCAN_IND": "Scannable advertising indication but not accepting connections",
        "ADV_EXT_IND": "BT 5.0, points to additional data on secondary channels",
        "AUX_ADV_IND": "BT 5.0, scannable advertising indication for all devices on secondary channels",
        "AUX_SCAN_REQ": "BT 5.0, sent by a device to receive more info from scannable advertisers on secondary channels",
        "AUX_SCAN_RSP": "BT 5.0, response to 0A containing more info",
        "AUX_CONNECT_REQ": "BT 5.0, sent to an advertiser to initiate a connection",
        "AUX_CHAIN_IND": "BT 5.0, chains advertising packets together when they get too big",
        "AUX_CONNECT_RSP": "BT 5.0, response to 0C",
        "Link Layer": "Link layer setup and operation",
        "Data": "General data and proprietary"
    }

    # Check if the msg_type exists in the dictionary, if yes return the description
    return packet_details.get(msg_type, "")


def parse_bluetooth_data(data):
    """Parses raw Bluetooth advertising data and identifies the device type"""
    
    if data.startswith("0x"):
        data = data[2:]

    try:
        raw_bytes = bytes.fromhex(data)
    except Exception as e:
        print(f"{color_red}Error converting data from hex: {str(e)}{color_reset}")
        raw_bytes = ''
    
    
    
    index = 0
    parsed_output = {}
    service_uuids = []
    device_name = ''

    while index < len(raw_bytes):
        length = raw_bytes[index]
        if length == 0 or index + length >= len(raw_bytes):
            break  # End of data
        
        ad_type = raw_bytes[index + 1]
        value = raw_bytes[index + 2 : index + 1 + length]
        type_description = AD_TYPE_MAP.get(ad_type, f"Unknown Type (0x{ad_type:02X})")

        # Manufacturer Data Parsing (0xFF)
        if ad_type == 0xFF and len(value) >= 2:
            company_id = struct.unpack("<H", value[:2])[0]
            CompanyName = COMPANY_IDS.get(company_id, f"Unknown Company (0x{company_id:04X})")
            parsed_output["Manufacturer Data"] = {
                "Company ID": f"0x{company_id:04X}",
                "Company Name": CompanyName,
                "Raw Manufacturer Data": value[2:].hex().upper(),
            }
        elif ad_type in [0x02, 0x03, 0x06, 0x07]:  # Service UUIDs
            for i in range(0, len(value), 2):
                service_uuids.append(value[i:i+2].hex().upper())
        elif ad_type == 0x09:  # Complete Local Name
            device_name = value.decode("utf-8", errors="ignore")  # Extract Device Name
            
        parsed_output[type_description] = value.hex().upper()
        index += length + 1

    CompanyName = parsed_output.get("Manufacturer Data", {}).get("Company Name", "Unknown")
    device_type = identify_device_type(CompanyName, service_uuids)

    if device_name:
        parsed_output["Device Name"] = device_name

    parsed_output["Identified Device Type"] = device_type
    # print(f'Bluetooth device_name = {device_name}')    # temp
    return parsed_output


def parse_hackRF(input_folder, output_xlsx, data):     # hackrf
    data2 = []
    mac_uniq = []
    
    mac_regex = r'(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}'
    
    # Read and process each .txt file in the folder
    for file_name in os.listdir(input_folder):
        file_path = os.path.join(input_folder, file_name)
        print(f'Reading {file_name}')    

        if os.path.getsize(file_path) == 0:  # Check if file is empty
            row_data = {}
            print(f'\tEmpty file: {os.path.basename(file_path)}')
            row_data["original_file"] = file_name
            row_data["Type"] = '_empty'
            data.append(row_data)
            
        elif file_name.lower().endswith('.csv') and os.path.exists(file_path):
            # with open(file_path, mode='r', newline='', encoding='utf-8') as file:
            with open(file_path, mode='r', newline='', encoding='ISO-8859-1') as file:
                reader = csv.reader(file)
                print(f'reading {file}') 
                # next(reader)  # Skip the header row
                for row in reader:
                    timestamp, MAC, Name, msg_type, dta, hits = '', '', '', '', '', ''
                    dB, channel, length, CompanyName, device_type, p_description = '', '', '', '', '', ''
                    Tag, Altitude, Latitude, Longitude = '', '', '', ''

                    row_data = {}
                    row_data2 = {}

                    try:
                        timestamp = row[0] if len(row) > 0 else ''
                        MAC = row[1] if len(row) > 1 else ''
                        Name = row[2] if len(row) > 2 else ''
                        msg_type = row[3] if len(row) > 3 else ''
                        dta = row[4] if len(row) > 4 else ''
                        hits = row[5] if len(row) > 5 else ''
                        dB = row[6] if len(row) > 6 else ''
                        channel = row[7] if len(row) > 7 else ''  # Only assign if there are at least 8 columns

                        MAC = MAC[:17]                        
                    except Exception as e:
                        print(f"Error processing line : {e}")

                    # if dta.startswith('0x0'):
                        # print(f'data = {dta}')  # temp

                    if Name.startswith('0x1'):
                        Name = ''
                    CompanyName = company_lookup(MAC)
                    if msg_type != '':
                        p_description = packet_lookup(msg_type)
                    if MAC != '':
                        description = (f'{p_description}\nMAC:{MAC}')  
                    if MAC != '':
                        description = (f'{p_description}\nMAC:{MAC}')  


                    if Tag == '':
                        (Tag, Name) =  protectList_check(MAC, Name)
                    if Tag == '':
                        (Tag, Name) =  watchList_check(MAC, Name)

                    if MAC not in mac_uniq:    
                        mac_uniq.append(MAC)

                        # row_data2["Time"] = timestamp
                        # row_data2["MAC"] = MAC
                        # row_data2["Name"] = Name
                        # row_data2["Hits"] = hits
                        # row_data2["dB"] = dB
                        # row_data2["Channel"] = channel  
                        # row_data2["CompanyName"] = CompanyName
                        # row_data2["Device Type"] = device_type
                        # row_data2["original_file"] = file_name
                        # row_data2["Tag"] = Tag    
                        # row_data2["file_name"] = file_name
                        # row_data2["Tag"] = Tag
                        # data2.append(row_data2)
                    if 1==1:
                    # if "imestamp" not in timestamp:
                        row_data["Time"] = timestamp    # 
                        row_data["MAC"] = MAC   # 
                        row_data["Name"] = Name # 
                        row_data["Packet Type"] = msg_type
                        row_data["Data"] = dta
                        row_data["Hits"] = hits
                        row_data["dB"] = dB
                        row_data["Channel"] = channel  # 
                        row_data["length"] = length
                        row_data["CompanyName"] = CompanyName   #
                        row_data["Type"] = device_type  #
                        row_data["Description"] = p_description #                        
                        row_data["original_file"] = file_name   #
                        row_data["Tag"] = Tag   #
                        row_data["Latitude"] = Latitude
                        row_data["Longitude"] = Longitude
                        row_data["AltitudeMeters"] = Altitude 

                        data.append(row_data)                            
                            
                            
        elif file_name.lower().endswith('.txt'):  # Check for .txt extension
            # print(f' mac_uniq  2 = {data2} filename = {file_name}')   # temp  
            with open(os.path.join(input_folder, file_name), "r") as file:
                for line in file:
                    row_count2 = 0
                    line = line.strip()
                    
                    row_data = {}
                    row_data2 = {}
                    (timestamp, MAC, Name, msg_type, dta, hits)  = ('', '', '', '', '', '')
                    (dB, channel, length, CompanyName, device_type, p_description) = ('', '', '', '', '', '')
                    (length2, Type, company_id, raw_data) = ('', 'plane', '', '')
                    (Tag, parsed_data) = ('', '')
                    parts = line.strip().split(" ")
                    
                    if line.strip() == '':
                        blah = 'blah'
                    elif file_name == 'ADSB.TXT':
                        
                        (timestamp, Latitude, Longitude, Coordinate, Type, original_file) = ('', '', '', '', 'ADS-B', '')
                        (Plate, Direction, Icon, ICAO, Name, note, AltitudeMeters) = ('', '', '', '', '', '', '')
                        (hit, lvl, speed, amp, age) = ('', '', '', '', '')
                        (HexID) = ('')
                        print(f'bobs your uncle')   # temp
                        row_data["original_file"] = file_name  
                        data.append(row_data)
                        
                        match1 = re.search(r'Alt:(\d+)', line)
                        match2 = re.search(r'Lat:([-+]?\d*\.\d+|\d+)', line)
                        match3 = re.search(r'Lon:([-+]?\d*\.\d+|\d+)', line)

                        if match1:
                            AltitudeMeters = int(match1.group(1)) 

                        if match2:
                            Latitude = str(match2.group(1))
                            dB = '-50'

                        if match3:
                            Longitude = float(match3.group(1))

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = parts[1] if len(parts) > 1 else ""
                            Name = parts[2] if len(parts) > 2 else ""
                            if dB != '':
                                MAC = dta
                            
                            if 'Alt:' in Name:
                                Name = ''
                            timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        except IndexError:

                            blah = 'blah'
                        if dta != '':
                            p_description = (f'Data = {dta}')
                        
                        
                        if 'ICAO:' in dta:
                            HexID = dta.split('ICAO:')[1]
                            p_description = (f'\nICAO = {HexID}')
                            dta = dta.split('ICAO:')[0] 
                            if Name == '':
                                Name = HexID 
                        p_description = p_description.strip()
                        
                        row_data["Time"] = timestamp    # 
                        row_data["MAC"] = MAC   # 
                        row_data["Name"] = Name # 
                        row_data["Packet Type"] = msg_type
                        row_data["Data"] = dta
                        row_data["Hits"] = hits
                        row_data["dB"] = dB
                        row_data["Channel"] = channel  # 
                        row_data["length"] = length
                        row_data["CompanyName"] = CompanyName   #
                        row_data["Type"] = device_type  #
                        row_data["original_file"] = file_name  
                        row_data["Description"] = p_description #                        
                        row_data["original_file"] = file_name   #
                        row_data["Tag"] = Tag   #
  
                        row_data["Latitude"] = Latitude
                        row_data["Longitude"] = Longitude
                        # row_data["Altitude"] = Altitude                        

                        data.append(row_data)
                        
                    elif file_name == 'APRS.TXT':
                        (timestamp, Name, Type, dta, MAC) = ('', '', 'APRS', '', '')

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = parts[1] if len(parts) > 1 else ""
                            Name = parts[2] if len(parts) > 2 else ""
                        except IndexError:
                            blah = 'blah'
                        MAC = dta
                        row_data["Time"] = timestamp
                        row_data["MAC"] = MAC   # 
                        row_data["Name"] = Name
                        row_data["Data"] = dta  
                        row_data["original_file"] = file_name
                        row_data["Type"] = Type                         
                        data.append(row_data)

                    elif file_name == 'AFSK.TXT':
                        (timestamp, Name, Type, dta, MAC) = ('', '', 'AFSK', '', '')

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = " ".join(parts[3]) if len(parts) > 3 else ""  # Join everything from parts[1] onward
                        except IndexError:
                            blah = 'blah'

                        row_data["Time"] = timestamp
                        # row_data["Name"] = Name
                        row_data["MAC"] = dta   # 
                        row_data["Data"] = dta  
                        row_data["original_file"] = file_name
                        row_data["Type"] = Type                         
                        data.append(row_data)                        

                    elif file_name == 'ERT.TXT':
                        timestamp, channel, device_type, Type, dta, p_description = '', '', '', '', 'ERT', ''
                                                
                        match5 = re.search(r'ID:(\d+)', line)
                        if match5:
                            Name = int(match5.group(1)) 
                        
                        timestamp = parts[0] if len(parts) > 0 else ""
                        try:
                            p_description = parts[3] if len(parts) > 3 else ""
                            channel = parts[2] if len(parts) > 2 else ""
                            dta = parts[4] if len(parts) > 4 else ""
                        except IndexError:
                            blah = 'blah'
                        if "/"in dta:
                            dta_temp = dta.split("/")
                            device_type = dta_temp[0]

                            try:
                                # Convert from hex to bytes
                                bytes_value = bytes.fromhex(device_type)
                                device_type = ''.join(c if c in string.printable else '.' for c in bytes_value.decode('ascii', errors='replace'))
                            except ValueError:
                                print(f'device_type is not a valid hex string: {device_type}')
                            
                        row_data["Time"] = timestamp
                        row_data["Channel"] = channel
                        row_data["Device Type"] = device_type
                        row_data["Name"] = Name
                        # row_data["MAC"] = timestamp   # requires a unique id or it will delete it
                        row_data["Description"] = p_description 
                        row_data["Data"] = dta  
                        row_data["original_file"] = file_name
                        row_data["Type"] = Type                         
                        data.append(row_data)
                        
                    elif file_name == 'TPMS.TXT':
                        (timestamp, Name, Type, dta, CompanyName) = ('', '', 'TPMS', '', '')

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = line.strip()
                            if '  ' in dta:
                                dta = dta.split('  ')[1]
                            Name = parts[6] if len(parts) > 6 else ""
                            CompanyName = parts[7] if len(parts) > 7 else ""
                            if '/' in CompanyName:
                                CompanyName = CompanyName.split('/')[0]
                        except IndexError:
                            blah = 'blah'
                        
                        if ' FSK ' in line:
                            p_description = "FSK"
                        timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        
                        row_data["Time"] = timestamp
                        row_data["Name"] = Name
                        row_data["Data"] = dta  
                        row_data["original_file"] = file_name
                        row_data["Description"] = p_description 
                        row_data["CompanyName"] = CompanyName 
                        row_data["Type"] = Type                         
                        data.append(row_data)
                    
                    elif len(parts) >= 5:

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            msg_type = parts[1] if len(parts) > 1 else ""
                            length = parts[2].split(":")[1] # if len(parts) > 3 and ":" in parts[3] else ""

                            MAC = parts[3].lstrip('Mac:')  # .split(":")[1] if len(parts) > 5 and ":" in parts[5] else ""
                            dta = parts[4].lstrip('Data:') # .split(":")[1] if len(parts) > 7 and ":" in parts[7] else ""

                            timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        except IndexError:
                            blah = 'blah'
                            if dta == "":
                                dta = line  # test


                        if MAC == '':
                            print(f'need to find MAC : {line}') # temp
                            
                        if MAC not in mac_uniq:
                            mac_uniq.append(MAC)

                        CompanyName = company_lookup(MAC)    

                        if re.fullmatch(r"[0-9a-fA-F]+", dta):    
                            parsed_data = parse_bluetooth_data(dta)

                            MfgrId = parsed_data.get('Manufacturer Data', {}).get('Company ID', '')
                            CompanyName = parsed_data.get('Manufacturer Data', {}).get('Company Name', '')
                            Type = parsed_data.get('Identified Device Type', 'Unknown')
                        
                        if 'unknown company' in CompanyName.lower():
                            CompanyName == ''

                        length2, ad_type, company_id, raw_data = BLE_Data_Translate(dta)
                      
                      
                        if CompanyName != '':
                            p_description = (f'{p_description}CompanyName: {CompanyName}')
                        if msg_type != '':
                            p_description = (f'{p_description}\nMsg Type: {msg_type}')
                        if length != '':
                            p_description = (f'{p_description}\nLength: {length}')
                                                
                        p_description = (f'{p_description}\nParsed data: {parsed_data}')
                        p_description = p_description.strip()
                        
  
                        # if MAC not in mac_uniq:
                            # mac_uniq.append(MAC)

                            # row_data2["Time"] = timestamp
                            # row_data2["MAC"] = MAC
                            # row_data2["Name"] = Name
                    
                            # row_data2["Hits"] = hits                        
                            # row_data2["dB"] = dB                        
                            # row_data2["Channel"] = channel  
                            
                            # row_data2["MfgrId"] = MfgrId
                            # row_data2["CompanyName"] = CompanyName
                            # row_data2["Device Type"] = device_type
                            # row_data2["file_name"] = file_name
                            # row_data2["Tag"] = Tag
                            # data2.append(row_data2)
                        
                        row_data["Time"] = timestamp
                        row_data["MAC"] = MAC
                        row_data["Name"] = Name
                        row_data["Packet Type"] = msg_type
                        row_data["Data"] = dta                        
                        row_data["Hits"] = hits                        
                        row_data["dB"] = dB                        
                        row_data["Channel"] = channel  
                        row_data["length"] = length
                        row_data["MfgrId"] = MfgrId
                        row_data["CompanyName"] = CompanyName
                        row_data["file_name"] = file_name
                        row_data["Device Type"] = device_type
                        row_data["original_file"] = file_name
                        row_data["Description"] = p_description
                        row_data["Length2"] = length2                        
                        row_data["Type"] = ad_type 
                        row_data["company_id"] = company_id 
                        row_data["raw_data"] = raw_data
                        row_data["Tag"] = Tag
                        data.append(row_data)

                    else:
                        # for files that start with BLELOG_*.txt
                        (timestamp, Name, Type, dta, CompanyName, MAC) = ('', '', 'misc', '', '', '')
                        (description) = ('')    
                        match4 = re.search(mac_regex, line)    # mac address
                        if match4:
                            MAC = {match4.group()}
                            print(f"Extracted MAC Address: {match4.group()}")    
                            if MAC != '':
                                CompanyName = company_lookup(MAC)
                            if CompanyName != '':
                                description = (f'{description}\nCompanyName:{CompanyName}')        
                            
                        try:
                            dta = line
                            timestamp = parts[0] if len(parts) > 0 else ""
                        except IndexError:

                            blah = 'blah'
                            if dta == "":
                                dta = line  # test
                                
                        timestamp = iso8601_timestamp(timestamp).replace('T', ' ')

                        # if MAC not in mac_uniq:
                            # mac_uniq.append(MAC)
                        
                        if 'AFSK' in file_name:
                            # print(f'___________{file_name}___________ AFSK found') # temp
                            Type = 'AFSK'
                        description = description.strip()
                        
                        row_data["Time"] = timestamp
                        row_data["description"] = description
                        row_data["MAC"] = MAC
                        row_data["Data"] = dta  
                        row_data["original_file"] = file_name
                        row_data["Type"] = Type
                        row_data["CompanyName"] = CompanyName
                        
                        data.append(row_data)                        
             
    # print(f' mac_uniq  2 = {data2}')   # temp  
    print(f'\n{len(mac_uniq)} uniq mac addresses found')    
    # write_xlsx(data, data2)
    write_xlsx(data,output_xlsx)

 
def process_wigle_file(filename, data):
    if filename.endswith('.gz'):
        unzipped_filename = filename[:-3]  # Remove .gz extension

        try:
            with gzip.open(filename, 'rb') as f_in:
                with open(unzipped_filename, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            message = (f'Unzipped: {filename} -> {unzipped_filename}')
            message_square(message, color_green)
            filename = unzipped_filename
        except Exception as e:
            message = (f'Error unzipping file: {e}')
            message_square(message, color_red)            
            return
    
    if not os.path.isfile(filename):
        print(f"Error: File '{filename}' not found or is not a valid file.")
        message = (f"Error: File {filename} not found or is not a valid file.")
        message_square(message, color_red)        
        return
    
    if not filename.startswith('WigleWifi') or not filename.endswith('.csv'):
        message = (f"Invalid file: Filename must start with 'WigleWifi' and end with '.csv'")
        message_square(message, color_red)        
        return
    else:
        
        output_xlsx = (f'{filename}.xlsx')

        # csv_file = open(filename)
        csv_file = open(filename, "r", encoding="utf-8")

        
        
        source_file = filename
        row_count = 0
        for row in csv_file:
            row = row.split(',')
            row_data = {}
            description, group, subgroup, type_data, Name, Type = '', '', '', '', '', ''
            Tag, CompanyName, country, source, Icon = '', '', '', 'Wigle', ''
       
            try:
                MAC = row[0] if len(row) > 0 else ''
                SSID = row[1] if len(row) > 1 else ''
                AuthMode = row[2] if len(row) > 2 else ''
                Time = row[3] if len(row) > 3 else ''
                Channel = row[4] if len(row) > 4 else ''
                Frequency = row[5] if len(row) > 5 else ''
                dB = row[6] if len(row) > 6 else ''
                latitude = row[7] if len(row) > 7 else ''
                longitude = row[8] if len(row) > 8 else ''
                AltitudeMeters = row[9] if len(row) > 9 else ''
                AccuracyMeters = row[10] if len(row) > 10 else ''
                RCOIs = row[11] if len(row) > 11 else ''
                MfgrId = row[12] if len(row) > 12 else ''
                subgroup = row[13] if len(row) > 13 else ''
                Type = row[13] if len(row) > 13 else ''
                
            except Exception as e:
                print(f"Error processing line : {e}")

            if MfgrId != '':
                CompanyName = MfgrId2Company(MfgrId)
            if latitude != '' and longitude != '':
                Coordinate = (f'{latitude},{longitude}')


            if AuthMode == 'Misc':  
                AuthMode = ''
            elif 'uncategorized' in AuthMode.lower():
                AuthMode = ''
            elif 'misc' in AuthMode.lower():
                AuthMode = ''


            subgroup = subgroup.strip()    
            Type = Type.strip()
            # SSID = SSID.replace('\"', '')

            if AuthMode == 'LTE;us':
                Type = 'Tower-LTE'
                subgroup = 'LTE'            
                country = 'US'
                Icon = 'Tower'

            elif ' tv' in SSID.lower() or '(tv)' in SSID.lower():  # todo doesn't match (tv)
                subgroup = 'Display'
                Type = 'Display/Speaker'
                Icon = 'Display'

            elif 'LTE;' in AuthMode:
                subgroup = 'LTE' 
                Type = 'Tower-LTE'
                Icon = 'Tower'
            elif AuthMode == 'GSM':
                Type = 'Tower-GSM'
                subgroup = 'GSM' 
                Icon = 'Tower'
  
            elif "Desktop;" in AuthMode:
                subgroup = 'Desktop'
                Type = 'Desktop'
                Icon = 'BT'
            elif 'Display/Speaker' in AuthMode:
                Type = 'Display/Speaker'
                Icon = 'Display'
                AuthMode = ''
                if 'speaker' in SSID.lower():
                    subgroup = 'Speaker'
                    Icon = 'Display'
                elif ' tv ' in SSID.lower():
                    subgroup = 'Display'
                    Icon = 'Display'
                else:
                    subgroup = 'Display/Speaker'
                    Icon = 'Display'
            elif 'speaker' in SSID.lower() or 'soundbar' in SSID.lower():
                Type = 'Display/Speaker'
                subgroup = 'Speaker'
                Icon = 'Display'




            elif "(oven)" in SSID.lower():
                subgroup = 'Oven'
                # Type = 'Display/Speaker'
                Icon = 'BT'


                    
            # elif Type == 'BT': # todo
            elif "BT" in Type or "BLE" in Type:
                subgroup = Type
                Type = 'BlueTooth'
                Icon = "BT"
                if 'oven' in SSID.lower():
                    subgroup = 'Oven'
                    AuthMode = ''
                elif 'qled' in SSID.lower():
                    subgroup = 'Display'                
                    Type = 'Display/Speaker'
                    Icon = 'Display'
                    AuthMode = ''
                elif 'sound' in SSID.lower():
                    subgroup = 'Speaker'                
                    Type = 'Display/Speaker'
                    Icon = 'Display'
                elif 'officejet' in SSID.lower() or 'deskjet' in SSID.lower():
                    subgroup = 'Printer'                
                elif 'dryer' in SSID.lower():
                    subgroup = 'Dryer' 
                    AuthMode = ''
                elif 'washer' in SSID.lower():
                    subgroup = 'Washer' 
                elif 'lamp' in SSID.lower() or ' light' in SSID.lower():
                    subgroup = 'Light' 
                elif 'tv' in SSID.lower():
                    Type = 'Display/Speaker'
                    subgroup = 'Display'
                    Icon = 'Display'                    

            elif "WPA2" in AuthMode:
                subgroup = 'WPA2'
                Icon = 'WIFI'
                Type = 'WIFI'
            elif AuthMode == "[ESS]":
                Type = 'WIFI'
                subgroup = 'ESS'
                Icon = 'WIFI-open' 
                if 'officejet' in SSID.lower():
                    subgroup = 'Printer'
                if 'thermostat' in SSID.lower():
                    subgroup = 'Thermostat'                    
                    
            if Tag == '':
                (Tag, Name) =  protectList_check(MAC, Name)
            if Tag == '':
                (Tag, Name) =  watchList_check(MAC, Name)
                    
            SSID = sanitize_string(SSID)
            type_data = Type
            
            if Icon == '':
                Icon = Type
            subgroup = subgroup.strip()
            group = group.strip()
            
            if SSID != '':
                description = (f'{description}\nSSID:{SSID}')            
            if MAC != '':
                description = (f'{description}\nMAC:{MAC}')             
            if AuthMode != '':
                description = (f'{description}\nAuthMode:{AuthMode}')             
            if Frequency != '':
                description = (f'{description}\nFrequency:{Frequency}')             
            if Channel != '':
                description = (f'{description}\nChannel:{Channel}')             
            if dB != '':
                description = (f'{description}\ndB:{dB}')           
            # if Type != '':
                # description = (f'{description}\nType:{Type}')
            # description = description.strip()   
            # if group != '':
                # description = (f'{description}\nGroup:{group}')            
            # if subgroup != '':
                # description = (f'{description}\nSubgroup:{subgroup}')                

            if CompanyName == '':
                CompanyName = company_lookup(MAC)
            if CompanyName != '':
                description = (f'{description}\nCompanyName:{CompanyName}')

            # Apply sanitization
            description = sanitize_string(description) 
            description = description.strip()


            if Coordinate == 'Latitude,Longitude':
                blah = 'blah'
            elif 'FirstSeen' in Time or 'release' in Time:

                blah = 'blah'        
            elif any(char.isalpha() for char in Time):
                # print(f'Alpha characters in {Time} Time') # temp
                blah = 'blah'
   
            else:
                row_count += 1
                row_data["#"] = SSID
                row_data["Time"] = Time
                row_data["Latitude"] = latitude
                row_data["Longitude"] = longitude 
                # row_data["Address"] = address
                row_data["Group"] = group
                row_data["Subgroup"] = subgroup
                row_data["Description"] = description
                row_data["Type"] = type_data
                row_data["Tag"] = Tag        
                row_data["Source"] = source
                row_data["Source file information"] = source_file
                row_data["Name"] = Name
                # row_data["business"] = business 
                # row_data["number"] = number 
                # row_data["street"] = street
                # row_data["city"] = city 
                # row_data["county"] = county 
                # row_data["state"] = state 
                # row_data["zipcode"] = zipcode
                row_data["country"] = country 
                # row_data["fulladdress"] = fulladdress
                # row_data["query"] = query
                # row_data["Plate"] = plate
                # row_data["Capture Time"] = capture_time
                # row_data["Highway Name"] = hwy
                row_data["Coordinate"] = Coordinate
                # row_data["Direction"] = direction
                # row_data["End time"] = end_time
                # row_data["Category"] = category
                # row_data["Time Original"] = time_orig
                # row_data["Timezone"] = timezone
                # row_data["PlusCode"] = PlusCode
                # row_data["Radius"] = Radius
                row_data["Icon"] = Icon
                row_data["original_file"] = filename # test


                row_data["MAC"] = MAC
                row_data["SSID"] = SSID
                row_data["AuthMode"] = AuthMode
                row_data["Channel"] = Channel  
                row_data["Frequency"] = Frequency        
                row_data["dB"] = dB
                row_data["AltitudeMeters"] = AltitudeMeters
                row_data["AccuracyMeters"] = AccuracyMeters
                row_data["RCOIs"] = RCOIs  
                row_data["filename"] = filename
                row_data["MfgrId"] = MfgrId 
                row_data["CompanyName"] = CompanyName 
                

                data.append(row_data)
        print(f'Processed {row_count} rows')
        if row_count > 3000:
            message = (f'You have too many rows for Google Earth. Manually go delete the # (labels) and maybe the Icon row before running GPS2Address.py')
            message_square(message, color_red)           
            print(f'Google Earth only likes up to 2000 labels (#). adjust the xlsx accordingly')
        output_xlsx = filename
        output_xlsx = output_xlsx.replace('.csv', '.xlsx')  # task  # it'sonly changing the local output_xlsx


        write_xlsx(data,output_xlsx)


def read_xlsx(input_xlsx):

    """Read data from an xlsx file and return as a list of dictionaries.
    Read XLSX Function: The read_xlsx() function reads data from the input 
    Excel file using the openpyxl library. It extracts headers from the 
    first row and then iterates through the data rows.    
    """

    if USERNAME == '':
        msg_blurb = (f"you need to enter your Wigle API USERNAME key, or this won't work")
        message_square(msg_blurb, color_red)        
        exit()
    if len(PASSWORD) < 15:
        msg_blurb = (f'you need to enter your Wigle API PASSWORD key')
        message_square(msg_blurb, color_red)
        exit()
    
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []
    datatype = input_xlsx
    datatype = datatype.replace('.xlsx', '')
    
    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        msg_blurb = (f'No data found in the Excel file: {input_xlsx}')
        message_square(msg_blurb, color_red)
    
        exit()
        return None

# active sheet (current sheet)
    active_sheet = wb.active
    global active_sheet_title
    active_sheet_title = active_sheet.title    


    for row_index, row_data in enumerate(data):
        (Time, time_orig, timezone, response, tag, Description) = ('', '', '', '', '', '')
        (Latitude, Longitude, Coordinate, raw_data, Time) = ('', '', '', '', '')
        (city, state, country, Altitude, note, MAC) = ('', '', '', '', '', '')
        (source, source_file, original_file, Icon, zipcode, Channel) = ('', '', '', '', '', '')
        (number, street, Subgroup, fulladdress, SSID, CompanyName) = ('', '', '', '', '', '')
# type_data
        Description = row_data.get("Description")
        if Description is None:
            Description = ''          
        
# type_data
        type_data = row_data.get("Type")
        if type_data is None:
            type_data = ''        

# Data
        raw_data = row_data.get("Data")
        if raw_data is None:
            raw_data = ''               
       
# MAC
        MAC = row_data.get("MAC")
        if MAC is None:
            MAC = ''         
        if MAC != '':
            print(f'{color_blue}MAC={MAC}{color_reset}' )
            if type_data == 'WIFI' or type_data == 'BT' or type_data == 'BLE':
                print(f'{color_yellow}sleeping for 40 seconds{color_reset}' )
                time.sleep(40)            
                result = query_wigle_wifi_api(USERNAME, PASSWORD, type_data, MAC)
   
                response = json.dumps(result)  # Convert dictionary to JSON string
                parsed_response = response_parse_wifi(response)

                Description = (f'{Description}{parsed_response}')
                Data = response

                
                print(parsed_response)

                if parsed_response:
                    Latitude = parsed_response["Latitude"]
                    Longitude = parsed_response["Longitude"]
                    SSID = parsed_response["SSID"]
                    Time = parsed_response["Time"]
                    number = parsed_response["number"]
                    street = parsed_response["street"]
                    city = parsed_response["city"]
                    state = parsed_response["state"]
                    country = parsed_response["country"]
                    zipcode = parsed_response["zipcode"]
                    MAC = parsed_response["MAC"]
                    Subgroup = parsed_response["Subgroup"]
                    Channel = parsed_response["Channel"]
                    SSID = parsed_response["SSID"]
                    Description = (f'Latitude: {Latitude}\nLongitude: {Longitude}\nTime: {Time}\nNumber: {number}\nstreet: {street}\ncity: {city}\nState: {state}\ncountry: {country}\nZipcode: {zipcode}\nMAC: {MAC}\nSubgroup: {Subgroup}\nChannel: {Channel})')
                
        if Latitude is None or Latitude == '':
            Latitude = row_data.get("Latitude")
            Longitude = row_data.get("Longitude")
            if Latitude is None:
                Latitude == ''
                Longitude == ''

# state
        if state == '':
            state = row_data.get("state")
        if state is None:
            state = ''

# city
        if city == '':
            city = row_data.get("city")
        if city is None:
            city = ''

# country
        if country == '':
            country = row_data.get("country")
        if country is None:
            country = ''


# fulladdress
        if fulladdress == '':
        # if number == '' and street == '' and city == '' and state == '' and zipcode == '':
            fulladdress = (f'{number} {street}, {city}, {state} {zipcode}')
        if fulladdress == ' , ,  ':
            fulladdress = ''
            
            
# source
        if source == '':
            source = row_data.get("Source")
        if source is None:
            source = ''


# source file
        if source_file == '':
            source_file = row_data.get("Source file information")
        if source_file is None:
            source_file = ''

# original_file
        if original_file == '':
            original_file = row_data.get("original_file")
        if original_file is None or original_file == "":
            original_file = input_xlsx
          
# Icon    
        if Icon == '':
            Icon = row_data.get("Icon")
        # Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''
        if Icon == "":
            if "Searched" in original_file:
               Icon = "Searched"            
            elif "Chats" in original_file:
               Icon = "Chats"  


# Time
        if Time == '':        
            Time = row_data.get("Time")
        if Time is None:
            Time = ''
        Time = Time.replace('T', ' ').replace('.000Z', '')

        if (Coordinate == '' or Coordinate is None) and Altitude == '':
            if Latitude is None:
                Latitude == ''
            if Longitude is None:
                Longitude == ''                
            else:    
                Coordinate = (f'{Latitude},{Longitude}')
            if 'None' in Coordinate:
                Coordinate == ''

# convert time
        output_format = "%Y-%m-%d %H:%M:%S "    # ISO 8601
        # output_format = "%Y-%m-%dT%H:%M:%SZ"    # Google Earth format
        # pattern = r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$'
        pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'  # ISO military time

        if time_orig == '' and Time != '': # copy the original time
            time_orig = Time
        # try:
            # (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
            # Time = Time.strftime(output_format)
            # if Time is None:
                # Time = ''              
            
        # except ValueError as e:
            # if Time != "":
                # print(f"Error time2: {e} - {Time}")
                # Time = ''    # temp rem of this

        if Coordinate == 'None,None':
            Coordinate == ''

# CompanyName
        CompanyName = company_lookup(MAC) 
        if CompanyName != '':
            print(f'CompanyName = {CompanyName}')   # temp
    
                
# write rows to data

        row_data["country"] = country        
        row_data["Description"] = Description  
        row_data["note"] = note
        row_data["MAC"] = MAC    
        row_data["number"] = number   
        row_data["street"] = street  
        row_data["Subgroup"] = Subgroup          
        row_data["city"] = city 
        row_data["zipcode"] = zipcode
        row_data["state"] = state 
        row_data["fulladdress"] = fulladdress 
        row_data["Time"] = Time
        row_data["Latitude"] = Latitude  
        row_data["Longitude"] = Longitude  
        row_data["Coordinate"] = Coordinate  
        row_data["Source file information"] = source_file     
        row_data["original_file"] = original_file     
        # row_data["Tag"] = tag     
        row_data["Type"] = type_data     
        row_data["Channel"] = Channel 
        row_data["Icon"] = Icon 
        row_data["Data"] = raw_data 
        row_data["SSID"] = SSID         
        row_data["#"] = SSID
        row_data["Manually decoded"] = str(parsed_response)
        row_data["CompanyName"] = CompanyName        

    return data
    

def remove_duplicate_macs(data):
    """
    Removes rows with duplicate MAC addresses and returns unique data.
    
    :param data: List of dictionaries containing the dataset
    :return: Filtered list with unique MAC addresses
    """
    unique_macs = set()
    filtered_data = []
    row_count = 0
    for row in data:
        MAC = row.get("MAC", "").strip()

        if MAC and MAC not in unique_macs:
            unique_macs.add(MAC)
            filtered_data.append(row)
            row_count += 1
    print(f'Found {row_count} unique MACs')
    return filtered_data


def query_wigle_wifi_api(username: str, password: str, type_data, MAC):
    if ":" in MAC:
        MAC = MAC.replace(":","%3A")
    if type_data.lower() == "wifi":
        url = (f"https://api.wigle.net/api/v3/detail/wifi/{MAC}")
    elif type_data.lower() == "bt" or type_data.lower() == "ble":
        url = (f"https://api.wigle.net/api/v3/detail/bt/{MAC}")
    # else:
        # url = "https://api.wigle.net/api/v2/profile/user"
        # print(f'fialed')
    # print(f'url =  {url}')

    headers = {"Accept": "application/json"}
    
    response = requests.get(url, headers=headers, auth=HTTPBasicAuth(username, password))
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error_wigle: {response.status_code} - {response.text}")
        # return None
        return response.text
        
        

def sanitize_string(text):
    # Normalize Unicode to decompose special characters
    text = unicodedata.normalize("NFKC", text)

    # Remove control characters (non-printable ASCII)
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    
    # Optionally replace problematic characters like brackets
    text = text.replace("[", "(").replace("]", ")")
    
    return text.strip()


        
def response_parse_wifi(response):
    parsed_data = {}
    try:
        data_wigle = json.loads(response)
        if "too many DETAIL queries today." in data_wigle:
            # print(f'too many DETAIL queries today.')
            msg_blurb = (f'too many DETAIL queries today. Sleeping for 2 minutes.')
            msg_blurb_square(msg_blurb, color_red)
            # exit()
            time.sleep(120)
        elif not isinstance(data_wigle, dict):  # Ensure data is a dictionary
            print("Error: Parsed data is not a dictionary.")
            return ''        
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
        return ''

    try:
        parsed_data = {
            "Latitude": data_wigle.get("trilateratedLatitude") if data_wigle.get("trilateratedLatitude") else '',
            "Longitude": data_wigle.get("trilateratedLongitude") if data_wigle.get("trilateratedLongitude") else '',
            "SSID": data_wigle.get("name") if data_wigle.get("name") else '',        
            "Time": data_wigle.get("lastUpdate") if data_wigle.get("lastUpdate") else '',
            "number": data_wigle.get("streetAddress", {}).get("housenumber") if data_wigle.get("streetAddress") else '',
            "street": data_wigle.get("streetAddress", {}).get("road") if data_wigle.get("streetAddress") else '',
            "city": data_wigle.get("streetAddress", {}).get("city") if data_wigle.get("streetAddress") else '',
            "state": data_wigle.get("streetAddress", {}).get("region") if data_wigle.get("streetAddress") else '',
            "country": data_wigle.get("streetAddress", {}).get("country") if data_wigle.get("streetAddress") else '',
            "zipcode": data_wigle.get("streetAddress", {}).get("postalcode") if data_wigle.get("streetAddress") else '',
            "MAC": data_wigle.get("networkId")  if data_wigle.get("networkId") else '',        
            "Subgroup": data_wigle.get("encryption")  if data_wigle.get("encryption") else '',
            "Channel": data_wigle.get("channel")  if data_wigle.get("channel") else '',
            "SSID": data_wigle.get("locationClusters", [{}])[0].get("clusterSsid", "") if data_wigle.get("locationClusters") else 'test',       
        }

    except Exception as e:
        print(f"{color_red}Error: {str(e)}{color_reset}")

    
    return parsed_data 
    

    
def protectList_check(MAC, Name):
    # Check if the CSV file exists
    Tag =  ''
    protectList_file = 'protectList.csv'
    if os.path.exists(protectList_file):
        # Read the CSV file and update the companies dictionary
        with open(protectList_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                mac = row[0]
                if MAC == mac:
                    Tag = 'protectList'
                    if Name == '':
                        Name = row[1]
    return Tag, Name
    

def write_xlsx(data,output_xlsx):
    '''
    The write_locations() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''
    
    # print(f'data = {data}') # temp
    
    
    message = (f'Writing {output_xlsx}')
    message_square(message, color_green)

    try:
        data = sorted(data, key=lambda x: (x.get("SSID", ""), x.get("dB", ""), x.get("MAC", "")))
        print(f'Sorting by MAC with the strongest signal')
    except TypeError as error:
        print(f'{color_red}{error}{color_reset}')

    data = remove_duplicate_macs(data)

    try:
        data = sorted(data, key=lambda x: (x.get("Time", ""), x.get("MAC", "")))
        print(f'Sorting by Time')
    except TypeError as error:
        print(f'{color_red}{error}{color_reset}')    
    
    
    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Locations'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    log_headers = [
        "Date", "Subject", "Requesting Agency", "Requesting Agent", "Case"
        , "Summary of Findings", "Source", "Notes"
    ]


    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [3, 4, 5, 6, 49, 50]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [7,8, 13, 14, 15, 29, 30, 35, 36, 37, 38, 39, 40, 41, 42, 43]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill
        # elif col_index == 27:  # Red for column 27
            # fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
            # cell.fill = fill

    ## Excel column width

    worksheet.column_dimensions['A'].width = 25 # 
    worksheet.column_dimensions['B'].width = 20 # 
    worksheet.column_dimensions['C'].width = 20 # 
    worksheet.column_dimensions['D'].width = 25 # 
    worksheet.column_dimensions['E'].width = 1 # 
    worksheet.column_dimensions['F'].width = 10 # Group
    worksheet.column_dimensions['G'].width = 14 # Subgroup
    worksheet.column_dimensions['H'].width = 35 # Description
    worksheet.column_dimensions['I'].width = 15 # Type
    worksheet.column_dimensions['J'].width = 8 # Source
    worksheet.column_dimensions['K'].width = 1 # Deleted
    worksheet.column_dimensions['L'].width = 5 # Tag
    worksheet.column_dimensions['M'].width = 20 # Source file information
    worksheet.column_dimensions['N'].width = 1 # 
    worksheet.column_dimensions['O'].width = 1 #
    worksheet.column_dimensions['P'].width = 10 # 
    worksheet.column_dimensions['Q'].width = 1 # 
    worksheet.column_dimensions['R'].width = 1 # 
    worksheet.column_dimensions['S'].width = 1 # 
    worksheet.column_dimensions['T'].width = 1 #
    worksheet.column_dimensions['U'].width = 1 # 
    worksheet.column_dimensions['V'].width = 1 # 
    worksheet.column_dimensions['W'].width = 1 # 
    worksheet.column_dimensions['X'].width = 1 # 
    worksheet.column_dimensions['Y'].width = 1 # 
    worksheet.column_dimensions['Z'].width = 1 # 
    worksheet.column_dimensions['AA'].width = 1 # 
    worksheet.column_dimensions['AB'].width = 1 # 
    worksheet.column_dimensions['AC'].width = 1 # 
    worksheet.column_dimensions['AD'].width = 1 # 
    worksheet.column_dimensions['AE'].width = 1 # 
    worksheet.column_dimensions['AF'].width = 22 # Coordinate
    worksheet.column_dimensions['AG'].width = 1 # 
    worksheet.column_dimensions['AH'].width = 1 # 
    worksheet.column_dimensions['AI'].width = 1 # 
    worksheet.column_dimensions['AJ'].width = 1 # 
    worksheet.column_dimensions['AK'].width = 1 # 
    worksheet.column_dimensions['AL'].width = 1 # 
    worksheet.column_dimensions['AM'].width = 1 # 
    worksheet.column_dimensions['AN'].width = 1 # 
    worksheet.column_dimensions['AO'].width = 1 # 
    worksheet.column_dimensions['AP'].width = 1 # 
    worksheet.column_dimensions['AQ'].width = 1 # 
    worksheet.column_dimensions['AR'].width = 1 # 
    worksheet.column_dimensions['AS'].width = 1 # 
    worksheet.column_dimensions['AT'].width = 12 # icon
    worksheet.column_dimensions['AU'].width = 22 # original_file
    worksheet.column_dimensions['AV'].width = 1 # 
    worksheet.column_dimensions['AW'].width = 1 # 
    worksheet.column_dimensions['AX'].width = 1 # 
    worksheet.column_dimensions['AY'].width = 1 # 
    worksheet.column_dimensions['AZ'].width = 1 # 
    worksheet.column_dimensions['BA'].width = 1 # 
    worksheet.column_dimensions['BB'].width = 1 # 
    worksheet.column_dimensions['BC'].width = 1 # 
    worksheet.column_dimensions['BD'].width = 1 # 
    worksheet.column_dimensions['BE'].width = 1 # 
    worksheet.column_dimensions['BF'].width = 1 # 
    worksheet.column_dimensions['BG'].width = 1 # 
    worksheet.column_dimensions['BH'].width = 1 # 


    worksheet.column_dimensions['BI'].width = 20 # MAC
    worksheet.column_dimensions['BJ'].width = 15 # SSID
    
    worksheet.column_dimensions['BS'].width = 30 # companyName

    for i in range(len(data)):
        if data[i] is None:
            data[i] = ''


    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            try:
                cell_data = row_data.get(col_name)
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    # Create a new worksheet for color codes
    color_worksheet = workbook.create_sheet(title='ColorCode')
    color_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    color_worksheet.column_dimensions['A'].width = 14# Color
    color_worksheet.column_dimensions['B'].width = 20# Description


    # Excel row height
    color_worksheet.row_dimensions[2].height = 22  # Adjust the height as needed
    color_worksheet.row_dimensions[3].height = 22
    color_worksheet.row_dimensions[4].height = 23
    color_worksheet.row_dimensions[5].height = 23
    color_worksheet.row_dimensions[6].height = 40   # truck

    color_worksheet.cell(row=1, column=1).value = 'Color'
    color_worksheet.cell(row=1, column=2).value = 'description'
    color_worksheet.cell(row=2, column=1).value = 'Red'
    color_worksheet.cell(row=3, column=1).value = 'Orange'
    color_worksheet.cell(row=4, column=1).value = 'Green'
    color_worksheet.cell(row=5, column=1).value = 'Yellow'

    color_worksheet.cell(row=7, column=1).value = 'ABBREVIATIONS'
    color_worksheet.cell(row=8, column=1).value = 'AKA'
    color_worksheet.cell(row=9, column=1).value = 'DOB'
    color_worksheet.cell(row=10, column=1).value = 'VIS'
    color_worksheet.cell(row=11, column=1).value = 'VIN'
    color_worksheet.cell(row=12, column=1).value = 'VYR'
    color_worksheet.cell(row=13, column=1).value = 'VMA'
    color_worksheet.cell(row=14, column=1).value = 'LIC'
    color_worksheet.cell(row=15, column=1).value = 'LIY'
    color_worksheet.cell(row=16, column=1).value = 'DLN'
    color_worksheet.cell(row=17, column=1).value = 'DLS'

       
    color_worksheet.cell(row=2, column=2).value = 'Bad Intel or dead link'
    color_worksheet.cell(row=3, column=2).value = 'Research'
    color_worksheet.cell(row=4, column=2).value = 'Good Intel'
    color_worksheet.cell(row=5, column=2).value = 'Highlighted'

    color_worksheet.cell(row=8, column=2).value = 'Also Known As (Alias)'
    color_worksheet.cell(row=9, column=2).value = 'Date of Birth'
    color_worksheet.cell(row=10, column=2).value = 'Vehicle State'
    color_worksheet.cell(row=11, column=2).value = 'Vehicle Identification Number'
    color_worksheet.cell(row=12, column=2).value = 'Vehicle Year'
    color_worksheet.cell(row=13, column=2).value = 'Vehicle Make'
    color_worksheet.cell(row=14, column=2).value = 'License'
    color_worksheet.cell(row=15, column=2).value = 'License Year'
    color_worksheet.cell(row=16, column=2).value = 'Drivers License Number'
    color_worksheet.cell(row=17, column=2).value = 'Drivers License State'


    # colored fills
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    # Apply the orange fill to the cell in row 2, column 2
    color_worksheet.cell(row=2, column=2).fill = red_fill
    color_worksheet.cell(row=3, column=2).fill = orange_fill
    color_worksheet.cell(row=4, column=2).fill = green_fill
    color_worksheet.cell(row=5, column=2).fill = yellow_fill


    # Create a new worksheet for logs
    log_worksheet = workbook.create_sheet(title='Log')
    log_worksheet.freeze_panes = 'B2'  # Freeze cells

# Date, Subject, Requesting Agency, Requesting Agent, Case, Summary of Findings, Source, Notes, Requestor

    # Excel column width
    log_worksheet.column_dimensions['A'].width = 14# Date
    log_worksheet.column_dimensions['B'].width = 20# Subject
    log_worksheet.column_dimensions['C'].width = 24# Requesting Agency
    log_worksheet.column_dimensions['D'].width = 20# Requesting Agent
    log_worksheet.column_dimensions['E'].width = 14# Case
    log_worksheet.column_dimensions['F'].width = 20# Summary of Findings
    log_worksheet.column_dimensions['G'].width = 14# Source
    log_worksheet.column_dimensions['H'].width = 25# Notes

    log_worksheet.cell(row=1, column=1).value = 'Date'
    log_worksheet.cell(row=1, column=2).value = 'Subject'
    log_worksheet.cell(row=1, column=3).value = 'Requesting Agency'
    log_worksheet.cell(row=1, column=4).value = 'Requesting Agent'
    log_worksheet.cell(row=1, column=5).value = 'Case'
    log_worksheet.cell(row=1, column=6).value = 'Summary of Findings'
    log_worksheet.cell(row=1, column=7).value = 'Notes'



    workbook.save(output_xlsx)
    workbook.close()

    
def usage():
    file = sys.argv[0].split('\\')[-1]

    print(f'\nDescription: {color_green}{description}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}')
    print(f'\nExample:')
    print(f'  python {file} -b      # create a blank sheet')  
    print(f'  python {file} -C      # clear logs off the HackRF')     
    print(f'  python {file} -L      # log grabber (HackRF)')      
    print(f'  python {file} -p      # parse HackRF text') 
    print(f'  python {file} -p -I logs -O WarDrive_.xlsx ')  
    print(f'  python {file} -w -I WigleWifi_Neighborhood.csv.gz     # parse wigle log')  
    print(f'  python {file} -w -I WigleWifi_sample.csv ')   
    print(f'  python {file} -x -I input_.xlsx ')       
    print(f'\n{color_reset}')    


if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

1.0.4 - conbined with hackRf logs parser
1.0.2 - ADSB.TXT, AFSK.txt (useless), APRS.TXT, & BLELOG_*.TXT parsing
1.0.1 - protectList.csv and watchList.csv Tagging, keep the .gz filename
1.0.0 - removes dulicate MAC's and keeps the stongest signal
0.0.1 - convert MfgrId to a real company
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
replace any non utf8 characters like an icon
create a module that merges Timestamp:Coordinates logs with timestamp logs missing coordinates. 
    needs to pick out the closest timestamp. Update the missing coordinate.
add protectList/watchList check to hackrf parser
test -L on a real hackrf log set (it's not copying the logs from d to \Logs

some devices have "quote,quote" in them and it breaks the csv parsing
    sometimes time and latitude are wrong. [ESS]	-93

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

If you have 2000 or more items, delete the # (label) row before making the KML file 
or google earth will be unusable. Leave just the Tagged labels.

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
