import os
import re
import json
import hashlib
import mailbox
from datetime import datetime
from email import policy
from email.parser import BytesParser
from email.header import decode_header
from email.utils import parsedate_to_datetime
from bs4 import BeautifulSoup
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def clean_date(date_str):
    try:
        dt = parsedate_to_datetime(date_str)
        if dt:
            return dt.isoformat().replace('T', ' ')
    except:
        pass
    try:
        dt = parser.parse(date_str)
        return dt.isoformat().replace('T', ' ')
    except:
        match = re.match(r"(\d{4})\.(\d{2})\.(\d{2})-(\d{2})\.(\d{2})\.(\d{2})", date_str)
        if match:
            dt = datetime(*map(int, match.groups()))
            return dt.isoformat().replace('T', ' ') + 'Z'
    return ''

def sha256_hash(path):
    sha256 = hashlib.sha256()
    try:
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b''):
                sha256.update(chunk)
        return sha256.hexdigest()
    except:
        return ''

def sanitize(text):
    if not isinstance(text, str): return ''
    text = text.strip().strip('"')
    return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', text)

def decode_header_str(header_obj):
    if not header_obj: return ""
    decoded_parts = decode_header(header_obj)
    return ''.join(
        part.decode(encoding or 'utf-8', errors='replace') if isinstance(part, bytes) else str(part)
        for part, encoding in decoded_parts
    ).strip()

def decode_payload(payload):
    for encoding in ['utf-8', 'latin1']:
        try:
            return payload.decode(encoding)
        except:
            continue
    return payload.decode(errors='ignore')

def extract_body(msg):
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                return decode_payload(part.get_payload(decode=True))
            elif part.get_content_type() == 'text/html':
                html = decode_payload(part.get_payload(decode=True))
                return BeautifulSoup(html, 'lxml').get_text()
    else:
        return decode_payload(msg.get_payload(decode=True) or b'')
    return ''

def get_attachments(msg):
    return "; ".join([
        part.get_filename() for part in msg.walk()
        if part.get_content_disposition() == 'attachment' and part.get_filename()
    ])

def deduplicate_by_sha256(data, contacts):
    seen = set()
    deduped_data = []
    deduped_contacts = []

    for item in data:
        hash_val = item.get("sha256")
        if hash_val and hash_val not in seen:
            seen.add(hash_val)
            deduped_data.append(item)

    for contact in contacts:
        hash_val = contact.get("sha256") or contact.get("query")
        if hash_val and hash_val in seen:
            deduped_contacts.append(contact)

    return deduped_data, deduped_contacts
    
def detect_spam(body, email):
    spam_keywords = ["Unsubscribe", "unsubscribe"]
    spam_email_keywords = ["donotreply", "info@", "no-response", "notifications@", "postmaster@", "support@", "verify@"]
    if any(k in body for k in spam_keywords) or any(k in email for k in spam_email_keywords):
        return "Spam"
    return ""

def extract_contact(from_field):
    if ' <' in from_field:
        name, email = from_field.split(' <')
        return name.strip().replace('"', ''), email.replace('>', '').strip()
    return from_field.strip().replace('"', ''), from_field.replace('>', '').strip()

def process_eml_folder(file_paths, output_file=None, progress_display=None):
    data, contacts = [], []


    for path in file_paths:
        file_name = os.path.basename(path)
        sha256 = sha256_hash(path)
        zip_origin = path.split("_unzipped")[0] if "_unzipped" in path else ""
        source = f"{path} (from {os.path.basename(zip_origin)})" if zip_origin else path

        try:
            if file_name.lower().endswith(".eml"):
                with open(path, 'rb') as f:
                    msg = BytesParser(policy=policy.default).parse(f)
                sender = sanitize(msg.get("From"))
                recipient = sanitize(msg.get("To"))
                subject = sanitize(msg.get("Subject"))
                labels = sanitize(msg.get("X-Gmail-Labels"))
                date_obj = clean_date(msg.get("Date") or "")
                body = sanitize(extract_body(msg))
                attachments = sanitize(get_attachments(msg))
                tag = detect_spam(body, sender)
                fullname, email = extract_contact(sender)

                data.append({
                    "Time": date_obj, "From": sender, "To": recipient, "Subject": subject,
                    "Labels": labels, "Body": body, "Attachments": attachments, "Tag": tag,
                    "original_file": file_name, "Source": source, "case": "", "Precedence": "",
                    "List-Unsubscribe": "", "X-Mailer": "", "fullname": fullname, "email": email,
                    "sha256": sha256
                })
                contacts.append({
                    "query": email, "ranking": "3 - contacts", "fullname": fullname,
                    "email": email, "note": recipient, "original_file": file_name,
                    "Source": source, "Tag": tag
                })
                print(f'✅ {file_name}')
            elif file_name.lower().endswith(".mbox"):
                mbox = mailbox.mbox(path)
                for msg in mbox:
                    sender = decode_header_str(msg['from'] or '')
                    recipient = decode_header_str(msg['to'] or '')
                    subject = sanitize(decode_header_str(msg['subject'] or ''))
                    labels = decode_header_str(msg.get('X-Gmail-Labels', ''))
                    date_obj = clean_date(msg.get('date', ''))
                    body = sanitize(extract_body(msg))
                    attachments = get_attachments(msg)
                    tag = detect_spam(body, sender)
                    fullname, email = extract_contact(sender)

                    data.append({
                        "Time": date_obj, "From": sender, "To": recipient, "Subject": subject,
                        "Labels": labels, "Body": body, "Attachments": attachments, "Tag": tag,
                        "original_file": file_name, "Source": source, "case": "", "Precedence": "",
                        "List-Unsubscribe": "", "X-Mailer": "", "fullname": fullname, "email": email,
                        "sha256": sha256
                    })
                    contacts.append({
                        "query": email, "ranking": "3 - contacts", "fullname": fullname,
                        "email": email, "note": recipient, "original_file": file_name,
                        "Source": source, "Tag": tag
                    })

            elif file_name.lower().endswith(".json"):
                with open(path, "r", encoding="utf-8") as f:
                    messages_json = json.load(f)
                for msg in messages_json.get("messages", []):
                    sender_email = msg.get("creator", {}).get("email", "")
                    fullname = msg.get("creator", {}).get("name", "")
                    body = msg.get("text", "")
                    date_obj = clean_date(msg.get("created_date", ""))
                    attachment_data = msg.get("attached_files", "")
                    if isinstance(attachment_data, list):
                        attachments = ", ".join([a.get("export_name", "") for a in attachment_data if isinstance(a, dict)])
                    elif isinstance(attachment_data, dict):
                        attachments = attachment_data.get("export_name", "")
                    else:
                        attachments = ""

                    tag = detect_spam(body, sender_email)

                    data.append({
                        "Time": date_obj, "From": sender_email, "To": "", "Subject": "",
                        "Labels": "", "Body": body, "Attachments": attachments, "Tag": tag,
                        "original_file": file_name, "Source": source, "case": "", "Precedence": "",
                        "List-Unsubscribe": "", "X-Mailer": "", "fullname": fullname, "email": sender_email,
                        "sha256": sha256
                    })
                    contacts.append({
                        "query": sender_email, "ranking": "3 - contacts", "fullname": fullname,
                        "email": sender_email, "note": "", "original_file": file_name,
                        "Source": source, "Tag": tag
                    })

        except Exception as e:
            if progress_display:
                progress_display.insert("end", f"Error parsing {file_name}: {e}\n")

    if output_file:
        write_xlsx(data, contacts, output_file)
    return data, contacts


def write_xlsx(data, contacts, file_path):
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Email"
    ws1.freeze_panes = "B2"

    ws2 = wb.create_sheet("Contacts")
    ws2.freeze_panes = "B2"

    ws3 = wb.create_sheet("Spam")
    ws3.freeze_panes = "B2"

    headers1 = ["Time", "From", "To", "Subject", "Body", "Attachments", "Labels", "Tag",
                "original_file", "Source", "case", "Precedence", "List-Unsubscribe", "X-Mailer",
                "fullname", "email", "sha256"]
    headers2 = ["query", "ranking", "fullname", "email", "note", "original_file", "Source", "Tag"]

    # Format headers for Eml sheet
    for col_index, header in enumerate(headers1):
        cell = ws1.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1, 2, 3, 4, 6]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Format headers for Contacts sheet
    for col_index, header in enumerate(headers2):
        cell = ws2.cell(row=1, column=col_index + 1)
        cell.value = header

    # Format headers for Spam sheet
    for col_index, header in enumerate(headers1):
        cell = ws3.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1, 2, 3, 4, 6]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")



    # Write data rows to Eml sheet
    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers1):
            try:
                ws1.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name, "")
            except Exception as e:
                print(f"Error writing Eml row {row_index}: {e}")

    # Write data rows to Contacts sheet
    for row_index, row_data in enumerate(contacts):
        for col_index, col_name in enumerate(headers2):
            try:
                ws2.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name, "")
            except Exception as e:
                print(f"Error writing Contacts row {row_index}: {e}")

    try:
        wb.save(file_path)
        print(f"✅ Excel file saved: {file_path}")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")


if __name__ == "__main__":
    test_folder = os.path.abspath("email")
    test_output = "email.xlsx"
    test_files = []

    for root, _, files in os.walk(test_folder):
        for file in files:
            if file.lower().endswith((".eml", ".mbox", ".json")):
                test_files.append(os.path.join(root, file))

    print(f"Found {len(test_files)} email files")
    process_eml_folder(test_files, test_output)


__all__ = ["process_eml_folder", "deduplicate_by_sha256"]
