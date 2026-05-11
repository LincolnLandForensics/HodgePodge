import os
import re
import json
import hashlib
import mailbox
import extract_msg
import zipfile
import tkinter as tk
from datetime import datetime
from email import policy
from email.parser import BytesParser
from email.header import decode_header
from email.utils import parsedate_to_datetime
from bs4 import BeautifulSoup
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog, scrolledtext

Version = "1.0.5"
description = "Convert email (.eml, .msg, .mbox and .json) files to xlsx"
# --- CORE PARSING LOGIC ---

def clean_date(date_str):
    try:
        dt = parsedate_to_datetime(date_str)
        if dt:
            return dt.isoformat().replace('T', ' ')
    except:
        pass
    try:
        dt = parser.parse(date_str)
        return dt.isoformat().replace('T', ' ')
    except:
        match = re.match(r"(\d{4})\.(\d{2})\.(\d{2})-(\d{2})\.(\d{2})\.(\d{2})", date_str)
        if match:
            dt = datetime(*map(int, match.groups()))
            return dt.isoformat().replace('T', ' ') + 'Z'
    return ''

def sha256_hash(path):
    sha256 = hashlib.sha256()
    try:
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b''):
                sha256.update(chunk)
        return sha256.hexdigest()
    except:
        return ''

def sanitize(text):
    if not isinstance(text, str): return ''
    text = text.strip().strip('"')
    return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', text)

def decode_header_str(header_obj):
    if not header_obj: return ""
    decoded_parts = decode_header(header_obj)
    return ''.join(
        part.decode(encoding or 'utf-8', errors='replace') if isinstance(part, bytes) else str(part)
        for part, encoding in decoded_parts
    ).strip()

def decode_payload(payload):
    for encoding in ['utf-8', 'latin1']:
        try:
            return payload.decode(encoding)
        except:
            continue
    return payload.decode(errors='ignore')

def extract_body(msg):
    body = ""
    if msg.is_multipart():
        # Prefer plain text
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                body = decode_payload(part.get_payload(decode=True))
                break
        
        # If no plain text found, try HTML
        if not body:
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    html = decode_payload(part.get_payload(decode=True))
                    body = BeautifulSoup(html, 'lxml').get_text(separator='\n')
                    break
    else:
        payload = decode_payload(msg.get_payload(decode=True) or b'')
        if msg.get_content_type() == 'text/html':
            body = BeautifulSoup(payload, 'lxml').get_text(separator='\n')
        else:
            body = payload
    
    if body:
        # Remove excess blank lines (3 or more newlines become 2)
        body = re.sub(r'\n\s*\n\s*\n+', '\n\n', body)
        return body.strip()
    return ''

def get_attachments(msg):
    filenames = []
    for part in msg.walk():
        filename = part.get_filename()
        if filename:
            filenames.append(filename)
    return "; ".join(filenames)

def deduplicate_by_sha256(data, contacts):
    seen = set()
    deduped_data = []
    deduped_contacts = []

    for item in data:
        hash_val = item.get("sha256")
        if hash_val and hash_val not in seen:
            seen.add(hash_val)
            deduped_data.append(item)

    for contact in contacts:
        hash_val = contact.get("sha256") or contact.get("query")
        if hash_val and hash_val in seen:
            deduped_contacts.append(contact)

    return deduped_data, deduped_contacts
    
def detect_spam(body, email, list_unsubscribe=""):
    tag = ""
    spam_keywords = ["Unsubscribe", "unsubscribe"]
    spam_email_keywords = ["donotreply", "info@", "no-reply", "noreply", "no-response", "notifications@", "postmaster@", "support@", "verify@"]
    if any(k in body for k in spam_keywords) or any(k in email for k in spam_email_keywords):
        tag = "Spam"
    
    if body and "unsubscribe" in body.lower() and list_unsubscribe == '':
        list_unsubscribe = '_unsubscribe'
    
    if email and 'CashApp' in email:
        tag = 'CashApp'
    
    return tag, list_unsubscribe

def extract_contact(sender_raw):
    if not sender_raw:
        return ("", "")

    s = sender_raw.strip().replace("\n", " ").replace("\r", " ")
    email_pattern = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"

    email_match = re.search(email_pattern, s)
    email = email_match.group(0) if email_match else ""

    fullname = s.replace(email, "").replace("<", "").replace(">", "").strip()
    fullname = fullname.rstrip() if fullname else ""
    fullname = fullname.strip().strip('\'"')

    if not email and "<" in s and ">" not in s:
        possible = s.split("<", 1)[1]
        if "@" in possible:
            email = possible.strip(" >")

    return (fullname, email)

def process_eml_folder(file_paths, output_file=None, progress_display=None, case_id=""):
    data = []
    contacts_unique = {}

    for path in file_paths:
        file_name = os.path.basename(path)
        sha256 = sha256_hash(path)
        zip_origin = path.split("_unzipped")[0] if "_unzipped" in path else ""
        source = f"{path} (from {os.path.basename(zip_origin)})" if zip_origin else path
        # source = os.path.basename(source) # this works
        
        try:
            if file_name.lower().endswith((".eml", ".emlx")):
                with open(path, 'rb') as f:
                    content = f.read()
                match = re.match(br'^(\d+)\s*\r?\n', content)
                if match:
                    byte_count_len = len(match.group(0))
                    byte_count = int(match.group(1))
                    eml_content = content[byte_count_len:byte_count_len+byte_count]
                    msg = BytesParser(policy=policy.default).parsebytes(eml_content)
                else:
                    msg = BytesParser(policy=policy.default).parsebytes(content)

                sender = sanitize(msg.get("From"))
                recipient = sanitize(msg.get("To"))
                subject = sanitize(msg.get("Subject"))
                labels = sanitize(msg.get("X-Gmail-Labels"))
                
                # Enhanced label extraction for .emlx from plist
                if not labels and file_name.lower().endswith(".emlx") and match:
                    try:
                        import plistlib
                        plist_content = content[byte_count_len + byte_count:]
                        if plist_content:
                            plist_data = plistlib.loads(plist_content.strip())
                            if "gmail-label-ids" in plist_data:
                                labels = f"Gmail IDs: {', '.join(map(str, plist_data['gmail-label-ids']))}"
                    except:
                        pass

                date_obj = clean_date(msg.get("Date") or "")
                body = sanitize(extract_body(msg))
                attachments = sanitize(get_attachments(msg))
                tag, list_unsubscribe = detect_spam(body, sender)
                fullname, email = extract_contact(msg.get("From"))

                data.append({
                    "Time": date_obj, "From": sender, "To": recipient, "Subject": subject,
                    "Labels": labels, "Body": body, "Attachments": attachments, "Tag": tag,
                    "original_file": file_name, "Source": source, "case": case_id, "Precedence": "",
                    "List-Unsubscribe": list_unsubscribe, "X-Mailer": "", "fullname": fullname, "email": email,
                    "sha256": sha256
                })
                if email and email not in contacts_unique:
                    ranking = "9 - contacts" if "Spam" in tag else "3 - contacts"
                    contacts_unique[email] = {
                        "query": email, "ranking": ranking, "fullname": fullname,
                        "email": email, "note": recipient, "original_file": file_name,
                        "Source": source, "Tag": tag, "case": case_id
                    }
                print(f'[DONE] {file_name}')

            elif file_name.lower().endswith(".msg"):
                msg = extract_msg.Message(path)
                sender = sanitize(msg.sender or "")
                recipient = sanitize(msg.to or "")
                subject = sanitize(msg.subject or "")
                labels = ""
                date_obj = clean_date(str(msg.date) if msg.date else "")
                
                if msg.body:
                    body = sanitize(msg.body)
                elif msg.htmlBody:
                    html_content = decode_payload(msg.htmlBody)
                    body = sanitize(BeautifulSoup(html_content, 'lxml').get_text())
                else:
                    body = ""
                
                attachments_list = []
                for att in msg.attachments:
                    if hasattr(att, 'longFilename') and att.longFilename:
                        attachments_list.append(att.longFilename)
                    elif hasattr(att, 'shortFilename') and att.shortFilename:
                        attachments_list.append(att.shortFilename)
                attachments = sanitize("; ".join(attachments_list))
                
                tag, list_unsubscribe = detect_spam(body, sender)
                fullname, email = extract_contact(msg.sender)
                
                data.append({
                    "Time": date_obj, "From": sender, "To": recipient, "Subject": subject,
                    "Labels": labels, "Body": body, "Attachments": attachments, "Tag": tag,
                    "original_file": file_name, "Source": source, "case": case_id, "Precedence": "",
                    "List-Unsubscribe": list_unsubscribe, "X-Mailer": "", "fullname": fullname, "email": email,
                    "sha256": sha256
                })
                if email and email not in contacts_unique:
                    ranking = "9 - contacts" if "Spam" in tag else "3 - contacts"
                    contacts_unique[email] = {
                        "query": email, "ranking": ranking, "fullname": fullname,
                        "email": email, "note": recipient, "original_file": file_name,
                        "Source": source, "Tag": tag, "case": case_id
                    }
                print(f'[DONE] {file_name}')

            elif file_name.lower().endswith(".mbox"):
                try:
                    mbox = mailbox.mbox(path)
                    for message in mbox:
                        sender_raw = message.get('from', '')
                        sender = decode_header_str(sender_raw)
                        recipient = decode_header_str(message.get('to', ''))
                        subject = decode_header_str(message.get('subject', ''))
                        
                        date_str = message.get('date', '')
                        date_obj = clean_date(date_str) if date_str else ''
                        
                        body = extract_body(message)
                        precedence = message.get('precedence', '')
                        list_unsubscribe = message.get('List-Unsubscribe', '')
                        x_mailer = message.get('X-Mailer', '')
                        attachments = get_attachments(message)
                        
                        fullname, email = extract_contact(sender_raw)
                        tag, list_unsubscribe = detect_spam(body, email, list_unsubscribe)
                        
                        data.append({
                            "Time": date_obj, "From": sender, "To": recipient, "Subject": subject,
                            "Labels": "", "Body": body, "Attachments": attachments, "Tag": tag,
                            "original_file": file_name, "Source": source, "case": case_id, "Precedence": precedence,
                            "List-Unsubscribe": list_unsubscribe, "X-Mailer": x_mailer, 
                            "fullname": fullname, "email": email, "sha256": sha256
                        })
                        
                        if email and email not in contacts_unique:
                            ranking = "9 - contacts" if "Spam" in tag else "3 - contacts"
                            contacts_unique[email] = {
                                "query": email, "ranking": ranking, "fullname": fullname,
                                "email": email, "note": recipient, "original_file": file_name,
                                "Source": source, "Tag": tag, "case": case_id
                            }
                except Exception as e:
                    if progress_display:
                        progress_display.insert("end", f"Error parsing mbox {file_name}: {e}\n")
                    print(f"Error parsing mbox {file_name}: {e}")

            elif file_name.lower().endswith(".json"):
                with open(path, "r", encoding="utf-8") as f:
                    messages_json = json.load(f)
                for msg in messages_json.get("messages", []):
                    sender_email = msg.get("creator", {}).get("email", "")
                    fullname = msg.get("creator", {}).get("name", "")
                    body = msg.get("text", "")
                    date_obj = clean_date(msg.get("created_date", ""))
                    attachment_data = msg.get("attached_files", "")
                    if isinstance(attachment_data, list):
                        attachments = ", ".join([a.get("export_name", "") for a in attachment_data if isinstance(a, dict)])
                    elif isinstance(attachment_data, dict):
                        attachments = attachment_data.get("export_name", "")
                    else:
                        attachments = ""

                    tag, list_unsubscribe = detect_spam(body, sender_email)

                    data.append({
                        "Time": date_obj, "From": sender_email, "To": "", "Subject": "",
                        "Labels": "", "Body": body, "Attachments": attachments, "Tag": tag,
                        "original_file": file_name, "Source": source, "case": case_id, "Precedence": "",
                        "List-Unsubscribe": list_unsubscribe, "X-Mailer": "", "fullname": fullname, "email": sender_email,
                        "sha256": sha256
                    })
                    if sender_email and sender_email not in contacts_unique:
                        ranking = "9 - contacts" if "Spam" in tag else "3 - contacts"
                        contacts_unique[sender_email] = {
                            "query": sender_email, "ranking": ranking, "fullname": fullname,
                            "email": sender_email, "note": "", "original_file": file_name,
                            "Source": source, "Tag": tag, "case": case_id
                        }

            elif file_name:
                attachments = file_name
                data.append({
                    "Time": "", "From": "", "To": "", "Subject": "",
                    "Labels": "", "Body": "", "Attachments": attachments, "Tag": 'Attachment',
                    "original_file": file_name, "Source": source, "case": case_id, "Precedence": "", 
                    "List-Unsubscribe": "", "X-Mailer": "", "fullname": "", "email": "", "sha256": sha256
                })

        except Exception as e:
            if progress_display:
                progress_display.insert("end", f"Error parsing {file_name}: {e}\n")

    contacts = list(contacts_unique.values())
    if output_file:
        write_xlsx(data, contacts, output_file)
    return data, contacts

def write_xlsx(data, contacts, file_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Email"
    ws1.freeze_panes = "B2"

    ws2 = wb.create_sheet("Contacts")
    ws2.freeze_panes = "B2"

    ws3 = wb.create_sheet("Spam")
    ws3.freeze_panes = "B2"

    headers1 = ["Time", "From", "To", "Subject", "Body", "Attachments", "Labels", "Tag",
                "original_file", "Source", "case", "Precedence", "List-Unsubscribe", "X-Mailer",
                "fullname", "email", "sha256"]
    headers2 = ["query", "ranking", "fullname", "url", "email", "user", "phone"
    , "business", "fulladdress", "city", "state", "country", "note", "AKA", "DOB"
    , "SEX", "info", "misc", "firstname", "middlename", "lastname", "associates"
    , "case", "sosfilenumber", "owner", "president", "sosagent", "managers", "Time"
    , "Latitude", "Longitude", "Coordinate", "original_file", "Source", "Source file information"
    , "Plate", "VIS", "VIN", "VYR", "VMA", "LIC", "LIY", "DLN", "DLS", "content", "referer"
    , "osurl", "titleurl", "pagestatus", "ip", "dnsdomain", "Tag", "Icon", "Type"]

    # Format headers for Email sheet
    for col_index, header in enumerate(headers1):
        cell = ws1.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1, 2, 3, 4, 5, 6]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    widths1 = [19, 30, 30, 30, 50, 20, 15, 15, 20, 15, 9, 12, 18, 15, 20, 25, 64]
    for i, w in enumerate(widths1):
        col_letter = chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
        ws1.column_dimensions[col_letter].width = w

    # Format headers for Contacts sheet
    for col_index, header in enumerate(headers2):
        cell = ws2.cell(row=1, column=col_index + 1)
        cell.value = header

    widths2 = [15,20,20,25,25,15,14,16,24,12,10,8,20,14,11,5,20,20,10,11,10,10,10,13,10,10,10,10,15,12,12,22,12,12,12,15,15]
    for i, w in enumerate(widths2):
        col_letter = chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
        ws2.column_dimensions[col_letter].width = w

    # Format headers for Spam sheet
    for col_index, header in enumerate(headers1):
        cell = ws3.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1, 2, 3, 4, 6]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for i, w in enumerate(widths1):
        col_letter = chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
        ws3.column_dimensions[col_letter].width = w

    # Write data rows to Email sheet
    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers1):
            try:
                ws1.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name, "")
            except Exception as e:
                print(f"Error writing Email row {row_index}: {e}")

    # Write data rows to Contacts sheet
    for row_index, row_data in enumerate(contacts):
        for col_index, col_name in enumerate(headers2):
            try:
                ws2.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name, "")
            except Exception as e:
                print(f"Error writing Contacts row {row_index}: {e}")

    try:
        wb.save(file_path)
        print(f"Excel file saved: {file_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

# --- GUI LOGIC ---

def extract_zip_recursive(zip_path, extract_to):
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
    except Exception as e:
        print(f"Error extracting {zip_path}: {e}")

def find_email_files(folder):
    email_files = []
    for root, _, files in os.walk(folder):
        for file in files:
            full_path = os.path.join(root, file)
            email_files.append(full_path)
    return email_files

def run_parser(input_folder, output_file, progress_display, dedup_enabled, case_id):
    temp_dir = os.path.join(input_folder, "_unzipped")
    print(f'saving unzipped files to {temp_dir}')
    os.makedirs(temp_dir, exist_ok=True)

    # Extract ZIPs
    for file in os.listdir(input_folder):
        if file.lower().endswith('.zip'):
            zip_path = os.path.join(input_folder, file)
            print(f'unzipping {zip_path}')
            zip_extract_path = os.path.join(temp_dir, os.path.splitext(file)[0])
            extract_zip_recursive(zip_path, zip_extract_path)
            progress_display.insert(tk.END, f"📦 Extracted ZIP: {file}\n")
            progress_display.see(tk.END)

    # Collect all email files
    all_emails = []
    for path in [input_folder, temp_dir]:
        if os.path.exists(path):
            all_emails.extend(find_email_files(path))

    if not all_emails:
        progress_display.insert(tk.END, "⚠️ No email files found.\n")
        progress_display.see(tk.END)
        return

    progress_display.insert(tk.END, f"📨 Found {len(all_emails)} email files\n")
    progress_display.see(tk.END)

    try:
        # Step 1: parse all files
        raw_data, raw_contacts = process_eml_folder(all_emails, output_file=None, progress_display=progress_display, case_id=case_id)
        
        # Step 2: deduplicate if needed
        if dedup_enabled:
            print(f'DeDuplicating')
            raw_data, raw_contacts = deduplicate_by_sha256(raw_data, raw_contacts)
            progress_display.insert(tk.END, f"🧹 Deduplicated to {len(raw_data)} unique emails\n")

        # Step 3: write to Excel
        write_xlsx(raw_data, raw_contacts, output_file)
        progress_display.insert(tk.END, f"\n✅ Export complete: {output_file}\n")
        progress_display.see(tk.END)
    except Exception as e:
        progress_display.insert(tk.END, f"❌ Error during parsing: {e}\n")
        progress_display.see(tk.END)

def launch_gui():
    root = tk.Tk()
    root.title(f"EML to XLSX Parser v{Version}")

    # Description
    tk.Label(root, text=description, font=("Arial", 10, "italic")).grid(row=0, column=0, columnspan=3, pady=(5, 10))

    # Input folder
    tk.Label(root, text="Input Folder:").grid(row=1, column=0, sticky="w")
    input_entry = tk.Entry(root, width=50)
    input_entry.insert(0, os.path.abspath("email"))
    input_entry.grid(row=1, column=1)

    def browse_input():
        folder = filedialog.askdirectory()
        if folder:
            input_entry.delete(0, tk.END)
            input_entry.insert(0, folder)

    tk.Button(root, text="Browse", command=browse_input).grid(row=1, column=2)

    # Output file
    tk.Label(root, text="Output File:").grid(row=2, column=0, sticky="w")
    output_entry = tk.Entry(root, width=50)
    output_entry.insert(0, "email.xlsx")
    output_entry.grid(row=2, column=1)

    # DeDuplicate checkbox
    dedup_var = tk.IntVar(value=0)
    tk.Checkbutton(root, text="DeDuplicate", variable=dedup_var).grid(row=3, column=0, sticky="w")

    # Case input
    tk.Label(root, text="Case:").grid(row=3, column=1, sticky="w")
    case_entry = tk.Entry(root, width=20)
    case_entry.grid(row=3, column=1, padx=(40, 0))

    # Progress display
    progress_display = scrolledtext.ScrolledText(root, width=80, height=20)
    progress_display.grid(row=5, column=0, columnspan=3, pady=10)

    # Start button
    def start():
        input_folder = input_entry.get()
        output_file = output_entry.get()
        case_id = case_entry.get()
        dedup_enabled = dedup_var.get() == 1
        progress_display.delete(1.0, tk.END)
        run_parser(input_folder, output_file, progress_display, dedup_enabled, case_id)

    tk.Button(root, text="Start Parsing", command=start).grid(row=4, column=1, pady=5)

    root.mainloop()

if __name__ == "__main__":
    launch_gui()
