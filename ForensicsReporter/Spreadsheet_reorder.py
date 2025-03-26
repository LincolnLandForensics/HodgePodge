#!/usr/bin/env python3
# coding: utf-8
"""
Read a case sheet and re-organize the order of the headers
It will read the first sheet
just change the order of headers on line 127
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import sys
import argparse
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Color support for Windows 11+
color_red = color_yellow = color_green = color_blue = color_purple = color_reset = ''
if sys.version_info > (3, 7, 9) and os.name == "nt":
    from colorama import Fore, Back, Style
    print(Back.BLACK)
    color_red, color_yellow, color_green = Fore.RED, Fore.YELLOW, Fore.GREEN
    color_blue, color_purple, color_reset = Fore.BLUE, Fore.MAGENTA, Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "Read a case sheet and re-organize the order of the headers"
version = '1.0.0'


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>



def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)    
    parser.add_argument('-b', '--blank', help='read xlsx', required=False, action='store_true')    
    parser.add_argument('-r', '--read', help='read xlsx', required=False, action='store_true')

    args = parser.parse_args()


    global input_file
    input_file = args.input if args.input else "ForensicCases.xlsx"

    global outuput_xlsx
    outuput_xlsx = args.output if args.output else "ForensicCases_new.xlsx"
    if args.blank:
        data = []
        write_xlsx(data, outuput_xlsx)

        msg_blurb = (f'Creating blank file: {outuput_xlsx}')
        msg_blurb_square(msg_blurb, color_green) 
    elif args.read:
        # create_xlsx()

        file_exists = os.path.exists(input_file)
        if file_exists == True:
            msg_blurb = (f'Reading {input_file}')
            msg_blurb_square(msg_blurb, color_green)    
            
            data = read_xlsx(input_file)
            # workbook.save(outuput_xlsx)
            # write_xlsx(data)
            write_xlsx(data, input_file)
            # workbook.close()
            msg_blurb = (f'Writing to {outuput_xlsx}')
            msg_blurb_square(msg_blurb, color_green)            

        else:
            msg_blurb = (f'{input_file} does not exist')
            msg_blurb_square(msg_blurb, color_red)      
            exit()

    else:
        usage()
    
    return 0


def msg_blurb_square(msg, color):
    border = f"+{'-' * (len(msg) + 2)}+"
    print(f"{color}{border}\n| {msg} |\n{border}{color_reset}")

def read_xlsx(file_path):
    """Read data from an XLSX file and return a list of dictionaries."""
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        row_data["zero"] = "bobs your uncle" if row_data.get("zero") == "test" else row_data.get("zero")
        row_data["one"] = "avocadoes rule" if row_data.get("one") == "TeslaSucks" else row_data.get("one")
        data.append(row_data)
    
    wb.close()
    return data


def write_xlsx(data, file_path):
    '''
    The write_xlsx() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Cases'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'


    headers = ["caseNumber", "exhibit", "caseName", "subjectBusinessName", "caseType"
    , "caseAgent", "forensicExaminer", "reportStatus", "notes", "summary", "tempNotes"
    , "exhibitType", "makeModel", "serial", "OS", "hostname", "userName", "userPwd"
    , "email", "emailPwd", "ip", "phoneNumber", "phoneIMEI", "phone2", "phoneIMEI2"
    , "mobileCarrier", "biosTime", "currentTime", "timezone", "shutdownMethod"
    , "shutdownTime", "seizureAddress", "seizureRoom", "dateSeized", "seizedBy"
    , "seizureStatus", "dateReceived", "receivedBy", "removalDate", "removalStaff"
    , "reasonForRemoval", "inventoryDate", "storageLocation", "status", "imagingTool"
    , "imagingType", "imageMD5", "imageSHA256", "imageSHA1", "verifyHash", "writeBlocker"
    , "imagingStarted", "imagingFinished", "storageType", "storageMakeModel"
    , "storageSerial", "storageSize", "evidenceDataSize", "analysisTool"
    , "analysisTool2", "exportLocation", "exportedEvidence", "qrCode", "operation"
    , "vaultCaseNumber", "vaultTotal", "caseNumberOrig", "Action", "priority", "temp"]




    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header

    # set colors of headers
    orange_columns = ['A', 'C', 'd', 'e', 'f', 'g', 'h']
    for col in orange_columns: 
        cell = worksheet[f"{col}1"]
        cell.fill = PatternFill(start_color='FFc000', end_color='FFc000', fill_type='solid')    #orange

    yellow_columns = ['B', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']
    for col in yellow_columns:
        cell = worksheet[f"{col}1"]
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    violet_columns = ['I', 'J', 'K']
    for col in violet_columns:
        cell = worksheet[f"{col}1"]
        cell.fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')    # purple

    green_columns = ['AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP' ]
    for col in green_columns:
        cell = worksheet[f"{col}1"]
        cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')    # green

    blue_columns = ['AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ']
    for col in blue_columns:
        cell = worksheet[f"{col}1"]
        cell.fill = PatternFill(start_color='66CCFF', end_color='66CCFF', fill_type='solid')    # blue

    pink_columns = ['BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ' ]
    for col in pink_columns:
        cell = worksheet[f"{col}1"]
        cell.fill = PatternFill(start_color='FF99FF', end_color='FF99FF', fill_type='solid')    # pink

    # Excel column width
    worksheet.column_dimensions['A'].width = 15 #  caseNumber
    worksheet.column_dimensions['B'].width = 7 #  exhibit
    worksheet.column_dimensions['C'].width = 16 #  caseName
    worksheet.column_dimensions['D'].width = 25 #  subjectBusinessName
    worksheet.column_dimensions['E'].width = 16 #  caseType
    worksheet.column_dimensions['F'].width = 25 #  caseAgent
    worksheet.column_dimensions['G'].width = 15 #  forensicExaminer
    worksheet.column_dimensions['H'].width = 13 #  reportStatus
    worksheet.column_dimensions['I'].width = 25 #  notes
    worksheet.column_dimensions['J'].width = 15 #  summary
    worksheet.column_dimensions['K'].width = 40 #  tempNotes
    worksheet.column_dimensions['L'].width = 12 #  exhibitType
    worksheet.column_dimensions['M'].width = 30 #  
    worksheet.column_dimensions['N'].width = 17 #  
    worksheet.column_dimensions['O'].width = 15 #  
    worksheet.column_dimensions['P'].width = 18 #  
    worksheet.column_dimensions['Q'].width = 12 # 
    worksheet.column_dimensions['R'].width = 12 #  
    worksheet.column_dimensions['S'].width = 20 #  
    worksheet.column_dimensions['T'].width = 12 #  
    worksheet.column_dimensions['U'].width = 14 #  
    worksheet.column_dimensions['V'].width = 14 #  
    worksheet.column_dimensions['W'].width = 16 #  
    worksheet.column_dimensions['X'].width = 16 #  
    worksheet.column_dimensions['Y'].width = 16 #  
    worksheet.column_dimensions['Z'].width = 15 #  
    worksheet.column_dimensions['AA'].width = 16 #  
    worksheet.column_dimensions['AB'].width = 16 #  
    worksheet.column_dimensions['AC'].width = 12 #  
    worksheet.column_dimensions['AD'].width = 15 #  
    worksheet.column_dimensions['AE'].width = 16 #  
    worksheet.column_dimensions['AF'].width = 15 #  
    worksheet.column_dimensions['AG'].width = 12 #  
    worksheet.column_dimensions['AH'].width = 16 #  
    worksheet.column_dimensions['AI'].width = 12 #  
    worksheet.column_dimensions['AJ'].width = 18 #  
    worksheet.column_dimensions['AK'].width = 16 #  
    worksheet.column_dimensions['AL'].width = 15 #  
    worksheet.column_dimensions['AM'].width = 16 #  
    worksheet.column_dimensions['AN'].width = 25 #  
    worksheet.column_dimensions['AO'].width = 18 #  
    worksheet.column_dimensions['AP'].width = 15 #  
    worksheet.column_dimensions['AQ'].width = 25 #  
    worksheet.column_dimensions['AR'].width = 12 #  
    worksheet.column_dimensions['AS'].width = 24 #  
    worksheet.column_dimensions['AT'].width = 15 #  
    worksheet.column_dimensions['AU'].width = 16 #  
    worksheet.column_dimensions['AV'].width = 15 #  
    worksheet.column_dimensions['AW'].width = 15 #  
    worksheet.column_dimensions['AX'].width = 11 #  
    worksheet.column_dimensions['AY'].width = 15 #  
    worksheet.column_dimensions['AZ'].width = 22 #  
    worksheet.column_dimensions['BA'].width = 16 #  
    worksheet.column_dimensions['BB'].width = 13 #  
    worksheet.column_dimensions['BC'].width = 23 #  
    worksheet.column_dimensions['BD'].width = 19 #  
    worksheet.column_dimensions['BE'].width = 14 #  
    worksheet.column_dimensions['BF'].width = 15 #  
    worksheet.column_dimensions['BG'].width = 23 #  
    worksheet.column_dimensions['BH'].width = 15 #  
    worksheet.column_dimensions['BI'].width = 25 #  
    worksheet.column_dimensions['BJ'].width = 15 #  
    worksheet.column_dimensions['BK'].width = 15 #  
    worksheet.column_dimensions['BL'].width = 15 #  
    worksheet.column_dimensions['BM'].width = 19 #  
    worksheet.column_dimensions['BN'].width = 15 #  
    worksheet.column_dimensions['BO'].width = 19 #      
    worksheet.column_dimensions['BP'].width = 10 #  
    worksheet.column_dimensions['BQ'].width = 9 #  
    worksheet.column_dimensions['BR'].width = 5 #

    for row_index, row_data in enumerate(data):
        # print(f'{color_purple}Processing row: {row_data}{color_reset}')  # Debugging output
        # print("row_index = %s" %(row_index))
        # print(f'{color_red}row = {row_index+1}{color_reset}')        

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    workbook.save(outuput_xlsx)
    
def usage():
    print(f"Usage: {sys.argv[0]} -r [-I input_case.xlsx] [-O output.xlsx]")
    print("Example:")
    print(f"   python {sys.argv[0]} -b")
    print(f"   python {sys.argv[0]} -r")    
    print(f"   python {sys.argv[0]} -r -I ForensicCases.xlsx -O ForensicCases_new.xlsx")

if __name__ == '__main__':
    main()
