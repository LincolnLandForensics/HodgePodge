#!/usr/bin/env python3

import os
import re
import sys
import random
import string
from faker import Faker # pip install faker
from argparse import ArgumentParser
from datetime import datetime
from openpyxl import Workbook   # pip install openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading

# Metadata
Author = "LincolnLandForensics"
version = "1.0.3"
description2 = 'Create fake data for research and testing. Never use live data in a test database!'
# Global color variables
color_red = "\033[31m"
color_yellow = "\033[33m"
color_green = "\033[32m"
color_reset = "\033[0m"

# <<<<<<<<<<<<<<<<<<<<<<<<<<      GUI Class      >>>>>>>>>>>>>>>>>>>>>>>>>>

class FakeGui:
    def __init__(self, root):
        self.root = root
        script_name = os.path.basename(sys.argv[0])
        self.root.title(f"{script_name} {version}")
        self.root.geometry("700x550")
        
        # Set Vista theme
        self.style = ttk.Style()
        try:
            self.style.theme_use('vista')
        except:
            pass # Fallback to default if vista is not available

        self.setup_gui()
        self.set_defaults()

    def setup_gui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title/Description
        ttk.Label(main_frame, text=f"{description2}", font=("Helvetica", 10, "italic"), wraplength=650).pack(pady=5)

        # Number of Identities
        num_frame = ttk.LabelFrame(main_frame, text="Number of Identities", padding="5")
        num_frame.pack(fill=tk.X, pady=5)
        
        self.num_var = tk.StringVar()
        ttk.Entry(num_frame, textvariable=self.num_var).pack(fill=tk.X, padx=5)

        # Output File
        output_frame = ttk.LabelFrame(main_frame, text="Output Excel File", padding="5")
        output_frame.pack(fill=tk.X, pady=5)
        
        self.output_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.output_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(output_frame, text="Browse", command=self.browse_output).pack(side=tk.RIGHT)

        # Progress Bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=10)

        # Status/Log Window
        self.status_box = ScrolledText(main_frame, height=15, state='disabled')
        self.status_box.pack(fill=tk.BOTH, expand=True, pady=5)

        # Extract Button
        self.btn_run = ttk.Button(main_frame, text="Fake it till you make it", command=self.start_processing)
        self.btn_run.pack(pady=10)

    def set_defaults(self):
        self.num_var.set("100")
        self.output_var.set("fake_identities.xlsx")

    def browse_output(self):
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.output_var.set(file)

    def log(self, message):
        self.status_box.config(state='normal')
        self.status_box.insert(tk.END, message + "\n")
        self.status_box.see(tk.END)
        self.status_box.config(state='disabled')
        self.root.update_idletasks()

    def start_processing(self):
        try:
            total = int(self.num_var.get())
        except ValueError:
            messagebox.showerror("Error", "Number of identities must be an integer.")
            return

        total = check_number(total)
        output_file = self.output_var.get()

        self.btn_run.config(state='disabled')
        self.progress_var.set(0)
        self.status_box.config(state='normal')
        self.status_box.delete(1.0, tk.END)
        self.status_box.config(state='disabled')
        
        self.log(f"Starting generation of {total} records...")

        # Start processing in a new thread
        thread = threading.Thread(target=self.run_process, args=(output_file, total))
        thread.daemon = True
        thread.start()

    def run_process(self, output_file, total):
        try:
            generate_fake_data(output_file, total, self.log, self.update_progress)
            # self.log(f"Output File: {os.path.abspath(output_file)}")
            self.log("\nDone.")
        except Exception as e:
            self.log(f"Critical Error: {e}")
        finally:
            self.btn_run.config(state='normal')

    def update_progress(self, current, total):
        if total > 0:
            percent = (current / total) * 100
            self.progress_var.set(percent)
        else:
            self.progress_var.set(100)


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu/Logic     >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    parser = ArgumentParser(description="Generate fake identities")
    parser.add_argument("-o", "--output", help="Output Excel file", required=False)
    parser.add_argument("-f", "--fakes", help="Generate fake IDs", action="store_true", required=False)
    parser.add_argument("-n", "--number", help="Number of records to generate", type=int, default=69)

    args = parser.parse_args()

    # If no arguments are provided, launch GUI
    if len(sys.argv) == 1:
        root = tk.Tk()
        gui = FakeGui(root)
        root.mainloop()
        return

    # CLI Mode
    output_file = args.output if args.output else "FakeIdentities.xlsx"
    total = check_number(args.number)

    if args.fakes:
        generate_fake_data(output_file, total)
    else:
        usage()


def check_number(number):
    if number > 1000000:
        print(f"{color_red}Number can't be larger than a million, I'm going to give you 50 for now.{color_reset}")
        number = 50  # Assign the default value
    return number


def fake_user(first_name, middle_initial, last_name):
    """
    Create 20 variations of fake usernames based on first_name, middle_initial, last_name, and a random word.
    """
    fake = Faker()
    usernames = [
        f"{first_name.lower()}.{last_name.lower()}",
        f"{first_name.lower()}_{last_name.lower()}{fake.random_int(1, 99)}",
        f"{first_name.lower()}{middle_initial.lower()}{last_name.lower()}",
        f"{first_name.lower()}.{middle_initial.lower()}.{last_name.lower()}",
        f"{first_name.lower()}_{fake.word()}_{last_name.lower()}",
        f"{first_name.lower()}{fake.random_int(100, 999)}",
        f"{last_name.lower()}{fake.random_int(10, 99)}",
        f"{first_name.lower()}_{last_name.lower()}_{fake.word()}",
        f"{first_name.lower()[0]}{last_name.lower()}{fake.random_int(1, 99)}",
        f"{last_name.lower()}_{first_name.lower()}{fake.random_int(1, 99)}",
        f"{first_name.lower()}-{last_name.lower()}",
        f"{first_name.lower()}{last_name.lower()}{fake.random_int(1000, 9999)}",
        f"{first_name.lower()}x{last_name.lower()}",
        f"{first_name.lower()}_{fake.word()}",
        f"{first_name.lower()}{middle_initial.lower()}{last_name.lower()}{fake.random_int(1, 99)}",
        f"{last_name.lower()}_{first_name.lower()}",
        f"{fake.word()}_{first_name.lower()}_{last_name.lower()}",
        f"{first_name.lower()}_{fake.word()}{fake.random_int(1, 99)}",
        f"{first_name.lower()}{last_name.lower()}_{fake.word()}",
        f"{first_name.lower()}.{last_name.lower()}{fake.random_int(10, 99)}"
    ]
    return random.choice(usernames)


    
# Generate fake data
def generate_fake_data(output_file, total, status_callback=None, progress_callback=None):
    # fake = Faker()
    fake = Faker('en_US')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FakeIdentities"

    # Freeze cell B2
    sheet.freeze_panes = "B2"  # Freeze the top row and the first column

    headers = [
        "query", "ranking", "fullname", "url", "email", "user", "phone",
        "business", "fulladdress", "city", "state", "country", "zipcode", "AKA",
        "DOB", "SEX", "info", "mothers_maiden_name", "firstname", "middlename", "lastname",
        "associates", "case", "sosfilenumber", "owner", "president", "sosagent",
        "managers", "Time", "Latitude", "Longitude", "Coordinate",
        "original_file", "Source", "Source file information", "Plate", "VIS", "VIN",
        "VYR", "VMA", "LIC", "LIY", "DLN", "DLS", "job", "referer", "osurl",
        "titleurl", "pagestatus", "ip", "dnsdomain", "Tag", "Icon", "Type"
    ]

    state_abb = [
        "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
        "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
        "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
        "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
        "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"
    ]


    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.fill = fill

    # Generate fake data rows
    for i in range(1, total + 1):
        row = i + 1
        ip = fake.ipv4()
        url = fake.url()
        fullname = fake.name()
        firstname, lastname = fullname.split()[:2]
        lastname = lastname.upper()
        
        middle_initial = random.choice(string.ascii_uppercase)
        ssn = fake.ssn()
        mothers_maiden = fake.last_name()
        gender = random.choice(["M", "F"])
        # email = fake.email()
        # username = f"{firstname.lower()}.{lastname.lower()}"
        username = fake_user(firstname, middle_initial, lastname)
        # username = f"{fake.first_name()}_{fake.word()}{fake.random_int(1, 99)}"
        
        country = 'US'
        dob = generate_random_dob()
        city = fake.city()
        state = fake.state_abbr()
        while state not in state_abb:
            state = fake.state_abbr()       
        
        phone = us_phone_number(fake)
        # while len(phone) != 10:
            # print(f'phone is not 10 long {phone}')  # temp
            # phone = us_phone_number(fake)        
        
        phone = phone_state_check(phone, state)
        # if len(phone) == 10:
            # phone = phone_state_check(phone, state)

        zip_code = fake.postcode()
        fulladdress = f"{fake.street_address()}, {city}, {state} {zip_code}"
        business = fake.company()
        email = generate_random_email(firstname, lastname, business)
        plate = fake.license_plate()
        associates = fake.name()
        owner = fake.name()
        president = fake.name()
        sosagent = fake.name()
        managers = fake.name()
        VYR = get_random_year()
        VIN = generate_fake_vin()
        DLN = generate_fake_license_number()
        job = fake.job()
        uuid = fake.uuid4()
        # country = fake.country()
        # country = 'US'
        # language = fake.language_name()
        # profile = fake.simple_profile()
        # user_agent = fake.user_agent()
        # isbn = fake.isbn13()
        # currency = fake.currency_name()
 
        data = [
            "", "99 - fake data", fullname, url, email, username, phone, business, fulladdress, city, state, country, zip_code, "", dob, gender, ssn, mothers_maiden, firstname, middle_initial, lastname, associates, "", "", owner, president, sosagent, managers, "", "", "", "", "", "", "", plate, state, VIN, VYR, "", "", "", DLN, state, job, "", "", "", "", ip, "", "", "", ""
            ]

        for col, value in enumerate(data, start=1):
            sheet.cell(row=row, column=col, value=value)

        if status_callback and i % 10 == 0:
            status_callback(f"Generated {i}/{total} identities...")
        if progress_callback:
            progress_callback(i, total)

    workbook.save(output_file)

    msg = f'{total} fake identities created. Saved to {output_file}'
    if status_callback: status_callback(msg)
    print(f'\n{color_green}{msg}{color_reset}\n')
import random

def generate_fake_license_number():
    """
    Generates a fake driver's license number in the format T555-5555-5555.
    'T' is a random uppercase letter, and each '5' represents a digit.
    
    Returns:
        str: A fake driver's license number.
    """
    # Generate a random uppercase letter for the first character
    letter = random.choice(string.ascii_uppercase)
    
    # Generate the numeric sections
    numeric_section = '-'.join(
        ''.join(random.choices(string.digits, k=4)) for _ in range(3)
    )
    
    # Combine the parts
    license_number = f"{letter}{numeric_section}"
    return license_number
    
def generate_fake_vin():
    """
    Generates a fake VIN (Vehicle Identification Number).
    A VIN is a 17-character alphanumeric string, excluding 'I', 'O', and 'Q'.
    """
    # Define valid characters for a VIN
    vin_characters = string.ascii_uppercase + string.digits
    vin_characters = vin_characters.replace('I', '').replace('O', '').replace('Q', '')
    
    # Generate a 17-character string using valid VIN characters
    fake_vin = ''.join(random.choices(vin_characters, k=17))
    return fake_vin
    
    
def generate_random_email(firstname, lastname, business):
    '''
    Randomly generate a variety of email addresses based on firstname, lastname, 
    first_initial, and a random list of domains such as example.com, example.net, 
    email.com, email.net. Some emails will have random numbers at the end.

    :param firstname: First name of the person
    :param lastname: Last name of the person
    :return: Randomly generated email address
    '''

    fake = Faker()
    # List of possible email domains
    domains = [fake.free_email_domain() for _ in range(10)]

    # Replace spaces with commas, remove 'llc' and 'plc', and add '.net'
    business = business.replace(" ", "")  # Replace spaces
    business = business.replace(",", "")  # Replace commas
    business = business.replace("llc", "")  # Remove 'llc'
    business = business.replace("plc", "")  # Remove 'plc'
    business = business.strip()  # Strip any leading/trailing spaces

    # Append '.net' to the end of the business name
    business += ".net"

    # Append the modified business name to the domains list
    domains.append(business)

    # Randomly select a domain
    domain = random.choice(domains)

    # Randomly choose a format for the email
    email_formats = [
        f"{firstname.lower()}.{lastname.lower()}@{domain}",           # firstname.lastname
        f"{firstname[0].lower()}{lastname.lower()}@{domain}",         # first initial + lastname
        f"{firstname.lower()}{lastname.lower()}@{domain}",            # firstname + lastname
        f"{firstname[0].lower()}{lastname.lower()}{random.randint(1, 99)}@{domain}",  # first initial + lastname + number
        f"{firstname.lower()}{random.randint(1, 99)}@{domain}",       # firstname + number
        f"{lastname.lower()}{random.randint(1, 99)}@{domain}"         # lastname + number
    ]

    # Select a random email format
    email = random.choice(email_formats)

    return email


def generate_random_dob():
    year = random.randint(1950, 2000)
    month = random.randint(1, 12)
    day = random.randint(1, 28)  # To avoid invalid dates
    return f"{year}-{month:02d}-{day:02d}"


def get_random_year():
    """Returns a random 4-digit year between 2005 and 2024."""
    return random.randint(2005, 2024)

def phone_state_check(phone, state):
    """
    Replace the first 3 digits of the number with a random area code from the same state.

    Args:
        phone (str): The phone number as a string.
        state (str): The state abbreviation.

    Returns:
        str: Updated phone number with a random area code from the same state.
    """

    area_codes_by_state = {
        "AL": ["205", "251", "256", "334", "938"],
        "AK": ["907"],
        "AZ": ["480", "520", "602", "623", "928"],
        "AR": ["479", "501", "870"],
        "CA": ["209", "213", "279", "310", "323", "341", "408", "415", "424", "442", "510", "530", "559", "562", "619", "626", "628", "650", "657", "661", "707", "714", "747", "760", "805", "818", "820", "831", "858", "909", "916", "925", "949", "951"],
        "CO": ["303", "719", "720", "970"],
        "CT": ["203", "475", "860", "959"],
        "DE": ["302"],
        "FL": ["239", "305", "321", "352", "386", "407", "561", "689", "727", "754", "772", "786", "813", "850", "863", "904", "941", "954"],
        "GA": ["229", "404", "470", "478", "678", "706", "762", "770", "912"],
        "HI": ["808"],
        "ID": ["208", "986"],
        "IL": ["217", "224", "309", "312", "331", "447", "464", "618", "630", "708", "730", "773", "779", "815", "847", "861", "872"], 
        "IN": ["219", "260", "317", "463", "574", "765", "812", "930"],
        "IA": ["319", "515", "563", "641", "712"],
        "KS": ["316", "620", "785", "913"],
        "KY": ["270", "364", "502", "606", "859"],
        "LA": ["225", "318", "337", "504", "985"],
        "ME": ["207"],
        "MD": ["240", "301", "410", "443", "667"],
        "MA": ["339", "351", "413", "508", "617", "774", "781", "857", "978"],
        "MI": ["231", "248", "269", "313", "517", "586", "616", "734", "810", "906", "947", "989"],
        "MN": ["218", "320", "507", "612", "651", "763", "952"],
        "MS": ["228", "601", "662", "769"],
        "MO": ["314", "417", "557", "573", "636", "660", "816", "975"],
        "MT": ["406"],
        "NE": ["308", "402", "531"],
        "NV": ["702", "725", "775"],
        "NH": ["603"],
        "NJ": ["201", "551", "609", "640", "732", "848", "856", "862", "908", "973"],
        "NM": ["505", "575"],
        "NY": ["212", "315", "329", "332", "347", "363", "516", "518", "585", "607", "631", "646", "680", "716", "718", "838", "845", "914", "917", "929", "934"], 
        "NC": ["252", "336", "704", "743", "828", "910", "919", "980", "984"],
        "ND": ["701"],
        "OH": ["216", "220", "234", "283", "326", "330", "380", "419", "440", "513", "567", "614", "740", "937"],
        "OK": ["405", "539", "580", "918"],
        "OR": ["458", "503", "541", "971"],
        "PA": ["215", "223", "267", "272", "412", "445", "484", "570", "610", "717", "724", "814", "878"],
        "RI": ["401"],
        "SC": ["803", "843", "854", "864"],
        "SD": ["605"],
        "TN": ["423", "615", "629", "731", "865", "901", "931"],
        "TX": ["210", "214", "254", "281", "325", "346", "361", "409", "430", "432", "469", "512", "682", "713", "737", "806", "817", "830", "832", "903", "915", "936", "940", "956", "972", "979"],
        "UT": ["385", "435", "801"],
        "VT": ["802"],
        "VA": ["276", "434", "540", "571", "703", "757", "804"],
        "WA": ["206", "253", "360", "425", "509", "564"],
        "WV": ["304", "681"],
        "WI": ["262", "414", "534", "608", "715", "920"],
        "WY": ["307"]
    }

    if state in area_codes_by_state:
        random_area_code = random.choice(area_codes_by_state[state])
        phone = f"{random_area_code}{phone[3:]}"
        
        return phone
    
    return phone  # Return original phone number if state is invalid

def us_phone_number(fake):
    # phone_number = fake.phone_number()
    phone_number = ''.join([str(random.randint(0, 9)) for _ in range(10)])
    
    # Check if the phone number starts with 0, and replace it if necessary
    if phone_number.startswith('0'):
        phone_number = str(random.randint(2, 9)) + phone_number[1:]
    return phone_number

  
# Usage instructions
def usage():
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description2}{color_reset}')
    print(f'{file} Version: {version} by {Author}')
    print(f"{color_yellow}Usage:{color_reset}")
    print(f"\tpython {sys.argv[0]} -f -n <number_of_records> -o <output_file>")
    print(f"{color_yellow}Example:{color_reset}")
    print(f"\tpython {sys.argv[0]} -f")
    print(f"\tpython {sys.argv[0]} -f -n 100 -o fake_identities.xlsx")

# Run the script
if __name__ == "__main__":
    main()


# <<<<<<<<<<<<<<<<<<<<<<<<<<Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

1.0.0 - inserted only legit states and changed the area codes to matcht the state
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<     notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<     The End        >>>>>>>>>>>>>>>>>>>>>>>>>>