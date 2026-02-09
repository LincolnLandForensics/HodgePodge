#!/usr/bin/env python3
# coding: utf-8
"""
Read and write XLSX files using openpyxl without Pandas.
Author: LincolnLandForensics
Version: 1.1.2
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import sys
import argparse
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from scipy.stats import chisquare   # pip install scipy

# colors
color_red = color_yellow = color_green = color_blue = color_purple = color_reset = ''
from colorama import Fore, Back, Style
print(Back.BLACK)
color_red, color_yellow, color_green = Fore.RED, Fore.YELLOW, Fore.GREEN
color_blue, color_purple, color_reset = Fore.BLUE, Fore.MAGENTA, Style.RESET_ALL


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "This script models Benford’s Law by generating and comparing authentic versus manipulated financial data. It outputs frequency distributions to Excel for forensic analysis, helping identify statistical anomalies suggestive of fraud."
version = '1.1.2'

# Informational blurb for More button
Blurb = """

The Manipulated Data section simulates financial values that are artificially rounded 
to the nearest thousand, creating unnaturally uniform distributions. These values do not 
follow the logarithmic pattern predicted by Benford's Law, which typically governs organic 
datasets. By comparing the first-digit frequency of this manipulated data to Benford's 
expected distribution, the model helps demonstrate how fabricated or tampered numbers 
diverge from statistical norms.


The Authentic Data section contains numerical values that are naturally occurring 
and statistically organic. These figures are either real-world samples or generated 
using randomization techniques that closely mimic legitimate datasets. They tend to 
follow Benford's Law, which predicts the frequency of first digits in many 
naturally formed datasets.

The Chi-Square Test is a statistical method used to evaluate whether observed frequencies 
in a dataset differ significantly from expected frequencies under a certain hypothesis. 
It's commonly applied to categorical data—like survey responses, classifications, or 
groupings—to test for independence or goodness-of-fit.
The bigger the number, the less likely it is that the differences occurred by chance, 
and the more likely it is that something meaningful is happening in the data.

Chi-Square Test ➤ Stat: 1114.61


git clone https://github.com/SheetJS/enron_xls.git

"""

# Global variables for GUI
root = None
input_file_entry = None
output_file_entry = None
column_entry = None
progress_bar = None
status_text = None
calculate_button = None

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>



def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False, default='benfordsLaw_tester.xlsx')
    parser.add_argument('-O', '--output', help='', required=False, default='benfordsLaw.xlsx')
    parser.add_argument('-b', '--benfords', help='read xlsx', required=False, action='store_true')
    # parser.add_argument('-c', '--column', help='choose column', required=False, action='store_true')
    parser.add_argument('-c', '--column', help='', required=False, default='A')
    args = parser.parse_args()

    # If no arguments provided, launch GUI
    if len(sys.argv) == 1:
        launch_gui()
        return 0

    global input_file
    # input_file = args.input if args.input else "benfordsLaw_tester.xlsx"
    input_file = args.input

    global output_file
    output_file = args.output

    global column_pick
    # column_pick = args.column if args.column else "A"
    column_pick = args.column
    print(f'Reading {input_file} column {column_pick}')

    
    if args.benfords:

        file_exists = os.path.exists(input_file)
        if file_exists == True:
            msg_blurb = (f'Reading {input_file} column {column_pick}')
            msg_blurb_square(msg_blurb, color_green)    
            
            # data = read_xlsx(input_file)
            benfords(input_file, output_file)

        else:
            msg_blurb = (f'{input_file} does not exist')
            msg_blurb_square(msg_blurb, color_red)      
            exit()

    else:
        usage()
    
    return 0


def launch_gui():
    """Launch the Tkinter GUI interface."""
    global root, input_file_entry, output_file_entry, column_entry
    global progress_bar, status_text, calculate_button
    
    root = tk.Tk()
    root.title(f"Benford's Law Tester {version}")
    root.geometry("700x500")
    
    # Apply vista theme
    try:
        style = ttk.Style()
        style.theme_use('vista')
    except:
        pass  # Theme not available on this system
    
    # Description Label
    desc_label = tk.Label(root, text="Benford's Law: Calculate frequency distributions of the first digit for forensic analysis",
                         font=("Arial", 10), wraplength=650, justify="center")
    desc_label.pack(pady=10)
    
    # Input File Section
    input_frame = tk.Frame(root)
    input_frame.pack(pady=5, padx=20, fill=tk.X)
    
    input_label = tk.Label(input_frame, text="Input File:", width=15, anchor='w')
    input_label.pack(side=tk.LEFT)
    
    input_file_entry = tk.Entry(input_frame, width=50)
    input_file_entry.insert(0, "benfordsLaw_tester.xlsx")
    input_file_entry.pack(side=tk.LEFT, padx=5)
    
    def browse_input():
        filename = filedialog.askopenfilename(
            title="Select Input File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            input_file_entry.delete(0, tk.END)
            input_file_entry.insert(0, filename)
    
    browse_input_btn = tk.Button(input_frame, text="Browse", command=browse_input)
    browse_input_btn.pack(side=tk.LEFT)
    
    # Output File Section
    output_frame = tk.Frame(root)
    output_frame.pack(pady=5, padx=20, fill=tk.X)
    
    output_label = tk.Label(output_frame, text="Output File:", width=15, anchor='w')
    output_label.pack(side=tk.LEFT)
    
    output_file_entry = tk.Entry(output_frame, width=50)
    output_file_entry.insert(0, "benfordsLaw.xlsx")
    output_file_entry.pack(side=tk.LEFT, padx=5)
    
    def browse_output():
        filename = filedialog.asksaveasfilename(
            title="Select Output File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            output_file_entry.delete(0, tk.END)
            output_file_entry.insert(0, filename)
    
    browse_output_btn = tk.Button(output_frame, text="Browse", command=browse_output)
    browse_output_btn.pack(side=tk.LEFT)
    
    # Column Selection Section
    column_frame = tk.Frame(root)
    column_frame.pack(pady=5, padx=20, fill=tk.X)
    
    column_label = tk.Label(column_frame, text="Column:", width=15, anchor='w')
    column_label.pack(side=tk.LEFT)
    
    column_entry = tk.Entry(column_frame, width=10)
    column_entry.insert(0, "A")
    column_entry.pack(side=tk.LEFT, padx=5)
    
    # Progress Bar
    progress_frame = tk.Frame(root)
    progress_frame.pack(pady=10, padx=20, fill=tk.X)
    
    progress_label = tk.Label(progress_frame, text="Progress:")
    progress_label.pack(anchor='w')
    
    progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
    progress_bar.pack(fill=tk.X, pady=5)
    
    # Status Window
    status_frame = tk.Frame(root)
    status_frame.pack(pady=5, padx=20, fill=tk.BOTH, expand=True)
    
    status_label = tk.Label(status_frame, text="Status:")
    status_label.pack(anchor='w')
    
    status_text = scrolledtext.ScrolledText(status_frame, height=10, width=80)
    status_text.pack(fill=tk.BOTH, expand=True)
    
    # Button Frame for Calculate and More buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)
    
    # Calculate Button
    calculate_button = tk.Button(button_frame, text="Calculate", command=start_calculation, 
                                 bg="#4CAF50", fg="white", font=("Arial", 12, "bold"),
                                 padx=20, pady=10)
    calculate_button.pack(side=tk.LEFT, padx=5)
    
    # More Button
    more_button = tk.Button(button_frame, text="More", command=show_more_info, 
                           bg="#2196F3", fg="white", font=("Arial", 12, "bold"),
                           padx=20, pady=10)
    more_button.pack(side=tk.LEFT, padx=5)
    
    root.mainloop()


def show_more_info():
    """Display the informational blurb in a separate window."""
    info_window = tk.Toplevel(root)
    info_window.title("About Benford's Law Analysis")
    info_window.geometry("700x600")
    
    # Add a scrolled text widget to display the blurb
    info_text = scrolledtext.ScrolledText(info_window, wrap=tk.WORD, font=("Arial", 10))
    info_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    # Insert the blurb content
    info_text.insert(tk.END, Blurb)
    info_text.config(state=tk.DISABLED)  # Make it read-only
    
    # Add a Close button
    close_button = tk.Button(info_window, text="Close", command=info_window.destroy,
                            bg="#757575", fg="white", font=("Arial", 10, "bold"),
                            padx=15, pady=5)
    close_button.pack(pady=10)


def log_message(message):
    """Add a message to the status window."""
    if status_text:
        status_text.insert(tk.END, message + "\n")
        status_text.see(tk.END)
        root.update_idletasks()


def start_calculation():
    """Start the Benford's Law calculation in a separate thread."""
    # Disable button during processing
    calculate_button.config(state=tk.DISABLED)
    
    # Clear previous status
    status_text.delete(1.0, tk.END)
    
    # Start progress bar
    progress_bar.start()
    
    # Run processing in background thread
    thread = threading.Thread(target=calculation_thread)
    thread.daemon = True
    thread.start()


def calculation_thread():
    """Background thread for running the calculation."""
    try:
        input_file = input_file_entry.get()
        output_file = output_file_entry.get()
        column = column_entry.get().upper()
        
        # Validate input file exists
        if not os.path.exists(input_file):
            root.after(0, calculation_error, f"Input file '{input_file}' does not exist")
            return
        
        root.after(0, log_message, f"Reading {input_file} column {column}")
        
        # Call the benfords function
        benfords_gui(input_file, output_file, column)
        
        # Update GUI from thread (use root.after for thread safety)
        root.after(0, calculation_complete, output_file)
    except Exception as e:
        root.after(0, calculation_error, str(e))


def calculation_complete(output_file):
    """Called when calculation completes successfully."""
    progress_bar.stop()
    calculate_button.config(state=tk.NORMAL)
    log_message(f"\nOutput saved to: {output_file}")
    log_message("\nDone")
    messagebox.showinfo("Complete", f"Analysis complete!\n\nOutput saved to:\n{output_file}")


def calculation_error(error_msg):
    """Called when an error occurs during calculation."""
    progress_bar.stop()
    calculate_button.config(state=tk.NORMAL)
    log_message(f"\nError: {error_msg}")
    messagebox.showerror("Error", error_msg)


def benfords(input_file, output_file=None):

    wb = load_workbook(input_file, data_only=True)
    ws = wb.active
    column_data = [cell.value for cell in ws[column_pick] if isinstance(cell.value, (int, float))]

    if not column_data:
        msg_blurb_square("No numeric data found in the selected Excel column. Skipping analysis.", color_yellow)
        return

    # Extract first digits from Excel data
    excel_first_digit = get_first_digit(column_data)

    # Benford expected distribution
    benford_dist = pd.DataFrame({
        'Digit': range(1, 10),
        'Expected': np.log10(1 + 1 / np.arange(1, 10))
    })

    # Convert digit list to DataFrame
    df = pd.Series(excel_first_digit).value_counts().sort_index().reset_index()
    df.columns = ['Digit', 'Freq']
    df['Prop'] = df['Freq'] / df['Freq'].sum()
    df['Type'] = 'Excel Data'
    df = pd.merge(df, benford_dist, on='Digit', how='left')

    # Chi-Square Test
    observed_counts = df.set_index('Digit').reindex(range(1, 10), fill_value=0)['Freq']
    expected_counts = benford_dist['Expected'] * len(excel_first_digit)
    chi2_stat, p_value = chisquare(f_obs=observed_counts, f_exp=expected_counts)

    # MAD Test
    mad = np.mean(np.abs(df['Prop'] - df['Expected']))

    # Plotting
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(df['Digit'], df['Prop'], width=0.5, alpha=0.7, label='Excel Data', color='orange')
    ax.plot(benford_dist['Digit'], benford_dist['Expected'],
            color='blue', linewidth=2, label="Benford's Law")

    ax.set_title(f"Benford's Law Analysis: {input_file}", fontsize=16, weight='bold')
    ax.set_xlabel('First Digit', fontsize=14)
    ax.set_ylabel('Proportion', fontsize=14)
    ax.set_xticks(range(1, 10))
    ax.legend()
    plt.grid(True, which='major', linestyle='--', alpha=0.4)

    # Add Chi-Square and MAD results below the chart
    stats_text = f"Chi-Square: {chi2_stat:.2f} (p={p_value:.4f})   |   MAD: {mad:.4f}"
    fig.text(0.5, 0.01, stats_text,
             ha='center', va='bottom',
             fontsize=12,
             bbox=dict(boxstyle='round,pad=0.4', facecolor='#f0f0f0', edgecolor='gray'))

    plt.tight_layout(rect=[0, 0.03, 1, 1])  # Leave space at bottom for stats
    plt.show()


def benfords_gui(input_file, output_file, column):
    """Benford's Law analysis for GUI mode with Excel output."""
    try:
        wb = load_workbook(input_file, data_only=True)
        ws = wb.active
        column_data = [cell.value for cell in ws[column] if isinstance(cell.value, (int, float))]
        
        if not column_data:
            root.after(0, log_message, "No numeric data found in the selected Excel column.")
            raise ValueError("No numeric data found in the selected column")
        
        root.after(0, log_message, f"Found {len(column_data)} numeric values")
        root.after(0, log_message, "Analyzing first digit distribution...")
        
        # Extract first digits from Excel data
        excel_first_digit = get_first_digit(column_data)
        
        # Benford expected distribution
        benford_dist = pd.DataFrame({
            'Digit': range(1, 10),
            'Expected': np.log10(1 + 1 / np.arange(1, 10))
        })
        
        # Convert digit list to DataFrame
        df = pd.Series(excel_first_digit).value_counts().sort_index().reset_index()
        df.columns = ['Digit', 'Freq']
        df['Prop'] = df['Freq'] / df['Freq'].sum()
        df['Type'] = 'Excel Data'
        df = pd.merge(df, benford_dist, on='Digit', how='left')
        
        # Chi-Square Test
        observed_counts = df.set_index('Digit').reindex(range(1, 10), fill_value=0)['Freq']
        expected_counts = benford_dist['Expected'] * len(excel_first_digit)
        chi2_stat, p_value = chisquare(f_obs=observed_counts, f_exp=expected_counts)
        
        # MAD Test
        mad = np.mean(np.abs(df['Prop'] - df['Expected']))
        
        root.after(0, log_message, f"Chi-Square: {chi2_stat:.2f} (p={p_value:.4f})")
        root.after(0, log_message, f"MAD: {mad:.4f}")
        root.after(0, log_message, f"Saving results to {output_file}...")
        
        # Save to Excel
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "Benfords Analysis"
        
        # Write headers
        headers = ['Digit', 'Frequency', 'Proportion', 'Expected (Benford)', 'Difference']
        for col_idx, header in enumerate(headers, start=1):
            cell = output_ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for idx, row in df.iterrows():
            output_ws.cell(row=idx+2, column=1, value=int(row['Digit']))
            output_ws.cell(row=idx+2, column=2, value=int(row['Freq']))
            output_ws.cell(row=idx+2, column=3, value=float(row['Prop']))
            output_ws.cell(row=idx+2, column=4, value=float(row['Expected']))
            output_ws.cell(row=idx+2, column=5, value=float(row['Prop'] - row['Expected']))
        
        # Add statistics at the bottom
        stats_row = len(df) + 4
        output_ws.cell(row=stats_row, column=1, value="Statistics:")
        output_ws.cell(row=stats_row, column=1).font = output_ws.cell(row=stats_row, column=1).font.copy(bold=True)
        
        output_ws.cell(row=stats_row+1, column=1, value="Chi-Square Statistic:")
        output_ws.cell(row=stats_row+1, column=2, value=f"{chi2_stat:.2f}")
        
        output_ws.cell(row=stats_row+2, column=1, value="P-Value:")
        output_ws.cell(row=stats_row+2, column=2, value=f"{p_value:.4f}")
        
        output_ws.cell(row=stats_row+3, column=1, value="MAD (Mean Absolute Deviation):")
        output_ws.cell(row=stats_row+3, column=2, value=f"{mad:.4f}")
        
        # Adjust column widths
        for col in output_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            output_ws.column_dimensions[column].width = adjusted_width
        
        output_wb.save(output_file)
        
        # Show plot
        root.after(0, log_message, "Displaying chart...")
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(df['Digit'], df['Prop'], width=0.5, alpha=0.7, label='Excel Data', color='orange')
        ax.plot(benford_dist['Digit'], benford_dist['Expected'],
                color='blue', linewidth=2, label="Benford's Law")
        
        ax.set_title(f"Benford's Law Analysis: {os.path.basename(input_file)}", fontsize=16, weight='bold')
        ax.set_xlabel('First Digit', fontsize=14)
        ax.set_ylabel('Proportion', fontsize=14)
        ax.set_xticks(range(1, 10))
        ax.legend()
        plt.grid(True, which='major', linestyle='--', alpha=0.4)
        
        # Add Chi-Square and MAD results below the chart
        stats_text = f"Chi-Square: {chi2_stat:.2f} (p={p_value:.4f})   |   MAD: {mad:.4f}"
        fig.text(0.5, 0.01, stats_text,
                 ha='center', va='bottom',
                 fontsize=12,
                 bbox=dict(boxstyle='round,pad=0.4', facecolor='#f0f0f0', edgecolor='gray'))
        
        plt.tight_layout(rect=[0, 0.03, 1, 1])
        plt.show()
        
    except Exception as e:
        raise Exception(f"Error during analysis: {str(e)}")


# 🧮 Convert digit list to DataFrame
def digit_df(digits, label):
    df = pd.Series(digits).value_counts().sort_index().reset_index()
    df.columns = ['Digit', 'Freq']
    df['Prop'] = df['Freq'] / df['Freq'].sum()
    df['Type'] = label
    return pd.merge(df, benford_dist, on='Digit', how='left')




# 🔍 First-digit extraction
def get_first_digit(arr):
    return [int(str(int(np.floor(x)))[0]) for x in arr if x > 0]



def msg_blurb_square(msg, color):
    border = f"+{'-' * (len(msg) + 2)}+"
    print(f"{color}{border}\n| {msg} |\n{border}{color_reset}")

def read_xlsx(file_path):
    """Read data from an XLSX file and return a list of dictionaries."""
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        row_data["zero"] = "bobs your uncle" if row_data.get("zero") == "test" else row_data.get("zero")
        row_data["one"] = "avocadoes rule" if row_data.get("one") == "TeslaSucks" else row_data.get("one")
        data.append(row_data)
    
    wb.close()
    return data

    
def usage():
    print(f"Usage: {sys.argv[0]} -b [-I input.xlsx]")
    print("Example:")
    print(f"    {sys.argv[0]} -b")
    print(f"    {sys.argv[0]} -b -I benfordsLaw_tester.xlsx")
    print(f"    {sys.argv[0]} -b -I benfordsLaw_tester.xlsx -c d")

if __name__ == '__main__':
    main()


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
0.1.0 - working copy

"""


# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
if there are no instances of a number, like 3, it blows an error
    raise ValueError(f'shapes {shape1} and {shape2} could not be '
ValueError: shapes (8,) and (9,) could not be broadcast together


"""


# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

# The Blurb variable is now defined in the Pre-Sets section at the top of the file

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>