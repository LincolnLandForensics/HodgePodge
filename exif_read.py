import os
import sys
import exifread
import openpyxl
import datetime

def convert_to_decimal(coord, ref):
    """Converts GPS coordinates to decimal format."""
    if not coord:
        return None
    degrees = float(coord[0].num) / float(coord[0].den)
    minutes = float(coord[1].num) / float(coord[1].den) / 60.0
    seconds = float(coord[2].num) / float(coord[2].den) / 3600.0
    decimal = degrees + minutes + seconds
    if ref in ['S', 'W']:
        decimal = -decimal
    return decimal

def read_exif_data(photo_path):
    """Reads selected EXIF data from an image file."""
    with open(photo_path, 'rb') as f:
        tags = exifread.process_file(f, stop_tag='UNDEFINED')

    # Extract GPS coordinates in decimal format if available
    gps_latitude = tags.get('GPS GPSLatitude')
    gps_latitude_ref = tags.get('GPS GPSLatitudeRef')
    gps_longitude = tags.get('GPS GPSLongitude')
    gps_longitude_ref = tags.get('GPS GPSLongitudeRef')
    
    # Convert GPS data to decimal if available
    latitude = convert_to_decimal(gps_latitude.values, gps_latitude_ref.values[0]) if gps_latitude and gps_latitude_ref else ""
    longitude = convert_to_decimal(gps_longitude.values, gps_longitude_ref.values[0]) if gps_longitude and gps_longitude_ref else ""

    exif_data = {
        'Name': os.path.basename(photo_path),
        'DateCreated': tags.get('Image DateTime'),
        'DateTimeOriginal': tags.get('EXIF DateTimeOriginal'),
        'FileCreateDate': datetime.datetime.fromtimestamp(os.path.getctime(photo_path)).strftime('%Y-%m-%d %H:%M:%S'),
        'FileModifyDate': datetime.datetime.fromtimestamp(os.path.getmtime(photo_path)).strftime('%Y-%m-%d %H:%M:%S'),
        'Timezone': tags.get('EXIF OffsetTime'),
        'DeviceManufacturer': tags.get('Image Make'),
        'FileSize': os.path.getsize(photo_path),
        'FileType': os.path.splitext(photo_path)[1].replace('.', '').upper(),
        'FileTypeExtension': os.path.splitext(photo_path)[1].replace('.', '').lower(),
        'LensMake': tags.get('EXIF LensMake'),
        'Software': tags.get('Image Software'),
        'HostComputer': tags.get('Image HostComputer'),
        'LensInfo': tags.get('EXIF LensInfo'),
        'LensModel': tags.get('EXIF LensModel'),
        'Model': tags.get('Image Model'),
        'Altitude': tags.get('GPS GPSAltitude'),
        'Latitude': latitude,
        'Longitude': longitude,
        'Coordinate': f"{latitude},{longitude}" if latitude and longitude else "",
        'NumberOfImages': tags.get('Exif NumberOfImages'),
        'ExifToolVersion': tags.get('Exif ExifToolVersion'),
        'Icon': 'Images',
        'Type': 'Images',
        'Time': tags.get('EXIF DateTimeOriginal')
        # 'Description': f"{tags.get('LensMake', '')}"

        # 'Description': f"{tags.get('LensMake', '')}, {tags.get('LensModel', '')}, {tags.get('Software', '')}"
    }
    
    # Convert EXIF values to strings, replacing None with ""
    for key, value in exif_data.items():
        exif_data[key] = str(value) if value else ""
    
    return exif_data

def main():
    # Folder containing photos
    photos_folder = 'photos'  # Change this to your folder path
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
    output_file = f'exif_data_{timestamp}.xlsx'

    # Create a new workbook and add basic formatting
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "File Metadata"
    
    # Define headers based on EXIF data keys
    headers = [
        'Name', 'DateCreated', 'DateTimeOriginal', 'FileCreateDate', 'FileModifyDate', 'Timezone',
        'DeviceManufacturer', 'ExifToolVersion', 'FileSize', 'FileType', 'FileTypeExtension', 'LensMake',
        'Software', 'HostComputer', 'LensInfo', 'LensModel', 'Model', 'NumberOfImages', 
        'Altitude', 'Latitude', 'Longitude', 'Coordinate', 'Time', 'Type', 'Icon', 'Description'
    ]
    sheet.append(headers)

    # Format headers as bold and set column widths
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header).font = openpyxl.styles.Font(bold=True)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Freeze the top row at cell B2
    sheet.freeze_panes = 'B2'

    # Check if the path exists
    if not os.path.exists(photos_folder):
        print(f"The '{photos_folder}' folder does not exist.")
        sys.exit(1)  # Exit the program with an error code


    # Iterate over files in the photos folder
    for filename in os.listdir(photos_folder):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.tiff')):
            photo_path = os.path.join(photos_folder, filename)
            try:
                exif_data = read_exif_data(photo_path)
                # Append the EXIF data to the sheet
                sheet.append([exif_data.get(header, "") for header in headers])
            except Exception as e:
                print(f"Error processing {filename}: {e}")

    # Save the workbook
    workbook.save(output_file)
    print(f"File metadata exported to {output_file}")

if __name__ == "__main__":
    main()
