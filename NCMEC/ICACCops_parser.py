#!/usr/bin/python
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import re
import sys
import argparse  # for menu system
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk
import threading

# Colors & GUI Globals
color_red = ""
color_green = ""
color_reset = ""
gui_active = False
text_status = None


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "Read a folder full of ICACCops .txt log files and export them to .xlsx"
version = '1.0.3'


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global Row
    global spreadsheet
    global filename
    global sheet_format

    Row = 1

    DEFAULT_INPUT = r"C:\Forensics\scripts\python\ICACCops"
    DEFAULT_OUTPUT = "ICACCopsLookup_.xlsx"

    if len(sys.argv) == 1:
        launch_gui()
        return

    parser = argparse.ArgumentParser(description=description)

    parser.add_argument('-I', '--input', help='Input folder', required=False)
    parser.add_argument('-O', '--output', help='Output XLSX file', required=False)
    parser.add_argument('-r', '--read', help='Read and parse files', action='store_true')

    args = parser.parse_args()

    sheet_format = "Sheet"

    # Apply defaults AFTER declaring globals
    input_folder = args.input if args.input else DEFAULT_INPUT
    spreadsheet = args.output if args.output else DEFAULT_OUTPUT

    if args.read:
        parse_folder(input_folder)

    try:
        workbook.close()
    except NameError:
        pass



# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>




def log_to_gui(msg):
    print(msg)
    if gui_active and text_status:
        text_status.insert(tk.END, str(msg) + "\n")
        text_status.see(tk.END)

def msg_blurb_square(msg_blurb, color=""):
    print(f"{color}{msg_blurb}{color_reset}")
    print(f'')
    if gui_active and text_status:
        text_status.insert(tk.END, str(msg_blurb) + "\n\n")
        text_status.see(tk.END)

def launch_gui():
    global gui_active, text_status, progress, btn_extract
    global entry_input, entry_output, root

    DEFAULT_INPUT = r"C:\Forensics\scripts\python\ICACCops"
    DEFAULT_OUTPUT = "ICACCopsLookup_.xlsx"

    gui_active = True
    root = tk.Tk()
    
    script_name = os.path.basename(sys.argv[0])
    root.title(f"{script_name} {version}")
    root.geometry("650x550")
    
    style = ttk.Style()
    if 'vista' in style.theme_names():
        style.theme_use('vista')

    lbl_desc = tk.Label(root, text=description, font=("Arial", 12, "bold"))
    lbl_desc.pack(pady=10)

    # Input Frame
    frame_input = tk.Frame(root)
    frame_input.pack(fill=tk.X, padx=20, pady=5)
    tk.Label(frame_input, text="Input Folder:", width=15, anchor="e").pack(side=tk.LEFT)
    entry_input = tk.Entry(frame_input)
    entry_input.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

    # INSERT DEFAULT INPUT
    entry_input.insert(0, DEFAULT_INPUT)

    def browse_input():
        folder = filedialog.askdirectory()
        if folder:
            entry_input.delete(0, tk.END)
            entry_input.insert(0, folder)
    tk.Button(frame_input, text="Browse", command=browse_input).pack(side=tk.LEFT)

    # Output Frame
    frame_output = tk.Frame(root)
    frame_output.pack(fill=tk.X, padx=20, pady=5)
    tk.Label(frame_output, text="Output File (.xlsx):", width=15, anchor="e").pack(side=tk.LEFT)
    entry_output = tk.Entry(frame_output)
    entry_output.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

    # INSERT DEFAULT OUTPUT
    entry_output.insert(0, DEFAULT_OUTPUT)

    def browse_output():
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            entry_output.delete(0, tk.END)
            entry_output.insert(0, file)
    tk.Button(frame_output, text="Browse", command=browse_output).pack(side=tk.LEFT)

    # Extract Data Button
    btn_extract = tk.Button(root, text="Extract Data", command=start_processing_thread,
                            font=("Arial", 10, "bold"), bg="#4CAF50", fg="white")
    btn_extract.pack(pady=15)

    # Progress Bar
    progress = ttk.Progressbar(root, mode='indeterminate')
    progress.pack(fill=tk.X, padx=20, pady=5)

    # Status Window
    tk.Label(root, text="Status Output:").pack(anchor="w", padx=20)
    text_status = scrolledtext.ScrolledText(root, height=15)
    text_status.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

    root.mainloop()


def start_processing_thread():
    input_folder = entry_input.get().strip()
    output_file = entry_output.get().strip()
    
    if not input_folder or not output_file:
        log_to_gui("Error: Please select both an input folder and an output file.")
        return
        
    btn_extract.config(state=tk.DISABLED)
    progress.start()
    text_status.delete(1.0, tk.END)
    
    log_to_gui(f"Input folder used: {input_folder}")
    
    thread = threading.Thread(target=process_data, args=(input_folder, output_file))
    thread.daemon = True
    thread.start()

def process_data(input_folder, output_file):
    try:
        global spreadsheet
        spreadsheet = output_file
        
        parse_folder(input_folder)
        
        root.after(0, processing_done, output_file)
    except Exception as e:
        root.after(0, processing_error, str(e))

def processing_done(output_file):
    progress.stop()
    btn_extract.config(state=tk.NORMAL)
    log_to_gui(f"Output file name: {output_file}")
    log_to_gui("Extraction Complete.")

def processing_error(err):
    progress.stop()
    btn_extract.config(state=tk.NORMAL)
    log_to_gui(f"Error during processing: {err}")

def parse_folder(input_folder):
    """
    Iterate through every file in input_folder.
    For each file, call it Source and route to the correct parser.
    Accumulate parsed data and write a final XLSX.
    """

    data = []

    for filename in os.listdir(input_folder):
        Source = os.path.join(input_folder, filename)

        # Skip directories
        if os.path.isdir(Source):
            continue

        ext = filename.lower().split('.')[-1]


        if ext == "txt":
            data = read_txt(data, Source)
        else:
            # Optional: log or ignore unsupported files
            print(f"Skipping unsupported file: {filename}")

    write_xlsx(data)

def read_txt(data, Source):
    '''
    read the file and parse out the data
    '''
    SourceTemp = os.path.basename(Source)
    
    txt_file = open(Source)
    TorrentInfoHash, IP, Port = '', '', ''
    print(f'testing {Source}')
    
    pattern1 = re.compile(
    r'^(?P<Time>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) - File (?P<Index>\d+): '
    r'SHA1=(?P<SHA1>[0-9a-fA-F]{40}) '
    r'\((?P<SHA1base32>[A-Z2-7]+)\), '
    r'MD5=(?P<MD5>[0-9a-fA-F]{32})')

    pattern2 = re.compile(
        r'^(?P<Time>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) - File index (?P<Index>\d+) '
        r'is named "(?P<Filename>.+?)" in the torrent')    

    pattern3 = re.compile(
        r'^File (?P<Index>\d+) '
        r'\((?P<Bytes>\d+) bytes\) '
        r'has name: (?P<Filename>.+)$')

    pattern4 = re.compile(
        r'^Piece (?P<Index>\d+) SHA1 hash: (?P<SHA1>[0-9a-fA-F]{40})$'    )

    pattern5 = re.compile(
        r'^(?P<Time>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) - '
        r'Piece (?P<Index>\d+) has expected SHA1 hash: '
        r'(?P<SHA1>[0-9a-fA-F]{40})$')

    pattern6 = re.compile(
        r'^File (?P<Index>\d+) '
        r'\((?P<Bytes>\d+) bytes\): '
        r'defined by pieces ')

    pattern7 = re.compile(
        r'^(?P<Time>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) - '
        r'File (?P<Index>\d+): no pieces written$')



    DatePattern1 = re.compile(
        r'^(?P<Time>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) - '   )

    MatchList1 = [
        "Torrential Downpour version", "Torrent has ", "Piece size: "
        ,"Started download thread at ", "Current local time is ", "Attempting to negotiate message stream encryption "
        ,"Torrent defines ", "Encryption successfully negotiated:", "Peer id we sent to remote client:"
        ,"Total file bytes: ", "Sent encrypted handshake to client", "Remote client "
        ,"Sent extended handshake message", "Sent have-none message", "Received an extended handshake message"
        ,"xtended handshake", "bitfield message"
        
    ]

    
    for line in txt_file:
        Text = line
        FOI, Index, Filename, SHA1base16, SHA1base32, MD5 = '', '', '', '', '', ''
        Time, Bytes, Temp = '', '', ''
        row_data = {}
        print(f'{line}')
        if 'Torrent info hash (hexadecimal): ' in line:
            try:
                TorrentInfoHash = line.split('Torrent info hash (hexadecimal): ')[1]
            except:pass
        elif 'Torrent info hash: ' in line:
            try:
                TorrentInfoHash = line.split('Torrent info hash: ')[1]
            except:pass
        elif 'Remote client located at IP address ' in line:
            try:
                IP = line.split('Remote client located at IP address ')[1].replace(', port ',':')
                if ':' in IP:
                    PortTemp = IP.split(':')
                    IP = PortTemp[0]
                    Port = PortTemp[1]
            except:pass
        elif pattern1.search(line):
            m = pattern1.search(line)

            Time = m.group("Time")
            Index = m.group("Index")
            SHA1 = m.group("SHA1")
            SHA1base32 = m.group("SHA1base32")
            MD5 = m.group("MD5")

            row_data["Time"] = Time
            row_data["Index"] = Index
            row_data["SHA1base16"] = SHA1
            row_data["SHA1base32"] = SHA1base32
            row_data["MD5"] = MD5
            row_data["Source"] = SourceTemp
            row_data["Text"] = Text
            row_data["Temp"] = 'pattern1'
            data.append(row_data.copy())

        elif pattern2.search(line):
            m = pattern2.search(line)

            Time = m.group("Time")
            Index = m.group("Index")
            Filename = m.group("Filename")
            Filename = os.path.basename(Filename)
            row_data["Filename"] = Filename
            row_data["Time"] = Time
            row_data["Index"] = Index
            row_data["IP"] = IP  
            row_data["Port"] = Port  
            row_data["Source"] = SourceTemp
            row_data["TorrentInfoHash"] = TorrentInfoHash            
            row_data["Text"] = Text
            row_data["Temp"] = 'pattern2'
            data.append(row_data.copy())

        elif pattern3.search(line):
            m = pattern3.search(line)

            Index = m.group("Index")
            Bytes = m.group("Bytes")
            Filename = m.group("Filename")
            Filename = os.path.basename(Filename)
            row_data["Filename"] = Filename
            row_data["Index"] = Index
            row_data["Bytes"] = Bytes
            row_data["IP"] = IP  
            row_data["Port"] = Port             
            row_data["Source"] = SourceTemp
            row_data["TorrentInfoHash"] = TorrentInfoHash
            row_data["Text"] = Text
            row_data["Temp"] = 'pattern3'
            
            data.append(row_data.copy())

        elif pattern4.search(line):
            m = pattern4.search(line)

            Index = m.group("Index")
            SHA1 = m.group("SHA1")

            row_data["Index"] = Index
            row_data["SHA1base16"] = SHA1
            row_data["IP"] = IP   
            row_data["Port"] = Port             
            row_data["Source"] = SourceTemp
            row_data["TorrentInfoHash"] = TorrentInfoHash
            row_data["Text"] = Text
            row_data["Temp"] = 'pattern4'
            data.append(row_data.copy())

        elif pattern5.search(line):
            m = pattern5.search(line)

            Time = m.group("Time")
            Index = m.group("Index")
            SHA1 = m.group("SHA1")

            row_data["Time"] = Time
            row_data["Index"] = Index
            row_data["SHA1base16"] = SHA1
            row_data["IP"] = IP   
            row_data["Port"] = Port             
            row_data["Source"] = SourceTemp
            row_data["TorrentInfoHash"] = TorrentInfoHash
            row_data["Text"] = Text
            row_data["Temp"] = 'pattern5'
            data.append(row_data.copy())

        elif pattern6.search(line):
            m = pattern6.search(line)

            Index = m.group("Index")
            Bytes = m.group("Bytes")

            row_data["Index"] = Index
            row_data["Bytes"] = Bytes
            row_data["IP"] = IP 
            row_data["Port"] = Port             
            row_data["Source"] = SourceTemp
            row_data["TorrentInfoHash"] = TorrentInfoHash
            row_data["Text"] = Text
            row_data["Temp"] = 'pattern6'
            data.append(row_data.copy())


        elif DatePattern1.search(line):
            m = DatePattern1.search(line)
            Time = m.group("Time")

            if pattern7.search(line):
                m = pattern7.search(line)

                Time = m.group("Time")
                Index = m.group("Index")

                row_data["Time"] = Time
                row_data["Index"] = Index
                row_data["IP"] = IP 
                row_data["Port"] = Port                 
                row_data["Source"] = SourceTemp
                row_data["TorrentInfoHash"] = TorrentInfoHash
                row_data["Text"] = Text
                row_data["Temp"] = 'pattern7'
                data.append(row_data.copy())


            elif any(x in line for x in MatchList1):
                # your logic here

                row_data["Time"] = Time
                row_data["IP"] = IP  
                row_data["Port"] = Port                 
                row_data["Source"] = SourceTemp
                row_data["TorrentInfoHash"] = TorrentInfoHash
                row_data["Text"] = Text
                row_data["Temp"] = 'MatchList1'
                data.append(row_data.copy())
  
    return data

def write_xlsx(data):
    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'ICACCops_TOI'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'C2'  # Freeze cells
    worksheet.selection = 'C2'

    headers = ["FOI", "Index", "Filename", "SHA1base16", "SHA1base32", "MD5"
    , "Time", "IP", "Port", "TorrentInfoHash", "Source", "Bytes", "Text"
    # , "Temp"
    ]


    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.value = header

    # Excel column width
    worksheet.column_dimensions['A'].width = 4 # 
    worksheet.column_dimensions['B'].width = 6 # 
    worksheet.column_dimensions['C'].width = 25 # 
    worksheet.column_dimensions['D'].width = 40 # 
    worksheet.column_dimensions['E'].width = 40 # 
    worksheet.column_dimensions['F'].width = 20 #   
    worksheet.column_dimensions['G'].width = 20 #   
    worksheet.column_dimensions['H'].width = 16 #
    worksheet.column_dimensions['I'].width = 6 #    
    worksheet.column_dimensions['J'].width = 40 #   
    worksheet.column_dimensions['K'].width = 15 #   
    worksheet.column_dimensions['L'].width = 12 #   
    worksheet.column_dimensions['M'].width = 65 #   
    worksheet.column_dimensions['N'].width = 11 #   


    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                if isinstance(cell_data, list):
                    cell_data = str(cell_data)  # Convert lists to strings
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    workbook.save(spreadsheet)



def usage():
    file = sys.argv[0].split('\\')[-1]
    print("\nDescription: " + description)
    print(file + " Version: %s by %s" % (version, author))
    print("\nExample:")
    print("\t" + file + " -r -I C:\Forensics\scripts\python\ICACCops -O out_.xlsx\t\t")
    print("\t" + file + " -r -I C:\Forensics\scripts\python\ICACCops\t\t")
    # print("\t" + File +" -s -I nodes.txt -O out_second.xls")


if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
1.0.0 - a functional copy

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
TBD


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""



"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
