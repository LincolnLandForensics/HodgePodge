#!/usr/bin/python
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>>
import os
import sys
import chardet  # For dynamic encoding detection
import argparse

import openpyxl
from openpyxl import load_workbook, Workbook    # pip install openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# <<<<<<<<<<<<<<<<<<<<<<<<<<     Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = "Convert a folder of .vcf files to an intel sheet, or visa versa"
version = '1.1.1'

# Colorize section
global color_red
global color_green
global color_reset
color_green = ''
color_red = ''
color_reset = ''


if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
  
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

    global headers_intel
    headers_intel = [
        "query", "ranking", "fullname", "url", "email", "user", "phone",
        "business", "fulladdress", "city", "state", "country", "note", "AKA",
        "DOB", "SEX", "info", "misc", "firstname", "middlename", "lastname",
        "associates", "case", "sosfilenumber", "owner", "president", "sosagent",
        "managers", "Time", "Latitude", "Longitude", "Coordinate",
        "original_file", "Source", "Source file information", "Plate", "VIS", "VIN",
        "VYR", "VMA", "LIC", "LIY", "DLN", "DLS", "content", "referer", "osurl",
        "titleurl", "pagestatus", "ip", "dnsdomain", "Tag", "Icon", "Type"
    ]


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    """
    Main function to parse arguments and initiate file conversion.
    """
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='Input folder path', required=False)
    parser.add_argument('-O', '--output', help='Output file path', required=False)
    parser.add_argument('-c', '--contacts', help='Read contacts from .vcf files and create an Excel sheet', required=False, action='store_true')
    parser.add_argument('-x', '--xlsx', help='Read contacts from .xlsx files and create .vcf files', required=False, action='store_true')

    args = parser.parse_args()

    global output_xlsx


    if args.contacts:
        input_folder = args.input or "LogsVCF"
        output_xlsx = args.output or "contacts_Apple.xlsx"

        if not os.path.exists(input_folder):
            print(f"Error: Input folder '{input_folder}' does not exist.")
            sys.exit(1)

        data = parse_vcf_files(input_folder)
        write_intel(data)
    elif args.xlsx:
        input_xlsx_path = args.input or "contacts_Apple.xlsx"
        output_dir = args.output or "VCF_Files"
        excel_to_vcf(input_xlsx_path, output_dir)


        message = (f'reading contacts from {input_xlsx_path}')
        message_square(message, color_green)  


        message = (f'vCards have been generated in the directory: {output_dir}')
        message_square(message, color_green)  

    else:
        parser.print_help()
        Usage()

    return 0

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

def convert_phone_number(phone_number):
    """
    Remove all non-digit characters and format the phone number.
    """
    digits = ''.join(filter(str.isdigit, phone_number))
    if digits.startswith('1'):
        digits = digits[1:]
    phone = f"{digits[:3]}{digits[3:6]}{digits[6:]}"
    return phone

def detect_encoding(content):
    """
    Detect the encoding of the file content using chardet.
    """
    result = chardet.detect(content)
    return result['encoding'] or 'utf-8'

def excel_to_vcf(excel_file, output_dir):
    """
    Convert an Excel file to .vcf files.

    Args:
        excel_file (str): Path to the Excel file.
        output_dir (str): Directory to save .vcf files.
    """

    message = (f'reading {excel_file}')
    message_square(message, color_green)    

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Load Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Extract headers from the first row
    headers = [cell.value for cell in sheet[1]]

    # Iterate over rows (skip the header row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        entry = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        create_vcf(entry, output_dir)

    message = (f'outputing .vcf files to {output_folder}')
    message_square(message, color_green)  

def create_vcf(entry, output_folder):
    """
    Create a .vcf file for a single entry.

    Args:
        entry (dict): A dictionary containing contact details.
        output_folder (str): The directory to save the .vcf file.
    """
    vcf_content = ["BEGIN:VCARD", "VERSION:3.0"]

    # Add fields to vCard
    if entry.get("fullname"):
        vcf_content.append(f"FN:{entry['fullname']}")
    if entry.get("firstname") or entry.get("lastname"):
        vcf_content.append(f"N:{entry.get('lastname', '')};{entry.get('firstname', '')};{entry.get('middlename', '')};;")
    if entry.get("email"):
        vcf_content.append(f"EMAIL:{entry['email']}")
    if entry.get("phone"):
        vcf_content.append(f"TEL:{entry['phone']}")
    if entry.get("business"):
        vcf_content.append(f"ORG:{entry['business']}")
    if entry.get("fulladdress"):
        vcf_content.append(f"ADR:;;{entry['fulladdress']}" + (f";{entry['city']};{entry['state']};{entry['country']}" if entry.get("city") or entry.get("state") or entry.get("country") else ""))
    if entry.get("note"):
        vcf_content.append(f"NOTE:{entry['note']}")
    
    vcf_content.append("END:VCARD")

    # Save the vCard
    filename = entry.get("fullname", "Unknown").replace(" ", "_") + ".vcf"
    filepath = os.path.join(output_folder, filename)
    with open(filepath, "w", encoding="utf-8") as vcf_file:
        vcf_file.write("\n".join(vcf_content))


def excel_to_vcf(input_xlsx, output_folder):
    """
    Convert an Excel file to .vcf files.

    Args:
        input_xlsx (str): Path to the Excel file.
        output_folder (str): Directory to save .vcf files.
    """
    # Ensure output directory exists
    os.makedirs(output_folder, exist_ok=True)

    # Load Excel file
    workbook = openpyxl.load_workbook(input_xlsx)
    sheet = workbook.active

    # Extract headers from the first row
    headers = [cell.value for cell in sheet[1]]

    # Iterate over rows (skip the header row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        entry = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        create_vcf(entry, output_folder)



def extract_field(line):
    """
    Extract field name and value from a line.
    """
    if ':' in line:
        parts = line.split(':', 1)
        return parts[0].strip(), parts[1].strip()
    return '', ''

def extract_phone_type(line):
    """
    Extract phone type from a line.
    """
    if 'type=' in line:
        parts = line.split(';')
        return [p.split('=')[1] for p in parts[1:]]
    return ''    

def extract_email_type(line):
    """
    Extract email type from a line.
    """
    if 'type=' in line:
        parts = line.split(';')
        return [p.split('=')[1] for p in parts[1:]]
    # return None
    return ''   

def parse_vcf_content(lines):
    """
    Parse lines from a .vcf file and extract contact information.

    :param lines: List of decoded lines from the .vcf file.
    :return: Dictionary with extracted contact information.
    """
    contact = {}
    contact['ranking'] = '5 - Contacts from VCF'
    (user, info) = ('', '')
    contact['Icon'] = ('Intel')    
    contact['Type'] = ('Intel')
    for line in lines:
        line = line.strip()
        field, value = extract_field(line)
        
        if field == 'FN':
            contact['fullname'] = value
        elif field == 'N':
            name_parts = value.split(';')
            contact['lastname'] = name_parts[0].strip() if len(name_parts) > 0 else ''
            contact['firstname'] = name_parts[1].strip() if len(name_parts) > 1 else ''
        elif 'TEL' in line:
            contact['query'] = (value)
            contact['phone'] = convert_phone_number(value)
            phone_types = extract_phone_type(line)
            contact['misc'] = ', '.join(phone_types) if phone_types else ''  # Join list to a single string

        elif 'ADR;' in line:
            contact['fulladdress'] = value.replace(';', ' ').strip()
        elif 'EMAIL' in line:
            contact['email'] = value
            email_types = extract_email_type(line)
            contact['misc'] = ', '.join(email_types) if email_types else ''  # Join list to a single string
        elif field == 'NICKNAME':
            contact['AKA'] = value
        elif field == 'NOTE':
            note = value
            note = note.replace("\\n","")
            contact['note'] = note
        elif 'X-SOCIALPROFILE' in line:
            contact['content'] = line
            if 'x-user=' in line:
                user = line.split('x-user=')[1].split(';')[0]
                contact['user'] = user
            if 'type=' in line:
                info = line.split('type=')[1].split(';')[0]
                
                if info == 'facebook' and user != '' and user is not None:
                    url = f"https://www.facebook.com/{user}"

                    contact['url'] = url
                contact['info'] = info
        elif 'URL:' in line:
            contact['url'] = value  # if url is not blank, add append it
        elif 'BDAY' in line:
            contact['DOB'] = value
        elif field == 'ORG':
            contact['business'] = value

    return contact

def parse_vcf_files(input_folder):
    """
    Parse .vcf files in the input folder and extract contact information.
    """

    message = (f'Converting phone contacts (*.vcf) to Excel from {input_folder} folder')
    message_square(message, color_green)
    
    vcf_files = [f for f in os.listdir(input_folder) if f.endswith('.vcf')]
    if not vcf_files:
        print("Error: No .vcf files found in the input folder.")
        sys.exit(1)

    contacts = []

    for vcf_file in vcf_files:
        file_path = os.path.join(input_folder, vcf_file)
        with open(file_path, 'rb') as file:
            content = file.read()
            encoding = detect_encoding(content)

            try:
                lines = content.decode(encoding).splitlines()
            except UnicodeDecodeError:
                print(f"{color_red}Error: Unable to decode file '{vcf_file}' with detected encoding '{encoding}'.{color_reset}")
                continue

            contact = parse_vcf_content(lines)
            contact['original_file'] = vcf_file
            contacts.append(contact)

    return contacts


def message_square(message, color):
    horizontal_line = f"+{'-' * (len(message) + 2)}+"
    empty_line = f"| {' ' * (len(message))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {message} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def write_intel(data):
    '''
    The write_locations() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''
    message = (f'Writing {output_xlsx}')
    message_square(message, color_green)

    try:
        data = sorted(data, key=lambda x: (x.get("fullname", ""), x.get("query", "")))

        # data = sorted(data, key=lambda x: (x.get("ranking", ""), x.get("fullname", ""), x.get("query", "")))
        print(f'sorted by fullname')
    except TypeError as error:

        print(f'{color_red}{error}{color_reset}')

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Intel'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    log_headers = [
        "Date", "Subject", "Requesting Agency", "Requesting Agent", "Case"
        , "Summary of Findings", "Source", "Notes"
    ]


    # Write headers to the first row
    for col_index, header in enumerate(headers_intel):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [3, 4, 5, 6, 49, 50]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [7,8, 13, 14, 15, 29, 30, 35, 36, 37, 38, 39, 40, 41, 42, 43]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill
        # elif col_index == 27:  # Red for column 27
            # fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
            # cell.fill = fill




    ## Excel column width

    worksheet.column_dimensions['A'].width = 15 # query
    worksheet.column_dimensions['B'].width = 20 # ranking
    worksheet.column_dimensions['C'].width = 20 # fullname
    worksheet.column_dimensions['D'].width = 25 # url
    worksheet.column_dimensions['E'].width = 25 # email
    worksheet.column_dimensions['F'].width = 15 # user
    worksheet.column_dimensions['G'].width = 14 # phone
    worksheet.column_dimensions['H'].width = 16 # business
    worksheet.column_dimensions['I'].width = 24 # fulladdress
    worksheet.column_dimensions['J'].width = 12 # city
    worksheet.column_dimensions['K'].width = 10 # state
    worksheet.column_dimensions['L'].width = 8 # country
    worksheet.column_dimensions['M'].width = 20 # note
    worksheet.column_dimensions['N'].width = 14 # AKA
    worksheet.column_dimensions['O'].width = 11 # DOB
    worksheet.column_dimensions['P'].width = 5 # SEX
    worksheet.column_dimensions['Q'].width = 20 # info
    worksheet.column_dimensions['R'].width = 20 # misc
    worksheet.column_dimensions['S'].width = 10 # firstname
    worksheet.column_dimensions['T'].width = 11 # middlename
    worksheet.column_dimensions['U'].width = 10 # lastname
    worksheet.column_dimensions['V'].width = 10 # associates
    worksheet.column_dimensions['W'].width = 10 # case
    worksheet.column_dimensions['X'].width = 13 # sosfilenumber
    worksheet.column_dimensions['Y'].width = 10 # owner
    worksheet.column_dimensions['Z'].width = 10 # president
    worksheet.column_dimensions['AA'].width = 10 # sosagent
    worksheet.column_dimensions['AB'].width = 10 # managers
    worksheet.column_dimensions['AC'].width = 15 # Time
    worksheet.column_dimensions['AD'].width = 12 # Latitude
    worksheet.column_dimensions['AE'].width = 12 # Longitude
    worksheet.column_dimensions['AF'].width = 22 # Coordinate
    worksheet.column_dimensions['AG'].width = 12 # original_file
    worksheet.column_dimensions['AH'].width = 12 # Source
    worksheet.column_dimensions['AI'].width = 12 # Source file information
    worksheet.column_dimensions['AJ'].width = 10 # Plate
    worksheet.column_dimensions['AK'].width = 10 # VIS
    worksheet.column_dimensions['AL'].width = 10 # VIN
    worksheet.column_dimensions['AM'].width = 10 # VYR
    worksheet.column_dimensions['AN'].width = 10 # VMA
    worksheet.column_dimensions['AO'].width = 10 # LIC
    worksheet.column_dimensions['AP'].width = 10 # LIY
    worksheet.column_dimensions['AQ'].width = 10 # DLN
    worksheet.column_dimensions['AR'].width = 10 # DLS
    worksheet.column_dimensions['AS'].width = 10 # content
    worksheet.column_dimensions['AT'].width = 10 # referer
    worksheet.column_dimensions['AU'].width = 10 # osurl
    worksheet.column_dimensions['AV'].width = 10 # titleurl
    worksheet.column_dimensions['AW'].width = 12 # pagestatus
    worksheet.column_dimensions['AX'].width = 16 # ip
    worksheet.column_dimensions['AY'].width = 15 # dnsdomain

    for i in range(len(data)):
        if data[i] is None:
            data[i] = ''


    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers_intel):
            try:
                cell_data = row_data.get(col_name)
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    # Create a new worksheet for color codes
    color_worksheet = workbook.create_sheet(title='ColorCode')
    color_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    color_worksheet.column_dimensions['A'].width = 14# Color
    color_worksheet.column_dimensions['B'].width = 20# Description


    # Excel row height
    color_worksheet.row_dimensions[2].height = 22  # Adjust the height as needed
    color_worksheet.row_dimensions[3].height = 22
    color_worksheet.row_dimensions[4].height = 23
    color_worksheet.row_dimensions[5].height = 23
    color_worksheet.row_dimensions[6].height = 40   # truck

    color_worksheet.cell(row=1, column=1).value = 'Color'
    color_worksheet.cell(row=1, column=2).value = 'description'
    color_worksheet.cell(row=2, column=1).value = 'Red'
    color_worksheet.cell(row=3, column=1).value = 'Orange'
    color_worksheet.cell(row=4, column=1).value = 'Green'
    color_worksheet.cell(row=5, column=1).value = 'Yellow'

    color_worksheet.cell(row=7, column=1).value = 'ABBREVIATIONS'
    color_worksheet.cell(row=8, column=1).value = 'AKA'
    color_worksheet.cell(row=9, column=1).value = 'DOB'
    color_worksheet.cell(row=10, column=1).value = 'VIS'
    color_worksheet.cell(row=11, column=1).value = 'VIN'
    color_worksheet.cell(row=12, column=1).value = 'VYR'
    color_worksheet.cell(row=13, column=1).value = 'VMA'
    color_worksheet.cell(row=14, column=1).value = 'LIC'
    color_worksheet.cell(row=15, column=1).value = 'LIY'
    color_worksheet.cell(row=16, column=1).value = 'DLN'
    color_worksheet.cell(row=17, column=1).value = 'DLS'

       
    color_worksheet.cell(row=2, column=2).value = 'Bad Intel or dead link'
    color_worksheet.cell(row=3, column=2).value = 'Research'
    color_worksheet.cell(row=4, column=2).value = 'Good Intel'
    color_worksheet.cell(row=5, column=2).value = 'Highlighted'

    color_worksheet.cell(row=8, column=2).value = 'Also Known As (Alias)'
    color_worksheet.cell(row=9, column=2).value = 'Date of Birth'
    color_worksheet.cell(row=10, column=2).value = 'Vehicle State'
    color_worksheet.cell(row=11, column=2).value = 'Vehicle Identification Number'
    color_worksheet.cell(row=12, column=2).value = 'Vehicle Year'
    color_worksheet.cell(row=13, column=2).value = 'Vehicle Make'
    color_worksheet.cell(row=14, column=2).value = 'License'
    color_worksheet.cell(row=15, column=2).value = 'License Year'
    color_worksheet.cell(row=16, column=2).value = 'Drivers License Number'
    color_worksheet.cell(row=17, column=2).value = 'Drivers License State'


    # colored fills
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


    # Apply the orange fill to the cell in row 2, column 2
    color_worksheet.cell(row=2, column=2).fill = red_fill
    color_worksheet.cell(row=3, column=2).fill = orange_fill
    color_worksheet.cell(row=4, column=2).fill = green_fill
    color_worksheet.cell(row=5, column=2).fill = yellow_fill


    # Create a new worksheet for logs
    log_worksheet = workbook.create_sheet(title='Log')
    log_worksheet.freeze_panes = 'B2'  # Freeze cells

# Date, Subject, Requesting Agency, Requesting Agent, Case, Summary of Findings, Source, Notes, Requestor

    # Excel column width
    log_worksheet.column_dimensions['A'].width = 14# Date
    log_worksheet.column_dimensions['B'].width = 20# Subject
    log_worksheet.column_dimensions['C'].width = 24# Requesting Agency
    log_worksheet.column_dimensions['D'].width = 20# Requesting Agent
    log_worksheet.column_dimensions['E'].width = 14# Case
    log_worksheet.column_dimensions['F'].width = 20# Summary of Findings
    log_worksheet.column_dimensions['G'].width = 14# Source
    log_worksheet.column_dimensions['H'].width = 25# Notes

    log_worksheet.cell(row=1, column=1).value = 'Date'
    log_worksheet.cell(row=1, column=2).value = 'Subject'
    log_worksheet.cell(row=1, column=3).value = 'Requesting Agency'
    log_worksheet.cell(row=1, column=4).value = 'Requesting Agent'
    log_worksheet.cell(row=1, column=5).value = 'Case'
    log_worksheet.cell(row=1, column=6).value = 'Summary of Findings'
    log_worksheet.cell(row=1, column=7).value = 'Notes'



    workbook.save(output_xlsx)
    

def Usage():
    file = sys.argv[0].split('\\')[-1]
    print(f"\nDescription: {description2}")
    print(f"{file} Version: {version} by {author}")
    print(f"    {file} -c")
    print(f"    {file} -c -I LogsVCF -O contacts_Apple.xlsx")
    print(f"    {file} -x -O LogsVCF -I VCF_Files")

    
if __name__ == '__main__':
    main()


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>


"""

1.0.2 - created a -x version for xlsx 2 vcf conversions
1.0.1 - it works
"""


# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
apend multiple phone numbers and emails into one line (or two)
ask for the case number

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
