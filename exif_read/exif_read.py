#!/usr/bin/python
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import sys
import exifread   # pip install exifread
import hashlib   # pip install hashlib
import openpyxl   # pip install openpyx
import argparse # for menu system
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "exif data reader"
version = '0.1.0'

global image_types
image_types = ['.jpg', '.jpeg', '.heic', '.heif', '.png', '.tiff', '.tif', '.webp']

# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000:  # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}')  # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      GUI Class      >>>>>>>>>>>>>>>>>>>>>>>>>>

class ExifGui:
    def __init__(self, root):
        self.root = root
        script_name = os.path.basename(sys.argv[0])
        self.root.title(f"{script_name} {version}")
        self.root.geometry("700x550")
        
        # Set Vista theme
        self.style = ttk.Style()
        try:
            self.style.theme_use('vista')
        except:
            pass # Fallback to default if vista is not available

        self.setup_gui()
        self.set_defaults()

    def setup_gui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title/Description
        ttk.Label(main_frame, text=f"{description}", font=("Helvetica", 12, "bold")).pack(pady=5)

        # Input Folder
        input_frame = ttk.LabelFrame(main_frame, text="Input Folder", padding="5")
        input_frame.pack(fill=tk.X, pady=5)
        
        self.input_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.input_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(input_frame, text="Browse", command=self.browse_input).pack(side=tk.RIGHT)

        # Output File
        output_frame = ttk.LabelFrame(main_frame, text="Output Excel File", padding="5")
        output_frame.pack(fill=tk.X, pady=5)
        
        self.output_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.output_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(output_frame, text="Browse", command=self.browse_output).pack(side=tk.RIGHT)

        # Progress Bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=10)

        # Status/Log Window
        self.status_box = ScrolledText(main_frame, height=15, state='disabled')
        self.status_box.pack(fill=tk.BOTH, expand=True, pady=5)

        # Extract Button
        self.btn_extract = ttk.Button(main_frame, text="Extract MetaData", command=self.start_processing)
        self.btn_extract.pack(pady=10)

    def set_defaults(self):
        self.input_var.set("photos")
        timestamp = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
        self.output_var.set(f"exif_data_{timestamp}.xlsx")

    def browse_input(self):
        folder = filedialog.askdirectory()
        if folder:
            self.input_var.set(folder)

    def browse_output(self):
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.output_var.set(file)

    def log(self, message):
        self.status_box.config(state='normal')
        self.status_box.insert(tk.END, message + "\n")
        self.status_box.see(tk.END)
        self.status_box.config(state='disabled')
        self.root.update_idletasks()

    def start_processing(self):
        input_folder = self.input_var.get()
        output_file = self.output_var.get()

        if not os.path.exists(input_folder):
            messagebox.showerror("Error", f"Input folder '{input_folder}' does not exist.")
            return

        self.btn_extract.config(state='disabled')
        self.progress_var.set(0)
        self.status_box.config(state='normal')
        self.status_box.delete(1.0, tk.END)
        self.status_box.config(state='disabled')
        
        self.log(f"Input Folder: {os.path.abspath(input_folder)}")

        # Start processing in a new thread
        thread = threading.Thread(target=self.run_process, args=(input_folder, output_file))
        thread.daemon = True
        thread.start()

    def run_process(self, input_folder, output_file):
        try:
            process_exif(input_folder, output_file, self.log, self.update_progress)
            self.log(f"Output File: {output_file}")
            self.log("\nDone.")
        except Exception as e:
            self.log(f"Critical Error: {e}")
        finally:
            self.btn_extract.config(state='normal')

    def update_progress(self, current, total):
        if total > 0:
            percent = (current / total) * 100
            self.progress_var.set(percent)
        else:
            self.progress_var.set(100)


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu/Logic     >>>>>>>>>>>>>>>>>>>>>>>>>>

def process_exif(photos_folder, output_file, status_callback=None, progress_callback=None):
    """
    Core logic to extract EXIF data and save to Excel.
    """
    # Ensure the input folder exists
    if not os.path.exists(photos_folder):
        msg = f"Input folder {photos_folder} doesn't exist."
        if status_callback: status_callback(msg)
        msg_blurb_square(msg, color_red)
        return False

    # Create a new workbook and add basic formatting
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "FileMetadata"

    # Define headers
    headers = [
        'Name', 'DateCreated', 'DateTimeOriginal', 'FileCreateDate', 'FileModifyDate', 'Timezone',
        'DeviceManufacturer', 'ExifToolVersion', 'FileSize', 'FileType', 'FileTypeExtension', 'LensMake',
        'Software', 'HostComputer', 'LensInfo', 'LensModel', 'Model', 'NumberOfImages', 
        'Altitude', 'Latitude', 'Longitude', 'Coordinate', 'Time', 'Type', 'Icon', 'Description', 'MD5'
    ]
    sheet.append(headers)

    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header).font = openpyxl.styles.Font(bold=True)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    sheet.freeze_panes = 'B2'

    files = [f for f in os.listdir(photos_folder) if os.path.splitext(f)[1].lower() in image_types]
    total_files = len(files)
    
    msg = f'Analyzing {total_files} photos from {photos_folder} folder'
    if status_callback: status_callback(msg)
    msg_blurb_square(msg, color_green)

    for i, filename in enumerate(files, 1):
        file_path = os.path.join(photos_folder, filename)
        try:
            exif_data, Description = read_exif_data(file_path)
            print(f"{color_green}{filename}{color_reset}")
            if status_callback: status_callback(filename)
            sheet.append([exif_data.get(header, "") for header in headers])
        except Exception as e:
            err_msg = f"Error processing {filename}: {e}"
            print(f"{color_red}{err_msg}{color_reset}")
            if status_callback: status_callback(err_msg)
            exif_data = {'Name': os.path.basename(file_path)}
            sheet.append([exif_data.get(header, "") for header in headers])
        
        if progress_callback:
            progress_callback(i, total_files)

    workbook.save(output_file)
    msg = f"File metadata exported to {output_file}"
    if status_callback: status_callback(msg)
    msg_blurb_square(msg, color_green)
    return True

    
def main():
    """
    Main function to parse arguments and initiate file conversion.
    """
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='Input folder', required=False)
    parser.add_argument('-e', '--exif', help='extract exif data from photos', required=False, action='store_true')

    args = parser.parse_args()

    # If no arguments are provided, launch GUI
    if len(sys.argv) == 1:
        root = tk.Tk()
        gui = ExifGui(root)
        root.mainloop()
        return

    # CLI Mode
    photos_folder = args.input if args.input else 'photos'
    timestamp = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
    output_file = f'exif_data_{timestamp}.xlsx'

    process_exif(photos_folder, output_file)


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def cleanup_description(description):
    cleaned_description = {k: v for k, v in description.items() if v not in ('', None) and k not in ('Icon', 'Type', 'Latitude', 'Longitude')}
    formatted_description = "\n".join(f"{k}: {v}" for k, v in cleaned_description.items())
    return formatted_description


def convert_to_decimal(coord, ref):
    """Converts GPS coordinates to decimal format."""
    if not coord:
        return None
    degrees = float(coord[0].num) / float(coord[0].den)
    minutes = float(coord[1].num) / float(coord[1].den) / 60.0
    seconds = float(coord[2].num) / float(coord[2].den) / 3600.0
    decimal = degrees + minutes + seconds
    if ref in ['S', 'W']:
        decimal = -decimal
    return decimal


def get_file_timestamps(file_path):
    """Returns the creation, last accessed, and last modified times of a file."""
    if not os.path.isfile(file_path):
        raise ValueError(f"The path {file_path} is not a valid file.")

    file_stats = os.stat(file_path)
    creation_time = datetime.fromtimestamp(file_stats.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
    access_time = datetime.fromtimestamp(file_stats.st_atime).strftime('%Y-%m-%d %H:%M:%S')
    modified_time = datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
    return creation_time, access_time, modified_time


def hashfile(file_path):
    """
    Computes and returns the MD5 hash of a file.
    """
    # Create an MD5 hash object
    hash_md5 = hashlib.md5()

    # Open the file in binary read mode
    with open(file_path, "rb") as f:
        # Read the file in chunks to avoid memory overload with large files
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    
    # Return the hexadecimal representation of the hash
    return hash_md5.hexdigest()
    

def read_exif_data(file_path):
    """Reads selected EXIF data from an image file."""
    with open(file_path, 'rb') as f:
        tags = exifread.process_file(f, stop_tag='UNDEFINED')

    gps_latitude = tags.get('GPS GPSLatitude')
    gps_latitude_ref = tags.get('GPS GPSLatitudeRef')
    gps_longitude = tags.get('GPS GPSLongitude')
    gps_longitude_ref = tags.get('GPS GPSLongitudeRef')
    
    latitude = convert_to_decimal(gps_latitude.values, gps_latitude_ref.values[0]) if gps_latitude and gps_latitude_ref else ""
    longitude = convert_to_decimal(gps_longitude.values, gps_longitude_ref.values[0]) if gps_longitude and gps_longitude_ref else ""

    # Generate MD5 hash for the file
    try:
        hash_md5 = hashfile(file_path)

    except Exception as e:
        print(f"Error hashing '{file_path}': {e}")        
        hash_md5 = ""
        
    exif_data = {
        'Name': os.path.basename(file_path),
        'DateCreated': tags.get('Image DateTime'),
        'DateTimeOriginal': tags.get('EXIF DateTimeOriginal'),
        'FileCreateDate': datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
        'FileModifyDate': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
        'Timezone': tags.get('EXIF OffsetTime'),
        'DeviceManufacturer': tags.get('Image Make'),
        'FileSize': os.path.getsize(file_path),
        'FileType': os.path.splitext(file_path)[1].replace('.', '').upper(),
        'FileTypeExtension': os.path.splitext(file_path)[1].replace('.', '').lower(),
        'LensMake': tags.get('EXIF LensMake'),
        'Software': tags.get('Image Software'),
        'HostComputer': tags.get('Image HostComputer'),
        'LensInfo': tags.get('EXIF LensInfo'),
        'LensModel': tags.get('EXIF LensModel'),
        'Model': tags.get('Image Model'),
        'Altitude': tags.get('GPS GPSAltitude'),
        'Latitude': latitude,
        'Longitude': longitude,
        'Coordinate': f"{latitude},{longitude}" if latitude and longitude else "",
        'NumberOfImages': tags.get('Exif NumberOfImages'),
        'ExifToolVersion': tags.get('Exif ExifToolVersion'),
        'Icon': 'Images',
        'Type': 'Images',
        'Time': tags.get('EXIF DateTimeOriginal'),
        'Description': '',        
        'MD5': hash_md5
    }

    for key, value in exif_data.items():
        exif_data[key] = str(value) if value else ""
        
    Description = cleanup_description(exif_data)
    return exif_data, Description


def msg_blurb_square(msg_blurb, color):
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"
    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def Usage():
    file = sys.argv[0].split('\\')[-1]

    print(f'\nDescription: {color_green}{description}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f"    {file}")
    print(f"    {file} -I photos")



if __name__ == "__main__":
    main()




# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>


"""
0.2.0 - added menu and usage
0.0.1 - exif data to xlsx
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>


"""

hash each file

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>


"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
