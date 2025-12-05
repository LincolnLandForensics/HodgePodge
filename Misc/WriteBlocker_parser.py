import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
from bs4 import BeautifulSoup

# --- Configuration ---

VERSION = "1.2" 

# Final set of headers
HEADERS = ["Date", "Test", "Type", "ID", "Device", "Serial #", "Status", "DFE", "Note", "Hostname or Location", "Firmware Version", "Case", "URL", "TestApp", "Interface", "Logfile"]

# UPDATED: Final set of column widths
COLUMN_WIDTHS = [21, 23, 14, 7, 33, 21, 7, 14, 30, 17, 20, 9, 20, 15, 9, 38] 

DEFAULT_OUTPUT_FILENAME = "Lab_Certification.xlsx"
DFE_OPTIONS = ["DFE SHERLOCK", "DFE HOLMES"]

# Help content
HELP_TEXT = '''
Download Write-Block Validation Utility from:
https://cdsg.com/products/usb-writeblocker#documents-downloads

Default log location:
C:\\Program Files (x86)\\CDSG\\WriteBlocking Validation Utility\\Test Results

ID is the unique ID (Ex. WB-03) that is written on each write blocker or Faraday bag/box.
'''

class LogImporterApp:
    def __init__(self, master):
        self.master = master
        # UPDATED: Set title using the VERSION variable
        master.title(f"WiebeTech Log Importer {VERSION}") 
        
        # Default values
        self.output_file = os.path.join(os.getcwd(), DEFAULT_OUTPUT_FILENAME)

        # Variables for GUI inputs
        self.logfile_path = tk.StringVar()
        self.id_var = tk.StringVar()
        self.outputfile_path_var = tk.StringVar(value=self.output_file)
        self.dfe_var = tk.StringVar(value=DFE_OPTIONS[0])
        self.tester_name_var = tk.StringVar()

        # Build the GUI
        self._create_widgets()

    def _create_widgets(self):
        # Frame for all inputs
        input_frame = ttk.LabelFrame(self.master, text=" Log Import Parameters ")
        input_frame.pack(padx=10, pady=10, fill="x")
        
        # Row 1: LogFile Input
        ttk.Label(input_frame, text="LogFile (HTML):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(input_frame, textvariable=self.logfile_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="Browse", command=self._browse_logfile).grid(row=0, column=2, padx=5, pady=5)

        # Row 2: ID Input - Includes example
        ttk.Label(input_frame, text="ID: (Ex. WB-03)").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(input_frame, textvariable=self.id_var, width=50).grid(row=1, column=1, padx=5, pady=5)

        # Row 3: OutputFile Input (with default)
        ttk.Label(input_frame, text="Output File (Excel):").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(input_frame, textvariable=self.outputfile_path_var, width=50).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="Browse", command=self._browse_outputfile).grid(row=2, column=2, padx=5, pady=5)

        # Row 4: DFE Selection (Pull-down and Custom Tester)
        ttk.Label(input_frame, text="DFE (Pull-down):").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        ttk.Combobox(input_frame, textvariable=self.dfe_var, values=DFE_OPTIONS, width=47).grid(row=3, column=1, padx=5, pady=5)
        
        # Row 5: Custom Tester Name
        ttk.Label(input_frame, text="Custom Tester Name (Optional):").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(input_frame, textvariable=self.tester_name_var, width=50).grid(row=4, column=1, padx=5, pady=5)

        # Button Frame for Run and Help
        button_frame = ttk.Frame(self.master)
        button_frame.pack(pady=10)

        # Run Button
        ttk.Button(button_frame, text="Import WiebeTech Validation Log", command=self._run_import).pack(side=tk.LEFT, padx=5)

        # Help Button
        ttk.Button(button_frame, text="Help", command=self._show_help).pack(side=tk.LEFT, padx=5)

        # Results Display (Text area)
        ttk.Label(self.master, text="--- Results ---").pack()
        self.results_text = tk.Text(self.master, height=10, width=80, state='disabled')
        self.results_text.pack(padx=10, pady=5)
        
    def _show_help(self):
        """Displays the help information in a separate message box."""
        messagebox.showinfo("Help - WiebeTech Log Importer", HELP_TEXT)

    def _browse_logfile(self):
        """Opens a file dialog for selecting the HTML log file, defaulting to the specified path."""
        default_dir = r"C:\Program Files (x86)\CDSG\WriteBlocking Validation Utility\Test Results"
        
        # Fall back to the current directory if the specific path doesn't exist
        initial_dir = default_dir if os.path.isdir(default_dir) else os.getcwd() 

        filename = filedialog.askopenfilename(
            defaultextension=".html",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
            initialdir=initial_dir 
        )
        if filename:
            self.logfile_path.set(filename)

    def _browse_outputfile(self):
        """Opens a file dialog for selecting or creating the Excel output file."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=os.path.basename(self.output_file)
        )
        if filename:
            self.outputfile_path_var.set(filename)

    def _update_results_display(self, text):
        """Enables, inserts text, and then disables the results text box."""
        self.results_text.config(state='normal')
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, text)
        self.results_text.config(state='disabled')

    def _get_div_text(self, soup, div_id):
        """Extracts text content from a specific div ID."""
        target_div = soup.find('div', id=div_id)
        if target_div:
            return target_div.text.strip()
        return ''

    def _parse_html_log(self, logfile_path):
        """Reads the HTML file, extracts data using BeautifulSoup for robustness."""
        
        extracted_data = {'URL': ''} # Initialize URL to empty string
        
        # Add Logfile name (no path)
        extracted_data['Logfile'] = os.path.basename(logfile_path)

        try:
            # Handle encoding error by ignoring bad characters, ensuring the file opens.
            with open(logfile_path, 'r', encoding='utf-8', errors='ignore') as f:
                html_content = f.read()
        except FileNotFoundError:
            raise FileNotFoundError(f"Error: Log file not found at {logfile_path}")
        except Exception as e:
            raise Exception(f"Error reading log file: {e}")

        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Helper function to extract a value from a table based on a label
        def get_table_value(table_id, label):
            target_div = soup.find('div', id=table_id)
            if target_div:
                table = target_div.find('table')
                if table:
                    for row in table.find_all('tr'):
                        cells = row.find_all('td')
                        if len(cells) >= 2 and cells[0].text.strip().lower() == label.lower():
                            return cells[1].text.strip()
            return ''

        # --- 1. Extract Hardware/System Information ---
        
        # Hostname (Computer Information: ID='machine_info')
        hostname = get_table_value('machine_info', 'Name:')
        extracted_data['Hostname or Location'] = hostname

        # Write Blocker Information (ID='wb_info')
        wb_name = get_table_value('wb_info', 'Name') 
        wb_serial = get_table_value('wb_info', 'Serial Number')
        wb_firmware = get_table_value('wb_info', 'Firmware')
        wb_interface = get_table_value('wb_info', 'Drive Interface') 

        extracted_data['Device'] = wb_name              
        extracted_data['Serial #'] = wb_serial          
        extracted_data['Firmware Version'] = wb_firmware 
        extracted_data['Interface'] = wb_interface
        
        # --- 2. Extract Test Information (Status, Note, Date, TestApp) ---

        # Status Check: check for key success phrase anywhere in the log
        pass_phrase_key = "No sectors on the drive were modified during the test."
        fail_phrase_key1 = "Sectors on the drive were modified during the test."
        
        
        if pass_phrase_key in html_content:
            extracted_data['Status'] = 'Pass'
        elif fail_phrase_key1 in html_content:
            extracted_data['Status'] = 'Fail'
        else:
             extracted_data['Status'] = ''
        
        # Robust Note Extraction
        note_div = soup.find('div', id='notes_info')
        note_content = ''
        if note_div:
            # Get all text and split into lines
            lines = [line.strip() for line in note_div.get_text().splitlines() if line.strip()]
            
            if lines:
                # If the first line starts with a common header, remove it
                if lines[0].lower().startswith('notes'):
                    # Keep the content after the header
                    lines = lines[1:]
                
                # Join remaining lines into a single string, replacing multiple spaces/newlines with a single space
                note_content = ' '.join(lines).strip()

        extracted_data['Note'] = note_content
        
        # TestApp (The whole line including version)
        test_app_version_text = self._get_div_text(soup, 'version_info')
        if "Version:" in test_app_version_text:
            version_number = test_app_version_text.split('Version:')[-1].strip()
            extracted_data['TestApp'] = f"WiebeTech WriteBlocking Validation Utility, version {version_number}"
            extracted_data['Test'] = 'WriteBlocking Validation'
            extracted_data['Type'] = 'Write blocker'
            
            # NEW: Set URL based on TestApp
            if "WiebeTech WriteBlocking Validation Utility" in extracted_data['TestApp']:
                extracted_data['URL'] = "https://cdsg.com/products/usb-writeblocker#documents-downloads"


        # Date
        # Search the raw HTML content for the starting timestamp
        date_match = re.search(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} [AP]M) \| Starting test...', html_content)
        if date_match:
            try:
                # Parse the specific AM/PM format
                date_obj = datetime.strptime(date_match.group(1), '%Y-%m-%d %I:%M:%S %p')
                extracted_data['Date'] = date_obj.strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                extracted_data['Date'] = date_match.group(1).strip() # Fallback

        return extracted_data

    def _append_to_excel(self, data, output_path):
        """Appends the extracted data to the specified Excel file and applies formatting."""
        
        # Create a new workbook if the file doesn't exist
        if not os.path.exists(output_path):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Lab"
            ws.append(HEADERS)
            # Apply initial formatting to the first row (headers)
            for i, width in enumerate(COLUMN_WIDTHS):
                col_letter = get_column_letter(i + 1)
                if width is not None:
                    ws.column_dimensions[col_letter].width = width
            ws.freeze_panes = "B2"
            wb.save(output_path)
            
        # Load the existing workbook
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Apply formatting (re-apply in case it was opened/saved by another program)
        ws.freeze_panes = "B2"
        for i, width in enumerate(COLUMN_WIDTHS):
            col_letter = get_column_letter(i + 1)
            if width is not None:
                ws.column_dimensions[col_letter].width = width
            
        # Prepare the row of data based on the HEADEERS configuration
        row_data = []
        for header in HEADERS:
            # Use data.get(header, '') to handle missing fields gracefully
            row_data.append(data.get(header, ''))
            
        # Append the new row
        ws.append(row_data)
        
        # Save the workbook
        wb.save(output_path)
        
        return len(ws['A']) - 1 # Return the row number where data was added

    def _run_import(self):
        """Main method to execute the log import process."""
        
        log_path = self.logfile_path.get()
        output_path = self.outputfile_path_var.get()
        tester_name = self.tester_name_var.get().strip()
        
        # Use custom tester name if provided, otherwise use the DFE selection
        dfe_value = tester_name if tester_name else self.dfe_var.get()

        if not log_path or not os.path.exists(log_path):
            messagebox.showerror("Input Error", "Please select a valid LogFile (HTML).")
            return
        
        # --- 1. Parse HTML Log ---
        try:
            extracted_data = self._parse_html_log(log_path)
        except Exception as e:
            messagebox.showerror("Processing Error", str(e))
            return
        
        # --- 2. Add GUI/User Inputs to Data ---
        extracted_data['ID'] = self.id_var.get()
        extracted_data['DFE'] = dfe_value # Export the tester as DFE
        
        # --- 3. Append to Excel ---
        try:
            row_num = self._append_to_excel(extracted_data, output_path)
        except Exception as e:
            messagebox.showerror("Excel Error", f"Error writing to Excel file: {e}")
            return
            
        # --- 4. Display Results ---
        result_display = f"✅ SUCCESS: Log imported successfully!\n\n"
        result_display += f"Source Log: {os.path.basename(log_path)}\n"
        result_display += f"Output File: {os.path.basename(output_path)}\n"
        result_display += f"Data written to row {row_num}.\n\n"
        result_display += "--- Extracted Data ---\n"
        
        # Display the data that was written
        for header in HEADERS:
            value = extracted_data.get(header, '')
            result_display += f"{header}: {value}\n"
            
        self._update_results_display(result_display)
        messagebox.showinfo("Success", "Log imported successfully and results displayed.")

if __name__ == "__main__":
    # Reminder to install dependencies: pip install openpyxl beautifulsoup4
    
    root = tk.Tk()
    app = LogImporterApp(root)
    root.mainloop()