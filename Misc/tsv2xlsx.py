import os
import csv
import re
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.styles import Font, PatternFill, Alignment

version = 1.0.0

# Excel limits
MAX_EXCEL_ROWS = 1_048_576
csv.field_size_limit(MAX_EXCEL_ROWS)

_illegal_xml_chars = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

HEADER_MAP = {
    "First Name": "firstname", "Middle Name": "middlename", "Last Name": "lastname",
    "Phone Numbers": "phone", "Phone Number": "phone", "Email addresses": "email",
    "Email Address": "email", "Username": "user", "User Name": "user", "Heading": "Direction",
    "Timestamp": "Time", "Message Timestamp": "Time", "Key Timestamp": "Time", "Path": "Source",
    "Call Time": "Time",
    "Data": "note", "Place": "fulladdress", "URL": "url", "Message": "note",
    "Final Live Latitude": "Latitude", "Final Live Longitude": "Longitude", "Label": "note",
    "Source": "Source file information", "Origin": "Source file information",
    "SSID": "note", "Password": "info"
}

def clean_cell_value(value):
    if not isinstance(value, str):
        return value
    value = _illegal_xml_chars.sub('', value)
    value = re.sub(r'<[^>]+>', '', value)
    return value.strip()

def apply_header_style(ws):
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

def convert_tsv_to_xlsx(input_folder, output_folder, message_box):
    os.makedirs(output_folder, exist_ok=True)
    tsv_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".tsv")]
    total = len(tsv_files)

    message_box.insert(tk.END, f"📁 converting .tsv to .xlsx format in {input_folder}\n")

    for idx, filename in enumerate(tsv_files, 1):
        tsv_path = os.path.join(input_folder, filename)
        safe_name = os.path.splitext(filename)[0].replace(" ", "_")
        xlsx_name = safe_name + ".xlsx"
        xlsx_path = os.path.join(output_folder, xlsx_name)

        wb = Workbook()
        ws = wb.active
        ws.title = "Data_1"
        ws.freeze_panes = "B2"

        sheet_count = 1
        row_count = 0
        headers_written = False

        def new_sheet():
            nonlocal wb, sheet_count, ws, row_count, headers_written
            sheet_count += 1
            ws = wb.create_sheet(title=f"Data_{sheet_count}")
            ws.freeze_panes = "B2"
            row_count = 0
            headers_written = False

        try:
            with open(tsv_path, "r", encoding="utf-8", errors="replace") as tsv_file:
                clean_lines = (line.replace('\x00', '') for line in tsv_file)
                reader = csv.reader(clean_lines, delimiter='\t')

                for row in reader:
                    if not any(cell.strip() for cell in row):
                        continue
                    clean_row = [clean_cell_value(cell) for cell in row]

                    if not headers_written:
                        headers = [HEADER_MAP.get(cell.strip(), cell.strip()) for cell in clean_row]
                        ws.append(headers)
                        apply_header_style(ws)
                        row_count += 1
                        headers_written = True
                        continue

                    if row_count >= MAX_EXCEL_ROWS - 1:
                        new_sheet()
                        ws.append(headers)
                        apply_header_style(ws)
                        row_count += 1

                    try:
                        ws.append(clean_row)
                        for col_idx, _ in enumerate(clean_row, 1):
                            ws.cell(row=row_count + 1, column=col_idx).alignment = Alignment(vertical="top", horizontal="left")
                        row_count += 1
                    except IllegalCharacterError:
                        ws.append(["[ILLEGAL CHARACTER REMOVED]"] + clean_row)
                        for col_idx, _ in enumerate(clean_row, 1):
                            ws.cell(row=row_count + 1, column=col_idx).alignment = Alignment(vertical="top", horizontal="left")
                        row_count += 1

        except UnicodeDecodeError:
            message_box.insert(tk.END, f"⚠️ UnicodeDecodeError in {filename}, retrying with Latin-1...\n")
            with open(tsv_path, "r", encoding="latin-1", errors="replace") as tsv_file:
                clean_lines = (line.replace('\x00', '') for line in tsv_file)
                reader = csv.reader(clean_lines, delimiter='\t')
                for row in reader:
                    if not any(cell.strip() for cell in row):
                        continue
                    clean_row = [clean_cell_value(cell) for cell in row]
                    if not headers_written:
                        headers = [HEADER_MAP.get(cell.strip(), cell.strip()) for cell in clean_row]
                        ws.append(headers)
                        apply_header_style(ws)
                        row_count += 1
                        headers_written = True
                        continue
                    if row_count >= MAX_EXCEL_ROWS - 1:
                        new_sheet()
                        ws.append(headers)
                        apply_header_style(ws)
                        row_count += 1
                    try:
                        ws.append(clean_row)
                        for col_idx, _ in enumerate(clean_row, 1):
                            ws.cell(row=row_count + 1, column=col_idx).alignment = Alignment(vertical="top", horizontal="left")
                        row_count += 1
                    except IllegalCharacterError:
                        ws.append(["[ILLEGAL CHARACTER REMOVED]"] + clean_row)
                        for col_idx, _ in enumerate(clean_row, 1):
                            ws.cell(row=row_count + 1, column=col_idx).alignment = Alignment(vertical="top", horizontal="left")
                        row_count += 1

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in list(col)[:200]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column].width = min(max_length + 2, 50)

        try:
            wb.save(xlsx_path)
            # Rename if '_-_' is in filename
            if "_-_" in xlsx_name:
                new_name = xlsx_name.replace("_-_", "-")
                new_path = os.path.join(output_folder, new_name)
                os.rename(xlsx_path, new_path)
                xlsx_name = new_name
                xlsx_path = new_path
            message_box.insert(tk.END, f"✅ Converted: {filename} → {xlsx_name} ({sheet_count} sheet(s))\n")
        except Exception as e:
            message_box.insert(tk.END, f"❌ Failed to save {filename}: {e}\n")

# GUI setup
root = tk.Tk()
root.title("TSV to XLSX Converter")

tk.Label(root, text="📎 Convert .tsv files to .xlsx format", font=("Arial", 10, "bold")).pack(pady=(5, 0))
tk.Label(root, text="Input and Output folders default to current folder", font=("Arial", 9)).pack(pady=(0, 5))

frame = tk.Frame(root)
frame.pack(pady=5)

tk.Label(frame, text="Input Folder:").grid(row=0, column=0, sticky="e")
input_entry = tk.Entry(frame, width=50)
input_entry.insert(0, os.getcwd())
input_entry.grid(row=0, column=1)
tk.Button(frame, text="Browse", command=lambda: input_entry.delete(0, tk.END) or input_entry.insert(0, filedialog.askdirectory())).grid(row=0, column=2)

tk.Label(frame, text="Output Folder:").grid(row=1, column=0, sticky="e")
output_entry = tk.Entry(frame, width=50)
output_entry.insert(0, os.getcwd())
output_entry.grid(row=1, column=1)
tk.Button(frame, text="Browse", command=lambda: output_entry.delete(0, tk.END) or output_entry.insert(0, filedialog.askdirectory())).grid(row=1, column=2)

message_box = tk.Listbox(root, width=80, height=15)
message_box.pack(pady=5)

tk.Button(root, text="Convert", command=lambda: convert_tsv_to_xlsx(input_entry.get(), output_entry.get(), message_box), bg="lightblue").pack(pady=10)

root.mainloop()