import os
import sys
import hashlib
import datetime
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext

# ----------------------------------------------------------------------
# Metadata
# ----------------------------------------------------------------------
version = "v1.1"
description = "Recursive file hasher (MD5 + SHA-1 + SHA-256) with XLSX output"

DEFAULT_INPUT = "files"
DEFAULT_OUTPUT = "FileSummary.xlsx"

# Dynamic header list — reorder this anytime
HEADERS = ["Item", "Path", "FileType", "ExportID", "ItemPath", "OriginalPath"
, "ParentItem", "HeadOfFamily", "OriginalMD5", "MD5", "OriginalSHA1", "SHA1"
, "OriginalSHA256", "SHA256", "Name", "Size", "Created", "Modified"]


# ----------------------------------------------------------------------
# Hashing + file info
# ----------------------------------------------------------------------
def hash_file(path, block_size=65536):
    md5 = hashlib.md5()
    sha1 = hashlib.sha1()
    sha256 = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(block_size), b""):
            md5.update(block)
            sha1.update(block)
            sha256.update(block)
    return md5.hexdigest(), sha1.hexdigest(), sha256.hexdigest()


def get_file_info(path):
    stat = os.stat(path)
    created = datetime.datetime.fromtimestamp(stat.st_ctime).isoformat()
    modified = datetime.datetime.fromtimestamp(stat.st_mtime).isoformat()
    size = stat.st_size
    return created, modified, size


def collect_files(root_dir):
    file_list = []
    for root, dirs, files in os.walk(root_dir):
        for name in files:
            file_list.append(os.path.join(root, name))
    return file_list


# ----------------------------------------------------------------------
# XLSX writer (header‑driven)
# ----------------------------------------------------------------------
def write_xlsx(data, output_xlsx, headers):
    print(f"converting to {output_xlsx}")

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "FilesSummary"

    # Freeze at B2
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 22

    # -----------------------------
    # Write headers (dynamic order)
    # -----------------------------
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)

        # Color columns 1–9 orange
        if 1 <= col_index <= 18:
            cell.fill = PatternFill(start_color="FFA500",
                                    end_color="FFA500",
                                    fill_type="solid")

    # -----------------------------
    # Column width settings
    # -----------------------------
    widths = {
        "A": 6,  # Item
        "B": 38,  # Path
        "C": 9,  # FileType
        "D": 9,  # ExportID
        "E": 9,  # ItemPath
        "F": 14,  # OriginalPath
        "G": 11,  # ParentItem
        "H": 14,  # HeadOfFamily
        "I": 12,  # OriginalMD5
        "J": 32,  # MD5
        "K": 12,  # OriginalSHA1
        "L": 41,  # SHA1
        "M": 15,  # OriginalSHA256
        "N": 65,  # SHA256
        "O": 30,  # Name
        "P": 6,  # Size
        "Q": 27,  # Created
        "R": 27,  # Modified        
    }

								



    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # -----------------------------
    # Write data rows (dynamic order)
    # -----------------------------
    for r_index, row in enumerate(data, start=2):
        for c_index, column_name in enumerate(headers, start=1):
            ws.cell(row=r_index, column=c_index, value=row.get(column_name, ""))

    # -----------------------------
    # Color Code sheet
    # -----------------------------
    color_ws = workbook.create_sheet("ColorCode")
    color_ws.freeze_panes = "B2"
    color_ws.column_dimensions['A'].width = 14
    color_ws.column_dimensions['B'].width = 22

    color_ws["A1"] = "Color"
    color_ws["B1"] = "Description"
    color_ws["A2"] = "Red";    color_ws["B2"] = "Bad Intel or dead link"
    color_ws["A3"] = "Orange"; color_ws["B3"] = "Research"
    color_ws["A4"] = "Green";  color_ws["B4"] = "Good Intel"
    color_ws["A5"] = "Yellow"; color_ws["B5"] = "Highlighted"

    red_fill    = PatternFill(start_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', fill_type='solid')
    green_fill  = PatternFill(start_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', fill_type='solid')

    color_ws["B2"].fill = red_fill
    color_ws["B3"].fill = orange_fill
    color_ws["B4"].fill = green_fill
    color_ws["B5"].fill = yellow_fill

    # -----------------------------
    # Log sheet
    # -----------------------------
    log_ws = workbook.create_sheet("Log")
    log_ws.freeze_panes = "B2"

    headers_log = [
        "Date", "Subject", "Requesting Agency", "Requesting Agent",
        "Case", "Summary of Findings", "Notes"
    ]

    for idx, title in enumerate(headers_log, start=1):
        log_ws.cell(row=1, column=idx, value=title)

    # -----------------------------
    # Save workbook
    # -----------------------------
    workbook.save(output_xlsx)
    print(f"Saved {output_xlsx}")


# ----------------------------------------------------------------------
# Core processing
# ----------------------------------------------------------------------
def process_folder(input_folder, output_file, do_hashing=True, status_callback=None):
    if status_callback:
        status_callback(f"Collecting files from: {input_folder}\n")

    files = collect_files(input_folder)
    total = len(files)

    if status_callback:
        status_callback(f"Found {total} files.\n")

    data = []
    max_workers = min(8, os.cpu_count() or 4)

    def worker(path):
        
        Name = os.path.basename(path)
        FileType = os.path.splitext(path)[1].lower()
        if FileType == '':
            FileType = TypeFind(path)
        record = {
            "Path": path,
            "Name": Name,
            # "Name": os.path.basename(path),
            "FileType": FileType,
            # "FileType": os.path.splitext(path)[1].lower(),
            "Size": "",
            "Created": "",
            "Modified": "",
            "MD5": "",
            "SHA1": "",
            "SHA256": "",
            "Error": "",
        }
        try:
            created, modified, size = get_file_info(path)
            if do_hashing:
                md5, sha1, sha256 = hash_file(path)
            else:
                md5, sha1, sha256 = "", "", ""

            record["Created"] = created
            record["Modified"] = modified
            record["Size"] = size
            record["MD5"] = md5
            record["SHA1"] = sha1
            record["SHA256"] = sha256

        except Exception as e:
            record["Error"] = str(e)

        return record

    processed = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(worker, p): p for p in files}

        for future in as_completed(futures):
            record = future.result()
            data.append(record)

            processed += 1
            if status_callback and processed % 50 == 0:
                status_callback(f"Processed {processed}/{total} files...\n")

    write_xlsx(data, output_file, HEADERS)

    if status_callback:
        status_callback(f"\nDone. Wrote: {output_file}\n")

    # Terminal output
    if files:
        print(f"Input folder: {input_folder}")
    print(f"Output file: {output_file}")


# ----------------------------------------------------------------------
# Tkinter GUI
# ----------------------------------------------------------------------
gui_active = False
root = None
entry_input = None
entry_output = None
text_status = None
progress = None
btn_start = None
hash_files_var = None


def gui_log(message):
    if gui_active and text_status is not None:
        text_status.insert(tk.END, message)
        text_status.see(tk.END)


def start_processing_thread():
    input_folder = entry_input.get().strip()
    output_file = entry_output.get().strip()

    if not input_folder:
        input_folder = DEFAULT_INPUT
        entry_input.delete(0, tk.END)
        entry_input.insert(0, input_folder)

    if not output_file:
        output_file = DEFAULT_OUTPUT
        entry_output.delete(0, tk.END)
        entry_output.insert(0, output_file)

    btn_start.config(state=tk.DISABLED)
    progress.start(10)
    gui_log(f"Starting hashing...\nInput: {input_folder}\nOutput: {output_file}\n\n")

    def worker():
        try:
            process_folder(input_folder, output_file, do_hashing=hash_files_var.get(),
                           status_callback=lambda m: root.after(0, gui_log, m))
        finally:
            root.after(0, processing_done)

    threading.Thread(target=worker, daemon=True).start()

def TypeFind(file_path):
    FileType = ''
    Name = os.path.basename(file_path)
    
    magic_dict = {
        b'\xff\xd8\xff': '.jpg',
        b'\x89PNG\r\n\x1a\n': '.png',
        b'%PDF-': '.pdf',
        b'PK\x03\x04': '.zip',
        b'GIF87a': '.gif',
        b'GIF89a': '.gif',
        b'Rar!\x1a\x07\x00': '.rar',
        b'Rar!\x1a\x07\x01\x00': '.rar',
        b'\x1f\x8b\x08': '.gz',
        b'\x42\x4d': '.bmp',
        b'ID3': '.mp3',
        b'\x00\x00\x00\x18ftyp': '.mp4',
        b'\x00\x00\x00\x20ftyp': '.mp4',
        b'\x25\x21\x50\x53': '.ps',
        b'{\\rtf1': '.rtf',
        b'\x49\x49\x2A\x00': '.tiff',
        b'\x4D\x4D\x00\x2A': '.tiff',
        b'OggS': '.ogg',
        b'MZ': '.exe',
        b'\x7FELF': '.elf',
        b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1': '.doc',
        b'\x53\x51\x4C\x69\x74\x65\x20\x66\x6F\x72\x6D\x61\x74\x20\x33\x00': '.sqlite',
        b'\x3c\x3f\x78\x6d\x6c': '.xml',
        b'BM': '.bmp',
    }

    try:
        with open(file_path, 'rb') as f:
            header = f.read(32)
            
        if header.startswith(b'RIFF'):
            if header[8:12] == b'WAVE':
                return '.wav'
            elif header[8:12] == b'AVI ':
                return '.avi'
            elif header[8:12] == b'WEBP':
                return '.webp'
            
        for magic, ext in magic_dict.items():
            if header.startswith(magic):
                return ext
    except Exception as e:
        print(f"Error reading header for {Name}: {e}")
        
    return FileType

def processing_done():
    progress.stop()
    btn_start.config(state=tk.NORMAL)
    gui_log("Processing complete.\n")


def launch_gui():
    global gui_active, root, entry_input, entry_output, text_status, progress, btn_start, hash_files_var

    gui_active = True
    root = tk.Tk()

    script_name = os.path.basename(sys.argv[0])
    root.title(f"{script_name} {version}")
    root.geometry("750x550")

    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")

    lbl_desc = tk.Label(root, text=description, font=("Arial", 12, "bold"))
    lbl_desc.pack(pady=10)

    # Input folder
    frame_input = tk.Frame(root)
    frame_input.pack(fill=tk.X, padx=20, pady=5)
    tk.Label(frame_input, text="Input Folder:", width=15, anchor="e").pack(side=tk.LEFT)
    entry_input = tk.Entry(frame_input)
    entry_input.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    entry_input.insert(0, DEFAULT_INPUT)

    def browse_input():
        folder = filedialog.askdirectory()
        if folder:
            entry_input.delete(0, tk.END)
            entry_input.insert(0, folder)

    tk.Button(frame_input, text="Browse", command=browse_input).pack(side=tk.LEFT)

    # Output file
    frame_output = tk.Frame(root)
    frame_output.pack(fill=tk.X, padx=20, pady=5)
    tk.Label(frame_output, text="Output File (.xlsx):", width=15, anchor="e").pack(side=tk.LEFT)
    entry_output = tk.Entry(frame_output)
    entry_output.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    entry_output.insert(0, DEFAULT_OUTPUT)

    def browse_output():
        file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel files", "*.xlsx")])
        if file:
            entry_output.delete(0, tk.END)
            entry_output.insert(0, file)

    tk.Button(frame_output, text="Browse", command=browse_output).pack(side=tk.LEFT)

    # Options
    hash_files_var = tk.BooleanVar(value=True)
    chk_hash = tk.Checkbutton(root, text="Hash files", variable=hash_files_var)
    chk_hash.pack(pady=5)

    # Start button
    btn_start = tk.Button(root, text="Start", command=start_processing_thread,
                          font=("Arial", 10, "bold"), bg="#4CAF50", fg="white")
    btn_start.pack(pady=15)

    # Progress bar
    progress = ttk.Progressbar(root, mode="indeterminate")
    progress.pack(fill=tk.X, padx=20, pady=5)

    # Status window
    tk.Label(root, text="Status Output:").pack(anchor="w", padx=20)
    text_status = scrolledtext.ScrolledText(root, height=15)
    text_status.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

    root.mainloop()


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
if __name__ == "__main__":
    launch_gui()
