#!/usr/bin/env python3
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<     Copyright        >>>>>>>>>>>>>>>>>>>>>>>>>>

# Copyright (C) 2024 LincolnLandForensics
#
# This program is free software; you can redistribute it and/or modify it under
# the terms of the GNU General Public License version 2, as published by the
# Free Software Foundation
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details (http://www.gnu.org/licenses/gpl.txt).


'''
https://github.com/ufrisk/MemProcFS
cd C:\Forensics\scripts\python\git-repo\MemProcFS_files_and_binaries_v5.8.1-win_x64-20230910
.\MemProcFS.exe -device C:\temp\memdump.raw  -forensic 1
maps it all as M:\

pulls out all the cool bits and throw them into a spreadsheet for a quick triage and case notes
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import csv
import sys
import time
import openpyxl

# import argparse  # for menu system
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build
    # print(f'major version = {major_version} Build= {build_version} {version_info}')   # temp

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "process a memproc"
version = '0.0.2'


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global row
    row = 0  # defines arguments

    memprocfs()
    csv_dumps()
    
    
    if 'Sheet' in workbook.sheetnames:
        # Get a reference to the "Sheet" sheet
        sheet_to_delete = workbook['Sheet']
    
    workbook.save(outuput_xlsx)

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def cls():
    linux = 'clear'
    windows = 'cls'
    os.system([linux, windows][os.name == 'nt'])
    
def csv_dumps():
    # Define the paths
    csv_files = [
        r'M:\forensic\csv\findevil.csv',
        r'M:\forensic\csv\timeline_web.csv',
        r'M:\forensic\csv\timeline_net.csv',
        r'M:\forensic\csv\timeline_process.csv',
        r'M:\forensic\csv\timeline_task.csv',
        r'M:\forensic\csv\files.csv',        
        r'M:\forensic\csv\process.csv',          
        r'M:\forensic\csv\services.csv'
    ]

    # Iterate through each CSV file and create a worksheet for each
    for csv_file_path in csv_files:
        try:
            # Extract the sheet name from the file name (excluding the file extension)
            sheet_name = os.path.splitext(os.path.basename(csv_file_path))[0]
            # Read the CSV file and populate the worksheet
            with open(csv_file_path, 'r', newline='', encoding='utf-8') as csv_file:
                csv_reader = csv.reader(csv_file)
                worksheet = workbook.create_sheet(title=sheet_name)  # Create a new worksheet
                # active_sheet.freeze_panes = 'B2'
                for row in csv_reader:
                    worksheet.append(row)

                # Set the "B2" cell as the top-left cell of the frozen pane
                worksheet.freeze_panes = 'B2'

                # Set the width of the G column to 100
                worksheet.column_dimensions['G'].width = 100
        except Exception as e:
            print(f"An error occurred: {str(e)}")
        # Set the "B2" cell as the top-left cell of the frozen pane
        # worksheet.freeze_panes = 'B2'

        # Set the width of the G column to 100
        # worksheet.column_dimensions['G'].width = 100

    # Remove the default sheet created when the workbook was initialized
    # workbook.remove(workbook.active)

    print(f'CSV data has been saved to "{outuput_xlsx}" in separate sheets.')
    
    
def memprocfs():
    global outuput_xlsx
    outuput_xlsx = ('output_memprocfs.xlsx')
    (OS, currentTime, timezone, userName, tempNotes, hostname) = ('', '', '', '', 'tempNotes', '')
    (notes, ip) = ('', '')
    data = []
    # Your headers list
    headers = [
        "caseNumber", "exhibit", "caseName", "subjectBusinessName", "caseType", "caseAgent", 
        "forensicExaminer", "reportStatus", "notes", "summary", "exhibitType", "makeModel", 
        "serial", "OS", "phoneNumber", "phoneIMEI", "mobileCarrier", "biosTime", "currentTime", 
        "timezone", "shutdownMethod", "shutdownTime", "userName", "userPwd", "email", "emailPwd", 
        "ip", "seizureAddress", "seizureRoom", "dateSeized", "seizedBy", "dateReceived", "receivedBy", 
        "removalDate", "removalStaff", "reasonForRemoval", "inventoryDate", "seizureStatus", "status", 
        "imagingTool", "imagingType", "imageMD5", "imageSHA1", "imageSHA256", "writeBlocker", 
        "imagingStarted", "imagingFinished", "storageType", "storageMakeModel", "storageSerial", 
        "storageSize", "evidenceDataSize", "analysisTool", "analysisTool2", "exportLocation", 
        "exportedEvidence", "storageLocation", "caseNumberOrig", "priority", "operation", "Action", 
        "vaultCaseNumber", "qrCode", "vaultTotal", "tempNotes", "hostname", "phoneimei2"
    ]

    # Create a dictionary where keys are column names and values are the corresponding headers
    headers_dict = {col_name: col_name for col_name in headers}

# progress loop
    global progress_file_path
    # progress_file_path = r"M:\misc\procinfo\progress_percent.txt"
    progress_file_path = r"M:\forensic\progress_percent.txt"
    
    while True:
        progress = read_progress()
        if progress is not None:
            cls()
            print(f"Progress: {progress}%")
            if progress >= 100:
                break
        else:
            print("Error reading progress file.")
        
        time.sleep(10)

# OS
    file_OS = r"M:\sys\version.txt"

    try:
        with open(file_OS, 'r') as file:
            OS = (f'Windows {file.read().strip()}')
            print(f"OS:   {OS}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")



# currentTime
    file_currentTime = r"M:\sys\time-current.txt"

    try:
        with open(file_currentTime, 'r') as file:
            currentTime = file.read().strip()
            notes = (f'{notes}A memory dump was performed at {currentTime}.\n')
            print(f"currentTime:   {currentTime}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# timezone
    file_timezone = r"M:\sys\timezone.txt"
    try:
        with open(file_timezone, 'r') as file:
            timezone = file.read().strip()
            print(f"timezone:   {timezone}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        
# hostname file path
    file_hostname = r"M:\registry\HKLM\SYSTEM\ControlSet001\Control\ComputerName\ComputerName\ComputerName.txt"
    # M:\sys\computername.txt
    try:
        # Open the file in read mode
        with open(file_hostname, 'r') as file:
            # Read all lines into a list
            lines = file.readlines()

            # Check if there are at least three lines in the file
            if len(lines) >= 3:
                # Extract the third line (index 2)
                hostname = lines[2].strip()
                print(f"hostname:   {hostname}")

    except FileNotFoundError:
        print(f"Hostname file not found:    {file_hostname}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

    # Append the updated hostname to the data list
    # data.append({"hostname": hostname})

# userName
    file_user = r"M:\sys\users\users.txt"
    # Initialize a variable to store the file contents

    try:
        # Open the file in read mode
        with open(file_user, 'r') as file:
            userName = file.read().strip()
            print(f"userName:\n{userName}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    # data.append({"userName": userName})

# IP
    file_path_ip = r'M:\py\reg\net\tcpip_interfaces.txt'

    # Initialize a dictionary to store DhcpIPAddress values
    dhcp_ip_addresses = {}

    try:
        # Open the file for reading
        with open(file_path_ip, 'r') as file:
            lines = file.readlines()
            
            # Iterate through each line
            for line in lines:
                # Check if the line contains "DhcpIPAddress"
                if "DhcpIPAddress:" in line:
                    # Split the line by ":" and get the value part
                    parts = line.split(':')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        value = parts[1].strip()
                        dhcp_ip_addresses[key] = value
                        if not value.endswith('.1'):
                            ip = (f'{ip}, {value}')                         
                            # print(f'{value}')

    except FileNotFoundError:
        print(f"File not found: {file_path_ip}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    ip = ip.strip().lstrip(', ')
    print(f'ip: {ip}')
# tempNotes -web history
    file_web = r"M:\misc\web\web.txt"

    try:
        with open(file_web, 'r') as file:
            web = file.read().strip()
            tempNotes = (f'Web history: \n%s\n\n%s\n' %(web, tempNotes))

            # print(f"tempNotes:   {tempNotes}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# tempNotes -usb storage
    file_usb_store = r"M:\py\reg\usb\usb_devices.txt"

    try:
        with open(file_usb_store, 'r') as file:
            storage = file.read().strip()
            storage = str(storage)
            # print(f'{storage}')
            print('\nSee tempNotes for list of usb storage devices\n')
            tempNotes = (f'\n{tempNotes}\n usb storage: \n{storage}\n\n') 

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        
# notes -bluetooth  # task
    file_bluetooth = r"M:\py\reg\net\bth_devices.txt"
    # notes = (f'The following bluetooth devices were found in the memory dump (M:\py\reg\net\bth_devices.txt): \n %s' %(notes))
  
    try:
        with open(file_bluetooth, 'r') as filebth:
            bluetooth = filebth.read().strip()
            # print(f'\nSee notes for list of bluetooth devices\n')
            # notes = (r'bluetooth: M:\py\reg\net\bth_devices.txt\n%s\n\n%s\n' %(notes, bluetooth))
            # notes = (f'{notes}\n\nbluetooth: M:\py\reg\net\bth_devices.txt\n{bluetooth}\n')
    except Exception as e:
        print(f"An error occurred: {str(e)}")  
    notes = notes.strip()
    
# bitlocker fkey # task

    # Define the directory path
    bitlocker_path = r'M:\misc\bitlocker'

    # Define the file extension to look for
    file_extension = '.fvek'

    # List all files in the directory
    file_list = os.listdir(bitlocker_path)

    # Filter the files with the specified extension
    filtered_files = [file for file in file_list if file.endswith(file_extension)]

    # Check if any files with the extension were found
    if not filtered_files:
        print(f"No bitlocker files {file_extension} were found in {bitlocker_path}")
    else:
        print(f"Files with the bitlocker {file_extension} extension found in {bitlocker_path}:")
        tempNotes = (f'file: %s\n%s' %(file, tempNotes))
        for file in filtered_files:
            print(os.path.join(bitlocker_path, file))
            tempNotes = (f'%s\n%s' %(tempNotes, os.path.join(bitlocker_path, file)))

# run statements
    # Define the directory path where the .txt files are located
    run_path = r'M:\registry\HKLM\SOFTWARE\Microsoft\Windows\CurrentVersion\Run' 
    runs = ('M:\registry\HKLM\SOFTWARE\Microsoft\Windows\CurrentVersion\Run\n')
    # List all files in the directory
    run_file_list = os.listdir(run_path)

    # Filter the files to only include .txt files
    txt_files = [file for file in run_file_list if file.endswith('.txt')]
    notes = (f'{notes}\n    The following registry run statements were found in the memory dump (HKLM\SOFTWARE\Microsoft\Windows\CurrentVersion\Run): \n')
  
    # Iterate through each .txt file and print the third line
    for txt_file in txt_files:
        file_path = os.path.join(run_path, txt_file)
        
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                
                # Check if the file has at least three lines
                if len(lines) >= 3:
                    third_line = lines[2]  # Index 2 corresponds to the third line (0-based index)
                    # print(f"File: {txt_file}, Third Line: {third_line.strip()}")
                    print(f"Run command: {third_line.strip()}")
                    notes = (f'{notes}\n{third_line.strip()}')

        except FileNotFoundError:
            print(f"File not found: {file_path}")
        except Exception as e:
            print(f"An error occurred while processing {txt_file}: {str(e)}")
    notes = notes.strip()
    print(notes)    # temp
    # append data
    data.append({"notes": notes, "OS": OS, "currentTime": currentTime, "timezone": timezone, "userName": userName, "ip": ip, "tempNotes": tempNotes, "hostname": hostname})

    write_xlsx(data)

def read_progress():
    try:
        with open(progress_file_path, 'r') as file:
            return float(file.read().strip())
    except (FileNotFoundError, ValueError):
        return None
        
def write_xlsx(data):
    '''
    The write_xlsx() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''
    
    
    # print("Received data:", data)  # Debugging output
    # print(f'{color_blue}Received data: {color_yellow}{data}{color_reset}')  # Debugging output
    
    # Sort the data by the "exhibit" column
    # sorted_data = sorted(data, key=lambda x: (isinstance(x.get("exhibit"), int), x.get("exhibit")))

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active
    # workbook.remove(workbook.active)
    worksheet = workbook.create_sheet(title='memprocfs')  # Create a new worksheet
    active_sheet = workbook.active
    active_sheet.freeze_panes = 'B2'
    # Delete the default sheet
    # workbook.remove('Sheet')
    
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'C2'  # Freeze cells
    worksheet.selection = 'B2'

    # headers = data[0].keys()  # Get the keys (headers) from the first row of data

    headers = [
        "caseNumber", "exhibit", "caseName", "subjectBusinessName", "caseType", "caseAgent", 
        "forensicExaminer", "reportStatus", "notes", "summary", "exhibitType", "makeModel", 
        "serial", "OS", "phoneNumber", "phoneIMEI", "mobileCarrier", "biosTime", "currentTime", 
        "timezone", "shutdownMethod", "shutdownTime", "userName", "userPwd", "email", "emailPwd", 
        "ip", "seizureAddress", "seizureRoom", "dateSeized", "seizedBy", "dateReceived", "receivedBy", 
        "removalDate", "removalStaff", "reasonForRemoval", "inventoryDate", "seizureStatus", "status", 
        "imagingTool", "imagingType", "imageMD5", "imageSHA1", "imageSHA256", "writeBlocker", 
        "imagingStarted", "imagingFinished", "storageType", "storageMakeModel", "storageSerial", 
        "storageSize", "evidenceDataSize", "analysisTool", "analysisTool2", "exportLocation", 
        "exportedEvidence", "storageLocation", "caseNumberOrig", "priority", "operation", "Action", 
        "vaultCaseNumber", "qrCode", "vaultTotal", "tempNotes", "hostname", "phoneimei2"
    ]

    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [13, 18, 19, 22, 23, 26, 64, 65]:  # yellow_columns 'B', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'BN', 'BO'
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = fill

        # elif col_index in [0, 2, 3, 4, 5, 6, 7]:  # Orange Indices of columns A, C, D, E, F, G, H
            # fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            # cell.fill = fill
        elif col_index in [8]:  # violet_columns A, C, D, E, F, G, H
            fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
            cell.fill = fill
        # elif col_index in [0, 2, 3, 4, 5, 6, 7]:  # green_columns A, C, D, E, F, G, H
            # fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            # cell.fill = fill
        # elif col_index in [0, 2, 3, 4, 5, 6, 7]:  # blue_columns A, C, D, E, F, G, H
            # fill = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")
            # cell.fill = fill
        # elif col_index in [0, 2, 3, 4, 5, 6, 7]:  # pink_columns A, C, D, E, F, G, H
            # fill = PatternFill(start_color="FF99FF", end_color="FF99FF", fill_type="solid")
            # cell.fill = fill
            
    # Write headers to the first row
    # for col_index, header in enumerate(headers):
        # worksheet.cell(row=1, column=col_index + 1).value = header


    # Excel Header row
    # worksheet.cell(row=1, column=1).value = 'caseNumber_blah'   # todo

    # Excel column width
    worksheet.column_dimensions['A'].width = 15# caseNumber
    worksheet.column_dimensions['B'].width = 7# exhibit
    worksheet.column_dimensions['C'].width = 16# caseName
    worksheet.column_dimensions['D'].width = 20# subjectBusinessName
    worksheet.column_dimensions['E'].width = 16# caseType
    worksheet.column_dimensions['F'].width = 25# caseAgent
    worksheet.column_dimensions['G'].width = 15# forensicExaminer
    worksheet.column_dimensions['H'].width = 13# reportStatus
    worksheet.column_dimensions['I'].width = 25# notes
    worksheet.column_dimensions['J'].width = 15# summary
    worksheet.column_dimensions['K'].width = 12# exhibitType
    worksheet.column_dimensions['L'].width = 30# makeModel
    worksheet.column_dimensions['M'].width = 17# serial
    worksheet.column_dimensions['N'].width = 15# OS
    worksheet.column_dimensions['O'].width = 14# phoneNumber
    worksheet.column_dimensions['P'].width = 16# phoneIMEI
    worksheet.column_dimensions['Q'].width = 15# mobileCarrier
    worksheet.column_dimensions['R'].width = 16# biosTime
    worksheet.column_dimensions['S'].width = 16# currentTime
    worksheet.column_dimensions['T'].width = 12# timezone
    worksheet.column_dimensions['U'].width = 15# shutdownMethod
    worksheet.column_dimensions['V'].width = 16# shutdownTime
    worksheet.column_dimensions['W'].width = 12# userName
    worksheet.column_dimensions['X'].width = 12# userPwd
    worksheet.column_dimensions['Y'].width = 20# email
    worksheet.column_dimensions['Z'].width = 12# emailPwd
    worksheet.column_dimensions['AA'].width = 14# ip
    worksheet.column_dimensions['AB'].width = 15# seizureAddress
    worksheet.column_dimensions['AC'].width = 12# seizureRoom
    worksheet.column_dimensions['AD'].width = 16# dateSeized
    worksheet.column_dimensions['AE'].width = 12# seizedBy
    worksheet.column_dimensions['AF'].width = 16# dateReceived
    worksheet.column_dimensions['AG'].width = 15# receivedBy
    worksheet.column_dimensions['AH'].width = 16# removalDate
    worksheet.column_dimensions['AI'].width = 25# removalStaff
    worksheet.column_dimensions['AJ'].width = 18# reasonForRemoval
    worksheet.column_dimensions['AK'].width = 15# inventoryDate
    worksheet.column_dimensions['AL'].width = 18# seizureStatus
    worksheet.column_dimensions['AM'].width = 12# status
    worksheet.column_dimensions['AN'].width = 24# imagingTool
    worksheet.column_dimensions['AO'].width = 15# imagingType
    worksheet.column_dimensions['AP'].width = 16# imageMD5
    worksheet.column_dimensions['AQ'].width = 15# imageSHA1
    worksheet.column_dimensions['AR'].width = 15# imageSHA256  #25
    worksheet.column_dimensions['AS'].width = 15# writeBlocker
    worksheet.column_dimensions['AT'].width = 22# imagingStarted
    worksheet.column_dimensions['AU'].width = 16# imagingFinished
    worksheet.column_dimensions['AV'].width = 13# storageType
    worksheet.column_dimensions['AW'].width = 23# storageMakeModel
    worksheet.column_dimensions['AX'].width = 19# storageSerial
    worksheet.column_dimensions['AY'].width = 14# storageSize
    worksheet.column_dimensions['AZ'].width = 15# evidenceDataSize
    worksheet.column_dimensions['BA'].width = 23# analysisTool
    worksheet.column_dimensions['BB'].width = 15# analysisTool2
    worksheet.column_dimensions['BC'].width = 25# exportLocation
    worksheet.column_dimensions['BD'].width = 15# exportedEvidence
    worksheet.column_dimensions['BE'].width = 20# storageLocation
    worksheet.column_dimensions['BF'].width = 19# caseNumberOrig
    worksheet.column_dimensions['BG'].width = 9# priority
    worksheet.column_dimensions['BH'].width = 15# operation
    worksheet.column_dimensions['BI'].width = 10# Action
    worksheet.column_dimensions['BJ'].width = 19# vaultCaseNumber
    worksheet.column_dimensions['BK'].width = 15# qrCode
    worksheet.column_dimensions['BL'].width = 15# vaultTotal
    worksheet.column_dimensions['BM'].width = 40# tempNotes



    for row_index, row_data in enumerate(data):
        # print(f'{color_purple}Processing row: {row_data}{color_reset}')  # Debugging output
        # print("row_index = %s" %(row_index))
        # print(f'{color_red}row = {row_index+1}{color_reset}')        

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")




        # for col_index, cell_data in enumerate(row_data):
            # try:
                # worksheet.cell(row=row_index+1, column=col_index+1).value = row_data
            # except Exception as e:
                # print(f"{color_red}Error printing line: {str(e)}{color_reset}")

    # worksheet.cell(row=7, column=8).value = 'hello world'
    
    workbook.save(outuput_xlsx)
 
                
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
delete "Sheet"

bluetooth is broken
            if c == 1:  # fault tolerance if there aren't enough rows
                one = worksheet.cell_value(r, 1)

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
https://www.youtube.com/watch?v=hjWVUrf7Obk
https://github.com/ufrisk/MemProcFS



"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

