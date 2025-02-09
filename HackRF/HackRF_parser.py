 #!/usr/bin/env python3
# coding: utf-8
'''

'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

# import re
import os
import re
import csv
import sys
import glob  # Import the glob module for pattern matching
import shutil   
import binascii
from datetime import datetime
import argparse
import openpyxl # pip install openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Colorize section
global color_red, color_yellow, color_green, color_blue, color_purple, color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build


    if major_version >= 10 and build_version >= 22000:
        import colorama # pip install colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}')
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL
        
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>
author = 'LincolnLandForensics'
description = "parse HackRF Mac Address logs"
version = '0.1.4'

global hackRF_drive
hackRF_drive= 'H'

global logfolder
logfolder=r'D:\Forensics\scripts\python\Logs'


global companies
companies = {
        "74:EC:B2": "Amazon Technologies Inc.",
        "E4:F0:42": "Google, Inc.",
        "A0:D7:F3": "Samsung Electronics Co.,Ltd",
        "D4:3A:2C": "Google, Inc."
    }


# <<<<<<<<<<<<<<<<<<<<<<<<<<       Main        >>>>>>>>>>>>>>>>>>>>>>>>>>
def main():
    # global row
    # row = 0
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-b', '--blank', help='blank sheet', required=False, action='store_true')
    parser.add_argument('-C', '--clear', help='clear logs off the HackRF', required=False, action='store_true')    
    parser.add_argument('-L', '--logs', help='log grabber', required=False, action='store_true')
    parser.add_argument('-p', '--parse', help='parse text', required=False, action='store_true')    
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)

    args = parser.parse_args()

    global output_xlsx
    if not args.output: 
        output_xlsx = f"output_HackRF.xlsx"        
    else:
        output_xlsx = args.output    
    
    global input_folder
    if not args.input: 
        # input_folder = "logs"  
        input_folder = "D:\Forensics\scripts\python\Logs" 

        
    else:
        input_folder = args.input

    if args.blank:
        data = []
        data2 = []
        write_xlsx(data,data2)
    elif args.logs:
        log_grab()
    elif args.clear:
        clear_logs()        
    elif args.parse:
        companies_read()
        output_excel_path = 'output_pdfs.xlsx'


        if not os.path.exists(input_folder):
            print(f"Error: The directory '{input_folder}' does not exist.")
            # return
        else:
            msg_blurb = (f"Reading files in {input_folder}")
            msg_blurb_square(msg_blurb, color_green)             
        text_parse(input_folder, output_excel_path)

    else:
        usage()

    # return 0

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def BLE_Data_Translate(data):
    length, ad_type, company_id, raw_data = ('', '', '', '')
    
    try:
        # Fix odd-length hex strings by padding (if needed)
        if len(data) % 2 != 0:
            data = "0" + data  # Add leading zero   
            
        # Convert hex string to raw bytes (binary data)
        raw_bytes = binascii.unhexlify(data)

        # Extract key parts of the packet
        length = raw_bytes[0]  # First byte: Length
        ad_type = raw_bytes[1]  # Second byte: Advertising type (should be 0xFF for Manufacturer Data)
        company_id = int.from_bytes(raw_bytes[2:4], byteorder="little")  # Convert 2-byte ID to int

        payload = raw_bytes[4:]  # Remaining data (MAC, UUID, or sensor values)
        raw_data = payload

        # print(f"Raw Payload: {payload.hex().upper()}")  # Convert binary payload to hex string
        raw_data = payload.hex().upper()   # temp
    except Exception as e:
        # print(f"Error processing data '{data}': {e}")
        print(f'{e}')
    
    return length, ad_type, company_id, raw_data

def clear_logs():
    # Prompt user for drive letter, use default if left blank
    drive = input(f"\nIf you continue, you will delete HackRF logs!\n\nEnter the drive letter where HackRF is plugged in (default is {hackRF_drive}): \n").strip().lower()
    if not drive:
        drive = hackRF_drive.lower()

    # Paths for .csv and .txt files
    source_csv = f"{drive.upper()}:/BLERX/Lists/*.csv"
    source_txt_1 = f"{drive.upper()}:/BLERX/LOGS/*.TXT"
    source_txt_2 = f"{drive.upper()}:/LOGS/*.TXT"
    
    # Delete .csv files
    for file in glob.glob(source_csv):
        os.remove(file)  # Remove the file
        print(f"Deleted: {file}")

    # Delete .txt files from BLERX/LOGS
    for file in glob.glob(source_txt_1):
        os.remove(file)  # Remove the file
        print(f"Deleted: {file}")
        
    # Delete .txt files from LOGS
    for file in glob.glob(source_txt_2):
        os.remove(file)  # Remove the file
        print(f"Deleted: {file}")

    # Optional: Display a message indicating the logs have been deleted
    msg_blurb = (f'HackRF logs have been deleted from the drive {drive.upper()}')
    msg_blurb_square(msg_blurb, color_green)

    
def company_lookup(mac_address):
    company = ''

    # Extract the first 8 characters of mac_address
    mac_prefix = mac_address[:8]  # Get the first 8 characters

    # Return the company name if found, else an empty string
    return companies.get(mac_prefix, "")    


def companies_read():
    # Check if the CSV file exists
    csv_file = 'mac-vendors-export.csv'
    if os.path.exists(csv_file):
        # Read the CSV file and update the companies dictionary
        with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                mac_prefix = row[0]
                vendor_name = row[1]
                companies[mac_prefix] = vendor_name

                
def iso8601_timestamp(timestamp):
    # Convert string to datetime object
    try:
        dt = datetime.strptime(timestamp, "%Y%m%d%H%M%S")

        # Convert to ISO 8601 format
        iso_timestamp = dt.isoformat()
        return iso_timestamp
    except Exception as e:
        # print(f"Error processing timestamp {timestamp}: {e}")
        return timestamp

def log_grab():
    # Prompt user for drive letter, use default if left blank
    drive = input(f"\nEnter the drive letter where HackRF is plugged in (default is {hackRF_drive}): \n").strip().lower()
    if not drive:
        drive = hackRF_drive.lower()

    # Paths for .csv and .txt files
    source_csv = f"{drive.upper()}:/BLERX/Lists/*.csv"
    source_txt_1 = f"{drive.upper()}:/BLERX/LOGS/*.TXT"
    source_txt_2 = f"{drive.upper()}:/LOGS/*.TXT"
    
    # Create the log folder if it doesn't exist
    if not os.path.exists(logfolder):
        os.makedirs(logfolder)

    # Copy .csv files to log folder
    for file in glob.glob(source_csv):
        shutil.copy(file, logfolder)
        print(f"Copied: {file}")

    # Copy .txt files from BLERX/LOGS to log folder
    for file in glob.glob(source_txt_1):
        shutil.copy(file, logfolder)
        print(f"Copied: {file}")
        
    # Copy .txt files from LOGS to log folder
    for file in glob.glob(source_txt_2):
        shutil.copy(file, logfolder)
        print(f"Copied: {file}")

    msg_blurb = (f'HackRF logs copied to {logfolder}')
    msg_blurb_square(msg_blurb, color_green)        

        
def msg_blurb_square(msg_blurb, color):
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')
  
def packet_lookup(msg_type):
    # Dictionary of message types and corresponding descriptions
    packet_details = {
        "ADV_IND": "Scannable advertising indication for all devices",
        "ADV_DIRECT_IND": "Directed advertising to a specific device indicating that only that device can connect",
        "ADV_NONCONN_IND": "Advertising indication but not accepting connections or scans",
        "SCAN_REQ": "Sent by a device to receive more info from scannable advertisers",
        "SCAN_RSP": "Response to 03 containing more info",
        "CONNECT_REQ": "Sent to an advertiser to initiate a connection",
        "ADV_SCAN_IND": "Scannable advertising indication but not accepting connections",
        "ADV_EXT_IND": "BT 5.0, points to additional data on secondary channels",
        "AUX_ADV_IND": "BT 5.0, scannable advertising indication for all devices on secondary channels",
        "AUX_SCAN_REQ": "BT 5.0, sent by a device to receive more info from scannable advertisers on secondary channels",
        "AUX_SCAN_RSP": "BT 5.0, response to 0A containing more info",
        "AUX_CONNECT_REQ": "BT 5.0, sent to an advertiser to initiate a connection",
        "AUX_CHAIN_IND": "BT 5.0, chains advertising packets together when they get too big",
        "AUX_CONNECT_RSP": "BT 5.0, response to 0C",
        "Link Layer": "Link layer setup and operation",
        "Data": "General data and proprietary"
    }

    # Check if the msg_type exists in the dictionary, if yes return the description
    return packet_details.get(msg_type, "")

    
def text_parse(input_folder, output_excel_path):
    data = []
    data2 = []
    data3 = []
    mac_uniq = []
    
    # Read and process each .txt file in the folder
    for file_name in os.listdir(input_folder):
        file_path = os.path.join(input_folder, file_name)
        print(f'reading {file_name}')    # temp

            # file_name = 'mac-vendors-export.csv'
      
        
        
        if file_name.lower().endswith('.csv') and os.path.exists(file_path):
            # with open(file_path, mode='r', newline='', encoding='utf-8') as file:
            with open(file_path, mode='r', newline='', encoding='ISO-8859-1') as file:
                reader = csv.reader(file)
                # next(reader)  # Skip the header row
                for row in reader:
                    timestamp, mac_address, name, msg_type, dta, hits = '', '', '', '', '', ''
                    db, channel, length, company, device_type, p_description = '', '', '', '', '', ''
                    whitelist = ''

                    row_data = {}
                    row_data2 = {}
                    row_data3 = {}

                    timestamp = row[0]
                    mac_address = row[1]
                    name = row[2]
                    msg_type = row[3]
                    dta = row[4]
                    hits = row[5]                    
                    db = row[6]                    
                    channel = row[7]  
                    mac_address = mac_address[:17]
                    
                    try:
                        blah = 'blah'
                        # timestamp = row[0]
                        # mac_address = row[1]
                        # name = row[2]
                        # msg_type = row[3]
                        # dta = row[4]
                        # hits = row[5]                    
                        # db = row[6]                    
                        # channel = row[7]  
                        # mac_address = mac_address[:17]                        
                    except Exception as e:
                        print(f"Error processing line : {e}")

                    if name.startswith('0x1'):
                        name = ''
                    company = company_lookup(mac_address)
                    if msg_type != '':
                        p_description = packet_lookup(msg_type)                        
                    (whitelist) =  whitelist_check(mac_address, name)
                    if mac_address not in mac_uniq and "imestamp" not in timestamp:
                        # (whitelist) =  whitelist_check(mac_address, name)
                        mac_uniq.append(mac_address)

                        row_data2["Time"] = timestamp
                        row_data2["MAC Address"] = mac_address
                        row_data2["Name"] = name
                        # row_data2["Packet Type"] = msg_type
                        # row_data2["Data"] = dta
                        row_data2["Hits"] = hits
                        row_data2["dB"] = db
                        row_data2["Channel"] = channel  
                        # row_data2["length"] = length
                        row_data2["company"] = company
                        row_data2["Device Type"] = device_type
                        row_data2["origin_file"] = file_name
                        row_data2["whitelist"] = whitelist    
                        
                        row_data2["file_name"] = file_name
                        row_data2["whitelist"] = whitelist
                        data2.append(row_data2)

                    if "imestamp" not in timestamp:
                        row_data["Time"] = timestamp
                        row_data["MAC Address"] = mac_address
                        row_data["Name"] = name
                        row_data["Packet Type"] = msg_type
                        row_data["Data"] = dta
                        row_data["Hits"] = hits
                        row_data["dB"] = db
                        row_data["Channel"] = channel  
                        row_data["length"] = length
                        row_data["company"] = company
                        row_data["Device Type"] = device_type
                        row_data["origin_file"] = file_name
                        row_data["Packet Description"] = p_description    
                        row_data["file_name"] = file_name
                        row_data["whitelist"] = whitelist
                        data.append(row_data)                            
                            
                            
        elif file_name.lower().endswith('.txt'):  # Check for .txt extension
            with open(os.path.join(input_folder, file_name), "r") as file:
                for line in file:
                    line = line.strip()
                    
                    row_data = {}
                    row_data2 = {}
                    row_data3 = {}
                    (timestamp, mac_address, name, msg_type, dta, hits)  = ('', '', '', '', '', '')
                    (db, channel, length, company, device_type, p_description) = ('', '', '', '', '', '')
                    (length2, Type, company_id, raw_data) = ('', 'plane', '', '')
                    (whitelist) = ('')
                    parts = line.strip().split(" ")
                    
                    if line.strip() == '':
                        blah = 'blah'
                    elif file_name == 'ADSB.TXT':
                        (timestamp, Latitude, Longitude, Coordinate, Type, origin_file) = ('', '', '', '', 'ADS-B', '')
                        (Plate, Direction, Icon, ICAO, name, note, Altitude) = ('', '', '', '', '', '', '')
                        (hit, lvl, speed, amp, age, Callsign) = ('', '', '', '', '', '')
                        (HexID) = ('')
                        
                        match1 = re.search(r'Alt:(\d+)', line)
                        match2 = re.search(r'Lat:([-+]?\d*\.\d+|\d+)', line)
                        match3 = re.search(r'Lon:([-+]?\d*\.\d+|\d+)', line)

                        if match1:
                            Altitude = int(match1.group(1)) 

                        if match2:
                            Latitude = str(match2.group(1))

                        if match3:
                            Longitude = float(match3.group(1))

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = parts[1] if len(parts) > 1 else ""
                            name = parts[2] if len(parts) > 2 else ""
                            if 'Alt:' in name:
                                name = ''
                            timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        except IndexError:
                            print(f"Skipping malformed line: {line.strip()}")
                        if 'ICAO:' in dta:
                            HexID = dta.split('ICAO:')[1]
                            p_description = HexID
                            dta = dta.split('ICAO:')[0] 
                            
                        Callsign = name
                        
                        row_data["Time"] = timestamp
                        row_data["Name"] = name
                        row_data["Packet Description"] = p_description    
                        row_data["Data"] = dta  
                        row_data["origin_file"] = file_name
                        row_data["Type"] = Type                         
                        row_data["Latitude"] = Latitude
                        row_data["Longitude"] = Longitude
                        row_data["Altitude"] = Altitude                        

                        row_data["Callsign"] = Callsign
                        row_data["Hits"] = hits                        
                        row_data["dB"] = db                        
                        data.append(row_data)

                    elif file_name == 'APRS.TXT':
                        (timestamp, name, Type, dta) = ('', '', 'APRS', '')

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = parts[1] if len(parts) > 1 else ""
                            name = parts[2] if len(parts) > 2 else ""
                        except IndexError:
                            print(f"Skipping malformed line: {line.strip()}")

                        row_data["Time"] = timestamp
                        row_data["Name"] = name
                        row_data["Data"] = dta  
                        row_data["origin_file"] = file_name
                        row_data["Type"] = Type                         
                        data.append(row_data)

                    elif file_name == 'TPMS.TXT':
                        (timestamp, name, Type, dta, company) = ('', '', 'TPMS', '', '')

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            dta = line.strip()
                            if '  ' in dta:
                                dta = dta.split('  ')[1]
                            name = parts[6] if len(parts) > 6 else ""
                            company = parts[7] if len(parts) > 7 else ""
                            if '/' in company:
                                company = company.split('/')[0]
                        except IndexError:
                            print(f"Skipping malformed line: {line.strip()}")
                        
                        if ' FSK ' in line:
                            p_description = "FSK"
                        timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        
                        row_data["Time"] = timestamp
                        row_data["Name"] = name
                        row_data["Data"] = dta  
                        row_data["origin_file"] = file_name
                        row_data["Packet Description"] = p_description 
                        row_data["company"] = company 
                        row_data["Type"] = Type                         
                        data.append(row_data)
                    
                    elif len(parts) >= 5:

                        try:
                            timestamp = parts[0] if len(parts) > 0 else ""
                            msg_type = parts[1] if len(parts) > 1 else ""
                            length = parts[2].split(":")[1] # if len(parts) > 3 and ":" in parts[3] else ""

                            mac_address = parts[3].lstrip('Mac:')  # .split(":")[1] if len(parts) > 5 and ":" in parts[5] else ""
                            dta = parts[4].lstrip('Data:') # .split(":")[1] if len(parts) > 7 and ":" in parts[7] else ""

                            timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        except IndexError:
                            print(f"Skipping malformed line: {line.strip()}")

                        if msg_type != '':
                            p_description = packet_lookup(msg_type)

                        if mac_address not in mac_uniq:
                            mac_uniq.append(mac_address)
                        company = company_lookup(mac_address)    
                        # if mac_address not in mac_uniq:
                            # company = company_lookup(mac_address)

                        length2, ad_type, company_id, raw_data = BLE_Data_Translate(dta)
                        
                        if mac_address not in mac_uniq and "imestamp" not in timestamp:
                            mac_uniq.append(mac_address)
                            (whitelist) =  whitelist_check(mac_address, name)

                            row_data2["Time"] = timestamp
                            row_data2["MAC Address"] = mac_address
                            row_data2["Name"] = name
                    
                            row_data2["Hits"] = hits                        
                            row_data2["dB"] = db                        
                            row_data2["Channel"] = channel  
                            row_data2["company"] = company
                            row_data2["Device Type"] = device_type
                            row_data2["file_name"] = file_name
                            row_data2["whitelist"] = whitelist
                            data2.append(row_data2)
                        
                        
                        row_data["Time"] = timestamp
                        row_data["MAC Address"] = mac_address
                        row_data["Name"] = name
                        row_data["Packet Type"] = msg_type
                        row_data["Data"] = dta                        
                        row_data["Hits"] = hits                        
                        row_data["dB"] = db                        
                        row_data["Channel"] = channel  
                        row_data["length"] = length
                        row_data["company"] = company
                        row_data["file_name"] = file_name
                        row_data["Device Type"] = device_type
                        row_data["origin_file"] = file_name
                        row_data["Packet Description"] = p_description
                        row_data["Length2"] = length2                        
                        row_data["Type"] = ad_type 
                        row_data["company_id"] = company_id 
                        row_data["raw_data"] = raw_data
                        row_data["whitelist"] = whitelist
                        data.append(row_data)

                    else:
                        (timestamp, name, Type, dta, company) = ('', '', '', '', '')

                        try:
                            dta = line.strip()
                            timestamp = parts[0] if len(parts) > 0 else ""
                        except IndexError:
                            print(f"Skipping malformed line: {line.strip()}")

                        timestamp = iso8601_timestamp(timestamp).replace('T', ' ')
                        
                        row_data["Time"] = timestamp
                        row_data["Data"] = dta  
                        row_data["origin_file"] = file_name
                       
                        data.append(row_data)                        
                        
      
    print(f'\n{len(mac_uniq)} mac addresses found')    
    write_xlsx(data, data2)


def whitelist_check(mac_address, name):
    # Check if the CSV file exists
    whitelist, description =  '', ''
    whitelist_file = 'whitelist.csv'
    if os.path.exists(whitelist_file):
        # Read the CSV file and update the companies dictionary
        with open(whitelist_file, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                mac = row[0]
                description = row[1]
                if mac_address == mac:
                    whitelist = description

    return whitelist
    
def write_xlsx(data, data2):

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'BleRX'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = ["Time", "MAC Address", "Name", "Packet Description", "Data", "Hits", "dB", "Channel", "length", "company", "Device Type", "origin_file", "Packet Type", "Length2"
    ,"Type", "company_id", "raw_data", "whitelist", "Latitude", "Longitude", "Altitude"
    ]


    # for col_index, header in enumerate(headers):
        # cell = uniq_worksheet.cell(row=1, column=col_index + 1)
        # cell.value = header
        # if col_index in range(21):  # range(18) generates numbers from 0 to 25 inclusive orange
            # fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange
            # cell.fill = fill    
    
    
    
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in range(21):  # range(18) generates numbers from 0 to 25 inclusive orange
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange
            cell.fill = fill

    # Excel column width
    worksheet.column_dimensions['A'].width = 20 # 
    worksheet.column_dimensions['B'].width = 18 # 
    worksheet.column_dimensions['C'].width = 11 # 
    worksheet.column_dimensions['D'].width = 30 # 
    worksheet.column_dimensions['E'].width = 73 # 
    worksheet.column_dimensions['F'].width = 13 #   
    worksheet.column_dimensions['G'].width = 10 # 
    worksheet.column_dimensions['H'].width = 9 # 
    worksheet.column_dimensions['I'].width = 6 # 
    worksheet.column_dimensions['J'].width = 24 # 
    worksheet.column_dimensions['K'].width = 11  # 
    worksheet.column_dimensions['L'].width = 12  # 
    worksheet.column_dimensions['M'].width = 24  # 

    worksheet.column_dimensions['S'].width = 14  # 
    worksheet.column_dimensions['T'].width = 14  # 
    worksheet.column_dimensions['U'].width = 9  # 
    
    
    
    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                if isinstance(cell_data, list):
                    cell_data = str(cell_data)  # Convert lists to strings
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")




    # Create a new worksheet for color codes
    uniq_worksheet = workbook.create_sheet(title='Uniq MACs')
    uniq_worksheet.freeze_panes = 'B2'  # Freeze cells

    # Excel column width
    uniq_worksheet.column_dimensions['A'].width = 14# 
    uniq_worksheet.column_dimensions['B'].width = 20# 


    # Excel row height

    # Excel column width
    uniq_worksheet.column_dimensions['A'].width = 20 # 
    uniq_worksheet.column_dimensions['B'].width = 18 # 
    uniq_worksheet.column_dimensions['C'].width = 15 # 
    uniq_worksheet.column_dimensions['D'].width = 17 # 
    uniq_worksheet.column_dimensions['E'].width = 5 # 
    uniq_worksheet.column_dimensions['F'].width = 5 #   
    uniq_worksheet.column_dimensions['G'].width = 10 # 
    uniq_worksheet.column_dimensions['H'].width = 9 # 
    uniq_worksheet.column_dimensions['I'].width = 6 # 
    uniq_worksheet.column_dimensions['J'].width = 24 # 
    uniq_worksheet.column_dimensions['K'].width = 11  # 
    uniq_worksheet.column_dimensions['L'].width = 12  # 
    uniq_worksheet.column_dimensions['M'].width = 24  # 

    for col_index, header in enumerate(headers):
        cell = uniq_worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in range(21):  # range(18) generates numbers from 0 to 25 inclusive orange
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange
            cell.fill = fill

    
    for row_index, row_data in enumerate(data2):
        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                if isinstance(cell_data, list):
                    cell_data = str(cell_data)  # Convert lists to strings
                uniq_worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")                

    msg_blurb = (f'Writing to {output_xlsx}')
    msg_blurb_square(msg_blurb, color_green)

    workbook.save(output_xlsx)

def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}insert your .txt files into the {input_folder} folder')
    print(f'\nExample:')
    print(f'    {file} -b')  
    print(f'    {file} -C')     
    print(f'    {file} -L')      
    print(f'    {file} -p') 
    print(f'    {file} -p -I logs -O output_.xlsx ')     

if __name__ == "__main__":

    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
0.1.3 - create a list of unique Mac addresses, output to a second sheet
0.1.2 - whitelist known good mac_address to skip those whitelist.csv
0.1.1 - -C to clear HackRF logs. -L to copy the logs off of the HackRF
0.1.0 - working copy
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
something is wrong with Uniq Mac's. it's not getting them all.

uniq Mac's works if you just run .csv files. If you also do .txt logs it is blank.

swap the uniq sheet to be first

compare mac address to https://maclookup.app/downloads/csv-database  mac-vendors-export.csv

fix '2025-02-0621:41:40' timestamps from .csv files

convert company_id to a company name

 https://macaddress.io/
 https://macaddresschanger.com/bluetooth-mac-lookup/E4:F0:42

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
https://globe.adsbexchange.com/ 

Logs are saved in: H:\BLERX\Lists\*.csv

and

Logs are saved in: H:\BLERX\Logs\*.txt

other logs are in 

H:\LOGS\*.TXT


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
