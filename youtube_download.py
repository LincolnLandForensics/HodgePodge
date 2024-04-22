#!/usr/bin/python
# coding: utf-8

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import sys
import os
import os.path
import socket
import argparse
from tkinter import *
from tkinter import messagebox
from pytube import YouTube  # import pytube.exceptions
from pytube.exceptions import AgeRestrictedError  # Import AgeRestrictedError

from youtube_transcript_api import YouTubeTranscriptApi # pip install youtube-transcript-api # for windows

import time # for sleep timer

from datetime import date, datetime
from openpyxl.styles import Font 
from openpyxl import Workbook
from openpyxl.styles import Alignment

d = datetime.today()
Day = d.strftime("%d")
Month = d.strftime("%m")
Year = d.strftime("%Y")
# todaysDate = d.strftime("%m/%d/%Y")
# todaysDate = d.strftime("%Y-%m-%d")
todaysDate = d.strftime("%Y-%m-%d %H:%M:%S")

# todaysDateTime = d.strftime("%m_%d_%Y_%H-%M-%S")
todaysDateTime = d.strftime("%Y-%m-%d_%H-%M-%S")

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "Download a list of Youtube videos from videos.txt, save list in xlsx file"
version = '1.0.6'

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global filename
    filename = 'videos.txt'
    global Row
    Row = 2
    global spreadsheet
    spreadsheet = f'videos_{todaysDateTime}.xlsx'
    global sheet_format
    sheet_format = ''

    status = internet()
    if not status:
        noInternetMsg()
        print('\nCONNECT TO THE INTERNET FIRST\n')
        exit()
    else:
        create_xlsx()

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-y', '--youtube', help='youtube download', required=False, action='store_true')

    args = parser.parse_args()

    if args.input:
        filename = args.input
    if args.output:
        outputFile = args.output

    if args.youtube:
        youtube()
    else:
        youtube()

    workbook.save(spreadsheet)
    return 0

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>
def caption_sort(captions): #beta
    print(f'captions = type(captions)') # temp
    # Sort captions based on start times
    sorted_captions = sorted(captions, key=lambda x: x['start'])
    # sorted_captions = captions

    new_format_captions = '\n'.join([f"{caption['start']} : {caption['text']}" for caption in sorted_captions])
    result_string = f"caption = [\n{new_format_captions}\n]"

    return result_string



def check_keywords_in_captions(caption, keywords):
    keyword = ''
    """
    Check for the presence of keywords in the captions.

    Args:
    - captions (str): The captions text.
    - keywords (list of str): List of keywords to check.

    Returns:
    - found_keywords (list of str): List of keywords found in the captions.
    """
    found_keywords = []
    for keyword in keywords:
        try:
            if keyword.lower() in caption.lower():
                found_keywords.append(keyword)
        except:pass
    # Convert found_keywords to a string
    found_keywords_str = ", ".join(found_keywords)

    return found_keywords_str

def create_xlsx():
    global workbook
    workbook = Workbook()
    global Sheet1
    Sheet1 = workbook.active
    Sheet1.title = 'videos'
    # header_format = openpyxl.styles.Font(bold=True)
    # header_formatUrl = openpyxl.styles.Font(bold=True, color='FFc000')  # orange Case items

    Sheet1.freeze_panes = 'B2'
    # Sheet1['B2'].selected = True

    # Excel column width
    Sheet1.column_dimensions['A'].width = 45
    Sheet1.column_dimensions['B'].width = 45
    Sheet1.column_dimensions['C'].width = 11
    Sheet1.column_dimensions['D'].width = 9
    Sheet1.column_dimensions['E'].width = 14
    Sheet1.column_dimensions['F'].width = 18
    Sheet1.column_dimensions['G'].width = 6
    Sheet1.column_dimensions['H'].width = 30
    Sheet1.column_dimensions['I'].width = 12
    Sheet1.column_dimensions['J'].width = 14
    Sheet1.column_dimensions['K'].width = 18
    Sheet1.column_dimensions['L'].width = 15
    Sheet1.column_dimensions['M'].width = 60
    Sheet1.column_dimensions['N'].width = 2
    Sheet1.column_dimensions['O'].width = 15    # keywords
    # Sheet1.column_dimensions['P'].width = 10

    # Write column headers
    Sheet1['A1'] = 'url'
    Sheet1['B1'] = 'title'
    Sheet1['C1'] = 'description'
    Sheet1['D1'] = 'views'
    Sheet1['E1'] = 'author'
    Sheet1['F1'] = 'publish_date'
    Sheet1['G1'] = 'length'
    Sheet1['H1'] = 'download_name'
    Sheet1['I1'] = 'video_id'
    Sheet1['J1'] = 'thumbnail_url'
    Sheet1['K1'] = 'dateDownloaded'
    Sheet1['L1'] = 'error'
    Sheet1['M1'] = 'caption'
    # Sheet1['N1'] = 'rating'
    Sheet1['O1'] = 'keywords'


    
def finalMessage():
    window = Tk()
    window.geometry("1x1")
    w = Label(window, text='Youtube Downloader', font="100")
    w.pack()
    print('\nSaved to current folder\n\nSaved info to %s  \n' % (spreadsheet))
    messagebox.showinfo('information', '\nDownloaded all Youtube videos from %s  \n\nSaved to current folder\n\nSaved info to %s  \n' % (filename, spreadsheet))

def internet(host="youtube.com", port=443, timeout=3):
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error as ex:
        print(ex)
        return False

def noInternetMsg():
    window = Tk()
    window.geometry("1x1")
    w = Label(window, text='Youtube Downloader', font="100")
    w.pack()
    messagebox.showwarning("Warning", "CONNECT TO THE INTERNET FIRST")

def noVideos():
    window = Tk()
    window.geometry("1x1")
    w = Label(window, text='Youtube Downloader', font="100")
    w.pack()
    messagebox.showwarning("Warning", f"you are missing youtube videos in {filename}")

def youtube_transcript(video_id):
    (caption) = ('')
    
    try:
        srt = YouTubeTranscriptApi.get_transcript(video_id)
        for i in srt:
            caption_temp = ("{}\n".format(i))    # test
            caption = (f'{caption}{caption_temp}')

    except Exception as e:
        print(f"An error occurred : {str(e)}")

    return caption


def youtube():
    keywords_to_check = ["tax", "lambo", "clubhouse was packed"]



    if not os.path.exists(filename):
        print(f"\n\n\n\t {filename} does not exist\n\n\n\t")
        noVideos()
        exit()
    else:
        print('Downloading a list of Youtube videos from %s.\n\nThis can take a bit' % (filename))
        csv_file = open(filename)

    count = 0

    for each_line in csv_file:
        (url, title, description, views, author, publish_date) = ('', '', '', '', '', '')
        (length, rating, thumbnail_url, dateDownloaded, owner, caption) = ('', '', '', '', '', '')
        (owner_id, owner_url, captions, caption, download_name, error) = ('', '', '', '', '', '')
        (video_id, keywords) = ('', '')
        
        each_line = each_line.split(',')
        each_line = each_line[0].strip()

        link = each_line

        if "youtu" in link.lower():
            count += 1

            # try:
                #                 Extract the video ID from the URL
                # video_id = extract_video_id(link)            
            # except:
                # pass

            try:
                yt = YouTube(link)
                video_id = yt.video_id

                title = yt.title
                download_name = f'{title}.mp4'
                author = yt.author
                description = yt.description
                rating = yt.rating
                publish_date = yt.publish_date
                length = str(yt.length)
                thumbnail_url = yt.thumbnail_url
                views = yt.views

                yt.streams.first().download()
                mp4_files = yt.streams.filter(file_extension="mp4")
                mp4_369p_files = mp4_files.get_by_resolution("360p")
                mp4_369p_files.download("")  # <Download folder path>
                dateDownloaded = todaysDate

                if len(video_id) > 5:
                    caption = youtube_transcript(video_id)
                    caption = caption.strip()

               # count += 1  # Increment count only if the video is successfully processed

            except AgeRestrictedError as e:
                print(f"age restricted and can't be accessed without logging in.")
                error = (f"age restricted and can't be accessed without logging in. {download_name}")
                download_name = ''
                dateDownloaded = ''
            except Exception as e:
                print(f"An error occurred while processing video {link}: {str(e)}")
                error = ('Error processing video. {download_name} {str(e)')
                download_name = ''
                dateDownloaded = ''                
        else:
            error = ('error')
            dateDownloaded = ''

        if len(video_id) > 5:
            caption = youtube_transcript(video_id)
            caption = caption.replace('{\'text\': ', '')    # .replace(', \'start\': ', '^').replace(', \'duration\': ', '^')
            # if '^' in caption:
                # captiontemp = caption.split('^')
                # caption = (f'captiontemp[0]
            # print(f'video_id = {video_id}')

            # caption = caption_sort(caption)    #test


        # keyword hunt
        # if "tax" in caption:
            # keywords = (f'tax, {keywords}')

        # Check for keywords in captions
        keywords = check_keywords_in_captions(caption, keywords_to_check)

        print(f'{link}  {title}')
        time.sleep(30) #will sleep for 30 seconds
        write_xlsx(link, title, description, views, author, publish_date, length, download_name, rating, thumbnail_url, dateDownloaded, error, caption, video_id, owner, owner_id, owner_url, keywords)


    if count == 0:
        noVideos()
        # exit()
    else:
        finalMessage()

def write_xlsx(link, title, description, views, author, publish_date, length, download_name, rating, thumbnail_url, dateDownloaded, error, caption, video_id, owner, owner_id, owner_url, keywords):

    global Row

    Sheet1.cell(row=Row, column=1, value=link)
    Sheet1.cell(row=Row, column=2, value=title)
    Sheet1.cell(row=Row, column=3, value=description)
    Sheet1.cell(row=Row, column=4, value=str(views))
    Sheet1.cell(row=Row, column=5, value=author)
    Sheet1.cell(row=Row, column=6, value=str(publish_date))
    Sheet1.cell(row=Row, column=7, value=length)
    Sheet1.cell(row=Row, column=8, value=download_name)  # download_name?
    Sheet1.cell(row=Row, column=9, value=str(video_id))
    Sheet1.cell(row=Row, column=10, value=thumbnail_url)
    Sheet1.cell(row=Row, column=11, value=dateDownloaded)

    # if download_name != '':
        # Sheet1.cell(row=Row, column=11, value=dateDownloaded)
    Sheet1.cell(row=Row, column=12, value=error)
    Sheet1.cell(row=Row, column=13, value=caption)    
    Sheet1.cell(row=Row, column=15, value=keywords)    
        
    # Save caption tracks info to cell
    # if caption:
        # print(f'caption = {caption}')
        # caption_info = f"Language: {caption.language}\nName: {caption.name}\nCode: {caption.code}"
        # Sheet1.cell(row=Row, column=13, value=caption_info)
    # Sheet1.cell(row=Row, column=14, value=rating)
    # Sheet1.cell(row=Row, column=15, value=owner_id) # not assigned
    # Sheet1.cell(row=Row, column=16, value=owner_url) # not assigned
    
    Row += 1

def usage():
    file = sys.argv[0].split('\\')[-1]
    print("\nDescription: " + description)
    print(file + f" Version: {version} by {author}")
    print("\nExample:")
    print(f"\t{file} -y\t\t")
    print(f"\t{file} -y -I videos.txt")

if __name__ == '__main__':
    main()


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>


"""

1.0.5 - transcript saved as caption, look for keywords like tax
1.0.4 - functional copy
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>


"""
download a list of all urls the user has

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>


"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
