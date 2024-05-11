#!/usr/bin/python
# coding: utf-8

'''
This script is used to translate the contents of an Excel spreadsheet from many 
languages to English. It uses the openypyl library to read the input file, googletrans  
to perform the translation, (requests as a backup module) and openypyl to write 
the translated contents to a new Excel file.
'''

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>
import re
import sys
from time import sleep
import os.path
from openpyxl import load_workbook

import requests
import argparse  # for menu system

from googletrans import Translator  # pip install googletrans>=4.0.0-rc1

# import requests.packages.urllib3
# requests.packages.urllib3.disable_warnings()  # Disable SSL verification warnings

import urllib3

# Disable SSL certificate verification
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings()

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>
author = 'LincolnLandForensics'
description = "Read input_translate.xlsx filled with another language and translate it to english"
version = '1.1.2'

# global variables
global auto_list
auto_list = ['!','?']

# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>
def main():
    output_xlsx = ('translation_.xlsx') 


    
    parser = argparse.ArgumentParser(description='Translate Excel contents from various languages to English')
    parser.add_argument('-C','--copyright', help='print copyright', required=False, action='store_true')
    parser.add_argument('-I', '--input', help='Input Excel file', required=False)
    parser.add_argument('-O', '--output', help='Output Excel file', required=False, default=output_xlsx)
    parser.add_argument('-a','--arabic', help='arabic module', required=False, action='store_true')
    parser.add_argument('-d','--detect', help='detect language only', required=False, action='store_true')
    parser.add_argument('-H','--howto', help='help module', required=False, action='store_true')
    parser.add_argument('-V','--version', help='display script and googletrans version', required=False, action='store_true')

    args = parser.parse_args()

    source_language = '' 
    input_xlsx = ('input_translate.xlsx')

    if args.copyright:  # this section might be redundant
        print(f'{color_blue}{copyright}{color_reset}')
        return 0
        input(f"{color_green}Hit any key to continue{color_reset}")
        sys.exit() 
        
    if args.howto:  # this section might be redundant
        parser.print_help()
        usage()
        return 0
        input(f"{color_green}Hit any key to continue{color_reset}")
        sys.exit() 

    if args.version:
        file = sys.argv[0].split('\\')[-1]
        print(f'{file} {version}')
        googletrans_ver()
        return 0
        input(f"{color_green}Hit any key to continue{color_reset}")
        sys.exit() 
        
    if args.input:  # defaults to input_translate.xlsx
        input_xlsx = args.input
    if args.output:  # defaults to out_english_.xlsx
        output_xlsx = args.output   

    # make sure you have a good enough version of googletrans
    googletrans_ver()

    if args.detect:
        detect_language(input_xlsx, output_xlsx)
    elif args.arabic:
        check_internet_connection()        
        source_language = 'ar'
        translate_excel(input_xlsx, output_xlsx, source_language)
    else:
        check_internet_connection()        
        translate_excel(input_xlsx, output_xlsx, source_language)

    input(f"{color_green}Hit any key to continue{color_reset}")
    sys.exit()




# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>
def check_internet_connection():
    '''
    check internet connection 
    Try to make a request to a known website
    '''
    
    try:
        response = requests.get("http://www.google.com", timeout=5)

        response.raise_for_status()  # Raise an error for any HTTP error status
        msg_blurb = 'Internet connection is available.'

    except:
        msg_blurb = 'Internet connection is not available.'
        msg_blurb_square(msg_blurb, color_red)
        exit(1)  # Exit with error status 1

def content_length(original_content):
    """
    Returns the length of the original_content.

    Args:
    original_content (str): The input string.

    Returns:
    int: The length of the input string.
    """
    if original_content is None:
        return 0
    return len(original_content)


def detect_language(input_xlsx, output_xlsx):
    '''
    read each line and detect the language
    Does not require internet access
    sort by note afterwards to move english to the bottom
    '''

    # Regular expression pattern to match words
    word_pattern = re.compile(r'\b\w+\b')
    word_pattern2 = re.compile(r'\b\w+\b', flags=re.UNICODE)

    file_exists = os.path.exists(input_xlsx)
    if file_exists == True:
        msg_blurb = (f'Reading {input_xlsx} to detect languages only')
        msg_blurb_square(msg_blurb, color_green)
        workbook = load_workbook(input_xlsx)        
        sheet = workbook.active
    else:
        msg_blurb = (f'Create {input_xlsx} and insert foreign language lines in the first column')
        msg_blurb_square(msg_blurb, color_red) 
        input(f"{color_green}Hit any key to continue{color_reset}")
        sys.exit()

    sheet.cell(row=1, column=2, value='English')
    sheet.cell(row=1, column=3, value='Language')
    sheet.cell(row=1, column=4, value='Note') 
    # sheet.cell(row=1, column=5, value='Confidence')     
    sheet.cell(row=1, column=6, value='Length')     

    for row in sheet.iter_rows(min_row=2):
        (translation, note, e, confidence) = ('', '', '', '')
        (source_language, text, skipper, length) = ('', '', '', '')
        original_content = row[0].value
        
        source_language, confidence = language_detect(original_content)
        if source_language == 'en' or source_language == 'English':
            note = ''
        elif source_language == 'auto':
            note = ''
        elif source_language == 'ar' or source_language == 'Arabic':
            note = '.'
        elif source_language == 'zh-CN' or source_language == 'Chinese (Simplified)':
            note = '.'
        elif source_language == 'zh-TW' or source_language == 'Chinese (Traditional)':
            note = '.'
        elif source_language == 'ur' or source_language == 'Urdu':
            note = '.'
        elif source_language == 'fa' or source_language == 'Persian':
            note = '.'
        else:
            note = '..'

        length = content_length(original_content)  
        
        if original_content is None:
            note = ''
            source_language = ''
        elif original_content is not None and len(original_content) > 3660:
            note = '.Translation failed - too long'
        elif "@s.whatsapp.net left" in original_content:
            source_language = 'en'
            note = '...Whatsapp'
        elif "@s.whatsapp.net) added" in original_content:
            note = '...Whatsapp'
            # source_language = 'Whatsapp'
        elif original_content is not None and isinstance(original_content, str) and original_content.isalpha() and len(original_content) == 1:
            note = ''
            source_language = 'en'


        source_language = source_language_enhance(source_language)
        sheet.cell(row=row[0].row, column=2, value=translation)
        sheet.cell(row=row[0].row, column=3, value=source_language)
        sheet.cell(row=row[0].row, column=4, value=note)
        sheet.cell(row=row[0].row, column=5, value=confidence)
        sheet.cell(row=row[0].row, column=6, value=length)
         
    workbook.save(output_xlsx)

    msg_blurb = (f'Language detection saved to {output_xlsx}')
    msg_blurb_square(msg_blurb, color_green)

def googletrans_ver():
    import googletrans
    # Extract major and minor version from the module
    major_version, minor_version = map(int, googletrans.__version__.split('.')[:2])
    # major_version, minor_version = map(int, Translator.__version__.split('.')[:2])

    # Check if the version is greater than 4.1
    if major_version >= 4:
        print(f"googletrans version {major_version}.{minor_version}")
        return True
    else:
        print("Your version of Googletrans needs to be >=4")
        print("pip install googletrans>=4.0.0-rc1")
        print(f"The wont detect or translate")

        return False

def source_language_enhance(source_language):
    '''
    Convert 2 digit language code to a full name
    '''
    LANGUAGES = {
        'af': 'Afrikaans',
        'sq': 'Albanian',
        'am': 'Amharic',
        'ar': 'Arabic',
        'hy': 'Armenian',
        'az': 'Azerbaijani',
        'eu': 'Basque',
        'be': 'Belarusian',
        'bn': 'Bengali',
        'bs': 'Bosnian',
        'bg': 'Bulgarian',
        'ca': 'Catalan',
        'ceb': 'Cebuano',
        'ny': 'Chichewa',
        'zh-CN': 'Chinese (Simplified)',
        'zh-TW': 'Chinese (Traditional)',
        'co': 'Corsican',
        'hr': 'Croatian',
        'cs': 'Czech',
        'da': 'Danish',
        'nl': 'Dutch',
        'en': 'English',
        'eo': 'Esperanto',
        'et': 'Estonian',
        'tl': 'Filipino',
        'fi': 'Finnish',
        'fr': 'French',
        'fy': 'Frisian',
        'gl': 'Galician',
        'ka': 'Georgian',
        'de': 'German',
        'el': 'Greek',
        'gu': 'Gujarati',
        'ht': 'Haitian Creole',
        'ha': 'Hausa',
        'haw': 'Hawaiian',
        'iw': 'Hebrew',
        'hi': 'Hindi',
        'hmn': 'Hmong',
        'hu': 'Hungarian',
        'is': 'Icelandic',
        'ig': 'Igbo',
        'id': 'Indonesian',
        'ga': 'Irish',
        'it': 'Italian',
        'ja': 'Japanese',
        'jw': 'Javanese',
        'kn': 'Kannada',
        'kk': 'Kazakh',
        'km': 'Khmer',
        'rw': 'Kinyarwanda',
        'ko': 'Korean',
        'ku': 'Kurdish (Kurmanji)',
        'ky': 'Kyrgyz',
        'lo': 'Lao',
        'la': 'Latin',
        'lv': 'Latvian',
        'lt': 'Lithuanian',
        'lb': 'Luxembourgish',
        'mk': 'Macedonian',
        'mg': 'Malagasy',
        'ms': 'Malay',
        'ml': 'Malayalam',
        'mt': 'Maltese',
        'mi': 'Maori',
        'mr': 'Marathi',
        'mn': 'Mongolian',
        'my': 'Myanmar (Burmese)',
        'ne': 'Nepali',
        'no': 'Norwegian',
        'ps': 'Pashto',
        'fa': 'Persian',
        'pl': 'Polish',
        'pt': 'Portuguese',
        'pa': 'Punjabi',
        'ro': 'Romanian',
        'ru': 'Russian',
        'sm': 'Samoan',
        'gd': 'Scots Gaelic',
        'sr': 'Serbian',
        'st': 'Sesotho',
        'sn': 'Shona',
        'sd': 'Sindhi',
        'si': 'Sinhala',
        'sk': 'Slovak',
        'sl': 'Slovenian',
        'so': 'Somali',
        'es': 'Spanish',
        'su': 'Sundanese',
        'sw': 'Swahili',
        'sv': 'Swedish',
        'tg': 'Tajik',
        'ta': 'Tamil',
        'te': 'Telugu',
        'th': 'Thai',
        'tr': 'Turkish',
        'uk': 'Ukrainian',
        'ur': 'Urdu',
        'ug': 'Uyghur',
        'uz': 'Uzbek',
        'vi': 'Vietnamese',
        'cy': 'Welsh',
        'xh': 'Xhosa',
        'yi': 'Yiddish',
        'yo': 'Yoruba',
        'zu': 'Zulu',
        'auto': '.'
    }

    if source_language in LANGUAGES:
        source_language = LANGUAGES[source_language]

    else:
        source_language = ''
        return None

    return source_language


def language_detect(original_content):
    (source_language, confidence) = ('', '')
    if original_content not in auto_list:
        try:
            translator = Translator()
            detection = translator.detect(original_content)
            source_language = detection.lang
            # confidence = detection.confidence
        except:
            source_language = 'auto'

    else:
        source_language = 'auto'

    if original_content not in auto_list and source_language == 'auto':
        auto_list.append(original_content)

    return source_language, confidence


def msg_blurb_square(msg_blurb, color):
    '''
+----------------------------------+
|                                  |
| put a square around your message |
|                                  |
+----------------------------------+
    '''
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def translate_excel(input_xlsx, output_xlsx, source_language):

    row_count = 2
    word_pattern = re.compile(r'\b\w+\b')
    word_pattern2 = re.compile(r'\b\w+\b', flags=re.UNICODE)
    
    skip_characters = ['!', '?', ':)']  # Define the list of special characters
    
    target_language = 'en'
    file_exists = os.path.exists(input_xlsx)
    if file_exists == True:
        msg_blurb = (f'Reading {input_xlsx}')
        msg_blurb_square(msg_blurb, color_green)
        workbook = load_workbook(input_xlsx)        
        sheet = workbook.active
    else:
        msg_blurb = (f'Create {input_xlsx} and insert foreign language lines in the first column')
        msg_blurb_square(msg_blurb, color_red)  # Using ANSI escape code for color
        input(f"{color_green}Hit any key to continue{color_reset}")
        sys.exit()

        
    sheet.cell(row=1, column=2, value='English')
    sheet.cell(row=1, column=3, value='Language')
    sheet.cell(row=1, column=4, value='Note')    
    # sheet.cell(row=1, column=5, value='Confidence')   # task
    sheet.cell(row=1, column=6, value='Length')    
    
    for row in sheet.iter_rows(min_row=2):
        (translation, note, e, confidence, length) = ('', '', '', '', '')
        (source_language, text, skipper) = ('', '', '')
        original_content = row[0].value

        source_language, confidence = language_detect(original_content)
        length = content_length(original_content)
        
        if original_content is None:
            note = ''
            source_language = ''
        elif original_content is not None and len(original_content) > 3660:
            note = '.Translation failed - too long'
   
        elif original_content is not None and original_content != '' and source_language != 'auto'  and source_language != 'en':

            if isinstance(original_content, (int, float)):
                translation = original_content
            # elif re.search(word_pattern, original_content):
                # (translation, source_language, note) = translate_googletrans(original_content, source_language, target_language, note)

                # sleep(1)
                # if not translation:
                    # note = "Translation failed"
                    # sleep(2)
            elif re.search(word_pattern, original_content):
                (translation, source_language, note) = translate_request(original_content, source_language, target_language, note)   # works
                # source_language = source_language
                sleep(1)
                
                if not translation:
                    note = "Translation failed"
                    # source_language = source_language
                    sleep(2)
            source_language = source_language_enhance(source_language)

            print(f'\n{color_red}{row_count} {color_blue}{original_content}      {color_yellow}{translation}  {color_green}{source_language}  {color_red}{note}{color_reset}')

        else:
            source_language = source_language_enhance(source_language)
            translation = original_content
        # source_language = source_language_enhance(source_language)    
        row_count += 1
        sheet.cell(row=row[0].row, column=2, value=translation)
        sheet.cell(row=row[0].row, column=3, value=source_language)
        sheet.cell(row=row[0].row, column=4, value=note)
        sheet.cell(row=row[0].row, column=5, value=confidence)
        sheet.cell(row=row[0].row, column=6, value=length)

    workbook.save(output_xlsx)

    msg_blurb = (f'Saving to {output_xlsx}')
    msg_blurb_square(msg_blurb, color_green)

def translate_googletrans(text, source_language, target_language, note):
    '''
    use googletrans module, 60% of time, it works every time
    '''

    translator = Translator()
    if source_language == '':
        source_language = 'auto'
    target_language = "en"

    translation = ('')
    # source_language = source_language
    original_content = text
    retries = 3 # 3
    for _ in range(retries):  
        try:
            translation_result = translator.translate(original_content, src=source_language, dest='en')
            if translation_result and translation_result.text:
                translation = translation_result.text
                break

        except Exception as e:
            note = (f'.Error occurred while translating {e}')
            msg_blurb = (f'Error translating_: {e}')
            msg_blurb_square(msg_blurb, color_red)
            sleep(2)

    return (translation, source_language, note)


def translate_request(text, source_language, target_language, note):
    '''
    use requests to translate language
    '''

    if source_language == '':
        source_language = 'auto'
    (translation) = ('')
    # if source_language != 'auto':
        # print(f'translating with {source_language}')
    url = "https://translate.googleapis.com/translate_a/single?client=gtx&sl={}&tl={}&dt=t&q={}".format(
        source_language, target_language, text
    )

    user_agent = "Mozilla/5.0"  # Define custom user agent
    verify_ssl_cert = False  # Change to False if you don't want to verify SSL certificates

    headers = {
        "User-Agent": user_agent
    }

    try:
        response = requests.get(url, headers=headers, verify=verify_ssl_cert)
        
        if response.status_code == 200:
            data = response.json()
            translation = data[0][0][0] if data else ""

            source_language = data[2]
            note = ''
        else:
            print("Failed to translate. Status code:", response.status_code)
            note = (".Failed to translate. Status code:", response.status_code)
            source_language = ''
            
    except Exception as e:
        print(f'Error occurred while translating: {color_red}{e}{color_reset}')
        (translation, source_language) = ('', '')
        note = (f'.Error occurred while translating {e}')
        source_language = ''

    return (translation, source_language, note)
    
def usage():
    file = sys.argv[0].split('\\')[-1]

    print(f'\nDescription: {color_green}{description}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}insert your info into input_translate.xlsx')
    print(f'\nExample:')
    print(f'    {file} -a') # beta
    print(f'    {file}')
    print(f'    {file} -d')
    print(f'    {file} -I input_translate.xlsx')
    print(f'    {file} -C # print copyright') 

copyright = '''
Copyright (c) 2024 LincolnLandForensics

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''

if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>
"""
1.0.5 -fixed language module ar = Arabic
1.0.0 - use requests with googletrans as a backup (non-working) module.
0.2.0 - removed any switch requirements to make the exe version easier
0.1.0 - read xlsx, translate, export to xlsx
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>
"""
specify a language like arabic -a (ar)
requests doesn't work behind a proxy 
Make sure to handle potential issues like rate limiting, certificate verification, and unexpected input data gracefully. 

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Note            >>>>>>>>>>>>>>>>>>>>>>>>>>
"""

GoogleTrans works if it is 4.0.0-rc1 or later. 3.0 doesn't work


"""
