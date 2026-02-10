#!/usr/bin/python
# coding: utf-8

import sys
import os
import socket
import argparse
import re
import time
from tkinter import *
from tkinter import messagebox, filedialog, scrolledtext
import tkinter.ttk as ttk
import threading

from pytube import YouTube # pip install --upgrade yt-dlp pytube
from pytube.exceptions import AgeRestrictedError

import yt_dlp   # pip install yt-dlp

from youtube_transcript_api import (
    YouTubeTranscriptApi,
    TranscriptsDisabled,
    NoTranscriptFound,
    CouldNotRetrieveTranscript
)

from datetime import datetime
from openpyxl import Workbook

from openpyxl.utils.exceptions import IllegalCharacterError

# --------------------------------------------------------------------------
# Globals
# --------------------------------------------------------------------------
d = datetime.today()
# todaysDate = d.strftime("%Y-%m-%d %H:%M:%S")
todaysDate = d.strftime("%Y-%m-%d")
todaysDateTime = d.strftime("%Y-%m-%d_%H-%M-%S")

author = 'LincolnLandForensics'
description = "Download a list of Youtube videos from videos.txt, save list in xlsx file"
version = '1.2.5'

# will be set in main()
filename = None
spreadsheet = None
Row = None
workbook = None
Sheet1 = None

# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------
def main():
    global filename, Row, spreadsheet

    filename = 'videos.txt'
    Row = 2
    spreadsheet = f'videos_{todaysDateTime}.xlsx'

    status = internet()
    if not status:
        # noInternetMsg() # GUI will handle this if run from GUI
        print('\nCONNECT TO THE INTERNET FIRST\n')
        # If GUI is starting, we might want to let it show the error in the status window
    else:
        create_xlsx()

    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='input file (default videos.txt)', required=False)
    parser.add_argument('-O', '--output', help='output Excel filename', required=False)
    parser.add_argument('-y', '--youtube', help='youtube download', required=False, action='store_true')
    parser.add_argument('--sleep', type=int, default=30, help='seconds to sleep between downloads')
    parser.add_argument('--headless', action='store_true', help='disable GUI popups')
    parser.add_argument('--resolution', type=str, default="360p", help='video resolution to download (e.g., 360p, 720p)')

    # Check if any arguments were provided (other than the script name)
    if len(sys.argv) > 1:
        args = parser.parse_args()
        if args.input:
            filename = args.input
        if args.output:
            spreadsheet = args.output
        
        if not status:
            noInternetMsg(headless=args.headless)
            exit()

        youtube(args)
        workbook.save(spreadsheet)
    else:
        # No arguments, launch GUI
        run_gui()

    return 0

# --------------------------------------------------------------------------
# GUI Class
# --------------------------------------------------------------------------
class YouTubeDownloaderGUI:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{author} - YouTube Download {version}")
        self.root.geometry("600x500")
        
        # Use Vista theme
        self.style = ttk.Style()
        try:
            self.style.theme_use('vista')
        except Exception:
            self.style.theme_use('clam')

        self.input_file = StringVar(value='videos.txt')
        self.processing = False

        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=BOTH, expand=True)

        # Title and Description
        ttk.Label(main_frame, text=f"YouTube Download {version}", font=("Helvetica", 16, "bold")).pack(pady=(0, 5))
        ttk.Label(main_frame, text=description, wraplength=550).pack(pady=(0, 15))

        # Input File Section
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=X, pady=5)
        ttk.Label(input_frame, text="Input File:").pack(side=LEFT, padx=(0, 5))
        self.input_entry = ttk.Entry(input_frame, textvariable=self.input_file)
        self.input_entry.pack(side=LEFT, fill=X, expand=True, padx=5)
        ttk.Button(input_frame, text="Browse", command=self.browse_input).pack(side=LEFT)

        # Start Button
        self.start_btn = ttk.Button(main_frame, text="Start Processing", command=self.start_processing)
        self.start_btn.pack(pady=15)

        # Progress Bar
        self.progress = ttk.Progressbar(main_frame, orient=HORIZONTAL, mode='determinate')
        self.progress.pack(fill=X, pady=(0, 10))

        # Status Window
        ttk.Label(main_frame, text="Status Output:").pack(anchor=W)
        self.status_box = scrolledtext.ScrolledText(main_frame, height=12, state='disabled', font=("Consolas", 9))
        self.status_box.pack(fill=BOTH, expand=True)

    def browse_input(self):
        filename_selected = filedialog.askopenfilename(initialdir=".", title="Select videos.txt",
                                                        filetypes=(("Text files", "*.txt"), ("all files", "*.*")))
        if filename_selected:
            self.input_file.set(filename_selected)

    def log(self, message):
        self.status_box.config(state='normal')
        self.status_box.insert(END, message + "\n")
        self.status_box.see(END)
        self.status_box.config(state='disabled')
        # Also print to terminal as requested
        print(message)

    def update_progress(self, value):
        self.progress['value'] = value
        self.root.update_idletasks()

    def start_processing(self):
        if self.processing:
            return
        
        self.processing = True
        self.start_btn.config(state='disabled')
        self.input_entry.config(state='disabled')
        self.progress['value'] = 0
        
        # Start threading
        thread = threading.Thread(target=self.worker_task, daemon=True)
        thread.start()

    def worker_task(self):
        global filename, spreadsheet
        filename = self.input_file.get()
        
        if not internet():
            self.log("ERROR: No internet connection.")
            self.root.after(0, self.finish_processing)
            return

        if not os.path.exists(filename):
            self.log(f"ERROR: {filename} does not exist.")
            self.root.after(0, self.finish_processing)
            return

        # Prepare arguments object for the existing youtube function
        class Args:
            def __init__(self, sleep=30, headless=True, resolution="360p"):
                self.sleep = sleep
                self.headless = headless
                self.resolution = resolution

        args = Args()
        
        try:
            self.log(f"Starting process with input: {filename}")
            self.log("TIP: If downloads fail, try: pip install --upgrade yt-dlp pytube")
            # Run the existing logic but with GUI hooks if possible
            youtube(args, gui=self)
            
            workbook.save(spreadsheet)
            self.log(f"Processing complete.")
            self.log(f"Output saved to: {spreadsheet}")
        except Exception as e:
            self.log(f"CRITICAL ERROR: {str(e)}")
        finally:
            self.root.after(0, self.finish_processing)

    def finish_processing(self):
        self.processing = False
        self.start_btn.config(state='normal')
        self.input_entry.config(state='normal')

def run_gui():
    root = Tk()
    app = YouTubeDownloaderGUI(root)
    root.mainloop()

# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def check_keywords_in_captions(caption, keywords):
    """Check for keywords inside transcript text."""
    found_keywords = []
    for keyword in keywords:
        try:
            if re.search(rf"\b{re.escape(keyword)}\b", caption, re.IGNORECASE):
                found_keywords.append(keyword)
        except Exception:
            pass
    return ", ".join(found_keywords)

# Remove illegal characters for Excel
def clean_excel_value(val):
    if isinstance(val, str):
        return re.sub(r'[\x00-\x1F\x7F]', '', val)
    return val

def create_xlsx():
    global workbook, Sheet1
    workbook = Workbook()
    Sheet1 = workbook.active
    Sheet1.title = 'videos'
    Sheet1.freeze_panes = 'B2'

    # Excel column widths
    widths = {
        'A': 45, 'B': 45, 'C': 11, 'D': 9, 'E': 14,
        'F': 18, 'G': 6, 'H': 30, 'I': 12, 'J': 14,
        'K': 18, 'L': 15, 'M': 60, 'N': 2, 'O': 15
    }
    for col, w in widths.items():
        Sheet1.column_dimensions[col].width = w

    # Headers
    headers = [
        'url','title','description','views','author','publish_date',
        'length','download_name','video_id','thumbnail_url','dateDownloaded',
        'error','caption','(unused)','keywords'
    ]
    for idx, h in enumerate(headers, 1):
        Sheet1.cell(row=1, column=idx, value=h)

def finalMessage(headless=False):
    msg = f"\nSaved to current folder\n\nSaved info to {spreadsheet}\n"
    if headless:
        print(msg)
    else:
        try:
            # Only show popup if actually in a GUI environment and not headless
            window = Tk()
            window.withdraw() # Hide the main 1x1 window
            messagebox.showinfo('information',
                f'\nDownloaded all Youtube videos from {filename}\n\nSaved to {spreadsheet}\n')
            window.destroy()
            print(msg)
        except Exception:
            print(msg)

def internet(host="youtube.com", port=443, timeout=3):
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error as ex:
        # print(ex)
        return False

def noInternetMsg(headless=False):
    msg = "CONNECT TO THE INTERNET FIRST"
    if headless:
        print(msg)
    else:
        try:
            window = Tk()
            window.withdraw()
            messagebox.showwarning("Warning", msg)
            window.destroy()
        except Exception:
            print(msg)

def noVideos(headless=False):
    msg = f"Missing youtube videos in {filename}"
    if headless:
        print(msg)
    else:
        try:
            window = Tk()
            window.withdraw()
            messagebox.showwarning("Warning", msg)
            window.destroy()
        except Exception:
            print(msg)

def youtube_transcript(video_id):
    """Fetch transcript text, fallback if English not available."""
    caption = ""
    try:
        # Try English transcript
        srt = YouTubeTranscriptApi.get_transcript(video_id, languages=['en'])
    except (TranscriptsDisabled, NoTranscriptFound, CouldNotRetrieveTranscript):
        try:
            # Fallback: any available transcript
            transcript_list = YouTubeTranscriptApi.list_transcripts(video_id)
            srt = transcript_list.find_transcript([t.language_code for t in transcript_list]).fetch()
        except Exception:
            return ""   # No transcripts available
    except Exception as e:
        print(f"Transcript error for {video_id}: {str(e)}")
        return ""

    for part in srt:
        text = part.get('text', '').strip()
        if text.lower() == '[music]':
            continue
        caption += " " + text

    return caption.strip()

# --------------------------------------------------------------------------
# Main downloader
# --------------------------------------------------------------------------
def youtube(args, gui=None):
    keywords_to_check = ["tax", "lambo", "drug dealer"]

    if not os.path.exists(filename):
        msg = f"\n{filename} does not exist\n"
        if gui: gui.log(msg)
        else: print(msg)
        noVideos(headless=True if gui else args.headless)
        return

    msg = f'Downloading list of Youtube videos from {filename}. This can take a while...\n'
    if gui: gui.log(msg)
    else: print(msg)

    try:
        with open(filename, 'r') as csv_file:
            lines = [line.strip() for line in csv_file if line.strip()]
    except Exception as e:
        if gui: gui.log(f"Error reading file: {e}")
        return

    total_lines = len(lines)
    count = 0

    for idx, each_line in enumerate(lines, 1):
        link = each_line.split(',')[0].strip()
        if "youtu" not in link.lower():
            continue

        count += 1
        if gui:
            gui.log(f"Processing {idx}/{total_lines}: {link}")
            gui.update_progress((idx / total_lines) * 100)

        dateDownloaded = todaysDate

        # Call the existing metadata extraction and download
        # Note: yt__dlp and pytube use the 'args' object for resolution
        link, title, description_val, views, author_val, publish_date, length, download_name, rating, thumbnail_url, dateDownloaded, error, caption, video_id, owner, owner_id, owner_url, keywords = yt__dlp(link)
        
        if error != '':
            if gui: gui.log(f"  yt-dlp failed: {error}")
            if gui: gui.log(f"  trying pytube...")
            link, title, description_val, views, author_val, publish_date, length, download_name, rating, thumbnail_url, dateDownloaded, error, caption, video_id, owner, owner_id, owner_url, keywords = pytube(link, args)

        if gui:
            if error: gui.log(f"  Error: {error}")
            else: gui.log(f"  Downloaded: {title}")

        if not gui:
            print(f'{link}  {title}')
        
        write_xlsx(link, title, description_val, views, author_val, publish_date, length,
                   download_name, rating, thumbnail_url, dateDownloaded, error,
                   caption, video_id, owner, owner_id, owner_url, keywords)
        
        if idx < total_lines:
            if gui: gui.log(f"  Sleeping for {args.sleep} seconds...")
            time.sleep(args.sleep)

    if count == 0:
        noVideos(headless=True if gui else args.headless)
    else:
        if gui:
            gui.update_progress(100)
            gui.log("Done.")
        else:
            finalMessage(headless=args.headless)


def pytube(link, args):
    (title, description_val, views, author_val, publish_date) = ('', '', '', '', '')
    (length, rating, thumbnail_url, owner, caption) = ('', '', '', '', '')
    (owner_id, owner_url, download_name, error) = ('', '', '', '')
    (video_id, keywords) = ('', '')

    dateDownloaded = todaysDateTime

    if "youtu" in link.lower():
        try:
            yt = YouTube(link)
            video_id = yt.video_id
            title = yt.title
            download_name = f'{title}.mp4'
            author_val = yt.author
            description_val = yt.description
            rating = yt.rating
            publish_date = yt.publish_date
            length = str(yt.length)
            thumbnail_url = yt.thumbnail_url
            views = yt.views

            stream = yt.streams.filter(file_extension="mp4", res=args.resolution).first()
            if stream:
                stream.download()
            else:
                yt.streams.first().download()
            caption = youtube_transcript(video_id)

        except AgeRestrictedError:
            print(f"Video is age restricted: {link}")
            error = f"Age restricted (requires login). {download_name}"
        except Exception as e:
            print(f"An error occurred while processing video {link}: {str(e)}")
            error = f"Error processing video {download_name}: {str(e)}"

    return link, title, description_val, views, author_val, publish_date, length, download_name, rating, thumbnail_url, dateDownloaded, error, caption, video_id, owner, owner_id, owner_url, keywords


def yt__dlp(link):
   
    (title, description, views, author, publish_date) = ('', '', '', '', '')
    (length, rating, thumbnail_url, owner, caption) = ('', '', '', '', '')
    (owner_id, owner_url, download_name, error) = ('', '', '', '')
    (video_id, keywords) = ('', '')
    
    dateDownloaded = todaysDateTime

    
    ydl_opts = {
        'format': 'best[ext=mp4]/best',
        'outtmpl': '%(title)s.%(ext)s',
        'quiet': False,
        'no_warnings': False,
    }

    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(link, download=True)

            # Extract desired metadata
            title = info.get('title')
            download_name = f'{title}.mp4'
            description = info.get('description')
            views = info.get('view_count')
            creator = info.get('uploader')
            release_date = info.get('upload_date')
            length = info.get('duration')
            display_id = info.get('display_id')
            publish_date = info.get('upload_date')
            artist = info.get('uploader')
            video_id = info.get('id')
            location = info.get('location')  
            tags = info.get('tags')
            keywords = ' '.join(tags)

    except Exception as e:
        print(f"An error occurred while processing video {link}: {str(e)}")
        # title = description = views = creator = release_date = length = display_id = caption = tags = ''
        error = str(e)

    return link, title, description, views, author, publish_date, length, download_name, rating, thumbnail_url, dateDownloaded, error, caption, video_id, owner, owner_id, owner_url, keywords


# --------------------------------------------------------------------------
# Excel writer
# --------------------------------------------------------------------------
def write_xlsx(link, title, description, views, author, publish_date, length,
               download_name, rating, thumbnail_url, dateDownloaded, error,
               caption, video_id, owner, owner_id, owner_url, keywords):
    # print(f'keywords = {keywords}') # temp
    global Row, workbook, Sheet1, spreadsheet

    values = [
        link, title, description, str(views), author, str(publish_date),
        length, download_name, str(video_id), thumbnail_url, dateDownloaded,
        error, caption, '', str(keywords)
    ]

    for idx, val in enumerate(values, 1):
        try:
            Sheet1.cell(row=Row, column=idx, value=clean_excel_value(val))
        except IllegalCharacterError as e:
            print(f"Illegal character in column {idx}, row {Row}: {e}")
            Sheet1.cell(row=Row, column=idx, value="ERROR")

    Row += 1
    workbook.save(spreadsheet)  # Autosave after each row

# --------------------------------------------------------------------------
if __name__ == '__main__':
    main()



# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>


"""
1.2.1 - re-written by chatGPT   
1.0.5 - transcript saved as caption, look for keywords like lambo
1.0.4 - functional copy
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>


"""
remove redundant Global statements
download a list of all urls the user has
parse comments


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>


"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
