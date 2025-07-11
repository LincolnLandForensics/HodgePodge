 #!/usr/bin/env python3
# coding: utf-8
'''
read xlsx, write xlsx with only openpyxl
Pandas is too big
send data all at once so it can be sorted if needed
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>> 

import os
import re
import sys
import time
import openpyxl
import simplekml    # pip install simplekml
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import argparse  # for menu system
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = "convert Cellebrite contacts, account, web history, chats and call exports to intel format"
version = '1.1.3'

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-b', '--blank', help='create blank input sheet', required=False, action='store_true')
    parser.add_argument('-C', '--cellebrite', help='cellebrite contacts parse', required=False, action='store_true')

    args = parser.parse_args()
    data = []
    global output_xlsx
    output_xlsx = args.output

    if args.blank:
        data = []
        print(f'{color_green}Writing to {output_xlsx} {color_reset}')
        write_xlsx(data)

    elif args.cellebrite:
        if not args.input: 
            input_xlsx = "Contacts.xlsx"        
        else:
            input_xlsx = args.input
            
        file_exists = os.path.exists(input_xlsx)

        datatype = input_xlsx
        datatype = datatype.replace('.xlsx', '')

        if not args.output: 
            output_xlsx = (f'intel_{datatype}.xlsx') 
  
        else:
            output_xlsx = args.output


        if file_exists == True:
            msg_blurb = (f'Reading {input_xlsx}')
            msg_blurb_square(msg_blurb, color_green)

            data = read_cellebrite(input_xlsx)
            write_xlsx(data)

            workbook.close()
            msg_blurb = (f'Writing to {output_xlsx}')
            msg_blurb_square(msg_blurb, color_green)            
            # print(f'{color_green}Writing to {output_xlsx} {color_reset}')

        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()

    else:
        usage()
    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def case_number_prompt():
    # Prompt the user to enter the case number
    case_number = input("Please enter the Case Number: ")
    # Assign the entered value to Case
    case_prompt = case_number
    return case_prompt
    

def clean_data(item):
    """
    Cleans the item by removing any unwanted newlines and other characters 
    that may interfere with Markdown tables.

    Returns:
        str: The cleaned cell without newlines or extra spaces.
    """
    # Remove newlines, extra spaces, and any non-numeric characters (if needed)
    item_cleaned = re.sub(r'[\n\r]+', ' ', item)  # Replaces newlines with space
    item_cleaned = item_cleaned.strip()  # Remove leading/trailing spaces
    item_cleaned = str(item_cleaned)
    return item_cleaned


def convert_timestamp(timestamp, time_orig, timezone):
    timezone = timezone or ''
    time_orig = time_orig or ''
    timestamp = str(timestamp).strip()

    # Regular expression to find all timezones
    timezone_pattern = r"([A-Za-z ]+)$"
    matches = re.findall(timezone_pattern, timestamp)

    # Extract the last timezone match
    if matches:
        timezone = matches[-1]
        timestamp = timestamp.replace(timezone, "").strip()
    else:
        timezone = ''
        
    if time_orig == "":
        time_orig = timestamp
    else:
        timezone = ''


    # timestamp = timestamp.replace(' at ', ' ')
    if "(" in timestamp:
        timestamp = timestamp.split('(')
        timezone = timestamp[1].replace(")", '')
        timestamp = timestamp[0]
    elif " CDT" in timestamp:
        timezone = "CDT"
        timestamp = timestamp.replace(" CDT", "")
    elif " CST" in timestamp:
        timezone = "CST"
        timestamp = timestamp.replace(" CST", "")




    formats = [
        "%B %d, %Y, %I:%M:%S %p %Z",    # June 13, 2022, 9:41:33 PM CDT (Flock)
        "%Y:%m:%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",  # timestamps without seconds
        "%m/%d/%Y %H:%M:%S",  # timestamps in military time without seconds
        "%m-%d-%y at %I:%M:%S %p %Z", # test 09-10-23 at 4:29:12 PM CDT
        "%m-%d-%y %I:%M:%S %p",
        "%B %d, %Y at %I:%M:%S %p %Z",
        "%B %d, %Y at %I:%M:%S %p",
        "%B %d, %Y %I:%M:%S %p %Z",
        "%B %d, %Y %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p %Z",
        "%Y-%m-%dT%H:%M:%SZ",  # ISO 8601 format with UTC timezone
        "%Y/%m/%d %H:%M:%S",  # 2022/06/13 21:41:33
        "%d-%m-%Y %I:%M:%S %p",  # 13-06-2022 9:41:33 PM
        "%d/%m/%Y %H:%M:%S",  # 13/06/2022 21:41:33
        "%Y-%m-%d %I:%M:%S %p",  # 2022-06-13 9:41:33 PM
        "%Y%m%d%H%M%S",  # 20220613214133
        "%Y%m%d %H%M%S",  # 20220613 214133
        "%m/%d/%y %H:%M:%S",  # 06/13/22 21:41:33
        "%d-%b-%Y %I:%M:%S %p",  # 13-Jun-2022 9:41:33 PM
        "%d/%b/%Y %H:%M:%S",  # 13/Jun/2022 21:41:33
        "%Y/%b/%d %I:%M:%S %p",  # 2022/Jun/13 9:41:33 PM
        "%d %b %Y %H:%M:%S",  # 13 Jun 2022 21:41:33
        "%A, %B %d, %Y %I:%M:%S %p %Z",  # Monday, June 13, 2022 9:41:33 PM CDT ?
        "%A, %B %d, %Y %I:%M:%S %p"     # Monday, June 13, 2022 9:41:33 PM CDT
    ]

    for fmt in formats:
        try:
            dt_obj = datetime.strptime(timestamp.strip(), fmt)
            timestamp = dt_obj
            return timestamp, time_orig, timezone
                        
        except ValueError:
            pass

    raise ValueError(f"{time_orig} Timestamp format not recognized")


def clean_phone(phone):
    """
    Sanitizes and validates phone numbers (7–15 digits, digits only).
    """
    regex_phone = r'^\d{7,15}$'  # Matches a digit-only number between 7 and 15 digits
    phone = re.sub(r'[^\d]', '', phone)  # Remove everything that's not a digit #optional?
    phone = phone.replace('-', '') #optional?
    phone = re.sub(r'[- \(\)]', '', phone).strip()  #optional?
    phone = phone.replace("+", "")          # E165 standard doesn't have a + (E.164 has a +)  #optional?
    phone = phone.replace("phone", "")  # task  #optional?
    phone = phone.replace(' ','').replace('+','').replace('.','') #optional?

    if phone.startswith('1'):
        phone = phone[1:]

    phone = ''.join(char for char in phone if char.isalnum())

    phone = phone.strip()  # Remove leading/trailing spaces
    phone = str(phone)        

    return phone if re.match(regex_phone, phone) else ""


def dedupe(value):
    '''
    Read each line in the value, remove duplicate lines, and return a cleaned string.
    '''
    lines = value.strip().splitlines()
    unique_lines = list(dict.fromkeys(line.strip() for line in lines if line.strip()))
    return '\n'.join(unique_lines)


def msg_blurb_square(msg_blurb, color):
    horizontal_line = f"+{'-' * (len(msg_blurb) + 2)}+"
    empty_line = f"| {' ' * (len(msg_blurb))} |"

    print(color + horizontal_line)
    print(empty_line)
    print(f"| {msg_blurb} |")
    print(empty_line)
    print(horizontal_line)
    print(f'{color_reset}')


def read_cellebrite(input_xlsx):

    """Read data from an xlsx file and return as a list of dictionaries.
    Read XLSX Function: The read_xlsx() function reads data from the input 
    Excel file using the openpyxl library. It extracts headers from the 
    first row and then iterates through the data rows.    
    """

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []
    datatype = input_xlsx
    datatype = datatype.replace('.xlsx', '')
    
    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    case_prompt = case_number_prompt()

# regex patterns
    regex_phone = r'^\d{7,15}$'  # Matches a digit-only number between 7 and 15 digits

    AKA_patterns = [
        r'User ID-Additional Name:\s*(\S+)',
        r'User ID-Push Name:\s*(\S+)',
        r'User ID-User Name:\s*(\S+)'
    ]

    email_patterns = [
        r'Email-:\s*(\S+)',
        r'Email-email:\s*(\S+)',  
        r'Email-Email:\s*(\S+)', 
        r'Email-General:\s*(\S+)', 
        r'Email-Home:\s*(\S+)', 
        r'Email-HOME:\s*(\S+)',    
        r'Email-iCloud:\s*(\S+)',
        r'Email-Email Address:\s*(\S+)',        
        r'Email-Professional:\s*(\S+)',
        r'Email-Work:\s*(\S+)', 
        r'Email-General:\s*(\S+)'
    ]

    dob_patterns = [
        r'-Birthday:\s*(\S+)',
        r'User ID-Birthday:s*(\S+)'
    ]

    misc_patterns = [
        r'User ID-Identifier: (\S+)',
        r'User ID-Instagram Id: (\S+)'
    ]
  
    user5_pattern = r'User ID-Facebook Id:\s*([^\s]+)'    # test

    user_patterns = [
        r'User ID-iMessage:\s*(.+)',
        r'User ID-WhatsApp User Id:\s*(\S+)',
        r'User ID-Id:\s*(\S+)', 
        r'User ID-Username:\s*(\S+)',  
        r'User ID-WeChat ID:\s*(\S+)',   
        r'User ID-Facebook Id:\s*([^\s]+)',
        r'User ID-cash tag:\s*([^\s]+)',        
    ]

    url_patterns = [
        r'Profile Picture-:\s*(\S+)',
        r'User ID-Tango ID:\s*(\S+)',
        r'Web address-Professional:\s*(\S+)'
    ]

    phone_patterns = [
        r'Phone 1 - Value\s*(.+)',
        r'Phone-\_\$\!<Mobile>\!\$_:\s*(.+)',  # Escaped properly
        r'Phone-\_\$\!<Home>\!\$_:\s*(.+)',    # Escaped properly
        r'Phone-\_\$\!<Work>\!\$_:\s*(.+)',
        r'Phone-:\s*(.+)',
        # r'phone-:\s*(.+)',        
        r'Phone-:\s*phone number is\s*(.+)',
        r'Phone-:\s*Phone:\s*(.+)',
        r'Phone-CELL:\s*(.+)',        
        r'Phone-General:\s*(.+)',
        r'Phone-Home:\s*(.+)',
        r'Phone-HOME:\s*(.+)',        
        r'Phone-iPhone:\s*(.+)',
        r'Phone-Mobile:\s*(.+)',
        r'Phone-mobile:\s*(.+)',        
        r'Phone-Phone Number:\s*(.+)',
        r'Phone-Phone:\s*(.+)',
        r'Phone-Secondary number:\s*(.+)',
        r'Phone-unknown:\s*(.+)',
        r'User ID-Additional Name:\s*(.+)',  # test      # duplicates
        r'Phone-Viber:\s*(.+)',
        r'Phone-Work:\s*(.+)',
        r'User ID-iMessage:\s*(.+)',
        r'User ID-SMS:\s*(.+)',
        r'User ID-User ID:\s*(.+)',         # test 
        r'User ID-WhatsApp User Id:\s*(\d+)@s\.whatsapp\.net'
        ]

    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        print(f"{color_red}No data found in the Excel file.{color_reset}")
        return None

# active sheet (current sheet)
    active_sheet = wb.active
    global active_sheet_title
    active_sheet_title = active_sheet.title    

    for row_index, row_data in enumerate(data):
        (fullname, url, phone, email, business, misc) = ('', '', '', '', '', '')
        (lastname, firstname, snippet) = ('', '', '')
        (Time, time_orig, timezone) = ('', '', '')
        (ranking, source_file, original_file) = (f'4 - {datatype}', '', '')
        (fulladdress , note, info, source, user, query) = ('', '', '', '', '', '')
        (AKA, DOB, otheremails, titleurl, dnsdomain, country) = ('', '', '', '', '', '')
        (Latitude, Longitude, Coordinate, fulladdress2, state, Time) = ('', '', '', '', '', '')
        (From, to, cc, bcc, subject, body) = ('', '', '', '', '', '')
        (status, priority, attachment, Altitude, receiver, Sender) = ('', '', '', '', '', '')
        (city, state, country, case, Direction) = ('', '', '', '', '')
        
# fullname        
        fullname = row_data.get("Name") or ''

        if fullname == 'None':
            fullname = ''
        elif isinstance(fullname, str) and fullname.startswith('+'):
            fullname = fullname.replace('+', '')
            phone = f'{phone}\n{fullname}'  # task
            fullname = ''
        else:
            fullname = str(fullname)

        # Sequential fallback logic
        if not fullname:
            fullname = row_data.get("Partner Name") or ''

        if not fullname:
            fullname = row_data.get("fullname") or ''
            
# lastname
        lastname = row_data.get("Last Name") or ''
        
# firstname
        firstname = row_data.get("First Name") or ''

# middlename
        middlename = row_data.get("Middle Name") or ''

# case  
        case = case or row_data.get("case") or case_prompt or ''  

# url   
        url = row_data.get("URL") or row_data.get("Website") or ''


# title
        title = row_data.get("Title") or ''

        if title:
            titleurl = title
            Time = row_data.get("Last Visited-Time") or ''
    

            
        if url is None:
            url = ''
        else:
            
            dnsdomain = re.compile(r":\/\/(.*?)\/")

            pattern = re.compile(r":\/\/(.*?)\/")
            dnsdomain = pattern.search(url)

            if dnsdomain:
                dnsdomain = dnsdomain.group(1)
 




# info        
     
        info = row_data.get("Entries") or row_data.get("Entry") or ''
        info2 = row_data.get("Body") or row_data.get("Message") or ''
        entry2 = row_data.get("Entry (2)") or ''
        entry3 = row_data.get("Entry (3)") or ''        
        


        if entry2:
            info += '\n' + entry2
        if entry3:
            info += '\n' + entry3

        info = (info or '').strip()
       

        cleaned_info = info.replace('\n', ' ')  # .replace('\r', '')

        if "marlboro" in info2.lower() or "menthol" in info2.lower():
            note = "Cigarette"

        elif "parliament" in info2.lower() or "newport" in info2.lower():
            note = "Cigarette"

        user5_match = re.search(user5_pattern, info)
        user5_match = re.search(user5_pattern, info.replace('\n', ' '))
        user5_match = re.search(user5_pattern, cleaned_info)

        user5 = user5_match.group(1) if user5_match else None


# body  
        infotemp = info #test

        
        if info == "" and info2 != "":
            info = info2
            ranking == "3 - Chats"
            Time = (row_data.get("Timestamp: Time") or '').strip()

            if info.startswith("http"):
                url = info

                # Parse the URL
                parsed_url = urlparse(url)

                # Extract address
                address_param = parse_qs(parsed_url.query).get('address', [''])[0]
                fulladdress2 = unquote(address_param)

                if ", IL " in fulladdress2:
                    state = "IL"
                elif ", MO " in fulladdress2:
                    state = "MO"
                elif ", CA " in fulladdress2:
                    state = "CA"
                elif ", Fl " in fulladdress2:
                    state = "FL"
                
                # Extract GPS coordinates
                if "google.com/maps" in url:
                    # Define a regular expression pattern to extract coordinates
                    patternMaps2 = re.compile(r'@([-+]?\d{1,2}\.\d+),([-+]?\d{1,3}\.\d+)')
                    # Search for the pattern in the URL
                    matchMaps2 = patternMaps2.search(url)

                    if matchMaps2:
                        note = (f'map,{note}')
                        Latitude = float(matchMaps2.group(1))
                        Longitude = float(matchMaps2.group(2))
                        Latitude = Latitude or ''
                        Longitude = Longitude or ''
 
                        Coordinate = (f'{Latitude},{Longitude}')

                elif "maps.apple.com" in url:
                    note = (f'map,{note}')
                    Coordinate = parse_qs(parsed_url.query).get('ll', [''])[0]
                    if 'None' in Coordinate:
                        # print(f'bobs your uncle')   # temp
                        Coordinate == ''
                    if "," in Coordinate:
                        Latitude, Longitude = map(str, Coordinate.split(','))
                        Latitude = Latitude or ''
                        Longitude = Longitude or ''
        
# phone
        # Define fallback keys in order of priority
        fallback_keys = [
            "Phone Number(s)",
            "_ChatId",
            "Partners",
            "Phone 1 - Value",
            "Mobile Phone",
            "Phone"
        ]

        phone = row_data.get("phone") or ''

        # Go through fallback keys only if phone is still empty
        for key in fallback_keys:
            if not phone:
                candidate = row_data.get(key)
                if candidate:
                    phone = str(candidate)
                    if key == "_ChatId":
                        type_data = 'Chat'
                    elif key == "Partners":
                        type_data = 'Calls'
                            
                    found_phones = set()  # Using a set ensures uniqueness
                    for pat in phone_patterns:
                        matches = re.findall(pat, info, re.MULTILINE)
                        for phone in matches:
                            if phone not in found_phones:
                                found_phones.add(phone)

                    # Convert back to string with one phone number per line
                    phone = '\n'.join(found_phones)
                  
                    phone = (row_data.get("phone") or '').strip()

# email
        email = row_data.get("email") or ''
        email_matches = []
        if email == '':
            for pat in email_patterns:
                email_match = re.search(pat, info)
                if email_match:
                    email_matches.append(email_match.group(1))

            # Combine all matched emails into a single string
            email = '\n'.join(email_matches).strip()


# Parties        
        Parties = row_data.get("Parties") or ''
        misc = row_data.get("Time") or '' if Parties else ''
            
        info = (f'{row_data.get("Direction")} {row_data.get("Status")} {row_data.get("Duration")}')
        if info == 'None None None':
            info = ''
        Parties = Parties.replace('From:  ', '').replace('To:  ', '').lstrip('+')
        if " " in Parties:
            Parties = Parties.split(' ', 1)
                
            if "@" in Parties[0]:
                email = Parties[0]
            elif ":" in Parties[0]:
                user = Parties[0]
            else:
                phone = Parties[0]
            if fullname == '':
                fullname = Parties[1]


# Direction 
        Direction = Direction or row_data.get("Direction") or row_data.get("ns2:course") or ''
                
# Position        
        Position = (row_data.get("Position") or '').strip()

        try:
            Position = Position.replace(' ', '')
        except Exception as e:
            print(f"{color_red}Error printing line: {str(e)}{color_reset}")
        
        if Position is None:
            Position = ''
        else:
            Position = Position.replace('(', '').replace(')', '').strip()
            Position = str(Position)
            Coordinate = Position
            if 'None' in Coordinate:
                Coordinate == ''
            
            
        if Latitude == '' and ',' in Position:
            Position = Position.split(',')
            Latitude = Position[0]
            Longitude= Position[1]
            # print(f'Latitude {Latitude}') # temp

        if not Latitude:
            Latitude = row_data.get("Latitude") or ''
            Longitude = row_data.get("Longitude") or ''

        if not Latitude:
            Latitude = (row_data.get("Destination Latitude") or '').strip()
            Longitude = (row_data.get("Destination Longitude") or '').strip()

        if Latitude is None or Latitude == '':
            Latitude = row_data.get("Destination Latitude")
            Longitude = row_data.get("Destination Longitude")
            if Latitude is None:
                Latitude == ''
                Longitude == ''  

        if Latitude is None or Latitude == '':  # GPX export from xml
            Latitude = row_data.get("lat")
            Longitude = row_data.get("lon")
            if Latitude is None:
                Latitude == ''
                Longitude == ''  


                
# ranking
        ranking2 = row_data.get("Service Type")
        if ranking2:
            ranking = f"{ranking} - {ranking2}"
            ranking = ranking.replace("4 -", "3 -")

        ranking = ranking or ''
        ranking = ranking.replace(" - None", "").strip()

# source
        source = row_data.get("Source")
        if source is None:
            source = ''
        elif ranking2 is not None:
            ranking = (f'{ranking} - {ranking2}')
            ranking = ranking.replace("4 -", "3 -")
        else:
            ranking = (f'{ranking} - {source}')
   
        if " - None" in ranking:
            ranking = ranking.replace(' - None', '')  
 
            
# user
        user = row_data.get("user") or ''

        found_users = set()  # Using a set ensures uniqueness
        for pat in user_patterns:
            matches = re.findall(pat, info, re.MULTILINE)
            for user in matches:
                # if user not in found_users and user != '_' and user != '.':
                if len(user) > 2 and user not in {'_', '.'} and user not in found_users:

                    found_users.add(user)

        # Convert back to string with one user number per line
        if found_users:
            user = '\n'.join(found_users)
        user = str(user or '')  

# url           
        if user5 is not None:
            url = (f'https://facebook.com/{user5}')
        elif 'Cash App' in source and user != '':
            url = (f'https://cash.app/{user}')
        elif 'Instagram' in ranking:
            url = (f'https://www.Instagram.com/{user}/')

        elif 'Snapchat' in source and user != '':
            url = (f'https://www.snapchat.com/add/{user}?')
        elif 'Telegram' in source and user != '':
            url = (f'https://t.me/{user}/')

        elif 'Threads' in source and user != '':
            url = (f'https://www.threads.net/@{user}')

        elif 'TikTok' in ranking:
            url = (f'https://www.tiktok.com/@{user}')
            
        elif 'Twitter' in source and user != '':
            url = (f'https://x.com/{user}')
        elif 'Venmo' in source and user != '':
            url = (f'https://account.venmo.com/u/{user}')


# username
        username = row_data.get("Username") or ''

        if username in ['local', '.']:
            username = ''
        elif '@' in username:
            email = username.strip()
        elif not user:
            user = username.strip()
    

# url
        url_matches = []
        for pat in url_patterns:
            url_match = re.search(pat, info)
            if url_match:
                url_matches.append(url_match.group(1))

        url2 = '\n'.join(url_matches)
        if url2:
            url = f"{url}\n{url2}"

        # if url2 != '':
            # print(f'url = {url2}')  # temp
        # if url1 is not None:
            # url = (f'{url}\n{url1}')
        # if url2 is not None:
            # url = (f'https://www.tango.me/stream/{url2}')
        if not url or url.strip().lower() == "none":
            url = ''

        url = url.strip()

        if active_sheet_title == 'Facebook Messenger Users Contac':
            misc = (row_data.get("User Key") or '').strip()
            
            Type = 'Intel'
            if username is not None and len(username) >3:
                url = (f'https://www.facebook.com/{username}/')
            elif len(misc) > 5:
                url = (f'https://www.facebook.com/{misc}/')
                
# DOB
        dob_matches = []
        for pat in dob_patterns:
            dob_match = re.search(pat, info)
            if dob_match:
                dob_matches.append(dob_match.group(1))
        DOB = '\n'.join(dob_matches)
        DOB = DOB.strip('')

# AKA
        AKA_matches = []
        for pat in AKA_patterns:
            AKA_match = re.search(pat, info)
            if AKA_match:
                AKA_matches.append(AKA_match.group(1))
        AKA = '\n'.join(AKA_matches)
        AKA = AKA.strip('')

# misc
        misc_matches = []
        for pat in misc_patterns:
            misc_match = re.search(pat, info)
            if misc_match:
                misc_matches.append(misc_match.group(1))

        misc = '\n'.join(misc_matches)

        if misc is not None and "@" in misc:
            email = misc   
            misc = ''
        # if phone == '' and misc.startswith("+"):
        phone = phone or ''
        if phone is not None and phone == '' and misc is not None and "+" in misc:
        # if phone == '' and "+" in misc:
            phone = misc   
            phone = phone.replace("+", '')
            misc = ''

        misc3 = (row_data.get("Participants") or '').strip()
        
        if misc == '' and misc3 != "":
            misc = misc3
        misc = misc.replace(' _x000D_', '').strip() #test

        # Sender
        Sender = (
            row_data.get("From") or
            row_data.get("Sender") or
            row_data.get("Sender Name") or
            ''
        ).strip()

        # Normalize Local User or phone-in-Sender formats
        if Sender.startswith('+') and phone:
            phone = Sender.replace('+', '').strip()
        elif Sender.startswith('Local User'):
            Sender = ''

        if Sender != '':
            # Split the string by the first space
            if email == '':
                if "@" in Sender and Sender is not None:
                    email = Sender.split(' ', 1)[0]
                
                elif phone == '':
                    phone = Sender.split(' ', 1)[0]                    
            if fullname == '':
                try:
                    fullname = Sender.split(' ', 1)[1]
                except:
                    fullname = Sender

        if active_sheet_title == 'Facebook Messenger Messages':
            fullname = fullname or row_data.get("Sender Name") or ''

# From, to, cc, bcc, subject, body

            sender_id = (row_data.get("Sender ID") or '').strip()

            receiver = receiver or row_data.get("Receiver Name") or ''
            if not receiver:
                receiver = row_data.get("To") or ''
            if not receiver:
                receiver = row_data.get("Recipient(s)") or ''
                if receiver.startswith('Local User <'):
                    receiver = ''
            
            cc = (cc or row_data.get("CC") or '').strip()

            bcc = str(bcc or row_data.get("BCC") or '').strip()
            if bcc.lower() == 'none':
                bcc = ''
            
            subject = (subject or row_data.get("Subject") or '').strip()            
            
            snippet = (
                row_data.get("Snippet") or
                row_data.get("Text") or
                row_data.get("Summary") or
                ''
            ).strip()

            if not url and sender_id:
                url = f'https://www.facebook.com/{sender_id}'

        if receiver and Sender:
            # Attachments
            attachment1 = row_data.get("Attachment #1") or row_data.get("Attachment Name(s)") or ''
            attachment1 = attachment1.strip()

            # Status
            status = row_data.get("Status") or row_data.get("Read") or ''
            status = status.strip()
            if status.lower() == 'yes':
                status = 'Read'
            elif status.lower() == 'no':
                status = 'Not Read'

            # Snippet fallback (if not already handled elsewhere)
            snippet = snippet or ''
                    
                        
            note = (f'''
FROM:{email}
TO:{receiver}
CC:{cc}
BCC:{bcc}
SUBJECT:{subject}
SNIPPET:{snippet}
ATTACHMENT1:{attachment1}
STATUS:{status}
{note}
''')
            note = note.strip()


# business        
        business = business or row_data.get("business") or row_data.get("Organizations") or ''
        business = business.strip()
        
# fulladdress        
        fulladdress = (
            fulladdress or
            row_data.get("fulladdress") or
            row_data.get("Addresses") or
            ''
        ).replace('Home: ', '').strip()

        if not fulladdress and fulladdress2:
            fulladdress = fulladdress2.strip()

        if not fulladdress:
            fulladdress = row_data.get("Location Address") or ''
            fulladdress = fulladdress.strip()

# state
        state = state or row_data.get("state") or row_data.get("State/Province") or ''

# city
        city = city or row_data.get("city") or row_data.get("City") or ''

# country
        country = country or row_data.get("country") or row_data.get("Country") or ''
            
# Map Address
        fulladdress3 = row_data.get("Map Address")
        if fulladdress3 is None:
            fulladdress3 = ''
            query = row_data.get("Value")
            if query is None:
                query = ''
        else:
            fulladdress3 = fulladdress3.strip()

        if len(fulladdress) < 2:
            fulladdress = fulladdress3

# Axiom address
        fulladdress = fulladdress or row_data.get("Address") or ''
                
# country        
        country = row_data.get("Country code") or ''
        country = country.strip()
        
# tag
        tag = (row_data.get("Tag") or '').strip()
        tag2 = (row_data.get("Tag Note - Chat") or '').strip()

            
        if tag == "" and tag2 != "":
            tag2 = tag2.replace('Tags: ', '').strip()

            tag = tag2
            if "Important" in tag:
                tag = "Important"
            elif "Review" in tag:
                tag = "Review"        

        tag = (tag or row_data.get("Tags") or '').strip()

        tag = (tag or row_data.get("Tag Note - Instant Message") or '').strip()

        if "Important" in tag:
            tag = "Important"
        elif "Review" in tag:
            tag = "Review"

# source file
        source_file = row_data.get("Source file information") or ''

# original_file
        original_file = row_data.get("original_file") or input_xlsx

# Altitude
        Altitude = Altitude or row_data.get("Altitude") or row_data.get("ns0:ele") or ''

# type_data
        type_data = row_data.get("Type") or ''
        Icon = row_data.get("Icon") or ''

        if type_data == "":
            if "Searched" in original_file:
               type_data = "Searched"
            elif "Chats" in original_file:
               type_data = "Chats"
            elif active_sheet_title == 'Facebook Messenger Messages':
                type_data = "Chats"

            elif active_sheet_title == 'Apple Maps Searches':
                type_data = "Searched"
                Icon = "Searched" 
            elif active_sheet_title == 'Apple Maps Trips':
                type_data = "Trip"
            
            type_data = (type_data or row_data.get("Application Name") or '').strip()

        type_data = (
            type_data or
            row_data.get("Call Type") or
            row_data.get("ns4:vehicleType") or
            ''
        ).strip()
        
# Icon    
        Icon = row_data.get("Icon") or ''
        if not Icon:
            for keyword in ("Searched", "Chats"):
                if keyword in original_file:
                    Icon = keyword
                    break

# Time
        # List of potential time fields in priority order
        time_fields = [
            "Time",
            "Timestamp: Time",
            "Timestamp-Time",
            "End Date/Time - UTC+00:00 (M/d/yyyy)",
            "Sent Date/Time - UTC+00:00 (M/d/yyyy)",
            "Timestamp Date/Time - UTC+00:00 (M/d/yyyy)",
            "Message Sent Date/Time - UTC+00:00 (M/d/yyyy)",
            "Call Date/Time - UTC+00:00 (M/d/yyyy)",
            "ns0:time"
        ]

        for field in time_fields:
            Time = row_data.get(field)
            if Time and Time != 'NoneType':
                Time = Time.strip()
                break
        else:
            Time = ''

        try:
            Time = Time.replace('T', ' ').strip()
        except AttributeError:
            print(f'Time type = {type(Time)}')  # temp
            
# coordinate
        if not Coordinate and not Altitude:
            Latitude = Latitude or ''
            Longitude = Longitude or ''
            
            if Latitude and Longitude:
                Coordinate = f'{Latitude},{Longitude}'

            if 'None' in Coordinate:
                Coordinate = ''

# convert time
        output_format = "%Y-%m-%d %H:%M:%S "    # ISO 8601
        # output_format = "%Y-%m-%dT%H:%M:%SZ"    # Google Earth format
        # pattern = r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$'
        pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'  # ISO military time

        if time_orig == '' and Time != '': # copy the original time
            time_orig = Time
        try:
            (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
            Time = Time.strftime(output_format)
            if Time is None:
                Time = ''              
            
        except ValueError as e:
            if Time != "":
                print(f"Error time2: {e} - {Time}")
                # Time = ''    # temp rem of this
            
# fullname cleanup
        fullname = fullname or ''

        if fullname == '':
           fullname, firstname, middlename, lastname = fullname_build(firstname, middlename, lastname) 
        elif ' ' in fullname:
            (fullname, firstname, middlename, lastname) = fullname_parse(fullname)

        fullname = fullname.replace("_$!<Other>!$_", "")
        if "@" in fullname and fullname is not None:
            fullname = ''
        if fullname is not None and '+1 (' in fullname:
            query = fullname
            fullname = ''

        if Coordinate == 'None,None':
            Coordinate == ''

# clean phone
        cleaned_phones = ""

        for line in phone.splitlines():
            cleaned = clean_phone(line)
            if cleaned:
                cleaned_phones += cleaned + "\n"

        # Optionally strip the trailing newline
        phone = cleaned_phones.strip()

        phone = dedupe(phone)  # test



# cleanup cells # test
        fullname = clean_data(fullname)
        # phone = clean_data(phone)
        email = clean_data(email)
        business = clean_data(business)
        fulladdress = clean_data(fulladdress)
        AKA = clean_data(AKA)
        tag = clean_data(tag)

# write rows to data
        row_data["query"] = query
        row_data["ranking"] = ranking
        row_data["fullname"] = fullname
        row_data["url"] = url
        row_data["phone"] = phone
        row_data["email"] = email
        row_data["user"] = user        
        row_data["business"] = business        
        row_data["fulladdress"] = fulladdress  
        row_data["country"] = country        
        row_data["note"] = note
        row_data["info"] = info # interaction status
        row_data["DOB"] = DOB
        row_data["AKA"] = AKA        
        row_data["misc"] = misc  
        row_data["lastname"] = lastname          
        row_data["firstname"] = firstname   
        row_data["middlename"] = middlename 
        row_data["case"] = case 
        row_data["otheremails"] = otheremails     
        row_data["city"] = city 
        row_data["state"] = state 
        row_data["dnsdomain"] = dnsdomain     
        row_data["titleurl"] = titleurl 
        row_data["Time"] = Time
        row_data["Latitude"] = Latitude  
        row_data["Longitude"] = Longitude  
        row_data["Coordinate"] = Coordinate  
        row_data["Direction"] = Direction         
        row_data["Altitude"] = Altitude         
        row_data["Source file information"] = source_file     
        row_data["original_file"] = original_file     
        row_data["Tag"] = tag     
        row_data["Type"] = type_data     
        row_data["Icon"] = Icon     
    return data

def fullname_build(firstname, middlename, lastname):
    try:
        lastname = lastname.upper()
        firstname = firstname.title()
        middlename = middlename.title()
    except:
        pass
    fullname = ''
    if firstname != '' and middlename != '' and lastname != '':
        fullname = (f'{firstname} {middlename} {lastname}')
    elif firstname != '' and lastname != '':
        fullname = (f'{firstname} {lastname}')
    elif lastname != '':
        fullname = (f'{lastname}')

    return fullname, firstname, middlename, lastname
    
def fullname_parse(fullname):
    (firstname, middlename, lastname) = ('', '', '')
    fullname = fullname.strip()
    if ' ' in fullname:
        fullname_parts = fullname.split(' ')
        firstname = fullname_parts[0].title()
        lastname = fullname_parts[-1].upper()
        
        if len(fullname_parts) == 2:
            fullname = (f'{firstname} {lastname}')
        elif len(fullname_parts) == 3:
            middlename = fullname_parts[1].title()
            fullname = (f'{firstname} {middlename} {lastname}')
            
    return fullname, firstname, middlename, lastname
    

def write_xlsx(data):
    '''
    The write_xlsx() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

# active sheet (current sheet)
    # active_sheet = workbook.active
    # active_sheet_title = active_sheet.title 
    print(f'Reading {active_sheet_title} sheet\n')
    worksheet.title = active_sheet_title
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = [
        "query", "ranking", "fullname", "url", "email", "user", "phone",
        "business", "fulladdress", "city", "state", "country", "note", "AKA",
        "DOB", "SEX", "info", "misc", "firstname", "middlename", "lastname",
        "associates", "case", "sosfilenumber", "owner", "president", "sosagent",
        "managers", "Time", "Latitude", "Longitude", "Coordinate",
        "original_file", "Source", "Source file information", "Plate", "VIS", "VIN",
        "VYR", "VMA", "LIC", "LIY", "DLN", "DLS", "content", "referer", "osurl",
        "titleurl", "pagestatus", "ip", "dnsdomain", "Tag", "Icon", "Type"
        ]

    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [1, 2, 4, 5, 6, 8, 9, 15, 16, 18, 19]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill


    ## Excel column width
    worksheet.column_dimensions['A'].width = 6# #
    worksheet.column_dimensions['B'].width = 20# 
    worksheet.column_dimensions['C'].width = 25# 
    worksheet.column_dimensions['D'].width = 25# 
    worksheet.column_dimensions['E'].width = 25#
    worksheet.column_dimensions['F'].width = 15# 
    worksheet.column_dimensions['G'].width = 18 # phone
    worksheet.column_dimensions['H'].width = 18# 
    worksheet.column_dimensions['I'].width = 16# 
    worksheet.column_dimensions['J'].width = 20 # fulladdress
    worksheet.column_dimensions['K'].width = 10# state
    worksheet.column_dimensions['L'].width = 10# country
    worksheet.column_dimensions['M'].width = 15# note
    worksheet.column_dimensions['N'].width = 13# 
    worksheet.column_dimensions['O'].width = 5# 
    worksheet.column_dimensions['P'].width = 4# 
    worksheet.column_dimensions['Q'].width = 18# 
    worksheet.column_dimensions['R'].width = 5# 
    worksheet.column_dimensions['S'].width = 18# 
    worksheet.column_dimensions['T'].width = 13#   
    worksheet.column_dimensions['U'].width = 10#     
    worksheet.column_dimensions['V'].width = 10#   
    worksheet.column_dimensions['W'].width = 14# 
    worksheet.column_dimensions['X'].width = 12#   
    worksheet.column_dimensions['Y'].width = 10#  
    worksheet.column_dimensions['Z'].width = 10# 
    worksheet.column_dimensions['AA'].width = 10# 
    worksheet.column_dimensions['AB'].width = 10# 
    worksheet.column_dimensions['AC'].width = 16# 
    worksheet.column_dimensions['AD'].width = 10# 
    worksheet.column_dimensions['AE'].width = 10#
    worksheet.column_dimensions['AF'].width = 16# 
    worksheet.column_dimensions['AG'].width = 20# 
    worksheet.column_dimensions['AH'].width = 10# 
    worksheet.column_dimensions['AI'].width = 16# 
    worksheet.column_dimensions['AJ'].width = 10# 
    worksheet.column_dimensions['AK'].width = 8#
    worksheet.column_dimensions['AL'].width = 8# 
    worksheet.column_dimensions['AM'].width = 8# 
    worksheet.column_dimensions['AN'].width = 8# 
    worksheet.column_dimensions['AO'].width = 8# 
    worksheet.column_dimensions['AP'].width = 8# 
    worksheet.column_dimensions['AQ'].width = 8# 
    worksheet.column_dimensions['AR'].width = 8# 
    worksheet.column_dimensions['AS'].width = 8# 
    worksheet.column_dimensions['AT'].width = 8# 
    worksheet.column_dimensions['AU'].width = 8# 
    worksheet.column_dimensions['AV'].width = 8# 
    worksheet.column_dimensions['AW'].width = 12# 
    worksheet.column_dimensions['AX'].width = 16# 
    worksheet.column_dimensions['AY'].width = 10# 
    worksheet.column_dimensions['AZ'].width = 10# 
    worksheet.column_dimensions['BA'].width = 10# 
    worksheet.column_dimensions['BB'].width = 12# 
    worksheet.column_dimensions['BC'].width = 10# 


    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")
    
    workbook.save(output_xlsx)


def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description2}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}export from Cellebrite categories')
    print(f'\nExample:')
   
    print(f'    {file} -b -O input_blank.xlsx') 
    print(f'    {file} -C -I Accounts.xlsx  ')  
    print(f'    {file} -C -I Calls.xlsx  ')      
    print(f'    {file} -C -I Chats.xlsx  ')       
    print(f'    {file} -C -I Contacts.xlsx  ') 
    # print(f'    {file} -C -I Journeys.xlsx  # beta')     
    print(f'    {file} -C -I SearchedItems.xlsx  ')   
    print(f'    {file} -C -I WebHistory.xlsx  ')  
 
                
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

1.1.1 - fixed regex for aka, misc, phone, user, email, dob, url
1.0.7 - Fixed phone numbers (was removing any 1's from phone numbers. (oops)
1.0.5 - asks for case number
1.0.4 - left strip 1 off phone numbers for uniformity.
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
if tik tok and user, create a url
info needs to have new lines so things like phone regex need to be done in a for loop by line
if phone = '' and misc is a phone number: phone = misc
AKA needs work
skip case sensative regex, for phone at least
if user = 'facebookuser': user = ''


if instagram id found create a url
Icon, type_data and origin update

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

