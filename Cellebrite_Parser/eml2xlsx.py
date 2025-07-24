#!/usr/bin/env python3
# coding: utf-8
"""
Author: LincolnLandForensics
Version: 1.0.2
Parses .eml or .mbox files in a folder, extracts metadata, and exports to Excel.
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>> #
import os
import re
import sys
import hashlib
import argparse
import mailbox
import email
from email import policy
from datetime import datetime
from openpyxl import Workbook
from bs4 import BeautifulSoup
from email.parser import BytesParser
from email.header import decode_header
from openpyxl.styles import PatternFill
from email.utils import parsedate_to_datetime

# <<<<<<<<<<<<<<<<<<<<<<<<      Color Support        >>>>>>>>>>>>>>>>>>>>>>>> #
color_red = color_yellow = color_green = color_blue = color_purple = color_reset = ''
if sys.version_info > (3, 7, 9) and os.name == "nt":
    from colorama import Fore, Back, Style
    print(Back.BLACK)
    color_red, color_yellow, color_green = Fore.RED, Fore.YELLOW, Fore.GREEN
    color_blue, color_purple, color_reset = Fore.BLUE, Fore.MAGENTA, Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<      Main Function        >>>>>>>>>>>>>>>>>>>>>>>> #
def main():
    parser = argparse.ArgumentParser(description="Parse .eml files into Excel")
    parser.add_argument('-I', '--input', help='Input folder of .eml files', required=False)
    parser.add_argument('-O', '--output', help='Output Excel file', required=False)
    parser.add_argument('-E', '--eml', help='Enable .eml parsing', action='store_true')
    parser.add_argument('-M', '--mbox', help='Enable .mbox parsing', action='store_true')

    args = parser.parse_args()
    input_folder = args.input if args.input else r"C:\Forensics\scripts\python\eml"
    output_xlsx = args.output if args.output else "email.xlsx"

    global data 
    data = []



    if args.eml:
        if not os.path.exists(input_folder):
            msg_blurb_square(f'{input_folder} does not exist', color_red)
            exit()
        msg_blurb_square(f'Reading .eml files from {input_folder}', color_green)
        data, contacts_data = read_eml(input_folder)
        msg_blurb_square(f'Writing to {output_xlsx}', color_green)
        write_xlsx(data, contacts_data, output_xlsx)
    elif args.mbox:
        if not os.path.exists(input_folder):
            msg_blurb_square(f'{input_folder} does not exist', color_red)
            exit()
        msg_blurb_square(f'Reading .mbox files from {input_folder}', color_green)
        data, contacts_data = read_mbox(input_folder)
        msg_blurb_square(f'Writing to {output_xlsx}', color_green)
        write_xlsx(data, contacts_data, output_xlsx)
    else:
        usage()
    return 0

# <<<<<<<<<<<<<<<<<<<<<<<<      subroutines        >>>>>>>>>>>>>>>>>>>>>>>> #

def case_number_prompt():
    # Prompt the user to enter the case number
    case_number = input("Please enter the Case Number: ")
    # Assign the entered value to Case
    case_prompt = case_number
    return case_prompt


def clean_date(date_str):
    try:
        dt = parsedate_to_datetime(date_str)
        if dt is None:
            return ''
        tz = dt.tzinfo
        if tz:
            return dt.isoformat().replace('T', ' ')
        else:
            return dt.isoformat().replace('T', ' ') + ' ' + 'Z'  # Append Z to indicate UTC
    except Exception:
        match = re.match(r"(\d{4})\.(\d{2})\.(\d{2})-(\d{2})\.(\d{2})\.(\d{2})", date_str)
        if match:
            year, month, day, hour, minute, second = map(int, match.groups())
            dt = datetime(year, month, day, hour, minute, second)
            return dt.isoformat().replace('T', ' ') + 'Z'

        return ''

def count_eml_files(folder_path):
    return sum(1 for file_name in os.listdir(folder_path) if file_name.lower().endswith(".eml"))

def count_mbox_files(folder_path):
    return sum(1 for file_name in os.listdir(folder_path) if file_name.lower().endswith(".mbox"))


def decode_header_str(header_obj):  # MBOX 
    if not header_obj:
        return ""
    decoded_parts = decode_header(header_obj)
    header_str = ""
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            try:
                header_str += part.decode(encoding or 'utf-8', errors='replace')
            except Exception:
                header_str += part.decode('utf-8', errors='replace')
        else:
            header_str += str(part)
    return header_str.strip()


def decode_payload(payload):
    try:
        return payload.decode('utf-8')
    except UnicodeDecodeError:
        try:
            return payload.decode('latin1')
        except Exception:
            return payload.decode(errors='ignore')


def extract_body(msg):  # mbox
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_dispo = str(part.get('Content-Disposition'))
            if content_type == 'text/plain' and 'attachment' not in content_dispo:
                body += part.get_payload(decode=True).decode(errors='replace')
            elif content_type == 'text/html' and not body:
                html = part.get_payload(decode=True).decode(errors='replace')
                soup = BeautifulSoup(html, 'lxml')
                body = soup.get_text()
    else:
        payload = msg.get_payload(decode=True)
        body = payload.decode(errors='replace') if payload else ''
    return body.strip()


def get_attachments(msg):
    attachments = []
    for part in msg.walk():
        if part.get_content_disposition() == 'attachment':
            filename = part.get_filename()
            if filename:
                attachments.append(filename)
    return "; ".join(attachments)


def get_attachment_filenames(msg):
    filenames = []
    for part in msg.walk():
        content_disposition = part.get("Content-Disposition", "")
        if 'attachment' in content_disposition.lower():
            filename = part.get_filename()
            if filename:
                filenames.append(filename)
    return filenames


def get_body(msg):
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                payload = part.get_payload(decode=True)
                return decode_payload(payload) if payload else ""
    else:
        payload = msg.get_payload(decode=True)
        return decode_payload(payload) if payload else ""
    return ""


def msg_blurb_square(msg, color):
    border = f"+{'-' * (len(msg) + 2)}+"
    print(f"{color}{border}\n| {msg} |\n{border}{color_reset}")


def read_eml(folder_path):
    contacts = []
    case = case_number_prompt()
    source = source_prompt()

    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith(".eml"):
            full_path = os.path.join(folder_path, file_name)
            sha256 = sha256_hash(full_path)
            
            try:
                with open(full_path, 'rb') as f:
                    msg = BytesParser(policy=policy.default).parse(f)
                    from_field = sanitize_field(msg.get("From"))

                    if ' <' in from_field:
                        from_field = from_field.split(' <')
                        fullname = from_field[0].strip()
                        
                        fullname = fullname.replace('"', '')    # test
                        
                        if "CashApp" in fullname:
                            Tag = "CashApp"
                        email = from_field[1]
                        email = from_field[1].replace('>', '').strip()

                    else:
                        email = from_field.replace('>', '').strip()

                    email2 = sanitize_field(msg.get("To"))
                    
                    email_data = {
                        "Time": clean_date(msg.get("Date")),
                        "From": sanitize_field(msg.get("From")),
                        "To": sanitize_field(msg.get("To")),
                        "Subject": sanitize_field(msg.get("Subject")),
                        "Body": sanitize_field(get_body(msg)),
                        "Attachments": sanitize_field(get_attachments(msg)),
                        "Tag": "",
                        "original_file": file_name,
                        "Source": source,
                        "case": case,
                        "Precedence": sanitize_field(msg.get("Precedence")), 
                        "List-Unsubscribe": sanitize_field(msg.get("List-Unsubscribe")),   
                        "X-Mailer": sanitize_field(msg.get("X-Mailer")),
                        "fullname": fullname,
                        "email": email,
                        'sha256': sha256,
                    }
                    contacts_data = {
                        "fullname": fullname,
                        "email": email,
                        "note": email2,
                        "original_file": file_name,
                        "Source": source
                    }

                    data.append(email_data)
                    
                    contacts.append(contacts_data)
                    
            except Exception as e:
                print(f"{color_red}Error parsing {file_name}: {str(e)}{color_reset}")

    eml_count = count_eml_files(folder_path)
    print(f"Found {eml_count} .eml files in the {folder_path} folder.")

    return data, contacts_data
    
    
def read_mbox(folder_path):
    contacts_data = []  # TEMP
    contacts = []
    case = case_number_prompt()
    source = source_prompt()

    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith(".mbox"):
            full_path = os.path.join(folder_path, file_name)

            sha256 = sha256_hash(full_path)
          
            mbox = mailbox.mbox(full_path)

            for message in mbox:

                (date_obj, sender, recipient, subject, body, precedence) = ('', '', '', '', '', '')
                (listUnsubscribe, xmailer, tag, timezone, original_file, attachments) = ('', '', '', '', '', '')
                (fullname) = ('')
                try:
                    sender = decode_header_str(message['from'])
                    recipient = decode_header_str(message['to'])
                    subject = decode_header_str(message['subject'])
                    date = message['date']
                    date_obj = clean_date(date) if date else ''
                    body = extract_body(message)
                    precedence = message['precedence']
                    listUnsubscribe = message['List-Unsubscribe']
                    xmailer = message['X-Mailer']
                    attachments = get_attachment_filenames(message)
                    # sender = sanitize_field(body)
                    sender = sender.replace('"', '')
                    recipient = recipient.replace('"', '')
                    from_field = sender.strip('"')
                    
                    if ' <' in from_field:
                        from_field = from_field.split(' <')
                        fullname = from_field[0].strip()
                        
                        fullname = fullname.replace('"', '')    # test
                        
                        if "CashApp" in fullname:
                            Tag = "CashApp"
                        email = from_field[1]
                        email = from_field[1].replace('>', '').strip()

                    else:
                        email = from_field.replace('>', '').strip()


                except Exception as e:
                    print(f"Error parsing message: {e}")
                
                # cleanup
                body = sanitize_field(body)
                subject = sanitize_field(subject)

                if 'CashApp' in sender:
                    tag = 'CashApp'     
                    
                try:
                    if 'CashApp' in sender:
                        tag = 'CashApp'

                    if "unsubscribe" in body.lower() and listUnsubscribe == '':
                        listUnsubscribe = '_unsubscribe'
                        print(f'listUnsubscribe = {listUnsubscribe}')
                        
                except Exception as e:
                    print(f"Error tweaking message: {e}")
                # 'Attachments': ', '.join(attachments) if attachments else ''

                
                try:
                    data.append({
                        'Time': date_obj,
                        'From': sender,
                        'To': recipient,
                        'Subject': subject,
                        'Body': body,
                        'Attachments': ', '.join(attachments) if attachments else '',
                        'Tag': tag,           
                        'original_file': file_name,
                        'Source': source,
                        'case': case,
                        'Precedence': precedence,
                        'List-Unsubscribe': listUnsubscribe,
                        'X-Mailer': xmailer,
                        'fullname': fullname,                        
                        'email': email,                        
                        'sha256': sha256,

                    })
                    
                    # contacts_data = {                 
                        # "fullname": fullname,
                        # "email": email,
                        # "note": email2,
                        # "original_file": file_name,
                        # "Source": source
                    # }

                    # print(f'contacts_data = {contacts_data}')   # temp

                except Exception as e:
                    print(f"Error appending message: {e}")

    mbox_count = count_mbox_files(folder_path)
    print(f"Found {mbox_count} .mbox files")

    return data, contacts_data
    
    
def remove_illegal_chars(text):
    if not isinstance(text, str):
        return text
    # Remove control characters and illegal XML characters
    return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', text)


def sanitize_field(text):
    text = (text or '').strip().strip('"')

    return remove_illegal_chars(text.strip()) if isinstance(text, str) else text


def sha256_hash(full_path):
    '''
    Calculate the SHA-256 hash of the file at full_path
    '''
    sha256 = hashlib.sha256()
    try:
        with open(full_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b''):
                sha256.update(chunk)
        sha256 = sha256.hexdigest()

    except Exception as e:
        print(f"⚠️ Error hashing file: {e}")

    return sha256


def source_prompt():
    # Prompt the user to enter the source file, like the Google Return zip file
    source = input("Please enter the original (zip) file name: ")
    return source


def write_xlsx(data, contacts_data, file_path):
    workbook = Workbook()
    
    # Worksheet 1
    worksheet = workbook.active
    worksheet.title = 'Eml'
    worksheet.freeze_panes = 'B2'

    # Worksheet 2
    worksheet2 = workbook.create_sheet(title='Contacts')
    worksheet2.freeze_panes = 'B2'

    headers = [
        "Time", "From", "To", "Subject", "Body", "Attachments", "Tag",
        "original_file", "Source", "case", "Precedence", "List-Unsubscribe", "X-Mailer", "fullname", "email", "sha256"
    ]
    
    headers2 = [
        "query", "ranking", "fullname", "url", "email", "user", "phone", "business",
        "fulladdress", "city", "state", "country", "note", "AKA", "DOB", "SEX", "info",
        "misc", "firstname", "middlename", "lastname", "associates", "case",
        "sosfilenumber", "owner", "president", "sosagent", "managers", "Time",
        "Latitude", "Longitude", "Coordinate", "original_file", "Source"
    ]

    # Header formatting
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1, 2, 3, 4, 6]:
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = fill

    # Column widths
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 13    
    worksheet.column_dimensions['G'].width = 10    
    worksheet.column_dimensions['H'].width = 17    
    worksheet.column_dimensions['I'].width = 20    
    worksheet.column_dimensions['J'].width = 9    
        
    worksheet.column_dimensions['N'].width = 18    
    worksheet.column_dimensions['O'].width = 27    
    worksheet.column_dimensions['P'].width = 30    

    # Header formatting for worksheet 1
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1, 2, 3, 4, 6]:
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = fill

    # Header formatting for worksheet 2
    for col_index, header in enumerate(headers2):
        cell = worksheet2.cell(row=1, column=col_index + 1)
        cell.value = header

    # Data rows for worksheet 1
    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index + 2, column=col_index + 1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error writing line in email: {str(e)}{color_reset}")

    # Data rows for worksheet 2

    # for row_index, row_data in enumerate(contacts_data):
        # for col_index, col_name in enumerate(headers2):
            # try:
                # cell_data = row_data.get(col_name)
                # worksheet2.cell(row=row_index + 2, column=col_index + 1).value = cell_data
            # except Exception as e:
                # print(f"{color_red}Error writing line in Contacts: {str(e)}{color_reset}")

    workbook.save(file_path)


def usage():
    print(f"Usage: {sys.argv[0]} -E [-I input_folder] [-O output.xlsx]")
    print("Examples:")
    print(f"    {sys.argv[0]} -E")
    print(f"    {sys.argv[0]} -E -I C:\\emails -O parsed_emails.xlsx")
    print(f"    {sys.argv[0]} -M")    


# <<<<<<<<<<<<<<<<<<<<<<<<      Run Program        >>>>>>>>>>>>>>>>>>>>>>>> #
if __name__ == '__main__':
    main()
    


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
1.0.0 - .eml and .mbox parser

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
Contacts sheet with unique emails

.eml or .mbox without seperate switch
see if RLEAPP is doing anything different

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
    