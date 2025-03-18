#!/usr/bin/env python3
# coding: utf-8
"""
Read a Cellebrite email export parse it and export it out.
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import re
import sys
import argparse
import openpyxl
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Color support for Windows 11+
color_red = color_yellow = color_green = color_blue = color_purple = color_reset = ''
if sys.version_info > (3, 7, 9) and os.name == "nt":
    from colorama import Fore, Back, Style
    print(Back.BLACK)
    color_red, color_yellow, color_green = Fore.RED, Fore.YELLOW, Fore.GREEN
    color_blue, color_purple, color_reset = Fore.BLUE, Fore.MAGENTA, Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "Read a Cellebrite email export parse it and export it out."
version = '0.1.1'


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>


def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-r', '--read', help='read xlsx', required=False, action='store_true')

    args = parser.parse_args()


    global input_file
    input_file = args.input if args.input else "Cellebrite_Emails.xlsx"

    global outuput_xlsx
    outuput_xlsx = args.output if args.output else "Cellebrite_Emails_Parsed.xlsx"


    if args.read:
        # create_xlsx()

        file_exists = os.path.exists(input_file)
        if file_exists == True:
            data = read_xlsx(input_file)
            write_xlsx(data, outuput_xlsx)
        else:
            msg_blurb = (f'{input_file} does not exist')
            msg_blurb_square(msg_blurb, color_red)      
            exit()

    else:
        usage()
    
    return 0

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>


def convert_timestamp(timestamp, time_orig=None, timezone=None):
    if timezone is None:
        timezone = ''
    if time_orig is None:
        time_orig = timestamp

    timestamp = str(timestamp)

    # Regular expression to find the timezone
    timezone_pattern = r"([A-Za-z ]+)$"
    matches = re.findall(timezone_pattern, timestamp)
    
    if matches:
        timezone = matches[-1].strip()
        timestamp = timestamp.replace(timezone, "").strip()
    
    # Handling specific timezones in the timestamp
    if "(" in timestamp:
        timestamp, tz_info = timestamp.split('(')
        timezone = tz_info.replace(")", '').strip()
    elif " CDT" in timestamp:
        timezone = "CDT"
        timestamp = timestamp.replace(" CDT", "").strip()
    elif " CST" in timestamp:
        timezone = "CST"
        timestamp = timestamp.replace(" CST", "").strip()

    formats = [
        "%B %d, %Y, %I:%M:%S %p %Z",  # June 13, 2022, 9:41:33 PM CDT (Flock)
        "%Y:%m:%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",  # Timestamps without seconds
        "%m/%d/%Y %H:%M:%S",  # Military time without seconds
        "%m-%d-%y at %I:%M:%S %p %Z",  # e.g., 09-10-23 at 4:29:12 PM CDT
        "%m-%d-%y %I:%M:%S %p",
        "%B %d, %Y at %I:%M:%S %p %Z",
        "%B %d, %Y at %I:%M:%S %p",
        "%B %d, %Y %I:%M:%S %p %Z",
        "%B %d, %Y %I:%M:%S %p",
        "%Y-%m-%dT%H:%M:%SZ",  # ISO 8601 with UTC timezone
        "%Y/%m/%d %H:%M:%S",  # e.g., 2022/06/13 21:41:33
        "%d-%m-%Y %I:%M:%S %p",  # e.g., 13-06-2022 9:41:33 PM
        "%d/%m/%Y %H:%M:%S",  # e.g., 13/06/2022 21:41:33
        "%Y-%m-%d %I:%M:%S %p",  # e.g., 2022-06-13 9:41:33 PM
        "%Y%m%d%H%M%S",  # e.g., 20220613214133
        "%Y%m%d %H%M%S",  # e.g., 20220613 214133
        "%m/%d/%y %H:%M:%S",  # e.g., 06/13/22 21:41:33
        "%d-%b-%Y %I:%M:%S %p",  # e.g., 13-Jun-2022 9:41:33 PM
        "%d/%b/%Y %H:%M:%S",  # e.g., 13/Jun/2022 21:41:33
        "%Y/%b/%d %I:%M:%S %p",  # e.g., 2022/Jun/13 9:41:33 PM
        "%d %b %Y %H:%M:%S",  # e.g., 13 Jun 2022 21:41:33
        "%A, %B %d, %Y %I:%M:%S %p %Z",  # e.g., Monday, June 13, 2022 9:41:33 PM CDT
        "%A, %B %d, %Y %I:%M:%S %p"     # e.g., Monday, June 13, 2022 9:41:33 PM
    ]

    for fmt in formats:
        try:
            dt_obj = datetime.strptime(timestamp.strip(), fmt)
            return dt_obj, time_orig, timezone
        except ValueError:
            continue

    raise ValueError(f"Timestamp format not recognized for: {time_orig}")

def msg_blurb_square(msg, color):
    border = f"+{'-' * (len(msg) + 2)}+"
    print(f"{color}{border}\n| {msg} |\n{border}{color_reset}")

def read_xlsx(file_path):
    """Read data from an XLSX file and return a list of dictionaries."""
    msg_blurb = (f'Reading {input_file}')
    msg_blurb_square(msg_blurb, color_green) 
    
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        DateStamp, Discount100, Cash, FoodNet, FoodTax, HouseTotal = '', '', '', '', '', ''
        NetSales, DineIn, Togo, GrossSales, Voids, SmileDining = '', '', '', '', '', ''
        TotalCustomers, TimeOriginal, time_orig, timezone = '', '', '', ''
        business, TotalTogoTickets, TotalTogoTickets = '', '', ''
        CreditCard, Coupon, DicountTotal, DiscountEmployee, EmployeeDiscount = '', '', '', '', ''
        
        output_format = "%Y-%m-%d %H:%M:%S"  # Changed to ISO 8601 military time
        row_data = dict(zip(headers, row))

        # Body = row_data.get("Body = row_data.get("Body")") = row_data.get("Body = row_data.get("Body")")
        Time = row_data.get("Time")
        
        
        if Time != '':
            try:
                (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
                Time = Time.strftime(output_format)

                if Time is None:
                    Time = ''              
                
            except ValueError as e:
                print(f"Error time2: {e} - {Time}")
                Time = ''    # temp rem of this
                pass

        Body = row_data.get("Body")
        
        if Body is None:
            Body == 'test'
        else:

            # Ensure Body is a string
            if not isinstance(Body, str):
                Body = str(Body)

            print(f'Body type = {type(Body)}')

            # Extract variables using regex
            business_match = re.search(r"Shop Name:\s*(.+)", Body)
            DateStamp_match = re.search(r"CREDITCARD transactions on SmilePOS System have been settled on (.+?)\.", Body)
            Discount100_match = re.search(r"100% Off \(\d+\)\s+([\d,.]+)", Body)
            DiscountEmployee_match = re.search(r"Employee Discount 50% \(\d+\)\s+([\d,.]+)", Body)
            Cash_match = re.search(r"Cash Sales\s+\+?([\d,.]+)", Body)
            FoodNet_match = re.search(r"Food Net Sales\s+([\d,.]+)", Body)
            GiftCard_match = re.search(r"Gift Card Sales\s+([\d,.]+)", Body)
            GratuityChargeTax_match = re.search(r"Gratuity Charge Tax\s+([\d,.]+)", Body)
            CreditCard_match = re.search(r"Total Credit Card\s+([\d,.]+)", Body)
            Coupon_match = re.search(r"Total Coupon\s+([\d,.]+)", Body)
            DicountTotal_match = re.search(r"DISCOUNT\s+Total\s+([\d,.]+)", Body)
            PaymentsTotal_match = re.search(r"DISCOUNT\s+Total\s+([\d,.]+)", Body)
            FoodTax_match = re.search(r"Food Tax\s+([\d,.]+)", Body)
            HouseTotal_match = re.search(r"House Total \(\d+\)\s+([\d,.]+)", Body)
            NetSales_match = re.search(r"Net Sales\s+=\s+([\d,.]+)", Body)
            DineIn_match = re.search(r"Dine In\s+([\d,.]+)", Body)
            Togo_match = re.search(r"Togo\s+([\d,.]+)", Body)
            GrossSales_match = re.search(r"GROSS SALES BY REV CLASS\s+Total\s+([\d,.]+)", Body)
            Voids_match = re.search(r"VOIDS BY TYPE\s+Total\s+(-?[\d,.]+)", Body)
            SmileDining_match = re.search(r"Smile Dining\s+Total\s+([\d,.]+)", Body)
            TotalCustomers_match = re.search(r"Total Customers:\s+([\d,.]+)", Body)

            # Assign values to variables, handling missing cases
            business = business_match.group(1) if business_match else ""
            DateStamp = DateStamp_match.group(1) if DateStamp_match else ""
            Discount100 = Discount100_match.group(1) if Discount100_match else ""
            DiscountEmployee = DiscountEmployee_match.group(1) if DiscountEmployee_match else ""

            Cash = Cash_match.group(1) if Cash_match else ""
            FoodNet = FoodNet_match.group(1) if FoodNet_match else ""
            FoodTax = FoodTax_match.group(1) if FoodTax_match else ""
            HouseTotal = HouseTotal_match.group(1) if HouseTotal_match else ""
            NetSales = NetSales_match.group(1) if NetSales_match else ""
            DineIn = DineIn_match.group(1) if DineIn_match else ""
            Togo = Togo_match.group(1) if Togo_match else ""
            GrossSales = GrossSales_match.group(1) if GrossSales_match else ""
            Voids = Voids_match.group(1) if Voids_match else ""
            SmileDining = SmileDining_match.group(1) if SmileDining_match else ""
            TotalCustomers = TotalCustomers_match.group(1) if TotalCustomers_match else ""
            CreditCard = CreditCard_match.group(1) if CreditCard_match else ""
            Coupon = Coupon_match.group(1) if Coupon_match else ""
            DicountTotal = DicountTotal_match.group(1) if DicountTotal_match else ""
            PaymentsTotal = PaymentsTotal_match.group(1) if PaymentsTotal_match else ""

        if DateStamp != '':
            DateStamp = DateStamp.replace(' at ', ' ')
            try:
                (DateStamp, time_orig2, timezone) = convert_timestamp(DateStamp, time_orig, timezone)
                DateStamp = DateStamp.strftime(output_format)

                if DateStamp is None:
                    DateStamp = ''              
                
            except ValueError as e:
                print(f"Error time2: {e} - {DateStamp}")
                DateStamp = ''    # temp rem of this
                pass

        row_data["Time"] = Time  
        row_data["TimeOriginal"] = time_orig
        row_data["business"] = business
        row_data["Body"] = Body
        row_data["DateStamp"] = DateStamp
        row_data["Discount100%"] = Discount100
        row_data["Cash"] = Cash
        row_data["FoodNet"] = FoodNet
        row_data["FoodTax"] = FoodTax
        row_data["HouseTotal"] = HouseTotal
        row_data["NetSales"] = NetSales
        row_data["DineIn"] = DineIn
        row_data["Togo"] = Togo
        row_data["GrossSales"] = GrossSales
        row_data["Voids"] = Voids
        # row_data["SmileDining"] = SmileDining
        row_data["TotalCustomers"] = TotalCustomers       
        row_data["CreditCard"] = CreditCard       
        row_data["Coupon"] = Coupon       
        row_data["DicountTotal"] = DicountTotal       
        row_data["DiscountEmployee"] = DiscountEmployee  
        row_data["Payments"] = PaymentsTotal
        
        # row_data["EmployeeDiscount"] = EmployeeDiscount       

        data.append(row_data)
    
    wb.close()
    return data


def write_xlsx(data, file_path):
    '''
    The write_xlsx() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    msg_blurb = (f'Writing to {outuput_xlsx}')
    msg_blurb_square(msg_blurb, color_green)  

    worksheet.title = 'Sheet1'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = [
    "DateStamp", "GrossSales", "NetSales", "Payments", "Cash", "CreditCard", "Coupon", "DicountTotal", "Discount100%", "DiscountEmployee", "Voids", "FoodNet", "FoodTax", "HouseTotal", "DineIn", "Togo", "SmileDining", "TotalCustomers", "Body", "File", "Source file information", "MD5", "Source file information", "#", "Deleted", "Time", "From", "To", "Direction", "Date", "Subject", "Account", "Status", "Priority", "Folder", "Source", "Carved", "Manually decoded", "Investigation notes", "Email header", "Tag", "Type", "TimeOriginal", "business", "case"
    ]

    # Write headers to the first row0
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [6, 7, 8, 9, 10]:  #     # orange
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            cell.fill = fill
        elif col_index in [1, 2, 3, 4, 5, 11, 12, 13 , 14 , 15, 16]:  # Indices of columns A, C, D, E, F, G, H      # green
            fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            cell.fill = fill

    # Excel column width
    worksheet.column_dimensions['A'].width = 19# DateStemp
    worksheet.column_dimensions['B'].width = 11# GrossSales
    worksheet.column_dimensions['C'].width = 10# NetSales
    worksheet.column_dimensions['D'].width = 8# Cash
    worksheet.column_dimensions['E'].width = 11# 
    worksheet.column_dimensions['F'].width = 11# 
    worksheet.column_dimensions['G'].width = 8# 
    worksheet.column_dimensions['H'].width = 13# 
    worksheet.column_dimensions['I'].width = 13# 
    worksheet.column_dimensions['J'].width = 17# 
    worksheet.column_dimensions['N'].width = 11# 
    worksheet.column_dimensions['Q'].width = 13# 

    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f'error')

    workbook.save(outuput_xlsx)
    
def usage():
    print(f"Usage: {sys.argv[0]} -r [-I input.xlsx] [-O output.xlsx]")
    print("Example:")
    print(f"   python  {sys.argv[0]} -r")
    print(f"   python {sys.argv[0]} -r -I Cellebrite_Emails.xlsx -O Cellebrite_Emails_Parsed.xlsx")


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
0.0.1 - parse smilePOS emails as example 

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
delete the line count, first line that cellebrite puts in so the header row is the top row.


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>


if __name__ == '__main__':
    main()
