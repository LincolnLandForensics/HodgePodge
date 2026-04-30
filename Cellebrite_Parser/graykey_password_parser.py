#!/usr/bin/python
# coding: utf-8

import os
import re
import sys
import argparse
import openpyxl
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description = "convert graykey password file to xlsx"
version = '1.4.3'

headers = [
    "url", "user", "password", "note", "case", "exhibit", "protocol",
    "filetype", "encryption", "complexity", "hash", "pwd", "pwdumpformat", "length",
    "email", "ip", "created", "modified"
]


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>
import threading
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))
        self.widget.see("end")
        self.widget.configure(state="disabled")
        self.widget.update_idletasks()

    def flush(self):
        pass

def run_gui():
    root = tk.Tk()
    root.title(f"GrayKey Password Parser {version}")
    root.geometry("600x650")

    # Style
    style = ttk.Style()
    style.theme_use('clam')

    main_frame = ttk.Frame(root, padding="10")
    main_frame.pack(fill=tk.BOTH, expand=True)

    # Title Label
    ttk.Label(main_frame, text="Convert GrayKey Passwords.txt to Excel", font=("Helvetica", 14, "bold")).pack(pady=10)

    # Input File
    input_frame = ttk.LabelFrame(main_frame, text="Input File", padding="5")
    input_frame.pack(fill=tk.X, pady=5)
    
    input_entry = ttk.Entry(input_frame)
    input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    
    default_input = "sample_passwords.txt" 
    if os.path.exists(default_input):
        input_entry.insert(0, default_input)
    else:
        # try to find sample_passwords.txt in current directory if different
        current_dir_sample = os.path.join(os.getcwd(), "sample_passwords.txt")
        if os.path.exists(current_dir_sample):
             input_entry.insert(0, current_dir_sample)

    def browse_input():
        filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Input File",
                                            filetypes=(("Text files", "*.txt"), ("All files", "*.*")))
        if filename:
            input_entry.delete(0, tk.END)
            input_entry.insert(0, filename)
            
            # Auto-suggest output name if possible
            if not output_entry.get():
                base_name = os.path.splitext(os.path.basename(filename))[0]
                output_entry.insert(0, f"passwords_{base_name}.xlsx")

    ttk.Button(input_frame, text="Browse", command=browse_input).pack(side=tk.RIGHT, padx=5)

    # Case & Exhibit
    info_frame = ttk.Frame(main_frame)
    info_frame.pack(fill=tk.X, pady=5)
    
    ttk.Label(info_frame, text="Case Number:").pack(side=tk.LEFT, padx=5)
    case_entry = ttk.Entry(info_frame, width=15)
    case_entry.pack(side=tk.LEFT, padx=5)
    
    ttk.Label(info_frame, text="Exhibit:").pack(side=tk.LEFT, padx=5)
    exhibit_entry = ttk.Entry(info_frame, width=15)
    exhibit_entry.pack(side=tk.LEFT, padx=5)

    # Output File
    output_frame = ttk.LabelFrame(main_frame, text="Output File", padding="5")
    output_frame.pack(fill=tk.X, pady=5)
    
    output_entry = ttk.Entry(output_frame)
    output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    
    def browse_output():
        filename = filedialog.asksaveasfilename(initialdir=os.getcwd(), title="Select Output File",
                                              defaultextension=".xlsx",
                                              filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        if filename:
            output_entry.delete(0, tk.END)
            output_entry.insert(0, filename)

    ttk.Button(output_frame, text="Browse", command=browse_output).pack(side=tk.RIGHT, padx=5)
    
    # Update output filename dynamically when Case/Exhibit changes if output is empty or default format
    def update_output_name(*args):
        c = case_entry.get().strip()
        e = exhibit_entry.get().strip()
        current_out = output_entry.get().strip()
        if not current_out or current_out.startswith("passwords_"):
             if c or e:
                 new_name = f"passwords_{c}_Ex_{e}.xlsx"
                 output_entry.delete(0, tk.END)
                 output_entry.insert(0, new_name)
    
    case_entry.bind("<KeyRelease>", update_output_name)
    exhibit_entry.bind("<KeyRelease>", update_output_name)

    # Progress Bar
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(main_frame, variable=progress_var, maximum=100)
    progress_bar.pack(fill=tk.X, pady=10)

    # Message Window
    log_frame = ttk.LabelFrame(main_frame, text="Log", padding="5")
    log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
    
    log_text = scrolledtext.ScrolledText(log_frame, height=10, state='disabled')
    log_text.pack(fill=tk.BOTH, expand=True)

    # Convert Button
    def start_conversion():
        input_file_path = input_entry.get().strip()
        output_file_path = output_entry.get().strip()
        case_val = case_entry.get().strip()
        exhibit_val = exhibit_entry.get().strip()
        
        if not input_file_path:
            messagebox.showerror("Error", "Please select an input file.")
            return
        if not os.path.exists(input_file_path):
             messagebox.showerror("Error", f"Input file not found: {input_file_path}")
             return
        if not output_file_path:
            # Default fallback
            output_file_path = f"passwords_{case_val}_Ex_{exhibit_val}.xlsx"

        convert_btn.config(state=tk.DISABLED)
        progress_var.set(0)
        log_text.configure(state='normal')
        log_text.delete(1.0, tk.END)
        log_text.configure(state='disabled')
        
        # Redirect stdout
        old_stdout = sys.stdout
        sys.stdout = TextRedirector(log_text)
        
        def run_thread():
            try:
                print(f"Starting conversion...")
                print(f"Input: {input_file_path}")
                print(f"Output: {output_file_path}")
                
                # Create a callback to update progress
                # Since we don't know total lines easily without reading first, we'll pulse or just set to 50% then 100%
                # Or we can just use indeterminate. For now let's just use manual updates.
                progress_var.set(10)
                
                read_pwords(input_file_path, output_file_path, case_val, exhibit_val)
                
                progress_var.set(100)
                print("Done")
                # messagebox.showinfo("Success", "Conversion Complete!")
            except Exception as e:
                print(f"Error: {e}")
                messagebox.showerror("Error", str(e))
            finally:
                sys.stdout = old_stdout
                convert_btn.config(state=tk.NORMAL)
        
        threading.Thread(target=run_thread, daemon=True).start()

    convert_btn = ttk.Button(main_frame, text="Convert", command=start_conversion)
    convert_btn.pack(pady=10)

    root.mainloop()

def main():
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description=description)
        parser.add_argument('-I', '--input', help='', required=False)
        parser.add_argument('-O', '--output', help='', required=False)
        parser.add_argument('-b', '--blank', help='create blank sheet', required=False, action='store_true')
        # parser.add_argument('-p', '--passwords', help='passwords module', required=False, action='store_true')
        parser.add_argument('-c', '--convert', help='convert GrayKey passwords to Excel', required=False, action='store_true')
    
        args = parser.parse_args()
    
        input_f = args.input if args.input else "sample_passwords.txt"
    
        if args.convert:
            encoded_case = input("Enter Case: ").strip()
            encoded_exhibit = input("Enter Exhibit: ").strip()
            output_f = args.output if args.output else (f"passwords_{encoded_case}_Ex_{encoded_exhibit}.xlsx")
            read_pwords(input_f, output_f, encoded_case, encoded_exhibit)
        elif args.blank:
            output_f = 'blank_password_sheet.xlsx'
            write_xlsx([], [], output_f)
            sys.exit(0)
        else:
            usage()
    else:
        run_gui()

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def complexinator(password):
    if not password:
        return "blank"

    length_ok = len(password) >= 8
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(not c.isalnum() for c in password)
    complexity_criteria = sum([has_upper, has_lower, has_digit, has_special])

    return "complex" if length_ok and complexity_criteria >= 3 else "weak"


import re

def convert_ISO8601_basic(time_str: str) -> str:
    """
    Convert timestamps like:
        20250929194037.434012Z
        20250220053502.8158187Z
        20250929194037Z
        20250929194037
    into: YYYY-MM-DD HH:MM:SS
    """

    s = time_str.strip().rstrip("Z")

    # Regex to split main timestamp and fractional seconds
    m = re.match(r"^(\d{14})(?:\.(\d+))?$", s)
    if not m:
        raise ValueError(f"Malformed ISO8601 basic timestamp: {time_str}")

    main = m.group(1)
    frac = m.group(2)

    # Normalize fractional seconds to microseconds (6 digits)
    if frac:
        if len(frac) > 6:
            frac = frac[:6]          # truncate
        else:
            frac = frac.ljust(6, "0")  # pad
        clean = f"{main}.{frac}"
        fmt = "%Y%m%d%H%M%S.%f"
    else:
        clean = main
        fmt = "%Y%m%d%H%M%S"

    dt = datetime.strptime(clean, fmt)
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def flip_if_reverse_dns(s: str) -> str:
    parts = s.split('.')

    # Detect reverse-DNS style
    if len(parts) >= 3 and parts[0] in {"com", "net", "org", "edu", "gov", "io"}:
        return '.'.join(reversed(parts))

    return s



def message_square(message):
    print(f"| {message} |")

def read_pwords(in_file, out_file, case_val, exhibit_val):
    if not os.path.isfile(in_file):
        print(f"Error: Input file '{in_file}' does not exist.")
        # sys.exit(1) # Don't exit in GUI
        return
    # else:
        # message_square(f'Reading {in_file}')

    email_pattern = re.compile(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$')


    data, uniq = [], set()
    fileType = in_file # Just use filename or full path
    fileType = os.path.basename(fileType)
    pattern = re.compile(r'^\d{9}\.\d{6}$')
    known_bad_passwords = {
        'false', 'true', 'US', 'Secret', '0', '1', 'treeup', 'mobile', '""',
        'ATM,CHK', 'myPSKkey', 'PERSONAL', 'Registered', 'stayPaired', 'POH',
        'PER', 'PR', '10', '09EA', 'ATM+CHK', 'kcKeepDeviceTrusted', 'dummy_value',
        '2', '4', '[]', '{}', 'YES', 'prod', 'reinstall_value', 'IS_LATEST_KEY_V2',
        'comcast-business', 'VAL_KeychainCanaryPassword', 'TwitterKeychainCanaryPassword'
    }

    try:
        with open(in_file, 'r', encoding='utf-8') as f:
            content = f.read()
            # content = content.replace("IP: N/A", "----------")  # for intel.veraxity.org output
            content = re.sub(r'^(IP: .*)', r'\1----------', content, flags=re.MULTILINE)
    
            entries = content.split("----------")
    
            for block in entries:
                entry = {
                    "url": '', "user": '', "password": '', "note": block.strip(),
                    "case": case_val, "exhibit": exhibit_val, "protocol": '', "filetype": fileType,
                    "encryption": '', "complexity": '', "hash": '', "pwd": '',
                    "pwdumpformat": '', "length": '', "email": '', "ip": '', "created": '', "modified": ''
                }
    
                for line in block.strip().splitlines():
                    line = line.strip()
                    if line.startswith("Account:"):
                        user = line.split("Account:", 1)[1].strip()
                        entry["user"] = user
                        if email_pattern.match(user):
                            email = user
                            entry["email"] = email
                            
                            entry["protocol"] = "SMTP"
                    elif line.startswith("Creation Date: "):
                        try:
                            created = line.split("Creation Date: ", 1)[1].strip()
                            created = convert_ISO8601_basic(created)
                            entry["created"] = created
                        except Exception as e:
                            print(f"Error: {e}")                        
                                                
                    elif line.startswith("Modification Date: "):
                        try:
                            modified = line.split("Modification Date: ", 1)[1].strip() 
                            modified = convert_ISO8601_basic(modified) 
                            entry["modified"] = modified
                        except Exception as e:
                            print(f"Error: {e}")
                        
                    elif line.startswith("srvr: "):
                        url = line.split("srvr: ", 1)[1].strip()
                        url = flip_if_reverse_dns(url)
                        entry["url"] = url
                        
                    elif line.startswith("ptcl: "):
                        protocol = line.split("ptcl: ", 1)[1].strip()
                        if protocol != "0":
                            entry["protocol"] = protocol
                    elif line.startswith("Service: "):
                        url = line.split("Service: ", 1)[1].strip()
                        url = flip_if_reverse_dns(url)
                        entry["url"] = url
                        
                        
                    elif line.startswith("Item value:"):
                        pwd = line.replace("Item value:", '').strip()
                        if pwd in known_bad_passwords or \
                           pwd.endswith('.com') or \
                           pwd.startswith('[{') or \
                           pwd.startswith('{"') or \
                           pwd.startswith('|DYN') or \
                           pwd.startswith('us-east') or \
                           pwd.startswith('http') or \
                           pwd.endswith("=") or \
                           pwd.endswith("~~") or \
                           "whatsapp.net" in pwd or \
                           len(pwd) > 33 or pattern.match(pwd):
                            hash_val = pwd
                            if email_pattern.match(hash_val):
                                email = hash_val
                                entry["email"] = email
                                entry["user"] = email
                                hash_val = ''
                            entry["hash"] = hash_val
                        else:
                            entry["password"] = pwd
                    elif line.startswith("Username: "):
                        entry["user"] = line.split("Username: ", 1)[1].strip().replace('N/A','')
                    elif line.startswith("Email: "):
                        entry["email"] = line.split("Email: ", 1)[1].strip().replace('N/A','')
                    elif line.startswith("Password: "):
                        entry["password"] = line.split("Password: ", 1)[1].strip().replace('N/A','')
                    elif line.startswith("Origin: "):
                        entry["url"] = line.split("Origin: ", 1)[1].strip().replace('N/A','') 
                        entry["filetype"] = "intel.veraxity.org"
                    elif line.startswith("IP: "):
                        entry["ip"] = line.split("IP: ", 1)[1].strip().replace('N/A','')
    
                if entry["url"] == "AirPort":
                    entry["protocol"] = "AirPort"
                elif "com.apple.airplay" in entry["url"]:
                    entry["protocol"] = "AirPlay"
                elif entry["url"] == "GuidedAccess":
                    entry["url"] = "_phone pin code ***"                
                    entry["user"] = "" 
                    
                if any(k in entry["user"].lower() for k in ["apikey", "token", "sessionkey"]) or \
                   entry["user"].startswith('com.') or entry["user"] in ["UUID", "secretKey", "acquiredPackages"]:
                    entry["hash"] = entry["password"]
                    entry["password"] = ''
                    entry["user"] = ''
    
                if entry["password"]:
                    entry["length"] = len(entry["password"])
                    entry["complexity"] = complexinator(entry["password"])
                    if entry["password"] not in uniq:
                        uniq.add(entry["password"])
    
                data.append(entry)
    
        data = sorted(data, key=lambda x: (x["length"] if isinstance(x["length"], int) else 100))
        write_xlsx(data, sorted(uniq, key=len), out_file)
        
    except Exception as e:
        print(f"Error reading file: {e}")
        # In GUI mode, this print will go to the message box
        # traceback.print_exc() if imported


def write_xlsx(data, uniq_list, out_filename):
    # print(f'Writing {out_filename}')
    # message_square(f'Writing {out_filename}')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Passwords'
    worksheet.freeze_panes = 'B2'
    worksheet.selection = 'B2'

    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if header in ["user", "password", "exhibit", "case", "note"]:
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        elif header in ["url", "length", "complexity"]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    col_widths = [20, 20, 20, 35, 9, 6, 10, 20, 9, 12, 5, 17, 5, 8, 15, 15, 18, 18]
    for i, width in enumerate(col_widths):
        worksheet.column_dimensions[chr(65+i)].width = width


    for row_index, row_data in enumerate(data):
        for col_index, col_name in enumerate(headers):
            worksheet.cell(row=row_index + 2, column=col_index + 1).value = row_data.get(col_name, '')

    # Create second sheet with unique passwords
    # Uniq Sheet
    uniq_sheet = workbook.create_sheet(title="uniq")
    uniq_sheet.freeze_panes = 'B2' 
    uniq_sheet['A1'] = 'Unique Passwords (Sorted by Length)'
    uniq_sheet.column_dimensions['A'].width = 40        
    # uniq_sheet.append(["Password"])
    for password in uniq_list:
        uniq_sheet.append([password])

    workbook.save(out_filename)

def usage():
    file = os.path.basename(sys.argv[0])
    print("\nDescription: " + description)
    print(f"{file} Version: {version} by {author}")
    print("\nExample:")
    print(f"\t{file} -c -I sample_passwords.txt")
    print(f"\t{file} -c -I sample_passwords.txt -O passwords_sample_.xlsx")
    print(f"\t{file} -b -O blank_sheet.xlsx")
    print(f"\t{file} -v -I input.txt")
    
if __name__ == '__main__':
    main()


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
0.5.0 - created intel.veraxity.org parser
0.4.0 - create a seperate sheet for uniq passwords
0.2.2 - working prototype
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
also create an intel sheet



"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""



"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
