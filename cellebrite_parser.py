 #!/usr/bin/env python3
# coding: utf-8
'''
read xlsx, write xlsx with only openpyxl
Pandas is too big
send data all at once so it can be sorted if needed
'''
# <<<<<<<<<<<<<<<<<<<<<<<<<<      Imports        >>>>>>>>>>>>>>>>>>>>>>>>>>

import os
import re
import sys
import time
import openpyxl
import simplekml
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import argparse  # for menu system
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# Colorize section
global color_red
global color_yellow
global color_green
global color_blue
global color_purple
global color_reset
color_red = ''
color_yellow = ''
color_green = ''
color_blue = ''
color_purple = ''
color_reset = ''

if sys.version_info > (3, 7, 9) and os.name == "nt":
    version_info = os.sys.getwindowsversion()
    major_version = version_info.major
    build_version = version_info.build

    if major_version >= 10 and build_version >= 22000: # Windows 11 and above
        import colorama
        from colorama import Fore, Back, Style  
        print(f'{Back.BLACK}') # make sure background is black
        color_red = Fore.RED
        color_yellow = Fore.YELLOW
        color_green = Fore.GREEN
        color_blue = Fore.BLUE
        color_purple = Fore.MAGENTA
        color_reset = Style.RESET_ALL

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = "convert Cellebrite contacts, account, web history, chats and call exports to intel format"
version = '1.0.1'

# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    global row
    row = 0  # defines arguments
    # Row = 1  # defines arguments   # if you want to add headers 
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='', required=False)
    parser.add_argument('-O', '--output', help='', required=False)
    parser.add_argument('-b', '--blank', help='create blank input sheet', required=False, action='store_true')
    parser.add_argument('-C', '--cellebrite', help='cellebrite contacts parse', required=False, action='store_true')

    args = parser.parse_args()
    data = []
    global outuput_xlsx
    outuput_xlsx = args.output

    if args.blank:
        data = []
        print(f'{color_green}Writing to {outuput_xlsx} {color_reset}')
        write_xlsx(data)

    elif args.cellebrite:
        if not args.input: 
            input_xlsx = "Contacts.xlsx"        
        else:
            input_xlsx = args.input
            
        file_exists = os.path.exists(input_xlsx)

        datatype = input_xlsx
        datatype = datatype.replace('.xlsx', '')

        if not args.output: 
            outuput_xlsx = (f'intel_{datatype}.xlsx') 
  
        else:
            outuput_xlsx = args.output


        if file_exists == True:
            print(f'{color_green}Reading {input_xlsx} {color_reset}')

            data = read_cellebrite(input_xlsx)
            write_xlsx(data)

            workbook.close()
            print(f'{color_green}Writing to {outuput_xlsx} {color_reset}')

        else:
            print(f'{color_red}{input_xlsx} does not exist{color_reset}')
            exit()

    else:
        usage()
    
    return 0


# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def convert_timestamp(timestamp, time_orig, timezone):
    if timezone is None:
    # if timezone != "" and timezone is not None:
        timezone = ''

    # time_data = timestamp
    timestamp = str(timestamp)

    if time_orig == "":
        time_orig = timestamp

    timestamp = timestamp.replace(' at ', ' ')
    if "(" in timestamp:
        timestamp = timestamp.split('(')
        timezone = timestamp[1].replace(")", '')
        timestamp = timestamp[0]
    elif " CDT" in timestamp:
        timezone = "CDT"
        timestamp = timestamp.replace(" CDT", "")
    elif " CST" in timestamp:
        timezone = "CST"
        timestamp = timestamp.replace(" CST", "")

    formats = [
        "%B %d, %Y, %I:%M:%S %p %Z",
        "%Y:%m:%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",  # timestamps without seconds
        "%m/%d/%Y %H:%M:%S",  # timestamps in military time without seconds
        "%B %d, %Y at %I:%M:%S %p %Z",
        "%B %d, %Y at %I:%M:%S %p CST",
        "%B %d, %Y at %I:%M:%S %p",
        "%B %d, %Y %I:%M:%S %p %Z",
        "%B %d, %Y %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p %Z",
        "%Y-%m-%dT%H:%M:%SZ"  # ISO 8601 format with UTC timezone
    ]

    for fmt in formats:
        try:
            dt_obj = datetime.strptime(timestamp.strip(), fmt)
            timestamp = dt_obj
            # return dt_obj, time_orig, timezone
            return timestamp, time_orig, timezone
                        
        except ValueError:
            pass

    raise ValueError(f"{time_orig} Timestamp format not recognized")


def read_cellebrite(input_xlsx):

    """Read data from an xlsx file and return as a list of dictionaries.
    Read XLSX Function: The read_xlsx() function reads data from the input 
    Excel file using the openpyxl library. It extracts headers from the 
    first row and then iterates through the data rows.    
    """

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    data = []
    datatype = input_xlsx
    datatype = datatype.replace('.xlsx', '')
    
    # get header values from first row
    global headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)


# regex patterns
    phone1_pattern = r'Phone-Mobile: (\S+)'
    phone2_pattern = r'Phone-: (\S+)'
    phone3_pattern = r'Phone-Phone Number: (\S+)'
    phone4_pattern = r'Phone-Phone: (\S+)'
    phone5_pattern = r'Phone-General: (\S+)'
    phone6_pattern = r'Phone-Home: (\S+)'
    phone7_pattern = r'User ID-User ID: (\S+)'

    email1_pattern = r'Email-: (\S+)'
    email2_pattern = r'Email-Email Address: (\S+)'
    email3_pattern = r'Email-General: (\S+)'
     
     
    dob1_pattern = r'User ID-Birthday: (\S+)'
    alias1_pattern = r'User ID-Additional Name: (\S+)'
    alias2_pattern = r'User ID-Push Name: (\S+)'
    alias3_pattern = r'User ID-User Name: (\S+)'

    user1_pattern = r'User ID-WhatsApp User Id: (\S+)'
    user2_pattern = r'User ID-Id: (\S+)'
    user3_pattern = r'User ID-Username: (\S+)'    
    user4_pattern = r'User ID-WeChat ID: (\S+)'    
    user5_pattern = r'User ID-Facebook Id: (\S+)'    # test

    url1_pattern = r'Profile Picture-: (\S+)'
    url2_pattern = r'User ID-Tango ID: (\S+)'

    misc1_pattern = r'User ID-Identifier: (\S+)'
    misc2_pattern = r'User ID-Instagram Id: (\S+)'


    # get data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    if not data:
        print(f"{color_red}No data found in the Excel file.{color_reset}")
        return None

    for row_index, row_data in enumerate(data):
        (fullname, url, phone, email, business, misc) = ('', '', '', '', '', '')
        (lastname, firstname) = ('', '')
        (Time, time_orig, timezone) = ('', '', '')
        (ranking, source_file, origin_file) = (f'4 - {datatype}', '', '')
        (fulladdress , note, info, source, user, query) = ('', '', '', '', '', '')
        (aka, DOB, otheremails, titleurl, dnsdomain, country) = ('', '', '', '', '', '')
        (Latitude, Longitude, Coordinate, fulladdress2, state, Time) = ('', '', '', '', '', '')
        
        
# fullname        
        ## replace all None values with '' 
        fullname = row_data.get("Name")
        if fullname is None:
            fullname = ''
        elif fullname.startswith('+'): 
            fullname = fullname.replace('+', '')
            phone = (f'{phone}\n{fullname}')    # task
            fullname = ''

# url   
        url = row_data.get("URL")
        if url is None:
            url = ''        

# title
        title = row_data.get("Title")
        if title is None:
            title = ''
        else:
            # (ranking) = ('4 - Chats')
            titleurl = title
            Time = row_data.get("Last Visited-Time")
            
            if Time is None:
                Time = ''
            
        if url is None:
            url = ''
        else:
        
            dnsdomain = re.compile(r":\/\/(.*?)\/")

            pattern = re.compile(r":\/\/(.*?)\/")
            dnsdomain = pattern.search(url)

            if dnsdomain:
                dnsdomain = dnsdomain.group(1)
                # ranking = ("4 - WebHistory")    




# info        
        ## replace all None values with '' 
        
        info = row_data.get("Entries")
        info2 = row_data.get("Body")
        if info is None:
            info = ''
        if info2 is None:
            info2 = ''

        if "marlboro" in info2.lower() or "menthol" in info2.lower():
            note = "Cigarette"
        # elif "menthol" in info2.lower():
            # note = "Cigarette"
        elif "parliament" in info2.lower() or "newport" in info2.lower():
            note = "Cigarette"


        # Use regular expressions to find matches
        phone1_match = re.search(phone1_pattern, info)
        phone2_match = re.search(phone2_pattern, info)
        phone3_match = re.search(phone3_pattern, info)
        phone4_match = re.search(phone4_pattern, info)
        phone5_match = re.search(phone5_pattern, info)
        phone6_match = re.search(phone6_pattern, info)
        phone7_match = re.search(phone7_pattern, info)


        email1_match = re.search(email1_pattern, info)
        email2_match = re.search(email2_pattern, info)
        email3_match = re.search(email3_pattern, info)        
        
        dob1_match = re.search(dob1_pattern, info)
        alias1_match = re.search(alias1_pattern, info)
        alias2_match = re.search(alias2_pattern, info)
        alias3_match = re.search(alias3_pattern, info)
        misc1_match = re.search(misc1_pattern, info)
        misc2_match = re.search(misc2_pattern, info)

        user1_match = re.search(user1_pattern, info)
        user2_match = re.search(user2_pattern, info)
        user3_match = re.search(user3_pattern, info)
        user4_match = re.search(user4_pattern, info)
        user5_match = re.search(user5_pattern, info)

        url1_match = re.search(url1_pattern, info)
        url2_match = re.search(url2_pattern, info)


        # Extract the values if matches are found
        phone1 = phone1_match.group(1) if phone1_match else None
        phone2 = phone2_match.group(1) if phone2_match else None
        phone3 = phone3_match.group(1) if phone3_match else None
        phone4 = phone4_match.group(1) if phone4_match else None
        phone5 = phone5_match.group(1) if phone5_match else None
        phone6 = phone6_match.group(1) if phone6_match else None
        phone7 = phone7_match.group(1) if phone7_match else None


        email1 = email1_match.group(1) if email1_match else None
        email2 = email2_match.group(1) if email2_match else None
        email3 = email3_match.group(1) if email3_match else None

        dob1 = dob1_match.group(1) if dob1_match else None

        alias1 = alias1_match.group(1) if alias1_match else None
        alias2 = alias2_match.group(1) if alias2_match else None
        alias3 = alias3_match.group(1) if alias3_match else None

        misc1 = misc1_match.group(1) if misc1_match else None
        misc2 = misc2_match.group(1) if misc2_match else None

        user1 = user1_match.group(1) if user1_match else None
        user2 = user2_match.group(1) if user2_match else None
        user3 = user3_match.group(1) if user3_match else None
        user4 = user4_match.group(1) if user4_match else None
        user5 = user5_match.group(1) if user5_match else None

        url1 = url1_match.group(1) if url1_match else None
        url2 = url2_match.group(1) if url2_match else None

# body  
        if info == "" and info2 != "":
            info = info2
            ranking == "3 - Chats"
            Time = row_data.get("Timestamp: Time")
            if Time is None:
                Time = ''
            if info.startswith("http"):
                url = info

                # Parse the URL
                parsed_url = urlparse(url)

                # Extract address
                address_param = parse_qs(parsed_url.query).get('address', [''])[0]
                fulladdress2 = unquote(address_param)

                if ", IL " in fulladdress2:
                    state = "IL"
                elif ", MO " in fulladdress2:
                    state = "MO"
                elif ", CA " in fulladdress2:
                    state = "CA"
                elif ", Fl " in fulladdress2:
                    state = "FL"
                
                # Extract GPS coordinates
                if "google.com/maps" in url:
                    print(f'map = {url}')   # temp
                    # Define a regular expression pattern to extract coordinates
                    patternMaps2 = re.compile(r'@([-+]?\d{1,2}\.\d+),([-+]?\d{1,3}\.\d+)')
                    # Search for the pattern in the URL
                    matchMaps2 = patternMaps2.search(url)

                    if matchMaps2:
                        note = (f'map,{note}')
                        Latitude = float(matchMaps2.group(1))
                        Longitude = float(matchMaps2.group(2))
                        Coordinate = (f'{Latitude},{Longitude}')
                        print(f"Latitude: {Latitude}")  # temp
                        print(f"Longitude: {Longitude}")    # temp
    
    
                elif "maps.apple.com" in url:
                    note = (f'map,{note}')
                    Coordinate = parse_qs(parsed_url.query).get('ll', [''])[0]
                    if "," in Coordinate:
                        Latitude, Longitude = map(str, Coordinate.split(','))


# phone
        if phone1 is not None:
            phone = (f'{phone}\n{phone1}')
        elif phone2 is not None:
            phone = (f'{phone}\n{phone2}')
        elif phone3 is not None:
            phone = (f'{phone}\n{phone3}')
        elif phone4 is not None:
            phone = (f'{phone}\n{phone4}')
        elif phone5 is not None:
            phone = (f'{phone}\n{phone5}')
        elif phone6 is not None:
            phone = (f'{phone}\n{phone6}')
        elif phone7 is not None:
            phone = (f'{phone}\n{phone7}')            
            
        if phone is None:  
            phone = ''
        phone = phone.replace(' ','').replace('+','')
        phone = ''.join(char for char in phone if char.isalnum())
        if len(phone) >= 7:
            phone = phone   
        else:
            phone = ''

        phone = phone.replace('-', '')

# email
        if email1 is not None:
            email = (f'{email}\n{email1}')
        elif email2 is not None:
            email = (f'{email}\n{email2}')
        elif email3 is not None:
            email = (f'{email}\n{email3}')            
        email = email.strip()

# Parties        
        Parties = row_data.get("Parties")
        if Parties is None:
            Parties = ''
        else:
            misc = row_data.get("Time")
            if misc is None:
                misc = ''            
            
            info = (f'{row_data.get("Direction")} {row_data.get("Status")} {row_data.get("Duration")}')
            Parties = Parties.replace('From:  ', '').replace('To:  ', '').lstrip('+')
            if " " in Parties:
                Parties = Parties.split(' ', 1)
                
                if "@" in Parties[0]:
                    email = Parties[0]
                elif ":" in Parties[0]:
                    user = Parties[0]
                else:
                    phone = Parties[0]
                fullname = Parties[1]
                
# Position        
        Position = row_data.get("Position")
        if Position is None:
            Position = ''         
        
        try:
            Position = Position.replace(' ', '')
        except Exception as e:
            print(f"{color_red}Error printing line: {str(e)}{color_reset}")
        
        if Position is None:
            Position = ''
        else:
            Position = Position.replace('(', '').replace(')', '').strip()
            Position = str(Position)
            Coordinate = Position
        if Latitude == '' and ',' in Position:
            Position = Position.split(',')
            Latitude = Position[0]
            Longitude= Position[1]
            # print(f'Latitude {Latitude}') # temp

            
# user
        if user1 is not None:
            if "@s.whatsapp.net" in user1 and phone == '':
                otheremails = user1
                phone = user1.replace("@s.whatsapp.net","")
            else:
                user = (f'{user}\n{user1}')
        elif user2 is not None:
            user = (f'{user}\n{user2}')
        elif user3 is not None:
            user = (f'{user}\n{user3}')
        elif user4 is not None:
            user = (f'{user}\n{user4}')            
        elif user5 is not None:
            user = (f'{user}\n{user5}')  
            url = (f'https://facebook.com/{user5}')
        user = user.strip()
        if user is None or user == "None":
            user = ''

# username
        username = row_data.get("Username")
        if username is None or username == 'local':
            username = ''
        elif '@' in username:  # and email != "":
            email = username.strip('')
        elif user == '':
            user = username.strip('')

# url
        if url1 is not None:
            url = (f'{url}\n{url1}')
        if url2 is not None:
            url = (f'https://www.tango.me/stream/{url2}')
        if url is None or url == "None":
            url = ''

        url = url.strip()

        
# DOB
        if dob1 is not None:
            DOB = (f'{email}\n{dob1}')
        DOB = DOB.strip('')

# aka
        if alias1 is not None:
            aka = (f'{aka}\n{alias1}')
        elif alias2 is not None:
            aka = (f'{aka}\n{alias2}')
        elif alias3 is not None:
            aka = (f'{aka}\n{alias3}')
        aka = aka.strip('')

# misc

        if misc1 is not None:
            misc = (f'{misc}\n{misc1}')
        # elif misc2 is not None:
            # misc = (f'{misc}\n{misc2}')
        # misc = misc.strip()
        # if email == '' and "@" in misc:
        if misc is not None and "@" in misc:
            email = misc   
            misc = ''
        # if phone == '' and misc.startswith("+"):
        if phone is None:
            phone = ''
        if phone is not None and phone == '' and misc is not None and "+" in misc:
        # if phone == '' and "+" in misc:
            phone = misc   
            phone = phone.replace("+", '')
            misc = ''

        misc3 = row_data.get("Participants")
        if misc3 is None:
            misc3 = ''
        misc3 = misc3.strip('')
        
        if misc == '' and misc3 != "":
            misc = misc3

            
# business        
        ## replace all None values with '' 
        business = row_data.get("Organizations")
        if business is None:
            business = ''
        business = business.strip('')
        
# fulladdress        
        ## replace all None values with '' 
        fulladdress = row_data.get("Addresses")
        if fulladdress is None:
            fulladdress = ''
        fulladdress = fulladdress.replace('Home: ','').strip()
        # fulladdress = fulladdress.strip()

        if fulladdress == '' and fulladdress2 != '':
            fulladdress = fulladdress2

# Map Address
        fulladdress3 = row_data.get("Map Address")
        if fulladdress3 is None:
            fulladdress3 = ''
            query = row_data.get("Value")
            if query is None:
                query = ''
            
            
            # query = query.strip()
            
        else:
            fulladdress3 = fulladdress3.strip()




        if len(fulladdress) < 2:
            fulladdress = fulladdress3

# country        
        ## replace all None values with '' 
        country = row_data.get("Country code")
        if country is None:
            country = ''
        country = country.strip()
        
# tag
        tag = row_data.get("Tag")
        if tag is None:
            tag = ''       

        tag2 = row_data.get("Tag Note - Chat")
        if tag2 is None:
            tag2 = ''
        # else:
            
        if tag == "" and tag2 != "":
            tag2 = tag2.replace('Tags: ', '').strip()
            # print(f'tag2 = {tag2}') # temp
            tag = tag2
            if "Important" in tag:
                tag = "Important"
            elif "Review" in tag:
                tag = "Review"        

# ranking
        ranking2 = row_data.get("Service Type")

        if ranking2 is None:
            ranking = (f'{ranking} - {ranking2}')
            ranking = ranking.replace("4 -", "3 -")

        if ranking is None:
            ranking = ''
        ranking = ranking.replace(" - None ", "")

# source
        source = row_data.get("Source")
        if source is None:
            source = ''
        elif ranking2 is not None:
            ranking = (f'{ranking} - {ranking2}')
            ranking = ranking.replace("4 -", "3 -")
        else:
            ranking = (f'{ranking} - {source}')
   
        if " - None" in ranking:
            ranking = ranking.replace(' - None', '')    

# source file
        source_file = row_data.get("Source file information")
        if source_file is None:
            source_file = ''

# origin_file
        origin_file = row_data.get("origin_file")
        if origin_file is None or origin_file == "":
            origin_file = input_xlsx



# type_data
        type_data = row_data.get("Type")
        if type_data is None:
            type_data = ''
        if type_data == "":
            if "Searched" in origin_file:
               type_data = "Searched"
            elif "Chats" in origin_file:
               type_data = "Chats"


# Icon    
        Icon = row_data.get("Icon")
        if Icon is None:
            Icon = ''
        if Icon == "":
            if "Searched" in origin_file:
               Icon = "Searched"            
            elif "Chats" in origin_file:
               Icon = "Chats"  
               
# misc time cleanup

# Time
        Time = row_data.get("Time")
        if Time is None:
            Time = ''

        if Time == '':
            Time = row_data.get("Timestamp: Time")
            if Time is None:
                Time = ''

# convert time
        # output_format = "%m/%d/%Y %H:%M:%S"  # Changed to military time
        output_format = "%Y/%m/%d %H:%M:%S"  # Changed to ISO military time
        # output_format = "%Y-%m-%dT%H:%M:%SZ"    # Google Earth format

        # pattern = r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$'
        pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'  # ISO military time

        if time_orig == '' and Time != '': # copy the original time
            time_orig = Time
        try:
            (Time, time_orig, timezone) = convert_timestamp(Time, time_orig, timezone)
            Time = Time.strftime(output_format)
            if Time is None:
                Time = ''              
            
        except ValueError as e:
            if Time != "":
                print(f"Error time2: {e} - {Time}")
                # Time = ''    # temp rem of this
            
# fullname cleanup
        if fullname is None:
            fullname = ''
        
        
        fullname = fullname.replace("_$!<Other>!$_", "")
        if "@" in fullname and fullname is not None:
            fullname = ''
        if fullname is not None and '+1 (' in fullname:
            query = fullname
            fullname = ''
        elif fullname is not None:
            if ',' in fullname:
                fullname2 = fullname.split(',')
                fullname = (f'fullname2[1] fullname2[0]')
                
            cleaned_fullname = re.sub(r'[^a-zA-Z] ', '', fullname).strip()
            if " " in cleaned_fullname:
                cleaned_fullname = cleaned_fullname.split(' ')
                firstname = cleaned_fullname[0]
                lastname = cleaned_fullname[1]
                lastname = re.sub(r'[^a-zA-Z]', '', lastname).strip()
            else:
                firstname = cleaned_fullname
                firstname = re.sub(r'[^a-zA-Z]', '', firstname).strip()
                
# write rows to data
        row_data["query"] = query
        row_data["ranking"] = ranking
        row_data["fullname"] = fullname
        row_data["url"] = url
        row_data["phone"] = phone
        row_data["email"] = email
        row_data["user"] = user        
        row_data["business"] = business        
        row_data["fulladdress"] = fulladdress  
        row_data["country"] = country        
        row_data["note"] = note
        row_data["info"] = info # interaction status
        row_data["DOB"] = DOB
        row_data["aka"] = aka        
        row_data["misc"] = misc  
        row_data["lastname"] = lastname          
        row_data["firstname"] = firstname          
        row_data["otheremails"] = otheremails     
        row_data["state"] = state 

        row_data["dnsdomain"] = dnsdomain     
        row_data["titleurl"] = titleurl 
        row_data["Time"] = Time
        row_data["Latitude"] = Latitude  
        row_data["Longitude"] = Longitude  
        row_data["Coordinate"] = Coordinate  
        row_data["Source file information"] = source_file     
        row_data["origin_file"] = origin_file     
        row_data["Tag"] = tag     
        row_data["Type"] = type_data     
        row_data["Icon"] = Icon     




    return data


def write_xlsx(data):
    '''
    The write_xlsx() function receives the processed data as a list of 
    dictionaries and writes it to a new Excel file using openpyxl. 
    It defines the column headers, sets column widths, and then iterates 
    through each row of data, writing it into the Excel worksheet.
    '''

    global workbook
    workbook = Workbook()
    global worksheet
    worksheet = workbook.active

    worksheet.title = 'Contacts'
    header_format = {'bold': True, 'border': True}
    worksheet.freeze_panes = 'B2'  # Freeze cells
    worksheet.selection = 'B2'

    headers = [
        "query", "ranking", "fullname", "url", "email", "user", "phone", "ip"
        , "business", "fulladdress", "city", "state", "zip", "country"
        , "note", "aka", "DOB", "SEX", "info", "misc", "lastname", "firstname"
        , "middlename", "friend", "otherurls", "otherphones", "otheremails"
        , "case", "sosfilenumber", "president", "sosagent", "managers", "dnsdomain"
        , "dstip", "srcip", "content", "referer", "osurl", "titleurl", "pagestatus"
        , "Time", "Latitude", "Longitude", "Coordinate", "Source", "Source file information"
        , "origin_file", "Tag", "Type", "Icon"
    ]

    # Write headers to the first row
    for col_index, header in enumerate(headers):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.value = header
        if col_index in [0, 1]: 
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # orange?
            cell.fill = fill
        elif col_index in [1, 2, 4, 5, 6, 8, 9, 15, 16, 18, 19]:  # yellow headers
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Use yellow color
            cell.fill = fill


    ## Excel column width
    worksheet.column_dimensions['A'].width = 6# #
    worksheet.column_dimensions['B'].width = 20# 
    worksheet.column_dimensions['C'].width = 25# 
    worksheet.column_dimensions['D'].width = 25# 
    worksheet.column_dimensions['E'].width = 25#
    worksheet.column_dimensions['F'].width = 15# 
    worksheet.column_dimensions['G'].width = 18 # phone
    worksheet.column_dimensions['H'].width = 3# 
    worksheet.column_dimensions['I'].width = 16# 
    worksheet.column_dimensions['J'].width = 20 # fulladdress
    worksheet.column_dimensions['K'].width = 6# 
    worksheet.column_dimensions['L'].width = 6# 
    worksheet.column_dimensions['M'].width = 4# 
    worksheet.column_dimensions['N'].width = 13# 
    worksheet.column_dimensions['O'].width = 5# 
    worksheet.column_dimensions['P'].width = 15# 
    worksheet.column_dimensions['Q'].width = 15# 
    worksheet.column_dimensions['R'].width = 5# 
    worksheet.column_dimensions['S'].width = 30# 
    worksheet.column_dimensions['T'].width = 20#   
    worksheet.column_dimensions['Y'].width = 10#     
    worksheet.column_dimensions['V'].width = 10#   
    worksheet.column_dimensions['W'].width = 10# 
    worksheet.column_dimensions['X'].width = 10#   
    worksheet.column_dimensions['Y'].width = 10#  
    worksheet.column_dimensions['Z'].width = 10# 
    worksheet.column_dimensions['AA'].width = 10# 
    worksheet.column_dimensions['AB'].width = 10# 
    worksheet.column_dimensions['AC'].width = 10# 
    worksheet.column_dimensions['AD'].width = 10# 
    worksheet.column_dimensions['AE'].width = 10#
    worksheet.column_dimensions['AF'].width = 10# 
    worksheet.column_dimensions['AG'].width = 10# 
    worksheet.column_dimensions['AH'].width = 10# 
    worksheet.column_dimensions['AI'].width = 10# 
    worksheet.column_dimensions['AJ'].width = 10# 
    worksheet.column_dimensions['AK'].width = 10#
    worksheet.column_dimensions['AL'].width = 10# 
    worksheet.column_dimensions['AM'].width = 10# 
    worksheet.column_dimensions['AN'].width = 10# 
    worksheet.column_dimensions['AO'].width = 28# 
    worksheet.column_dimensions['AP'].width = 12# 
    worksheet.column_dimensions['AQ'].width = 12# 
    worksheet.column_dimensions['AR'].width = 22# Coordinate
    worksheet.column_dimensions['AS'].width = 12# Coordinate
    worksheet.column_dimensions['AT'].width = 12# origin_file




    
    for row_index, row_data in enumerate(data):

        for col_index, col_name in enumerate(headers):
            cell_data = row_data.get(col_name)
            try:
                worksheet.cell(row=row_index+2, column=col_index+1).value = cell_data
            except Exception as e:
                print(f"{color_red}Error printing line: {str(e)}{color_reset}")
    
    workbook.save(outuput_xlsx)


def usage():
    '''
    working examples of syntax
    '''
    file = sys.argv[0].split('\\')[-1]
    print(f'\nDescription: {color_green}{description2}{color_reset}')
    print(f'{file} Version: {version} by {author}')
    print(f'\n    {color_yellow}export from Cellebrite categories')
    print(f'\nExample:')
   
    print(f'    {file} -b -O input_blank.xlsx') 
    print(f'    {file} -C -I Accounts.xlsx  ')  
    print(f'    {file} -C -I Calls.xlsx  ')      
    print(f'    {file} -C -I Chats.xlsx  ')       
    print(f'    {file} -C -I Contacts.xlsx  ')   
    print(f'    {file} -C -I SearchedItems.xlsx  ')   
    print(f'    {file} -C -I WebHistory.xlsx  ')  
 
                
if __name__ == '__main__':
    main()

# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>

"""

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
if instagram id found create a url
Icon, type_data and origin update
"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>

