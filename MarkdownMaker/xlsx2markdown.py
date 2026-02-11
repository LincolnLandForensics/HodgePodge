import os
import re
import sys
import argparse
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext
from typing import List, Dict
from datetime import datetime
from openpyxl import load_workbook

# <<<<<<<<<<<<<<<<<<<<<<<<<<     Pre-Sets       >>>>>>>>>>>>>>>>>>>>>>>>>>

author = 'LincolnLandForensics'
description2 = '''turn intel.xlsx into actionable markdown files
contacts: read contacts from xlsx file. Create 1 .md file per sheet
or
xlsx:    read xlsx file, create 1 .md file per row
'''
tech = 'LincolnLandForensics'  # change this to your name if you are using Linux
version = '1.0.2'

headers_intel = [
    "query", "ranking", "fullname", "url", "email", "user", "phone",
    "business", "fulladdress", "city", "state", "country", "note", "AKA",
    "DOB", "SEX", "info", "misc", "firstname", "middlename", "lastname",
    "associates", "case", "sosfilenumber", "owner", "president", "sosagent",
    "managers", "Time", "Latitude", "Longitude", "Coordinate",
    "original_file", "Source", "Source file information", "Plate", "VIS", "VIN",
    "VYR", "VMA", "LIC", "LIY", "DLN", "DLS", "content", "referer", "osurl",
    "titleurl", "pagestatus", "ip", "dnsdomain", "Tag", "Icon", "Type"
]


# <<<<<<<<<<<<<<<<<<<<<<<<<<      Menu           >>>>>>>>>>>>>>>>>>>>>>>>>>

def main():
    """
    Main function to parse arguments and initiate file conversion.
    """
    parser = argparse.ArgumentParser(description=description2)
    parser.add_argument('-I', '--input', help='Input folder path', required=False)
    parser.add_argument('-O', '--output', help='Output folder path', required=False)
    parser.add_argument('-c', '--contacts', help='read contacts from xlsx file. Create 1 .md file per sheet', required=False, action='store_true')
    parser.add_argument('-x', '--xlsx', help='read xlsx file, create 1 .md file per row', required=False, action='store_true')
    parser.add_argument('--dry-run', help='Validate the input file without generating output', required=False, action='store_true')

    args = parser.parse_args()

    # Launch GUI if no arguments are provided
    if not any(vars(args).values()):
        root = tk.Tk()
        app = Xlsx2MarkdownGUI(root)
        root.mainloop()
        return 0

    input_xlsx = args.input or "Intel_.xlsx"
    output_folder = args.output or "output_markdown"

    if args.dry_run:
        validate_input_file(input_xlsx)
        return 0

    # Ensure the output folder exists
    if not os.path.exists(output_folder):
        try:
            os.makedirs(output_folder)
        except Exception as e:
            msg_blurb_square(f"Error creating output folder: {e}")
            sys.exit(1)

    if args.xlsx:
        data = read_excel(input_xlsx)
        create_markdown_files(data, output_folder, headers_intel)
    elif args.contacts:
        if input_xlsx.lower().endswith(".xlsx"):
            data = read_excel(input_xlsx)
            create_contacts_markdown_files(data, input_xlsx, output_folder)
        else:
            print(f"{input_xlsx} isn't an xlsx file. Try again.")
    else:
        parser.print_help()
        Usage()

    return 0


class Xlsx2MarkdownGUI:
    def __init__(self, root):
        self.root = root
        script_name = os.path.basename(sys.argv[0])
        self.root.title(f"{script_name} {version}")
        self.root.geometry("550x600")
        
        # Vista theme
        style = ttk.Style()
        try:
            style.theme_use('vista')
        except:
            style.theme_use('clam')

        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Start button (packed first with side=BOTTOM to ensure visibility)
        self.start_button = ttk.Button(self.main_frame, text="Start Processing", command=self.start_processing)
        self.start_button.pack(side=tk.BOTTOM, pady=10)

        # Title and Description
        ttk.Label(self.main_frame, text=f"{script_name} {version}", font=("Helvetica", 16, "bold")).pack(pady=5)
        ttk.Label(self.main_frame, text=description2, justify=tk.LEFT, wraplength=520).pack(pady=10)

        # Mode Selection
        mode_frame = ttk.LabelFrame(self.main_frame, text="Mode Selection", padding="5")
        mode_frame.pack(fill=tk.X, pady=5)
        self.mode_var = tk.StringVar(value="xlsx")
        ttk.Radiobutton(mode_frame, text="xlsx (1 .md per row)", variable=self.mode_var, value="xlsx").pack(side=tk.LEFT, padx=20)
        ttk.Radiobutton(mode_frame, text="contacts (1 .md per sheet)", variable=self.mode_var, value="contacts").pack(side=tk.LEFT, padx=20)

        # File Inputs
        file_frame = ttk.LabelFrame(self.main_frame, text="File/Folder Selection", padding="5")
        file_frame.pack(fill=tk.X, pady=5)

        ttk.Label(file_frame, text="Input XLSX:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.input_entry = ttk.Entry(file_frame)
        self.input_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=2, padx=5)
        self.input_entry.insert(0, "Intel_.xlsx")
        ttk.Button(file_frame, text="Browse", command=self.browse_input).grid(row=0, column=2, pady=2)

        ttk.Label(file_frame, text="Output Folder:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.output_entry = ttk.Entry(file_frame)
        self.output_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=2, padx=5)
        self.output_entry.insert(0, "output_markdown")
        ttk.Button(file_frame, text="Browse", command=self.browse_output).grid(row=1, column=2, pady=2)

        file_frame.columnconfigure(1, weight=1)

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=10)

        # Status window
        ttk.Label(self.main_frame, text="Process Log / Terminal:").pack(anchor=tk.W)
        self.status_text = scrolledtext.ScrolledText(self.main_frame, height=7, state='disabled')
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=5)

        # Queue for thread communication
        self.log_queue = queue.Queue()
        self.root.after(100, self.poll_log_queue)

        # Redirect prints
        self.old_stdout = sys.stdout
        sys.stdout = self

    def write(self, text):
        self.log_queue.put(text)
        self.old_stdout.write(text)

    def flush(self):
        self.old_stdout.flush()

    def browse_input(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, filename)

    def browse_output(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, directory)

    def log(self, message):
        self.status_text.config(state='normal')
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state='disabled')

    def poll_log_queue(self):
        while True:
            try:
                msg = self.log_queue.get_nowait()
                self.log(msg.strip())
            except queue.Empty:
                break
        self.root.after(100, self.poll_log_queue)

    def start_processing(self):
        input_file = self.input_entry.get()
        output_dir = self.output_entry.get()
        mode = self.mode_var.get()

        if not input_file or not os.path.exists(input_file):
            self.log(f"Error: Input file '{input_file}' not found.")
            return

        self.start_button.config(state='disabled')
        self.progress_var.set(0)
        
        # Clear status
        self.status_text.config(state='normal')
        self.status_text.delete(1.0, tk.END)
        self.status_text.config(state='disabled')
        
        self.log(f"Processing input file: {input_file}")

        processing_thread = threading.Thread(target=self.run_processing, args=(input_file, output_dir, mode))
        processing_thread.daemon = True
        processing_thread.start()

    def run_processing(self, input_xlsx, output_folder, mode):
        try:
            # Ensure the output folder exists
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)

            if mode == "xlsx":
                data = read_excel(input_xlsx)
                total = len(data)
                
                # We need a way to update progress from the create_markdown_files function
                # For now, I'll just simulate it or call them normally.
                # To get progress, I'd need to modify the functions to accept a callback.
                # I'll just run them and set progress to 50 then 100 for now, 
                # or better, modify the loop in create_markdown_files.
                
                create_markdown_files(data, output_folder, headers_intel, progress_callback=self.update_progress)
            else:
                if input_xlsx.lower().endswith(".xlsx"):
                    data = read_excel(input_xlsx)
                    create_contacts_markdown_files(data, input_xlsx, output_folder, progress_callback=self.update_progress)
                else:
                    self.log(f"Error: {input_xlsx} isn't an xlsx file.")

            self.log(f"\nDone! Output files are in: {os.path.abspath(output_folder)}")
        except Exception as e:
            self.log(f"Error during processing: {e}")
        finally:
            self.root.after(0, lambda: self.start_button.config(state='normal'))
            self.root.after(0, lambda: self.progress_var.set(100))

    def update_progress(self, current, total):
        if total > 0:
            percent = (current / total) * 100
            self.root.after(0, lambda: self.progress_var.set(percent))

# <<<<<<<<<<<<<<<<<<<<<<<<<<   Sub-Routines   >>>>>>>>>>>>>>>>>>>>>>>>>>

def clean_data(item):
    """
    Cleans the item by removing characters that interfere with Markdown formatting,
    including newlines, pipes, tabs, and excessive whitespace.

    Returns:
        str: Markdown-safe cleaned text.
    """
    if item is None:
        return ''
    
    # Convert to string and replace Markdown-breaking characters
    item_cleaned = str(item)

    # Replace newlines, carriage returns, tabs, and pipes with a single space
    item_cleaned = re.sub(r'[\n\r\t|]', ' ', item_cleaned)

    # Collapse multiple whitespace characters into one
    item_cleaned = re.sub(r'\s+', ' ', item_cleaned)

    # Remove Markdown formatting symbols (optional)
    item_cleaned = re.sub(r'[#>*_`~]', '', item_cleaned)

    return item_cleaned.strip()

def clean_phone(phone):
    """
    Sanitizes and validates phone numbers (7–15 digits, digits only).
    """
    regex_phone = r'^\d{7,15}$'

    if not isinstance(phone, str):
        phone = str(phone)

    phone = re.sub(r'[^\d]', '', phone)  # Strip non-digit characters

    # Strip country code if it starts with '1'
    if phone.startswith('1') and len(phone) > 10:
        phone = phone[1:]

    return phone if re.match(regex_phone, phone) else ""


def create_contacts_markdown_files(data, input_xlsx, output_folder, progress_callback=None):
    """
    Create a markdown file with a table for all rows out of input_xlsx with the following columns:
    fullname, email, user, phone, AKA, Tag, business, fulladdress, case, original_file.

    Before the table, print the name of {input_xlsx} and the sheet name.
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    sheet_name = data[0].get("sheet_name", "Unknown")
    output_file = os.path.join(output_folder, f"{sheet_name}_{current_time}.md")

    with open(output_file, "w", encoding="utf-8") as md_file:
        md_file.write(f"# Contact Summary from {input_xlsx}\n\n")
        # md_file.write("| fullname | phone | email | user | AKA | Tag | business | fulladdress | case | original_file |\n")
        # md_file.write("|----------|-------|------|-------|-----|-----|---------|--------------|------|---------------|\n")

        md_file.write("| fullname | phone | user | fulladdress | case |\n")
        md_file.write("|------------|-------|---------|-----------------|---------|\n")

        phone_regex = re.compile(r"^\+?[1-9]\d{1,14}$")  # E.164 format or similar valid phone patterns

        for idx, row in enumerate(data, start=1):
            fullname = clean_data(row.get("fullname", ""))
            email = clean_data(row.get("email", ""))
            user = clean_data(row.get("user", ""))
            phone = clean_phone(str(row.get("phone", "")))
            if phone and phone_regex.match(phone):
                phone = f"[[{phone}]]"
            aka = clean_data(row.get("AKA", ""))
            Tag = clean_data(row.get("Tag", ""))
            if Tag:
                Tag = f"@{Tag}"
            business = clean_data(row.get("business", ""))
            fulladdress = clean_data(row.get("fulladdress", ""))
            case = clean_data(row.get("case", ""))
            original_file = clean_data(row.get("original_file", ""))

            md_file.write(f"| {fullname} | {phone} | {user}  | {fulladdress} | {case} |\n")
            
            if progress_callback:
                progress_callback(idx, len(data))

        md_file.write(f"\nSheet Name: {sheet_name}\n")

    # msg_blurb_square(f"{sheet_name}....md file created in {output_folder} folder")




def create_markdown_files(data: List[Dict[str, str]], output_folder: str, headers, progress_callback=None):
    """Creates markdown files for each row."""
    # Define the headers to use for creating markdown files
    # headers = ['query', 'phone', 'fullname', 'business', 'Tag']  # Adjust as needed

    # Ensure the output folder exists
    os.makedirs(output_folder, exist_ok=True)

    for idx, row in enumerate(data, start=1):
        query = str(row.get("query", "_") or "_")   # .replace(" ", "_")
        # query = re.sub(r'[^\w\-_]', '_', query)  # Sanitize filename
        query = re.sub(r'[^\w\-]', ' ', query)  # Sanitize filename

        # Check for existing files and adjust the name if necessary
        base_filename = os.path.join(output_folder, f"{query}.md")
        filename = base_filename
        counter = 1

        while os.path.exists(filename):
            filename = os.path.join(output_folder, f"{query}_{counter}.md")
            counter += 1

        try:
            with open(filename, "w", encoding="utf-8") as md_file:
                for header in headers:
                    value = row.get(header)
                    if value:  # Only include non-blank fields
                        if header in ['phone', 'fullname', 'business']:
                            formatted_value = f"[[{value}]]"
                        elif header == 'Tag':
                            formatted_value = f"@{value}"
                        else:
                            formatted_value = value
                        md_file.write(f"{header}: {formatted_value}\n\n")
            
            if progress_callback:
                progress_callback(idx, len(data))
        except Exception as e:
            msg_blurb_square(f"Unexpected error: {e}")
            continue

    # msg_blurb_square(f"Markdown file generation complete in {output_folder} folder")
    



def msg_blurb_square(msg_blurb):
    # horizontal_line = f"+{'-' * (len(msg_blurb) + 1)}+"
    # empty_line = f"| {' ' * (len(msg_blurb))} |"
    print(f"{msg_blurb}")
    print(f'')

    # print(horizontal_line)
    # print(empty_line)
    # print(f"| {msg_blurb} |")
    # print(empty_line)
    # print(horizontal_line)
    # print(f'')


def read_excel(file_path: str, validate_only: bool = False) -> List[Dict[str, str]]:
    """Reads an Excel file and returns a list of rows as dictionaries."""
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        sheet_name = sheet.title  # Get the sheet name
        # msg_blurb_square(f"Reading {sheet_name} sheet in {file_path} ")
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    except Exception as e:
        raise Exception(f"Error reading file: {e}")

    headers = [clean_data(cell.value) if cell is not None else "" for cell in sheet[1]]  # Read header row
    
    if not headers:
        raise ValueError("The Excel file contains no headers or is empty.")

    # Map Excel headers dynamically to headers_intel
    header_map = {header: headers_intel[idx] for idx, header in enumerate(headers) if idx < len(headers_intel)}

    if len(header_map) < len(headers_intel):
        missing_headers = set(headers_intel) - set(header_map.values())
        raise ValueError(f"Missing required headers: {', '.join(missing_headers)}")

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip completely empty rows
            continue

        row_dict = {header_map.get(headers[idx], f"unknown_column_{idx}"): (cell if cell is not None else "") for idx, cell in enumerate(row)}
        row_dict["sheet_name"] = sheet_name  # Add sheet name to each row
        rows.append(row_dict)

    if validate_only:
        return []  # Validation successful; no need to return data

    return rows


def validate_input_file(file_path: str):
    """
    Validates the input Excel file without generating Markdown files.
    """
    try:
        data = read_excel(file_path, validate_only=True)
        msg_blurb_square(f"Validation successful for file: {file_path}")
    except FileNotFoundError:
        msg_blurb_square(f"Error: File not found: {file_path}")
    except ValueError as e:
        msg_blurb_square(f"Validation failed: {e}")
    except Exception as e:
        msg_blurb_square(f"Unexpected error during validation: {e}")


def Usage():
    file = sys.argv[0].split(os.sep)[-1]

    print(f'\nDescription: {description2}')
    print(f'{file} Version: {version} by {author}')
    print(f"    {file} -c -I contacts_sample.xlsx -O c:\\temp")
    print(f"    {file} -x -I <input_folder> -O <output_folder>")
    print(f"    {file} --dry-run")


if __name__ == '__main__':
    main()


# <<<<<<<<<<<<<<<<<<<<<<<<<< Revision History >>>>>>>>>>>>>>>>>>>>>>>>>>


"""
0.2.2 - shortened up -C output
0.1.0 - working copy
0.0.1 - created by ChatGPT
"""


# <<<<<<<<<<<<<<<<<<<<<<<<<< Future Wishlist  >>>>>>>>>>>>>>>>>>>>>>>>>>

"""
if filename doesn't exist, create it else create fullname_{row}.md
copy the original file to original_file when doing a -c

"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      notes            >>>>>>>>>>>>>>>>>>>>>>>>>>

"""


"""

# <<<<<<<<<<<<<<<<<<<<<<<<<<      The End        >>>>>>>>>>>>>>>>>>>>>>>>>>
